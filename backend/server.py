from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Header, Depends, File, UploadFile, Form, Body
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, EmailStr, Field
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone, timedelta
import hashlib
import secrets
import base64
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
from fpdf import FPDF
from io import BytesIO
from PIL import Image
import asyncio
import unicodedata
import re

ROOT_DIR = Path(__file__).parent

# Fonction pour nettoyer les noms de fichiers (supprimer les caractères spéciaux)
def sanitize_filename(name: str) -> str:
    """Remove special characters from filename to avoid encoding issues"""
    # Normalize unicode characters
    name = unicodedata.normalize('NFKD', name)
    # Remove non-ASCII characters
    name = name.encode('ascii', 'ignore').decode('ascii')
    # Replace spaces and special chars with underscore
    name = re.sub(r'[^\w\s-]', '', name)
    name = re.sub(r'[\s]+', '_', name)
    return name or 'file'

# Background task for auto-regenerating translations
async def regenerate_translations_background(restaurant_id: str):
    """Background task to regenerate translations after menu changes"""
    try:
        # Wait a bit to batch multiple rapid changes
        await asyncio.sleep(2)
        
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        logging.info(f"Auto-regenerating translations for restaurant {restaurant_id}")
        
        # Get all menu sections for this restaurant
        sections = await menu_restaurant_sections_collection.find(
            {"restaurant_id": restaurant_id},
            {"_id": 0}
        ).to_list(500)
        
        # Get all items (excluding deleted)
        items = await menu_restaurant_items_collection.find(
            {
                "restaurant_id": restaurant_id,
                "status": {"$nin": ["a_supprimer", "deleted"]}
            },
            {"_id": 0}
        ).to_list(2000)
        
        if not sections and not items:
            logging.info(f"No items to translate for restaurant {restaurant_id}")
            return
        
        # Collect all texts to translate
        texts_to_translate = []
        text_mapping = {}
        
        # Add section names
        for sec in sections:
            name = sec.get("name", "")
            if name:
                idx = len(texts_to_translate)
                texts_to_translate.append(name)
                text_mapping[idx] = ("section", sec.get("section_id"), "name")
        
        # Add item names and descriptions
        for item in items:
            name = item.get("name", "")
            if name:
                idx = len(texts_to_translate)
                texts_to_translate.append(name)
                text_mapping[idx] = ("item", item.get("item_id"), "name")
            
            for desc_idx, desc in enumerate(item.get("descriptions", [])):
                if desc:
                    idx = len(texts_to_translate)
                    texts_to_translate.append(desc)
                    text_mapping[idx] = ("item", item.get("item_id"), f"description_{desc_idx}")
        
        if not texts_to_translate:
            return
        
        # Target languages
        target_languages = {
            'en': 'English',
            'es': 'Spanish', 
            'de': 'German',
            'it': 'Italian',
            'zh': 'Chinese',
            'ru': 'Russian',
            'pt': 'Portuguese'
        }
        
        all_translations = {}
        
        # Translate in batches
        BATCH_SIZE = 30
        for lang_code, lang_name in target_languages.items():
            all_translations[lang_code] = {}
            
            for i in range(0, len(texts_to_translate), BATCH_SIZE):
                batch = texts_to_translate[i:i + BATCH_SIZE]
                
                prompt = f"""Translate these French restaurant menu items to {lang_name}.
Return ONLY a JSON array with the translations in the same order.
Keep proper nouns, brand names unchanged.
Items: {batch}"""
                
                try:
                    chat = LlmChat(
                        api_key=os.environ.get('EMERGENT_LLM_KEY'),
                        session_id=f"translation_{restaurant_id}_{lang_code}",
                        system_message="You are a translator."
                    ).with_model("openai", "gpt-4.1-mini")
                    
                    user_message = UserMessage(text=prompt)
                    response = await chat.send_message(user_message)
                    
                    import json
                    import re
                    json_match = re.search(r'\[.*\]', response, re.DOTALL)
                    if json_match:
                        translations = json.loads(json_match.group())
                        
                        for j, translation in enumerate(translations):
                            original_idx = i + j
                            if original_idx in text_mapping:
                                type_, id_, field = text_mapping[original_idx]
                                key = f"{type_}_{id_}_{field}"
                                all_translations[lang_code][key] = translation
                except Exception as e:
                    logging.error(f"Translation batch error for {lang_code}: {e}")
                    continue
        
        # Save to database
        await translations_collection.update_one(
            {"restaurant_id": restaurant_id},
            {"$set": {
                "restaurant_id": restaurant_id,
                **all_translations,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
        
        logging.info(f"Auto-translation complete for restaurant {restaurant_id}: {len(texts_to_translate)} texts")
        
    except Exception as e:
        logging.error(f"Background translation error: {e}")

# Track pending translation tasks to avoid duplicates
_pending_translation_tasks: Dict[str, asyncio.Task] = {}

def trigger_translation_regeneration(restaurant_id: str):
    """Trigger background translation regeneration for a restaurant"""
    global _pending_translation_tasks
    
    # Cancel any pending task for this restaurant
    if restaurant_id in _pending_translation_tasks:
        task = _pending_translation_tasks[restaurant_id]
        if not task.done():
            task.cancel()
    
    # Create new background task
    loop = asyncio.get_event_loop()
    task = loop.create_task(regenerate_translations_background(restaurant_id))
    _pending_translation_tasks[restaurant_id] = task
    logging.info(f"Scheduled translation regeneration for restaurant {restaurant_id}")
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection with error handling
mongo_url = os.environ.get('MONGO_URL')
if not mongo_url:
    logging.error("MONGO_URL environment variable is not set!")
    raise ValueError("MONGO_URL environment variable is required")

try:
    client = AsyncIOMotorClient(mongo_url, serverSelectionTimeoutMS=5000)
    db_name = os.environ.get('DB_NAME', 'mise_en_place')
    db = client[db_name]
    print(f"[STARTUP] MongoDB connected to database: {db_name}")
except Exception as e:
    logging.error(f"Failed to initialize MongoDB client: {e}")
    raise

# Collections
restaurants_collection = db.mep_restaurants
users_collection = db.mep_users
categories_collection = db.mep_categories
task_templates_collection = db.mep_task_templates
daily_tasks_collection = db.mep_daily_tasks
history_collection = db.mep_task_history
sessions_collection = db.mep_sessions
notifications_collection = db.mep_notifications
password_reset_collection = db.mep_password_resets

# Menu Groupe Collections
menu_sections_collection = db.mep_menu_sections
menu_items_collection = db.mep_menu_items
group_reservations_collection = db.mep_group_reservations

# Permanent Tasks Collections
permanent_categories_collection = db.mep_permanent_categories
permanent_tasks_collection = db.mep_permanent_tasks
permanent_task_completions_collection = db.mep_permanent_task_completions

# Sub-tasks Collection
subtasks_collection = db.mep_subtasks
subtask_completions_collection = db.mep_subtask_completions

# Permanent Sub-tasks Collection
permanent_subtasks_collection = db.mep_permanent_subtasks
permanent_subtask_completions_collection = db.mep_permanent_subtask_completions

# Push Notifications Collection
push_tokens_collection = db.mep_push_tokens

# Invoice Counter Collection
invoice_counters_collection = db.mep_invoice_counters

# Invoices and Quotes Collection (Factures et Devis)
invoices_collection = db.mep_invoices  # Factures et Devis

# Préparation de Commande Collections (Order Preparation)
suppliers_collection = db.mep_suppliers
supplier_products_collection = db.mep_supplier_products
supplier_orders_collection = db.mep_supplier_orders

# Fiche Technique Collections (Recipe/Technical Sheet Management)
fiche_technique_sections_collection = db.mep_fiche_sections  # Sections (Cocktails, Entrée, etc.)
fiche_technique_products_collection = db.mep_fiche_products  # Products/Recipes (Mojito, etc.)
fiche_technique_ingredients_collection = db.mep_fiche_ingredients  # Ingredients with cost calculation

# Menu Restaurant Collections (Carte Food & Carte Boisson)
menu_restaurant_collection = db.mep_menu_restaurant  # Main menu document per restaurant
menu_restaurant_sections_collection = db.mep_menu_restaurant_sections  # Sections (Apéro, Entrée, Vins, etc.)
menu_restaurant_items_collection = db.mep_menu_restaurant_items  # Menu items with prices
menu_restaurant_notes_collection = db.mep_menu_restaurant_notes  # Menu notes (Happy Hour, etc.)

# Events Collections (Module Événements)
events_collection = db.mep_events  # Événements
event_providers_collection = db.mep_event_providers  # Prestataires
event_tasks_collection = db.mep_event_tasks  # Tâches d'événement
event_menu_sections_collection = db.mep_event_menu_sections  # Sections menu événement
event_menu_items_collection = db.mep_event_menu_items  # Items menu événement
event_price_packages_collection = db.mep_event_price_packages  # Packages de prix
event_drink_options_collection = db.mep_event_drink_options  # Options de boissons

# Privatisation Spaces Collection (Espaces de privatisation: Bibliothèque, Pergola, etc.)
privatisation_spaces_collection = db.mep_privatisation_spaces

# A L'ARDOISE Collection (Shareable Menu Board)
ardoise_collection = db.mep_ardoise  # Menu "A l'Ardoise" avec lien partageable
ardoise_sales_collection = db.mep_ardoise_sales  # Historique des ventes de l'ardoise
ardoise_planned_collection = db.mep_ardoise_planned  # Ardoises planifiées par date

# Translations Collection (for instant language switching)
translations_collection = db.mep_translations  # Store pre-computed translations

# Group Options Collection (Options configurables pour les réservations groupe)
group_options_collection = db.mep_group_options  # Options comme DJ, Gâteau anniversaire, etc.

# Events uploads directory
EVENTS_UPLOADS_DIR = "/app/backend/uploads/events"
os.makedirs(EVENTS_UPLOADS_DIR, exist_ok=True)

# SendGrid Configuration
SENDGRID_API_KEY = os.environ.get('SENDGRID_API_KEY')
FRONTEND_URL = os.environ.get('FRONTEND_URL')
if not FRONTEND_URL:
    print("WARNING: FRONTEND_URL not set - password reset emails will not work properly")
    FRONTEND_URL = ""

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# ==================== REQUEST/RESPONSE MODELS ====================

class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class RegisterAdminRequest(BaseModel):
    email: EmailStr
    password: str
    name: str
    restaurant_name: str

class RegisterHoldingRequest(BaseModel):
    email: EmailStr
    password: str
    name: str
    holding_name: Optional[str] = None  # Nom du groupe/holding (optionnel)

class LinkRestaurantRequest(BaseModel):
    restaurant_email: EmailStr
    restaurant_password: str

# ==================== SYSTÈME DE PERMISSIONS DÉTAILLÉ ====================
# Structure: module -> sous-module -> actions (ajouter, modifier, supprimer, etc.)

class ActionPermissions(BaseModel):
    """Permissions d'actions: Ajouter, Modifier, Supprimer"""
    ajouter: bool = False
    modifier: bool = False
    supprimer: bool = False

class ActionPermissionsWithView(BaseModel):
    """Permissions d'actions avec vue"""
    voir: bool = False
    ajouter: bool = False
    modifier: bool = False
    supprimer: bool = False

class TachesPermissions(BaseModel):
    """Permissions pour les Tâches-Catégories"""
    actif: bool = False
    categories: List[str] = []  # froid, chaud, salle, bar, fermeture
    editer: bool = False
    ajouter: bool = False
    supprimer: bool = False
    modeles_ajouter: bool = False
    modeles_modifier: bool = False
    modeles_supprimer: bool = False

class PrepCommandePermissions(BaseModel):
    """Permissions pour Préparation de Commande"""
    actif: bool = False
    section: str = "none"  # "none", "cuisine", "bar", "tous"
    fournisseur: ActionPermissions = ActionPermissions()
    produits: ActionPermissions = ActionPermissions()
    consignes: ActionPermissions = ActionPermissions()

class MenuRestaurantPermissions(BaseModel):
    """Permissions pour Menu Restaurant"""
    actif: bool = False
    lien_partage: bool = False
    section: ActionPermissions = ActionPermissions()
    produits: ActionPermissions = ActionPermissions()
    export_pdf: bool = False
    export_csv: bool = False
    import_csv: bool = False
    import_pdf: bool = False
    note: bool = False

class FicheTechniquePermissions(BaseModel):
    """Permissions pour Fiche Technique"""
    actif: bool = False
    section_access: str = "none"  # "none", "bar", "cuisine", "tous"
    export_pdf_excel: bool = False
    export_type: str = "tous"  # "tous", "avec_prix", "sans_prix"
    analyse_marges: bool = False
    section: ActionPermissions = ActionPermissions()
    produits: ActionPermissions = ActionPermissions()
    photo: ActionPermissions = ActionPermissions()

class MenuGroupePermissions(BaseModel):
    """Permissions pour Menu Groupe"""
    actif: bool = False
    bouton_lien: bool = False
    section: ActionPermissions = ActionPermissions()
    plats: ActionPermissions = ActionPermissions()
    reservation_creer: bool = False
    reservation_modifier: bool = False
    reservation_supprimer: bool = False
    statut: bool = False
    proposition: bool = False
    facture: bool = False

class EvenementPermissions(BaseModel):
    """Permissions pour Événements"""
    actif: bool = False
    ajouter: bool = False
    modifier: bool = False
    supprimer: bool = False
    archiver: bool = False
    prestataires: ActionPermissions = ActionPermissions()

class FacturationPermissions(BaseModel):
    """Permissions pour Facturation/BonBon"""
    actif: bool = False
    facture_deposer: bool = False
    facture_telecharger: bool = False
    devis_deposer: bool = False
    devis_telecharger: bool = False

class ArdoiseEditionPermissions(BaseModel):
    """Permissions pour l'édition de l'ardoise (onglet Édition)"""
    acces: bool = False
    mode: str = "lecture"  # "lecture" ou "modifier"

class ArdoiseVentesPermissions(BaseModel):
    """Permissions pour les ventes de l'ardoise (onglet Ventes)"""
    acces: bool = False
    mode: str = "lecture"  # "lecture" ou "modifier"

class ArdoiseRapportsPermissions(BaseModel):
    """Permissions pour les rapports de l'ardoise (onglet Rapports)"""
    acces: bool = False
    mode: str = "lecture"  # "lecture" ou "modifier"
    export_pdf: bool = False
    export_excel: bool = False

class ArdoisePermissions(BaseModel):
    """Permissions pour Ardoise - structure granulaire"""
    actif: bool = False
    edition: ArdoiseEditionPermissions = ArdoiseEditionPermissions()
    ventes: ArdoiseVentesPermissions = ArdoiseVentesPermissions()
    rapports: ArdoiseRapportsPermissions = ArdoiseRapportsPermissions()
    # Anciens champs conservés pour compatibilité
    menu: bool = False
    pdf: bool = False
    section: ActionPermissions = ActionPermissions()
    produits: ActionPermissions = ActionPermissions()
    packages_prix: ActionPermissions = ActionPermissions()

class DetailedUserPermissions(BaseModel):
    """Permissions détaillées complètes pour un utilisateur"""
    # Accès globaux
    parametres: bool = False
    equipe: bool = False
    
    # Modules avec sous-permissions
    taches: TachesPermissions = TachesPermissions()
    preparation_commande: PrepCommandePermissions = PrepCommandePermissions()
    menu_restaurant: MenuRestaurantPermissions = MenuRestaurantPermissions()
    fiche_technique: FicheTechniquePermissions = FicheTechniquePermissions()
    menu_groupe: MenuGroupePermissions = MenuGroupePermissions()
    evenement: EvenementPermissions = EvenementPermissions()
    facturation: FacturationPermissions = FacturationPermissions()
    ardoise: ArdoisePermissions = ArdoisePermissions()

# Ancien modèle pour compatibilité
class UserPermissions(BaseModel):
    menu_groupe: bool = False
    taches: bool = True
    preparation_commande: bool = False
    fiche_technique: bool = False
    fiche_technique_access: str = "none"  # "none", "bar", "cuisine", "both"
    categories: List[str] = []

class UpdateUserPermissionsRequest(BaseModel):
    menu_groupe: bool = False
    taches: bool = True
    preparation_commande: bool = False
    fiche_technique: bool = False
    fiche_technique_access: str = "none"
    categories: List[str] = []
    # Nouvelles permissions détaillées (optionnel)
    detailed_permissions: Optional[DetailedUserPermissions] = None

class CreateUserRequest(BaseModel):
    email: EmailStr
    password: str
    name: str
    phone: Optional[str] = None
    role: str = "staff"  # "admin", "associe", "staff"
    assigned_categories: List[str] = []
    permissions: Optional[UserPermissions] = None
    detailed_permissions: Optional[DetailedUserPermissions] = None

class UpdateUserRequest(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    assigned_categories: Optional[List[str]] = None
    notification_prefs: Optional[Dict[str, bool]] = None
    role: Optional[str] = None
    detailed_permissions: Optional[DetailedUserPermissions] = None

class ResetPasswordRequest(BaseModel):
    new_password: str

class UpdateRestaurantRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    logo_base64: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    # Nouvelles informations de contact
    address_street: Optional[str] = None
    address_postal_code: Optional[str] = None
    address_city: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    # Informations légales
    siret: Optional[str] = None
    rcs: Optional[str] = None
    # Réseaux sociaux
    facebook_url: Optional[str] = None
    instagram_url: Optional[str] = None
    # WiFi
    wifi_name: Optional[str] = None
    wifi_password: Optional[str] = None
    # Happy Hour
    happy_hour_enabled: Optional[bool] = None
    happy_hour_start: Optional[str] = None  # Format "HH:MM"
    happy_hour_end: Optional[str] = None    # Format "HH:MM"

class CreateCategoryRequest(BaseModel):
    name: str

class UpdateCategoryRequest(BaseModel):
    name: Optional[str] = None
    order: Optional[int] = None

class RecurrenceRule(BaseModel):
    """Règle de récurrence pour les tâches permanentes"""
    type: str = "daily"  # daily, weekly, monthly
    days_of_week: Optional[List[int]] = None  # 0=Lundi, 1=Mardi, ..., 6=Dimanche (pour weekly)
    days_of_month: Optional[List[int]] = None  # 1-31 (pour monthly)

class CreateTaskTemplateRequest(BaseModel):
    category_id: str
    title: str
    description: Optional[str] = None
    task_type: str = "manual"  # "manual" ou "permanent"
    recurrence_rule: Optional[RecurrenceRule] = None  # Pour les tâches permanentes

class UpdateTaskTemplateRequest(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None
    task_type: Optional[str] = None
    recurrence_rule: Optional[RecurrenceRule] = None

# ==================== SUBTASK MODELS ====================

class CreateSubtaskRequest(BaseModel):
    parent_template_id: str  # ID du template parent
    name: str
    quantity: Optional[int] = None  # Quantité optionnelle

class UpdateSubtaskRequest(BaseModel):
    name: Optional[str] = None
    quantity: Optional[int] = None
    is_active: Optional[bool] = None

# ==================== PERMANENT SUBTASK MODELS ====================

class CreatePermanentSubtaskRequest(BaseModel):
    parent_permanent_task_id: str  # ID de la tâche permanente parent
    name: str
    quantity: Optional[int] = None  # Quantité optionnelle

class UpdatePermanentSubtaskRequest(BaseModel):
    name: Optional[str] = None
    quantity: Optional[int] = None
    is_active: Optional[bool] = None

# ==================== PERMANENT TASKS MODELS ====================

class CreatePermanentCategoryRequest(BaseModel):
    name: str  # Ex: Ouverture, Fermeture, Livraison, etc.

class UpdatePermanentCategoryRequest(BaseModel):
    name: Optional[str] = None
    order: Optional[int] = None

class CreatePermanentTaskRequest(BaseModel):
    permanent_category_id: str
    title: str
    description: Optional[str] = None
    recurrence_rule: Optional[RecurrenceRule] = None  # Si None, c'est tous les jours (daily)

class UpdatePermanentTaskRequest(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None
    recurrence_rule: Optional[RecurrenceRule] = None

class TaskSelection(BaseModel):
    template_id: str
    assigned_user_id: Optional[str] = None  # Pour tagger une personne spécifique

class SelectTasksForDayRequest(BaseModel):
    date: str
    selections: List[TaskSelection]  # Liste avec possibilité de tagger

class CreatePunctualTaskRequest(BaseModel):
    category_id: str
    title: str
    description: Optional[str] = None
    date: str
    assigned_user_id: Optional[str] = None  # Pour tagger une personne spécifique

class SendDailyTasksRequest(BaseModel):
    date: str

class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordWithTokenRequest(BaseModel):
    token: str
    new_password: str

# ==================== SUPPLIER / ORDER PREPARATION MODELS ====================

class DeliverySchedule(BaseModel):
    """Planning de livraison d'un fournisseur"""
    delivery_days: List[int] = []  # 0=Lundi, 1=Mardi, ..., 6=Dimanche
    order_deadline_days: List[int] = []  # Jours de commande possibles (0=Lundi, etc.)
    order_deadline_time: str = "19:00"  # Heure limite de commande
    delivery_time: Optional[str] = None  # Heure de livraison approximative

class CreateSupplierRequest(BaseModel):
    name: str  # Nom du fournisseur (Metro, Transgourmet, etc.)
    phone: Optional[str] = None  # Numéro du commercial
    delivery_schedule: Optional[DeliverySchedule] = None
    notes: Optional[str] = None
    supplier_category: Optional[str] = None  # "bar", "cuisine", or "both" - for prep filtering

class UpdateSupplierRequest(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    delivery_schedule: Optional[DeliverySchedule] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None
    supplier_category: Optional[str] = None  # "bar", "cuisine", or "both"

class CreateSupplierProductRequest(BaseModel):
    supplier_id: str
    name: str  # Nom du produit
    product_type: str = "product"  # "product" ou "consigne" (pour les bouteilles vides, etc.)
    order: Optional[int] = None
    price_ht: Optional[float] = None  # Prix HT (surtout pour consignes et réclamations)

class UpdateSupplierProductRequest(BaseModel):
    name: Optional[str] = None
    product_type: Optional[str] = None  # "product" ou "consigne"
    order: Optional[int] = None
    is_active: Optional[bool] = None
    price_ht: Optional[float] = None  # Prix HT (surtout pour consignes et réclamations)

class OrderItem(BaseModel):
    product_id: Optional[str] = None  # Optionnel pour les réclamations
    product_name: str
    quantity: int
    product_type: str = "product"  # "product", "consigne", ou "reclamation"
    price_ht: Optional[float] = None  # Prix HT pour les réclamations
    is_refunded: bool = False  # Pour consignes et réclamations: produit remboursé

class CreateSupplierOrderRequest(BaseModel):
    supplier_id: str
    items: List[OrderItem]
    notes: Optional[str] = None
    order_type: Optional[str] = None  # "reclamation" pour les réclamations
    total_ht: Optional[float] = None  # Total HT pour les réclamations

class UpdateOrderStatusRequest(BaseModel):
    status: str  # Pour commandes: "to_order", "ordered", "delivered" 
                 # Pour consignes: "sent", "partial_refund", "full_refund"
                 # Pour réclamations: "to_request", "partial_refund", "full_refund"
    refunded_items: Optional[List[str]] = None  # Liste des product_name remboursés
    manual_full_refund: bool = False  # Marquer manuellement comme tout remboursé

# ==================== FICHE TECHNIQUE MODELS ====================

class CreateFicheSectionRequest(BaseModel):
    """Créer une section dans Bar ou Cuisine (ex: Cocktails, Entrée)"""
    category: str  # "bar" ou "cuisine"
    name: str  # Ex: Cocktails, Mocktails, Entrée, Viande
    order: Optional[int] = None
    is_preparations: bool = False  # True pour les sections de préparations

class UpdateFicheSectionRequest(BaseModel):
    name: Optional[str] = None
    order: Optional[int] = None
    is_active: Optional[bool] = None
    is_preparations: Optional[bool] = None

class MarginThresholds(BaseModel):
    """Seuils de marge pour une section (en pourcentage)"""
    low: float = 20.0  # En dessous = rouge (faible)
    high: float = 50.0  # Au-dessus = vert (bon), entre low et high = jaune (moyen)

class UpdateMarginThresholdsRequest(BaseModel):
    """Mettre à jour les seuils de marge d'une section"""
    low: float
    high: float

class FicheIngredient(BaseModel):
    """Ingrédient standard pour une fiche technique"""
    ingredient_type: str = "standard"  # "standard" ou "preparation"
    name: str  # Nom de l'ingrédient (ex: Sel, Rhum)
    quantity_used: float  # Quantité utilisée dans la recette
    unit_used: str  # cl, ml, l, g, kg
    quantity_purchased: float  # Quantité achetée
    unit_purchased: str  # cl, ml, l, g, kg
    purchase_price: float  # Prix d'achat
    # Champs pour les préparations
    preparation_id: Optional[str] = None  # ID de la préparation si ingredient_type == "preparation"
    preparation_name: Optional[str] = None  # Nom de la préparation

# === Nouveau modèle pour Boisson Multi-Formats ===
class SellingFormat(BaseModel):
    """Format de vente pour une boisson (ex: verre 14cl, pichet 50cl)"""
    format_id: Optional[str] = None
    name: str  # Ex: "Verre 14cl", "Pichet 50cl", "Bouteille"
    size: float  # Taille en unité (ex: 14, 50, 75)
    unit: str  # cl, ml, l
    selling_price: float  # Prix de vente
    cost: Optional[float] = None  # Coût calculé automatiquement

class PurchaseInfo(BaseModel):
    """Informations d'achat pour boisson multi-formats"""
    quantity: float  # Quantité (ex: 75 pour une bouteille 75cl)
    unit: str  # cl, ml, l
    price: float  # Prix d'achat

class CreateFicheProductRequest(BaseModel):
    """Créer un produit/recette dans une section"""
    section_id: str
    name: str  # Nom du produit (ex: Mojito, Salade César, Château Margaux)
    photo_base64: Optional[str] = None  # Photo optionnelle
    notes: Optional[str] = None  # Notes de préparation/instructions (optionnel)
    
    # Type de produit : "standard", "boisson_multi", "preparation"
    product_type: str = "standard"
    
    # === Champs pour produit STANDARD ===
    recipe_unit: Optional[str] = None  # cl, ml, l, g, kg
    multiplier: Optional[float] = None  # Multiplicateur pour prix de vente
    selling_price_override: Optional[float] = None  # Prix de vente manuel (écrase le calcul)
    ingredients: Optional[List[FicheIngredient]] = []  # Liste des ingrédients
    
    # === Champs pour PREPARATION (sous-recette) ===
    yield_quantity: Optional[float] = None  # Rendement (ex: 50 crêpes)
    yield_unit: Optional[str] = None  # Unité du rendement (ex: "crêpes", "portions")
    
    # === Champs pour BOISSON MULTI-FORMATS ===
    purchase_info: Optional[PurchaseInfo] = None  # Info d'achat (bouteille, fût)
    selling_formats: Optional[List[SellingFormat]] = []  # Formats de vente

class UpdateFicheProductRequest(BaseModel):
    name: Optional[str] = None
    photo_base64: Optional[str] = None
    product_type: Optional[str] = None
    notes: Optional[str] = None  # Notes de préparation/instructions
    
    # Champs standard
    recipe_unit: Optional[str] = None
    multiplier: Optional[float] = None
    selling_price_override: Optional[float] = None
    ingredients: Optional[List[FicheIngredient]] = None
    is_active: Optional[bool] = None
    archived: Optional[bool] = None  # Pour archiver/désarchiver un produit
    
    # Champs préparation
    yield_quantity: Optional[float] = None
    yield_unit: Optional[str] = None
    
    # Champs boisson multi
    purchase_info: Optional[PurchaseInfo] = None
    selling_formats: Optional[List[SellingFormat]] = None

# === IMPORT PDF MODELS ===

class ImportProductItem(BaseModel):
    """Un produit à importer depuis PDF"""
    product_name: str
    section_name: str
    category: str  # "bar" ou "cuisine"
    ingredients: List[Dict] = []  # Liste des ingrédients avec name, quantity, unit
    selling_price: Optional[float] = None
    instructions_or_notes: Optional[str] = None
    product_type: str = "standard"  # standard, preparation, boisson_multi

class ImportFicheProductsRequest(BaseModel):
    """Importer des produits en masse depuis PDF"""
    products: List[ImportProductItem]
    create_sections: bool = True  # Créer automatiquement les sections manquantes

# ==================== MENU RESTAURANT MODELS ====================

class MenuRestaurantFormat(BaseModel):
    """Format de prix pour un item (ex: 14cl, 50cl, 75cl)"""
    format_id: Optional[str] = None
    name: str  # Ex: "14cl", "50cl", "Bouteille 75cl"
    price: float  # Prix pour ce format
    happy_hour_price: Optional[float] = None  # Prix Happy Hour optionnel

class MenuRestaurantSuggestion(BaseModel):
    """Suggestion/Succession pour un plat (ex: Verre de vin)"""
    suggestion_id: Optional[str] = None
    name: str  # Ex: "Verre de vin blanc"
    price: float  # Prix du supplément

class MenuRestaurantSupplement(BaseModel):
    """Supplément pour boissons (ex: Coca, Limonade pour cocktails)"""
    supplement_id: Optional[str] = None
    name: str  # Ex: "Coca", "Limonade"
    price: float  # Prix du supplément

class MenuRestaurantOption(BaseModel):
    """Option payante pour un plat (ex: +2€ Jambon Serrano, +3€ Saumon fumé)"""
    option_id: Optional[str] = None
    name: str  # Ex: "Jambon Serrano", "Saumon fumé"
    price: float  # Prix de l'option (ex: 2.0, 3.0)

class CreateMenuRestaurantSectionRequest(BaseModel):
    """Créer une section dans Carte Food ou Carte Boisson"""
    menu_type: str  # "food" ou "boisson"
    name: str  # Ex: "Apéro", "Entrée", "Vins Rouges"
    parent_section_id: Optional[str] = None  # Pour les sous-sections
    order: Optional[int] = None
    color: Optional[str] = None  # Couleur personnalisée pour la section
    has_happy_hour: bool = False  # Si cette section a des prix Happy Hour

class UpdateMenuRestaurantSectionRequest(BaseModel):
    name: Optional[str] = None
    order: Optional[int] = None
    color: Optional[str] = None
    has_happy_hour: Optional[bool] = None
    is_active: Optional[bool] = None

# Liste des allergènes standards
ALLERGENS_LIST = [
    "gluten", "crustaces", "oeufs", "poissons", "arachides", 
    "soja", "lactose", "fruits_a_coque", "celeri", "moutarde",
    "sesame", "sulfites", "lupin", "mollusques"
]

# Liste des tags standards
TAGS_LIST = ["vegetarien", "vegan", "epice"]

# Statuts possibles pour un article (couleurs)
ITEM_STATUS_LIST = ["normal", "a_ajouter", "a_modifier", "a_supprimer"]  # normal, vert, violet, rouge

class CreateMenuRestaurantItemRequest(BaseModel):
    """Créer un item dans une section du menu"""
    section_id: str
    fiche_technique_product_id: Optional[str] = None  # Lien optionnel vers Fiche Technique
    name: str
    descriptions: List[str] = []  # Une ou plusieurs descriptions
    price: Optional[float] = None  # Prix simple TTC (pour Food principalement)
    happy_hour_price: Optional[float] = None  # Prix Happy Hour TTC
    tva_rate: Optional[float] = 10.0  # Taux de TVA: 10 ou 20 (défaut: 10%)
    formats: Optional[List[MenuRestaurantFormat]] = []  # Multi-formats (pour Boisson)
    suggestions: Optional[List[MenuRestaurantSuggestion]] = []  # Suggestions/Succession
    supplements: Optional[List[MenuRestaurantSupplement]] = []  # Suppléments
    options: Optional[List[MenuRestaurantOption]] = []  # Options payantes (+2€ Jambon, etc.)
    order: Optional[int] = None
    # Champs pour allergènes, tags et coloration Excel
    allergens: Optional[List[str]] = []  # Liste des allergènes (ex: ["gluten", "oeufs"])
    tags: Optional[List[str]] = []  # Liste des tags (ex: ["vegetarien", "epice"])
    excel_status: Optional[str] = "added"  # "added" (vert), "deleted" (rouge), "modified" (violet), "normal"
    modified_fields: Optional[List[str]] = []  # ["name", "price", "description"] - pour coloration Excel

class UpdateMenuRestaurantItemRequest(BaseModel):
    name: Optional[str] = None
    descriptions: Optional[List[str]] = None
    price: Optional[float] = None
    happy_hour_price: Optional[float] = None  # Prix Happy Hour TTC
    tva_rate: Optional[float] = None  # Taux de TVA: 10 ou 20
    formats: Optional[List[MenuRestaurantFormat]] = None
    suggestions: Optional[List[MenuRestaurantSuggestion]] = None
    supplements: Optional[List[MenuRestaurantSupplement]] = None
    options: Optional[List[MenuRestaurantOption]] = None  # Options payantes
    order: Optional[int] = None
    is_active: Optional[bool] = None
    # Champs pour allergènes, tags et coloration Excel
    allergens: Optional[List[str]] = None
    tags: Optional[List[str]] = None
    excel_status: Optional[str] = None  # "added" (vert), "deleted" (rouge), "modified" (violet), "normal"
    modified_fields: Optional[List[str]] = None  # ["name", "price", "description"] - pour coloration Excel

class CreateMenuRestaurantNoteRequest(BaseModel):
    """Créer une note de menu (ex: Happy Hour -20%)"""
    menu_type: str  # "food" ou "boisson"
    content: str  # Texte de la note
    order: Optional[int] = None

class UpdateMenuRestaurantNoteRequest(BaseModel):
    content: Optional[str] = None
    order: Optional[int] = None
    is_active: Optional[bool] = None

# ==================== MULTI-RESTAURANT MODELS ====================

class CreateRestaurantRequest(BaseModel):
    name: str
    description: Optional[str] = None

class DuplicateRestaurantRequest(BaseModel):
    source_restaurant_id: str
    new_restaurant_name: str

class SwitchRestaurantRequest(BaseModel):
    restaurant_id: str

# ==================== REORDERING MODELS ====================

class SetOrderRequest(BaseModel):
    """Requête pour définir l'ordre d'un élément"""
    order: int

# ==================== EVENT MODELS ====================

class CreateEventRequest(BaseModel):
    """Créer un nouvel événement"""
    title: str
    date: str  # Format: YYYY-MM-DD
    description: Optional[str] = None
    notes: Optional[str] = None
    assigned_team: List[str] = []  # Liste des user_id assignés à l'événement

class UpdateEventRequest(BaseModel):
    """Modifier un événement"""
    title: Optional[str] = None
    date: Optional[str] = None
    description: Optional[str] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None
    archived: Optional[bool] = None  # Pour archiver/désarchiver manuellement
    assigned_team: Optional[List[str]] = None  # Liste des user_id assignés

class DuplicateEventRequest(BaseModel):
    """Dupliquer un événement"""
    new_title: Optional[str] = None
    new_date: Optional[str] = None

class CreateProviderRequest(BaseModel):
    """Créer un prestataire pour un événement"""
    name: str
    contact_name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None

class UpdateProviderRequest(BaseModel):
    """Modifier un prestataire"""
    name: Optional[str] = None
    contact_name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None
    quote_status: Optional[str] = None
    invoice_status: Optional[str] = None
    payment_method: Optional[str] = None
    is_active: Optional[bool] = None

class ValidateQuoteRequest(BaseModel):
    """Valider un devis"""
    validated: bool = True

class UpdateInvoiceStatusRequest(BaseModel):
    """Mettre à jour le statut de la facture"""
    status: str
    payment_method: Optional[str] = None

class CreateEventTaskRequest(BaseModel):
    """Créer une tâche pour un événement"""
    title: str
    description: Optional[str] = None
    due_date: str
    assigned_user_id: Optional[str] = None

class UpdateEventTaskRequest(BaseModel):
    """Modifier une tâche d'événement"""
    title: Optional[str] = None
    description: Optional[str] = None
    due_date: Optional[str] = None
    assigned_user_id: Optional[str] = None
    status: Optional[str] = None
    is_active: Optional[bool] = None

class UpdateTaskStatusRequest(BaseModel):
    """Modifier uniquement le statut d'une tâche"""
    status: str

class CreateEventMenuSectionRequest(BaseModel):
    """Créer une section de menu pour l'événement"""
    name: str
    order: Optional[int] = None

class UpdateEventMenuSectionRequest(BaseModel):
    """Modifier une section de menu"""
    name: Optional[str] = None
    order: Optional[int] = None

class CreateEventMenuItemRequest(BaseModel):
    """Créer un item dans une section de menu"""
    section_id: str
    name: str
    description: Optional[str] = None
    order: Optional[int] = None

class UpdateEventMenuItemRequest(BaseModel):
    """Modifier un item de menu"""
    name: Optional[str] = None
    description: Optional[str] = None
    order: Optional[int] = None

class CreatePricePackageRequest(BaseModel):
    """Créer un package de prix"""
    name: str
    section_ids: List[str]
    price: float

class UpdatePricePackageRequest(BaseModel):
    """Modifier un package de prix"""
    name: Optional[str] = None
    section_ids: Optional[List[str]] = None
    price: Optional[float] = None

class CreateDrinkOptionRequest(BaseModel):
    """Créer une option de boisson"""
    name: str
    price: float

class UpdateDrinkOptionRequest(BaseModel):
    """Modifier une option de boisson"""
    name: Optional[str] = None
    price: Optional[float] = None
    is_selected: Optional[bool] = None

# ==================== HELPER FUNCTIONS ====================

def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    pwd_hash = hashlib.sha256((password + salt).encode()).hexdigest()
    return f"{salt}${pwd_hash}"

def verify_password(password: str, password_hash: str) -> bool:
    try:
        salt, pwd_hash = password_hash.split('$')
        calculated = hashlib.sha256((password + salt).encode()).hexdigest()
        print(f"[VERIFY] Salt: {salt[:20]}..., Expected: {pwd_hash[:20]}..., Calculated: {calculated[:20]}...")
        return calculated == pwd_hash
    except Exception as e:
        print(f"[VERIFY] Error: {e}")
        return False

async def get_current_user(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    token = authorization.replace("Bearer ", "")
    
    session_doc = await sessions_collection.find_one({"session_token": token}, {"_id": 0})
    if not session_doc:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    
    user_doc = await users_collection.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    return user_doc

async def get_current_user_optional(authorization: Optional[str] = Header(None)):
    """Version optionnelle qui retourne None au lieu de lever une exception"""
    if not authorization or not authorization.startswith("Bearer "):
        return None
    
    token = authorization.replace("Bearer ", "")
    return await get_user_from_token(token)

async def get_user_from_token(token: str):
    """Récupérer l'utilisateur à partir d'un token de session"""
    if not token:
        return None
    
    session_doc = await sessions_collection.find_one({"session_token": token}, {"_id": 0})
    if not session_doc:
        return None
    
    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        return None
    
    user_doc = await users_collection.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    return user_doc

async def get_next_invoice_number(restaurant_id: str) -> str:
    """
    Génère le prochain numéro de facture au format YYMMXXX
    YY = 2 derniers chiffres de l'année
    MM = mois (01-12)
    XXX = compteur séquentiel (001, 002, 003...)
    """
    now = datetime.now(timezone.utc)
    year_str = now.strftime("%y")  # 2 derniers chiffres de l'année (26 pour 2026)
    month_str = now.strftime("%m")  # Mois (01-12)
    period_key = f"{year_str}{month_str}"  # Ex: "2602" pour février 2026
    
    # Rechercher ou créer le compteur pour ce restaurant et cette période
    counter_doc = await invoice_counters_collection.find_one_and_update(
        {"restaurant_id": restaurant_id, "period": period_key},
        {"$inc": {"count": 1}},
        upsert=True,
        return_document=True
    )
    
    # Si c'est un nouveau document, count sera 1
    count = counter_doc.get("count", 1)
    
    # Formater le numéro: YYMMXXX (ex: 2602001)
    invoice_number = f"{period_key}{count:03d}"
    
    return invoice_number

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register-admin")
async def register_admin(request: RegisterAdminRequest):
    existing_user = await users_collection.find_one({"email": request.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    restaurant_id = f"rest_{uuid.uuid4().hex[:12]}"
    restaurant = {
        "restaurant_id": restaurant_id,
        "name": request.restaurant_name,
        "description": "Gestion des tâches cuisine",
        "logo_base64": None,
        "primary_color": "#26252D",
        "secondary_color": "#EAE6CA",
        "created_by": request.email,
        "created_at": datetime.now(timezone.utc)
    }
    await restaurants_collection.insert_one(restaurant)
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    user = {
        "user_id": user_id,
        "email": request.email,
        "password_hash": hash_password(request.password),
        "role": "admin",
        "restaurant_id": restaurant_id,
        "name": request.name,
        "phone": None,
        "assigned_categories": [],
        "notification_prefs": {"push": True, "email": False, "sms": False},
        "created_at": datetime.now(timezone.utc)
    }
    await users_collection.insert_one(user)
    
    session_token = secrets.token_urlsafe(32)
    session = {
        "session_token": session_token,
        "user_id": user_id,
        "expires_at": datetime.now(timezone.utc) + timedelta(days=30),
        "created_at": datetime.now(timezone.utc)
    }
    await sessions_collection.insert_one(session)
    
    restaurant_doc = await restaurants_collection.find_one({"restaurant_id": restaurant_id}, {"_id": 0})
    
    return {
        "session_token": session_token,
        "user": {
            "user_id": user_id,
            "email": request.email,
            "name": request.name,
            "role": "admin",
            "restaurant_id": restaurant_id,
            "assigned_categories": []
        },
        "restaurant": restaurant_doc
    }

@api_router.post("/auth/login")
async def login(request: LoginRequest):
    print(f"[LOGIN] Attempting login for: {request.email}")
    user_doc = await users_collection.find_one({"email": request.email}, {"_id": 0})
    if not user_doc:
        print(f"[LOGIN] User not found: {request.email}")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    print(f"[LOGIN] User found, checking password...")
    if not verify_password(request.password, user_doc["password_hash"]):
        print(f"[LOGIN] Password verification failed")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    print(f"[LOGIN] Password OK, creating session...")
    # Gérer le cas où il n'y a pas de restaurant (compte Holding)
    restaurant_doc = None
    if user_doc.get("restaurant_id"):
        restaurant_doc = await restaurants_collection.find_one(
            {"restaurant_id": user_doc["restaurant_id"]}, 
            {"_id": 0}
        )
    
    session_token = secrets.token_urlsafe(32)
    session = {
        "session_token": session_token,
        "user_id": user_doc["user_id"],
        "expires_at": datetime.now(timezone.utc) + timedelta(days=30),
        "created_at": datetime.now(timezone.utc)
    }
    await sessions_collection.insert_one(session)
    
    return {
        "session_token": session_token,
        "user": {
            "user_id": user_doc["user_id"],
            "email": user_doc["email"],
            "name": user_doc["name"],
            "role": user_doc["role"],
            "restaurant_id": user_doc.get("restaurant_id"),
            "restaurant_ids": user_doc.get("restaurant_ids", []),
            "holding_name": user_doc.get("holding_name"),
            "assigned_categories": user_doc.get("assigned_categories", []),
            "notification_prefs": user_doc.get("notification_prefs", {})
        },
        "restaurant": restaurant_doc
    }

@api_router.post("/auth/register-holding")
async def register_holding(request: RegisterHoldingRequest):
    """Créer un compte Holding (sans restaurant)"""
    existing_user = await users_collection.find_one({"email": request.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    user = {
        "user_id": user_id,
        "email": request.email,
        "password_hash": hash_password(request.password),
        "role": "holding",  # Nouveau rôle pour les comptes holding
        "restaurant_id": None,  # Pas de restaurant au départ
        "restaurant_ids": [],  # Liste vide des restaurants liés
        "name": request.name,
        "holding_name": request.holding_name,
        "phone": None,
        "assigned_categories": [],
        "notification_prefs": {"push": True, "email": False, "sms": False},
        "created_at": datetime.now(timezone.utc)
    }
    await users_collection.insert_one(user)
    
    session_token = secrets.token_urlsafe(32)
    session = {
        "session_token": session_token,
        "user_id": user_id,
        "expires_at": datetime.now(timezone.utc) + timedelta(days=30),
        "created_at": datetime.now(timezone.utc)
    }
    await sessions_collection.insert_one(session)
    
    return {
        "session_token": session_token,
        "user": {
            "user_id": user_id,
            "email": request.email,
            "name": request.name,
            "role": "holding",
            "restaurant_id": None,
            "restaurant_ids": [],
            "holding_name": request.holding_name,
            "assigned_categories": []
        },
        "restaurant": None
    }

@api_router.post("/auth/link-restaurant")
async def link_restaurant(
    request: LinkRestaurantRequest,
    current_user: dict = Depends(get_current_user)
):
    """Lier un restaurant existant au compte Holding en utilisant ses identifiants"""
    # Vérifier que l'utilisateur est un holding ou admin
    if current_user["role"] not in ["holding", "admin"]:
        raise HTTPException(status_code=403, detail="Seuls les comptes Holding peuvent lier des restaurants")
    
    # Trouver le compte admin du restaurant avec les identifiants fournis
    restaurant_admin = await users_collection.find_one({"email": request.restaurant_email})
    if not restaurant_admin:
        raise HTTPException(status_code=401, detail="Identifiants du restaurant invalides")
    
    if not verify_password(request.restaurant_password, restaurant_admin["password_hash"]):
        raise HTTPException(status_code=401, detail="Identifiants du restaurant invalides")
    
    # Vérifier que ce compte a un restaurant
    if not restaurant_admin.get("restaurant_id"):
        raise HTTPException(status_code=400, detail="Ce compte n'a pas de restaurant associé")
    
    restaurant_id = restaurant_admin["restaurant_id"]
    
    # Vérifier si le restaurant n'est pas déjà lié
    current_restaurant_ids = current_user.get("restaurant_ids", [])
    if current_user.get("restaurant_id"):
        if current_user["restaurant_id"] not in current_restaurant_ids:
            current_restaurant_ids.append(current_user["restaurant_id"])
    
    if restaurant_id in current_restaurant_ids:
        raise HTTPException(status_code=400, detail="Ce restaurant est déjà lié à votre compte")
    
    # Ajouter le restaurant à la liste
    current_restaurant_ids.append(restaurant_id)
    
    # Si c'est le premier restaurant, le définir comme actif
    new_active_restaurant_id = current_user.get("restaurant_id") or restaurant_id
    
    # Mettre à jour l'utilisateur
    await users_collection.update_one(
        {"user_id": current_user["user_id"]},
        {"$set": {
            "restaurant_ids": current_restaurant_ids,
            "restaurant_id": new_active_restaurant_id,
            "role": "admin" if current_user["role"] == "holding" else current_user["role"]  # Upgrade holding to admin
        }}
    )
    
    # Récupérer les infos du restaurant
    restaurant_doc = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    
    # Récupérer l'utilisateur mis à jour
    updated_user = await users_collection.find_one(
        {"user_id": current_user["user_id"]},
        {"_id": 0, "password_hash": 0}
    )
    
    return {
        "message": f"Restaurant '{restaurant_doc['name']}' lié avec succès !",
        "restaurant": restaurant_doc,
        "user": updated_user,
        "restaurant_ids": current_restaurant_ids
    }

@api_router.post("/auth/logout")
async def logout(authorization: Optional[str] = Header(None)):
    if authorization and authorization.startswith("Bearer "):
        token = authorization.replace("Bearer ", "")
        await sessions_collection.delete_one({"session_token": token})
    return {"message": "Logged out successfully"}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    restaurant_doc = None
    if current_user.get("restaurant_id"):
        restaurant_doc = await restaurants_collection.find_one(
            {"restaurant_id": current_user["restaurant_id"]},
            {"_id": 0}
        )
    
    return {
        "user": {
            "user_id": current_user["user_id"],
            "email": current_user["email"],
            "name": current_user["name"],
            "role": current_user["role"],
            "restaurant_id": current_user.get("restaurant_id"),
            "restaurant_ids": current_user.get("restaurant_ids", []),
            "holding_name": current_user.get("holding_name"),
            "assigned_categories": current_user.get("assigned_categories", []),
            "notification_prefs": current_user.get("notification_prefs", {})
        },
        "restaurant": restaurant_doc
    }

# ==================== PASSWORD RESET ENDPOINTS ====================

async def send_password_reset_email(email: str, reset_token: str, user_name: str, restaurant_name: str):
    """Send password reset email via SendGrid"""
    if not SENDGRID_API_KEY:
        logger.warning("SendGrid API key not configured, skipping email")
        return False
    
    reset_link = f"{FRONTEND_URL}?reset_token={reset_token}"
    
    html_content = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background-color: #26252D; padding: 20px; text-align: center;">
            <h1 style="color: #EAE6CA; margin: 0;">Mise en Place Pro</h1>
        </div>
        <div style="padding: 30px; background-color: #f9f9f9;">
            <h2 style="color: #26252D;">Réinitialisation de mot de passe</h2>
            <p>Bonjour <strong>{user_name}</strong>,</p>
            <p>Vous avez demandé la réinitialisation de votre mot de passe pour votre compte <strong>{restaurant_name}</strong>.</p>
            <p>Cliquez sur le bouton ci-dessous pour créer un nouveau mot de passe :</p>
            <div style="text-align: center; margin: 30px 0;">
                <a href="{reset_link}" style="background-color: #26252D; color: #EAE6CA; padding: 15px 30px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                    Réinitialiser mon mot de passe
                </a>
            </div>
            <p style="color: #666; font-size: 14px;">Ce lien expirera dans 1 heure.</p>
            <p style="color: #666; font-size: 14px;">Si vous n'avez pas demandé cette réinitialisation, ignorez cet email.</p>
        </div>
        <div style="text-align: center; padding: 20px; color: #888; font-size: 12px;">
            <p>© 2026 Mise en Place Pro - Gestion des tâches cuisine</p>
        </div>
    </body>
    </html>
    """
    
    # Use a verified sender email from SendGrid
    from_email = os.environ.get('SENDGRID_FROM_EMAIL', 'kajarupan10@gmail.com')
    
    message = Mail(
        from_email=from_email,
        to_emails=email,
        subject=f"Réinitialisation de mot de passe - {restaurant_name}",
        html_content=html_content
    )
    
    try:
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        response = sg.send(message)
        logger.info(f"Password reset email sent to {email}, status: {response.status_code}")
        return response.status_code == 202
    except Exception as e:
        logger.error(f"Failed to send password reset email: {e}")
        return False

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    """Request a password reset email"""
    # Find user by email
    user_doc = await users_collection.find_one({"email": request.email}, {"_id": 0})
    
    # Always return success to prevent email enumeration attacks
    if not user_doc:
        return {"message": "Si un compte existe avec cette adresse email, un lien de réinitialisation a été envoyé."}
    
    # Get restaurant info (if user has a restaurant)
    restaurant_name = "NeoChef"
    if user_doc.get("restaurant_id"):
        restaurant_doc = await restaurants_collection.find_one(
            {"restaurant_id": user_doc["restaurant_id"]},
            {"_id": 0, "name": 1}
        )
        if restaurant_doc:
            restaurant_name = restaurant_doc.get("name", "NeoChef")
    
    # Generate reset token
    reset_token = secrets.token_urlsafe(32)
    
    # Store reset request (expire in 1 hour)
    await password_reset_collection.delete_many({"user_id": user_doc["user_id"]})  # Remove old requests
    await password_reset_collection.insert_one({
        "reset_id": f"reset_{uuid.uuid4().hex[:12]}",
        "user_id": user_doc["user_id"],
        "email": request.email,
        "token": reset_token,
        "expires_at": datetime.now(timezone.utc) + timedelta(hours=1),
        "created_at": datetime.now(timezone.utc),
        "used": False
    })
    
    # Send email
    await send_password_reset_email(
        request.email,
        reset_token,
        user_doc.get("name", "Utilisateur"),
        restaurant_name
    )
    
    return {"message": "Si un compte existe avec cette adresse email, un lien de réinitialisation a été envoyé."}

@api_router.post("/auth/reset-password-with-token")
async def reset_password_with_token(request: ResetPasswordWithTokenRequest):
    """Reset password using a token from email"""
    # Find the reset request
    reset_doc = await password_reset_collection.find_one({
        "token": request.token,
        "used": False
    })
    
    if not reset_doc:
        raise HTTPException(status_code=400, detail="Lien invalide ou expiré")
    
    # Check expiration
    expires_at = reset_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Ce lien a expiré. Veuillez demander un nouveau lien.")
    
    # Validate password
    if len(request.new_password) < 4:
        raise HTTPException(status_code=400, detail="Le mot de passe doit contenir au moins 4 caractères")
    
    # Update password
    new_hash = hash_password(request.new_password)
    await users_collection.update_one(
        {"user_id": reset_doc["user_id"]},
        {"$set": {"password_hash": new_hash}}
    )
    
    # Mark token as used
    await password_reset_collection.update_one(
        {"token": request.token},
        {"$set": {"used": True, "used_at": datetime.now(timezone.utc)}}
    )
    
    # Invalidate all user sessions
    await sessions_collection.delete_many({"user_id": reset_doc["user_id"]})
    
    return {"message": "Mot de passe réinitialisé avec succès. Vous pouvez maintenant vous connecter."}

@api_router.get("/auth/verify-reset-token")
async def verify_reset_token(token: str):
    """Verify if a reset token is valid"""
    reset_doc = await password_reset_collection.find_one({
        "token": token,
        "used": False
    })
    
    if not reset_doc:
        raise HTTPException(status_code=400, detail="Lien invalide ou expiré")
    
    # Check expiration
    expires_at = reset_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Ce lien a expiré")
    
    # Get user email for display
    user_doc = await users_collection.find_one(
        {"user_id": reset_doc["user_id"]},
        {"_id": 0, "email": 1, "name": 1}
    )
    
    return {
        "valid": True,
        "email": user_doc.get("email") if user_doc else None,
        "name": user_doc.get("name") if user_doc else None
    }

@api_router.put("/restaurants/{restaurant_id}")
async def update_restaurant(
    restaurant_id: str,
    update_request: UpdateRestaurantRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if current_user["restaurant_id"] != restaurant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await restaurants_collection.update_one(
            {"restaurant_id": restaurant_id},
            {"$set": update_data}
        )
    
    restaurant_doc = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    
    return restaurant_doc

# ==================== PUBLIC CLIENT MENU ENDPOINTS (No auth required) ====================

# Duplicate routes to handle /api/api/ prefix bug from cached frontend
@api_router.get("/api/restaurants/{restaurant_id}/public")
async def get_restaurant_public_fix(restaurant_id: str):
    """Redirect handler for double /api prefix bug"""
    return await get_restaurant_public(restaurant_id)

@api_router.get("/api/menu-restaurant/public/{restaurant_id}")
async def get_public_menu_fix(restaurant_id: str, menu_type: Optional[str] = None):
    """Redirect handler for double /api prefix bug"""
    return await get_public_menu(restaurant_id, menu_type)

@api_router.get("/api/ardoise/by-restaurant/{restaurant_id}")
async def get_ardoise_by_restaurant_fix(restaurant_id: str):
    """Redirect handler for double /api prefix bug - redirects to ardoise/by-restaurant"""
    # Inline the logic since the function is defined later
    ardoise = await ardoise_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    if not ardoise:
        return {
            "entree": [],
            "plat": [],
            "dessert": [],
            "formule_prices": {
                "plat_du_jour": 15.90,
                "entree_plat": 18.90,
                "plat_dessert": 18.90,
                "entree_plat_dessert": 23.90
            }
        }
    return {
        "entree": ardoise.get("entree", []),
        "plat": ardoise.get("plat", []),
        "dessert": ardoise.get("dessert", []),
        "formule_prices": ardoise.get("formule_prices", {
            "plat_du_jour": 15.90,
            "entree_plat": 18.90,
            "plat_dessert": 18.90,
            "entree_plat_dessert": 23.90
        })
    }

@api_router.get("/api/public/translations/{restaurant_id}")
async def get_public_translations_fix(restaurant_id: str):
    """Redirect handler for double /api prefix bug - redirects to public/translations"""
    # Inline the logic since the function is defined later
    cached = await translations_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    if cached:
        return cached.get("translations", {})
    return {}

@api_router.get("/restaurants/{restaurant_id}/public")
async def get_restaurant_public(restaurant_id: str):
    """Get public restaurant info (for client QR code view)"""
    restaurant_doc = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {
            "_id": 0,
            "restaurant_id": 1,
            "name": 1,
            "logo_base64": 1,
            "primary_color": 1,
            "secondary_color": 1,
            "address_street": 1,
            "address_postal_code": 1,
            "address_city": 1,
            "email": 1,
            "phone": 1,
            "facebook_url": 1,
            "instagram_url": 1,
            "wifi_name": 1,
            "wifi_password": 1,
            "happy_hour_enabled": 1,
            "happy_hour_start": 1,
            "happy_hour_end": 1
        }
    )
    
    if not restaurant_doc:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    return restaurant_doc

@api_router.get("/menu-restaurant/public/{restaurant_id}")
async def get_public_menu(restaurant_id: str, menu_type: Optional[str] = None):
    """Get public menu sections and items (for client QR code view)"""
    # Verify restaurant exists
    restaurant_doc = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "restaurant_id": 1, "name": 1, "logo_base64": 1, "primary_color": 1, "secondary_color": 1}
    )
    if not restaurant_doc:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    # Build query for sections
    sections_query = {"restaurant_id": restaurant_id, "is_active": {"$ne": False}}
    if menu_type and menu_type in ["food", "boisson"]:
        sections_query["menu_type"] = menu_type
    
    # Get sections (excluding deleted)
    sections = await menu_restaurant_sections_collection.find(
        sections_query,
        {"_id": 0}
    ).sort("order", 1).to_list(200)
    
    # Get section IDs for filtering items
    section_ids = [s.get("section_id") for s in sections]
    
    # Get items (excluding deleted and status 'a_supprimer')
    items_query = {
        "restaurant_id": restaurant_id,
        "status": {"$nin": ["a_supprimer", "deleted"]},
        "is_active": {"$ne": False}
    }
    if section_ids:
        items_query["section_id"] = {"$in": section_ids}
    
    items = await menu_restaurant_items_collection.find(
        items_query,
        {
            "_id": 0,
            "item_id": 1,
            "name": 1,
            "descriptions": 1,
            "price": 1,
            "price_happy_hour": 1,
            "formats": 1,
            "allergens": 1,
            "tags": 1,
            "section_id": 1,
            "order": 1,
            "tva_rate": 1
        }
    ).sort("order", 1).to_list(500)
    
    return {
        "restaurant": restaurant_doc,
        "sections": sections,
        "items": items
    }

# ==================== TRANSLATION ENDPOINT ====================

class TranslationRequest(BaseModel):
    texts: List[str]
    target_language: str
    source_language: str = "fr"

SUPPORTED_LANGUAGES = {
    "fr": "French",
    "en": "English", 
    "es": "Spanish",
    "de": "German",
    "it": "Italian",
    "zh": "Chinese",
    "ru": "Russian",
    "pt": "Portuguese"
}

@api_router.get("/public/translations/{restaurant_id}")
async def get_cached_translations(restaurant_id: str):
    """Get all cached translations for a restaurant"""
    cached = await translations_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    if cached:
        return cached.get("translations", {})
    return {}

@api_router.post("/public/translations/{restaurant_id}/generate")
async def generate_restaurant_translations(restaurant_id: str):
    """Generate and cache all translations for a restaurant's menu"""
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    
    # Get all menu sections for this restaurant (use same collections as public API)
    sections = await menu_restaurant_sections_collection.find(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    ).to_list(500)
    
    # Get all items (excluding deleted)
    items = await menu_restaurant_items_collection.find(
        {
            "restaurant_id": restaurant_id,
            "status": {"$nin": ["a_supprimer", "deleted"]}
        },
        {"_id": 0}
    ).to_list(2000)
    
    logging.info(f"Generating translations for {len(sections)} sections and {len(items)} items")
    
    # Collect all texts to translate
    texts_to_translate = []
    text_mapping = {}  # Map index to (type, id, field)
    
    # Add section names
    for sec in sections:
        name = sec.get("name", "")
        if name:
            idx = len(texts_to_translate)
            texts_to_translate.append(name)
            text_mapping[idx] = ("section", sec.get("section_id"), "name")
    
    # Add item names and descriptions
    for item in items:
        # Item name
        name = item.get("name", "")
        if name:
            idx = len(texts_to_translate)
            texts_to_translate.append(name)
            text_mapping[idx] = ("item", item.get("item_id"), "name")
        
        # Item descriptions
        for desc_idx, desc in enumerate(item.get("descriptions", [])):
            if desc:
                idx = len(texts_to_translate)
                texts_to_translate.append(desc)
                text_mapping[idx] = ("item", item.get("item_id"), f"description_{desc_idx}")
    
    # Add Ardoise items (daily specials) for translation
    ardoise = await ardoise_collection.find_one({"restaurant_id": restaurant_id}, {"_id": 0})
    if ardoise:
        for category in ["entree", "plat", "dessert"]:
            ardoise_items = ardoise.get(category, [])
            for ardoise_idx, ardoise_item in enumerate(ardoise_items):
                ardoise_name = ardoise_item.get("name", "")
                if ardoise_name and ardoise_name.strip():
                    idx = len(texts_to_translate)
                    texts_to_translate.append(ardoise_name)
                    text_mapping[idx] = ("ardoise", f"{category}_{ardoise_idx}", "name")
                ardoise_desc = ardoise_item.get("description", "")
                if ardoise_desc and ardoise_desc.strip():
                    idx = len(texts_to_translate)
                    texts_to_translate.append(ardoise_desc)
                    text_mapping[idx] = ("ardoise", f"{category}_{ardoise_idx}", "description")
        logging.info(f"Added ardoise items for translation")
    
    if not texts_to_translate:
        return {"message": "No items to translate"}
    
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="Translation service not configured")
    
    all_translations = {}
    
    # Translate to each supported language (except French which is source)
    for lang_code, lang_name in SUPPORTED_LANGUAGES.items():
        if lang_code == "fr":
            continue
            
        try:
            chat = LlmChat(
                api_key=api_key,
                session_id=f"translate_{restaurant_id}_{lang_code}",
                system_message=f"""You are a professional translator specializing in restaurant menus.
Translate the following items from French to {lang_name}.
Keep food item names authentic when appropriate (e.g., 'Tiramisu' stays 'Tiramisu').
Maintain formatting and punctuation.
Return ONLY the translations, one per line, in the same order as the input.
Do not add numbers, bullets, or any extra formatting."""
            ).with_model("openai", "gpt-4.1-mini")
            
            # Batch translate (max 50 at a time)
            lang_translations = {}
            batch_size = 50
            
            for i in range(0, len(texts_to_translate), batch_size):
                batch = texts_to_translate[i:i + batch_size]
                input_text = "\n".join(batch)
                
                user_message = UserMessage(text=input_text)
                response = await chat.send_message(user_message)
                
                translations = response.strip().split("\n")
                
                # Map translations back to items
                for j, trans in enumerate(translations):
                    idx = i + j
                    if idx in text_mapping:
                        item_type, item_id, field = text_mapping[idx]
                        key = f"{item_type}_{item_id}_{field}"
                        lang_translations[key] = trans
            
            all_translations[lang_code] = lang_translations
            
        except Exception as e:
            logging.error(f"Translation error for {lang_code}: {str(e)}")
            continue
    
    # Save to database
    await translations_collection.update_one(
        {"restaurant_id": restaurant_id},
        {"$set": {
            "restaurant_id": restaurant_id,
            "translations": all_translations,
            "updated_at": datetime.now(timezone.utc)
        }},
        upsert=True
    )
    
    return {"message": "Translations generated", "languages": list(all_translations.keys())}

@api_router.post("/translate")
async def translate_menu_items(request: TranslationRequest):
    """Translate menu item names and descriptions using OpenAI GPT"""
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    
    if request.target_language not in SUPPORTED_LANGUAGES:
        raise HTTPException(status_code=400, detail=f"Unsupported language: {request.target_language}")
    
    if not request.texts or len(request.texts) == 0:
        return {"translations": []}
    
    # If source and target are the same, return original texts
    if request.target_language == request.source_language:
        return {"translations": request.texts}
    
    try:
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Translation service not configured")
        
        # Create LLM chat instance
        chat = LlmChat(
            api_key=api_key,
            session_id=f"translate_{uuid.uuid4().hex[:8]}",
            system_message=f"""You are a professional translator specializing in restaurant menus.
Translate the following items from {SUPPORTED_LANGUAGES[request.source_language]} to {SUPPORTED_LANGUAGES[request.target_language]}.
Keep food item names authentic when appropriate (e.g., 'Tiramisu' stays 'Tiramisu').
Maintain formatting and punctuation.
Return ONLY the translations, one per line, in the same order as the input.
Do not add numbers, bullets, or any extra formatting."""
        ).with_model("openai", "gpt-4.1-mini")
        
        # Batch translate (join texts with newlines)
        input_text = "\n".join(request.texts)
        
        user_message = UserMessage(text=input_text)
        response = await chat.send_message(user_message)
        
        # Split response into lines
        translations = response.strip().split("\n")
        
        # Ensure we have same number of translations as inputs
        if len(translations) != len(request.texts):
            # If mismatch, try to handle gracefully
            if len(translations) < len(request.texts):
                # Pad with original texts
                translations.extend(request.texts[len(translations):])
            else:
                # Truncate
                translations = translations[:len(request.texts)]
        
        return {"translations": translations}
        
    except Exception as e:
        logging.error(f"Translation error: {str(e)}")
        # Return original texts on error
        return {"translations": request.texts, "error": str(e)}

# ==================== MULTI-RESTAURANT ENDPOINTS ====================

@api_router.get("/restaurants/my-restaurants")
async def get_my_restaurants(current_user: dict = Depends(get_current_user)):
    """Liste tous les restaurants auxquels le gérant a accès"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Seuls les gérants peuvent accéder à plusieurs restaurants")
    
    # Récupérer la liste des restaurant_ids de l'utilisateur
    restaurant_ids = current_user.get("restaurant_ids", [current_user["restaurant_id"]])
    
    # Si pas encore de liste, créer une liste avec le restaurant actuel
    if not restaurant_ids:
        restaurant_ids = [current_user["restaurant_id"]]
    
    # Récupérer les informations de tous les restaurants
    restaurants = await restaurants_collection.find(
        {"restaurant_id": {"$in": restaurant_ids}},
        {"_id": 0}
    ).to_list(100)
    
    return {
        "restaurants": restaurants,
        "current_restaurant_id": current_user["restaurant_id"]
    }

@api_router.post("/restaurants/create")
async def create_new_restaurant(
    create_request: CreateRestaurantRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer un nouveau restaurant (uniquement pour les gérants)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Seuls les gérants peuvent créer un restaurant")
    
    # Créer le nouveau restaurant
    restaurant_id = f"rest_{uuid.uuid4().hex[:12]}"
    restaurant = {
        "restaurant_id": restaurant_id,
        "name": create_request.name,
        "description": create_request.description or "Gestion des tâches cuisine",
        "logo_base64": None,
        "primary_color": "#26252D",
        "secondary_color": "#EAE6CA",
        "created_by": current_user["email"],
        "created_at": datetime.now(timezone.utc)
    }
    await restaurants_collection.insert_one(restaurant)
    
    # Ajouter ce restaurant à la liste des restaurants de l'utilisateur
    current_restaurant_ids = current_user.get("restaurant_ids", [current_user["restaurant_id"]])
    if current_user["restaurant_id"] not in current_restaurant_ids:
        current_restaurant_ids.append(current_user["restaurant_id"])
    current_restaurant_ids.append(restaurant_id)
    
    # Mettre à jour l'utilisateur
    await users_collection.update_one(
        {"user_id": current_user["user_id"]},
        {"$set": {"restaurant_ids": current_restaurant_ids}}
    )
    
    # Retourner le nouveau restaurant
    restaurant_doc = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    
    return {
        "message": "Restaurant créé avec succès",
        "restaurant": restaurant_doc,
        "restaurant_ids": current_restaurant_ids
    }

@api_router.post("/restaurants/duplicate")
async def duplicate_restaurant(
    duplicate_request: DuplicateRestaurantRequest,
    current_user: dict = Depends(get_current_user)
):
    """Dupliquer un restaurant avec toutes ses données (uniquement pour les gérants)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Seuls les gérants peuvent dupliquer un restaurant")
    
    source_restaurant_id = duplicate_request.source_restaurant_id
    new_name = duplicate_request.new_restaurant_name
    
    # Vérifier que l'utilisateur a accès au restaurant source
    restaurant_ids = current_user.get("restaurant_ids", [current_user.get("restaurant_id")])
    if source_restaurant_id not in restaurant_ids:
        raise HTTPException(status_code=403, detail="Vous n'avez pas accès à ce restaurant")
    
    # Récupérer le restaurant source
    source_restaurant = await restaurants_collection.find_one(
        {"restaurant_id": source_restaurant_id},
        {"_id": 0}
    )
    if not source_restaurant:
        raise HTTPException(status_code=404, detail="Restaurant source non trouvé")
    
    # Créer le nouveau restaurant
    new_restaurant_id = f"rest_{uuid.uuid4().hex[:12]}"
    new_restaurant = {
        "restaurant_id": new_restaurant_id,
        "name": new_name,
        "description": source_restaurant.get("description", ""),
        "logo_base64": source_restaurant.get("logo_base64"),
        "primary_color": source_restaurant.get("primary_color", "#26252D"),
        "secondary_color": source_restaurant.get("secondary_color", "#EAE6CA"),
        "invoice_settings": source_restaurant.get("invoice_settings"),
        "happy_hour_start": source_restaurant.get("happy_hour_start"),
        "happy_hour_end": source_restaurant.get("happy_hour_end"),
        "created_by": current_user["email"],
        "created_at": datetime.now(timezone.utc),
        "duplicated_from": source_restaurant_id
    }
    await restaurants_collection.insert_one(new_restaurant)
    
    # Mapping des anciens IDs vers les nouveaux
    section_id_mapping = {}
    
    # Copier les sections du menu restaurant
    source_sections = await menu_restaurant_sections_collection.find(
        {"restaurant_id": source_restaurant_id},
        {"_id": 0}
    ).to_list(1000)
    
    for section in source_sections:
        old_section_id = section["section_id"]
        new_section_id = f"sec_{uuid.uuid4().hex[:12]}"
        section_id_mapping[old_section_id] = new_section_id
        
        new_section = {
            **section,
            "section_id": new_section_id,
            "restaurant_id": new_restaurant_id,
            "created_at": datetime.now(timezone.utc)
        }
        await menu_restaurant_sections_collection.insert_one(new_section)
    
    # Copier les items du menu restaurant
    source_items = await menu_restaurant_items_collection.find(
        {"restaurant_id": source_restaurant_id},
        {"_id": 0}
    ).to_list(5000)
    
    for item in source_items:
        new_item_id = f"item_{uuid.uuid4().hex[:12]}"
        old_section_id = item.get("section_id")
        new_section_id = section_id_mapping.get(old_section_id, old_section_id)
        
        new_item = {
            **item,
            "item_id": new_item_id,
            "section_id": new_section_id,
            "restaurant_id": new_restaurant_id,
            "created_at": datetime.now(timezone.utc)
        }
        await menu_restaurant_items_collection.insert_one(new_item)
    
    # Copier les notes du menu restaurant
    source_notes = await menu_restaurant_notes_collection.find(
        {"restaurant_id": source_restaurant_id},
        {"_id": 0}
    ).to_list(1000)
    
    for note in source_notes:
        new_note_id = f"note_{uuid.uuid4().hex[:12]}"
        new_note = {
            **note,
            "note_id": new_note_id,
            "restaurant_id": new_restaurant_id,
            "created_at": datetime.now(timezone.utc)
        }
        await menu_restaurant_notes_collection.insert_one(new_note)
    
    # Copier les espaces de privatisation
    source_spaces = await privatisation_spaces_collection.find(
        {"restaurant_id": source_restaurant_id},
        {"_id": 0}
    ).to_list(100)
    
    for space in source_spaces:
        new_space_id = f"space_{uuid.uuid4().hex[:12]}"
        new_space = {
            **space,
            "space_id": new_space_id,
            "restaurant_id": new_restaurant_id,
            "created_at": datetime.now(timezone.utc)
        }
        await privatisation_spaces_collection.insert_one(new_space)
    
    # Copier les catégories
    source_categories = await categories_collection.find(
        {"restaurant_id": source_restaurant_id},
        {"_id": 0}
    ).to_list(100)
    
    for category in source_categories:
        new_category_id = f"cat_{uuid.uuid4().hex[:12]}"
        new_category = {
            **category,
            "category_id": new_category_id,
            "restaurant_id": new_restaurant_id,
            "created_at": datetime.now(timezone.utc)
        }
        await categories_collection.insert_one(new_category)
    
    # Copier les templates de tâches
    source_templates = await task_templates_collection.find(
        {"restaurant_id": source_restaurant_id},
        {"_id": 0}
    ).to_list(500)
    
    for template in source_templates:
        new_template_id = f"tmpl_{uuid.uuid4().hex[:12]}"
        new_template = {
            **template,
            "template_id": new_template_id,
            "restaurant_id": new_restaurant_id,
            "created_at": datetime.now(timezone.utc)
        }
        await task_templates_collection.insert_one(new_template)
    
    # Copier les sections de menu groupe
    source_group_sections = await menu_sections_collection.find(
        {"restaurant_id": source_restaurant_id},
        {"_id": 0}
    ).to_list(500)
    
    group_section_mapping = {}
    for section in source_group_sections:
        old_id = section["section_id"]
        new_id = f"grpsec_{uuid.uuid4().hex[:12]}"
        group_section_mapping[old_id] = new_id
        
        # Mettre à jour le parent_id si présent
        old_parent = section.get("parent_id")
        new_parent = group_section_mapping.get(old_parent, old_parent)
        
        new_section = {
            **section,
            "section_id": new_id,
            "parent_id": new_parent,
            "restaurant_id": new_restaurant_id,
            "created_at": datetime.now(timezone.utc)
        }
        await menu_sections_collection.insert_one(new_section)
    
    # Copier les items de menu groupe
    source_group_items = await menu_items_collection.find(
        {"restaurant_id": source_restaurant_id},
        {"_id": 0}
    ).to_list(2000)
    
    for item in source_group_items:
        new_item_id = f"grpitem_{uuid.uuid4().hex[:12]}"
        old_section_id = item.get("section_id")
        new_section_id = group_section_mapping.get(old_section_id, old_section_id)
        
        new_item = {
            **item,
            "item_id": new_item_id,
            "section_id": new_section_id,
            "restaurant_id": new_restaurant_id,
            "created_at": datetime.now(timezone.utc)
        }
        await menu_items_collection.insert_one(new_item)
    
    # Ajouter le nouveau restaurant à la liste des restaurants de l'utilisateur
    current_restaurant_ids = current_user.get("restaurant_ids", [current_user.get("restaurant_id")])
    if current_user.get("restaurant_id") and current_user["restaurant_id"] not in current_restaurant_ids:
        current_restaurant_ids.append(current_user["restaurant_id"])
    current_restaurant_ids.append(new_restaurant_id)
    
    await users_collection.update_one(
        {"user_id": current_user["user_id"]},
        {"$set": {"restaurant_ids": current_restaurant_ids}}
    )
    
    # Retourner le résumé
    new_restaurant_doc = await restaurants_collection.find_one(
        {"restaurant_id": new_restaurant_id},
        {"_id": 0}
    )
    
    return {
        "message": f"Restaurant '{new_name}' créé avec succès à partir de '{source_restaurant.get('name')}'",
        "restaurant": new_restaurant_doc,
        "restaurant_ids": current_restaurant_ids,
        "stats": {
            "sections_menu_restaurant": len(source_sections),
            "items_menu_restaurant": len(source_items),
            "notes_menu_restaurant": len(source_notes),
            "espaces_privatisation": len(source_spaces),
            "categories": len(source_categories),
            "templates_taches": len(source_templates),
            "sections_menu_groupe": len(source_group_sections),
            "items_menu_groupe": len(source_group_items)
        }
    }

@api_router.post("/restaurants/switch")
async def switch_restaurant(
    switch_request: SwitchRestaurantRequest,
    current_user: dict = Depends(get_current_user)
):
    """Changer le restaurant actif (uniquement pour les gérants)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Seuls les gérants peuvent changer de restaurant")
    
    # Vérifier que l'utilisateur a accès à ce restaurant
    restaurant_ids = current_user.get("restaurant_ids", [current_user["restaurant_id"]])
    if switch_request.restaurant_id not in restaurant_ids:
        raise HTTPException(status_code=403, detail="Vous n'avez pas accès à ce restaurant")
    
    # Vérifier que le restaurant existe
    restaurant_doc = await restaurants_collection.find_one(
        {"restaurant_id": switch_request.restaurant_id},
        {"_id": 0}
    )
    if not restaurant_doc:
        raise HTTPException(status_code=404, detail="Restaurant non trouvé")
    
    # Mettre à jour le restaurant actif de l'utilisateur
    await users_collection.update_one(
        {"user_id": current_user["user_id"]},
        {"$set": {"restaurant_id": switch_request.restaurant_id}}
    )
    
    # Retourner les nouvelles informations
    user_doc = await users_collection.find_one(
        {"user_id": current_user["user_id"]},
        {"_id": 0, "password_hash": 0}
    )
    
    return {
        "message": "Restaurant changé avec succès",
        "user": user_doc,
        "restaurant": restaurant_doc
    }

@api_router.post("/restaurants/{restaurant_id}/upload-logo")
async def upload_logo(
    restaurant_id: str,
    logo: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload un logo pour le restaurant"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if current_user["restaurant_id"] != restaurant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Lire le fichier et convertir en base64
    contents = await logo.read()
    logo_base64 = base64.b64encode(contents).decode('utf-8')
    
    # Mettre à jour le restaurant
    await restaurants_collection.update_one(
        {"restaurant_id": restaurant_id},
        {"$set": {"logo_base64": logo_base64}}
    )
    
    restaurant_doc = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    
    return restaurant_doc

# Endpoint d'upload générique pour les images
UPLOADS_DIR = Path("/app/backend/uploads/images")
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

@api_router.post("/upload/image")
async def upload_image(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload une image et retourne l'URL publique"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier le type de fichier
    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Type de fichier non autorisé. Utilisez JPG, PNG, GIF ou WebP.")
    
    # Générer un nom de fichier unique
    file_ext = Path(file.filename).suffix.lower() if file.filename else ".jpg"
    if file_ext not in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
        file_ext = ".jpg"
    
    unique_filename = f"{uuid.uuid4().hex}{file_ext}"
    file_path = UPLOADS_DIR / unique_filename
    
    # Sauvegarder le fichier
    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Retourner l'URL publique
    frontend_url = os.environ.get('FRONTEND_URL', '')
    public_url = f"{frontend_url}/api/uploads/images/{unique_filename}"
    
    return {"url": public_url, "filename": unique_filename}

# Route pour servir les images uploadées
from fastapi.responses import FileResponse

@api_router.get("/uploads/images/{filename}")
async def serve_uploaded_image(filename: str):
    """Servir une image uploadée"""
    file_path = UPLOADS_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Image not found")
    
    # Déterminer le type MIME
    mime_types = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp"
    }
    ext = Path(filename).suffix.lower()
    media_type = mime_types.get(ext, "application/octet-stream")
    
    return FileResponse(file_path, media_type=media_type)

# ==================== USER ENDPOINTS ====================

@api_router.post("/users/create")
async def create_user(
    create_request: CreateUserRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    existing_user = await users_collection.find_one({"email": create_request.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    
    # Préparer les permissions détaillées
    detailed_perms = None
    if create_request.detailed_permissions:
        detailed_perms = create_request.detailed_permissions.dict()
    
    user = {
        "user_id": user_id,
        "email": create_request.email,
        "password_hash": hash_password(create_request.password),
        "role": create_request.role,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "phone": create_request.phone,
        "assigned_categories": create_request.assigned_categories,
        "notification_prefs": {"push": True, "email": False, "sms": False},
        "detailed_permissions": detailed_perms,
        "created_at": datetime.now(timezone.utc)
    }
    await users_collection.insert_one(user)
    
    return {
        "user_id": user_id,
        "email": create_request.email,
        "name": create_request.name,
        "role": create_request.role,
        "restaurant_id": current_user["restaurant_id"],
        "assigned_categories": create_request.assigned_categories,
        "detailed_permissions": detailed_perms
    }

@api_router.get("/users/list")
async def list_users(current_user: dict = Depends(get_current_user)):
    # Admin et Staff peuvent voir la liste des utilisateurs
    users = await users_collection.find(
        {"restaurant_id": current_user["restaurant_id"]},
        {"_id": 0, "password_hash": 0}
    ).to_list(100)
    
    return users

@api_router.put("/users/{user_id}")
async def update_user(
    user_id: str,
    update_request: UpdateUserRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {}
    for k, v in update_request.dict().items():
        if v is not None:
            if k == "detailed_permissions" and v:
                update_data[k] = v
            elif k != "detailed_permissions":
                update_data[k] = v
    
    if update_data:
        await users_collection.update_one(
            {"user_id": user_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    user_doc = await users_collection.find_one(
        {"user_id": user_id},
        {"_id": 0, "password_hash": 0}
    )
    
    return user_doc

@api_router.put("/users/{user_id}/reset-password")
async def reset_user_password(
    user_id: str,
    reset_request: ResetPasswordRequest,
    current_user: dict = Depends(get_current_user)
):
    """Réinitialiser le mot de passe d'un utilisateur (admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que l'utilisateur existe et appartient au même restaurant
    user_doc = await users_collection.find_one({
        "user_id": user_id,
        "restaurant_id": current_user.get("restaurant_id")
    })
    
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Ne pas permettre de réinitialiser le mot de passe d'un admin
    if user_doc["role"] == "admin":
        raise HTTPException(status_code=403, detail="Cannot reset admin password")
    
    # Mettre à jour le mot de passe
    new_hash = hash_password(reset_request.new_password)
    await users_collection.update_one(
        {"user_id": user_id},
        {"$set": {"password_hash": new_hash}}
    )
    
    # Supprimer toutes les sessions de cet utilisateur
    await sessions_collection.delete_many({"user_id": user_id})
    
    return {"message": "Password reset successfully"}

class ChangeEmailRequest(BaseModel):
    new_email: str
    current_password: str

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

@api_router.put("/users/{user_id}/change-email")
async def change_user_email(
    user_id: str,
    request: ChangeEmailRequest,
    current_user: dict = Depends(get_current_user)
):
    """Changer son propre email (nécessite le mot de passe actuel)"""
    # Vérifier que l'utilisateur modifie son propre compte
    if current_user["user_id"] != user_id:
        raise HTTPException(status_code=403, detail="Vous ne pouvez modifier que votre propre compte")
    
    # Vérifier le mot de passe actuel
    user_doc = await users_collection.find_one({"user_id": user_id}, {"_id": 0})
    if not user_doc or not verify_password(request.current_password, user_doc["password_hash"]):
        raise HTTPException(status_code=401, detail="Mot de passe actuel incorrect")
    
    # Vérifier que le nouvel email n'est pas déjà utilisé
    existing = await users_collection.find_one({"email": request.new_email, "user_id": {"$ne": user_id}})
    if existing:
        raise HTTPException(status_code=400, detail="Cet email est déjà utilisé")
    
    # Mettre à jour l'email
    await users_collection.update_one(
        {"user_id": user_id},
        {"$set": {"email": request.new_email}}
    )
    
    return {"message": "Email mis à jour avec succès", "new_email": request.new_email}

@api_router.put("/users/{user_id}/change-password")
async def change_user_password(
    user_id: str,
    request: ChangePasswordRequest,
    current_user: dict = Depends(get_current_user)
):
    """Changer son propre mot de passe (nécessite le mot de passe actuel)"""
    # Vérifier que l'utilisateur modifie son propre compte
    if current_user["user_id"] != user_id:
        raise HTTPException(status_code=403, detail="Vous ne pouvez modifier que votre propre compte")
    
    # Vérifier le mot de passe actuel
    user_doc = await users_collection.find_one({"user_id": user_id}, {"_id": 0})
    if not user_doc or not verify_password(request.current_password, user_doc["password_hash"]):
        raise HTTPException(status_code=401, detail="Mot de passe actuel incorrect")
    
    # Vérifier la longueur du nouveau mot de passe
    if len(request.new_password) < 6:
        raise HTTPException(status_code=400, detail="Le mot de passe doit contenir au moins 6 caractères")
    
    # Mettre à jour le mot de passe
    new_hash = hash_password(request.new_password)
    await users_collection.update_one(
        {"user_id": user_id},
        {"$set": {"password_hash": new_hash}}
    )
    
    # Ne pas supprimer la session actuelle
    return {"message": "Mot de passe mis à jour avec succès"}


@api_router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await users_collection.delete_one(
        {"user_id": user_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

@api_router.put("/users/{user_id}/permissions")
async def update_user_permissions(
    user_id: str,
    permissions: UpdateUserPermissionsRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour les permissions d'un utilisateur (Menu Groupe, Tâches, Préparation de Commande, Fiche Technique, Catégories)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que l'utilisateur existe et appartient au même restaurant
    target_user = await users_collection.find_one({
        "user_id": user_id,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Valider fiche_technique_access
    valid_access = ["none", "bar", "cuisine", "both"]
    ft_access = permissions.fiche_technique_access if permissions.fiche_technique_access in valid_access else "none"
    
    # Préparer les permissions
    permissions_data = {
        "menu_groupe": permissions.menu_groupe,
        "taches": permissions.taches,
        "preparation_commande": permissions.preparation_commande,
        "fiche_technique": permissions.fiche_technique,
        "fiche_technique_access": ft_access,
        "categories": permissions.categories
    }
    
    # Mettre à jour l'utilisateur
    await users_collection.update_one(
        {"user_id": user_id},
        {"$set": {
            "permissions": permissions_data,
            "assigned_categories": permissions.categories  # Garder la compatibilité
        }}
    )
    
    # Récupérer l'utilisateur mis à jour
    updated_user = await users_collection.find_one(
        {"user_id": user_id},
        {"_id": 0, "password_hash": 0}
    )
    
    return {
        "message": "Permissions updated successfully",
        "user": updated_user
    }

@api_router.get("/users/{user_id}/permissions")
async def get_user_permissions(
    user_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer les permissions d'un utilisateur"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    target_user = await users_collection.find_one({
        "user_id": user_id,
        "restaurant_id": current_user["restaurant_id"]
    }, {"_id": 0, "password_hash": 0})
    
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Retourner les permissions ou les valeurs par défaut
    permissions = target_user.get("permissions", {
        "menu_groupe": False,
        "taches": True,
        "preparation_commande": False,
        "fiche_technique": False,
        "fiche_technique_access": "none",
        "categories": target_user.get("assigned_categories", [])
    })
    
    return {
        "user_id": user_id,
        "name": target_user.get("name"),
        "permissions": permissions
    }

# ==================== CATEGORY ENDPOINTS ====================

@api_router.post("/categories/create")
async def create_category(
    create_request: CreateCategoryRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    existing_categories = await categories_collection.find(
        {"restaurant_id": current_user["restaurant_id"]}
    ).to_list(100)
    max_order = max([c.get("order", 0) for c in existing_categories], default=-1)
    
    category_id = f"cat_{uuid.uuid4().hex[:12]}"
    category = {
        "category_id": category_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "order": max_order + 1,
        "created_at": datetime.now(timezone.utc)
    }
    await categories_collection.insert_one(category)
    
    return {
        "category_id": category_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "order": max_order + 1
    }

@api_router.get("/categories/list")
async def list_categories(current_user: dict = Depends(get_current_user)):
    # Admin voit toutes les catégories, Staff voit seulement les siennes
    if current_user["role"] == "admin":
        categories = await categories_collection.find(
            {"restaurant_id": current_user["restaurant_id"]},
            {"_id": 0}
        ).sort("order", 1).to_list(100)
    else:
        assigned = current_user.get("assigned_categories", [])
        if not assigned:
            return []
        categories = await categories_collection.find(
            {
                "restaurant_id": current_user["restaurant_id"],
                "category_id": {"$in": assigned}
            },
            {"_id": 0}
        ).sort("order", 1).to_list(100)
    
    return categories

@api_router.put("/categories/{category_id}")
async def update_category(
    category_id: str,
    update_request: UpdateCategoryRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await categories_collection.update_one(
            {"category_id": category_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    category_doc = await categories_collection.find_one(
        {"category_id": category_id},
        {"_id": 0}
    )
    
    return category_doc

@api_router.delete("/categories/{category_id}")
async def delete_category(
    category_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await categories_collection.delete_one(
        {"category_id": category_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    await task_templates_collection.delete_many({"category_id": category_id})
    
    return {"message": "Category deleted successfully"}

class ReorderCategoriesRequest(BaseModel):
    category_ids: List[str]  # Liste des category_ids dans le nouvel ordre

@api_router.post("/categories/reorder")
async def reorder_categories(
    reorder_request: ReorderCategoriesRequest,
    current_user: dict = Depends(get_current_user)
):
    """Réorganiser l'ordre des catégories"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    restaurant_id = current_user["restaurant_id"]
    
    # Mettre à jour l'ordre de chaque catégorie
    for index, category_id in enumerate(reorder_request.category_ids):
        await categories_collection.update_one(
            {"category_id": category_id, "restaurant_id": restaurant_id},
            {"$set": {"order": index}}
        )
    
    return {"message": "Categories reordered successfully"}

# ==================== PERMANENT CATEGORIES ENDPOINTS ====================

@api_router.post("/permanent-categories/create")
async def create_permanent_category(
    create_request: CreatePermanentCategoryRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une catégorie permanente (Ouverture, Fermeture, Livraison, etc.)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    existing_categories = await permanent_categories_collection.find(
        {"restaurant_id": current_user["restaurant_id"]}
    ).to_list(100)
    max_order = max([c.get("order", 0) for c in existing_categories], default=-1)
    
    category_id = f"pcat_{uuid.uuid4().hex[:12]}"
    category = {
        "permanent_category_id": category_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "order": max_order + 1,
        "created_at": datetime.now(timezone.utc)
    }
    await permanent_categories_collection.insert_one(category)
    
    return {
        "permanent_category_id": category_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "order": max_order + 1
    }

@api_router.get("/permanent-categories/list")
async def list_permanent_categories(current_user: dict = Depends(get_current_user)):
    """Lister toutes les catégories permanentes"""
    categories = await permanent_categories_collection.find(
        {"restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    
    return categories

@api_router.put("/permanent-categories/{category_id}")
async def update_permanent_category(
    category_id: str,
    update_request: UpdatePermanentCategoryRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await permanent_categories_collection.update_one(
            {"permanent_category_id": category_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    category_doc = await permanent_categories_collection.find_one(
        {"permanent_category_id": category_id},
        {"_id": 0}
    )
    
    return category_doc

@api_router.delete("/permanent-categories/{category_id}")
async def delete_permanent_category(
    category_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await permanent_categories_collection.delete_one(
        {"permanent_category_id": category_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Permanent category not found")
    
    # Supprimer aussi les tâches permanentes de cette catégorie
    await permanent_tasks_collection.delete_many({"permanent_category_id": category_id})
    
    return {"message": "Permanent category deleted successfully"}

# ==================== PERMANENT TASKS ENDPOINTS ====================

def task_matches_date(recurrence_rule: dict, date_str: str) -> bool:
    """Vérifie si une tâche doit apparaître pour une date donnée basée sur sa règle de récurrence"""
    if not recurrence_rule:
        return True  # Pas de règle = tous les jours
    
    rule_type = recurrence_rule.get("type", "daily")
    
    if rule_type == "daily":
        return True
    
    # Parser la date (format YYYY-MM-DD)
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return False
    
    if rule_type == "weekly":
        # days_of_week: 0=Lundi, 1=Mardi, ..., 6=Dimanche
        days_of_week = recurrence_rule.get("days_of_week", [])
        if not days_of_week:
            return True  # Si aucun jour spécifié, tous les jours
        # Python weekday(): 0=Lundi, 6=Dimanche
        return date_obj.weekday() in days_of_week
    
    if rule_type == "monthly":
        # days_of_month: 1-31
        days_of_month = recurrence_rule.get("days_of_month", [])
        if not days_of_month:
            return True  # Si aucun jour spécifié, tous les jours
        return date_obj.day in days_of_month
    
    return True  # Type inconnu = tous les jours par défaut

def format_recurrence_display(recurrence_rule: dict) -> str:
    """Formate la règle de récurrence pour affichage"""
    if not recurrence_rule:
        return "Tous les jours"
    
    rule_type = recurrence_rule.get("type", "daily")
    
    if rule_type == "daily":
        return "Tous les jours"
    
    if rule_type == "weekly":
        days = recurrence_rule.get("days_of_week", [])
        day_names = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
        if not days:
            return "Tous les jours"
        return "Chaque " + ", ".join([day_names[d] for d in sorted(days) if 0 <= d <= 6])
    
    if rule_type == "monthly":
        days = recurrence_rule.get("days_of_month", [])
        if not days:
            return "Tous les jours"
        return "Le " + ", ".join([str(d) for d in sorted(days)]) + " du mois"
    
    return "Tous les jours"

@api_router.post("/permanent-tasks/create")
async def create_permanent_task(
    create_request: CreatePermanentTaskRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une tâche permanente avec règle de récurrence (quotidien, hebdomadaire ou mensuel)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que la catégorie permanente existe
    category_doc = await permanent_categories_collection.find_one(
        {"permanent_category_id": create_request.permanent_category_id, "restaurant_id": current_user["restaurant_id"]}
    )
    if not category_doc:
        raise HTTPException(status_code=404, detail="Permanent category not found")
    
    # Préparer la règle de récurrence
    recurrence = None
    if create_request.recurrence_rule:
        recurrence = {
            "type": create_request.recurrence_rule.type,
            "days_of_week": create_request.recurrence_rule.days_of_week,
            "days_of_month": create_request.recurrence_rule.days_of_month
        }
    
    task_id = f"ptask_{uuid.uuid4().hex[:12]}"
    task = {
        "permanent_task_id": task_id,
        "permanent_category_id": create_request.permanent_category_id,
        "restaurant_id": current_user["restaurant_id"],
        "title": create_request.title,
        "description": create_request.description,
        "recurrence_rule": recurrence,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await permanent_tasks_collection.insert_one(task)
    
    return {
        "permanent_task_id": task_id,
        "permanent_category_id": create_request.permanent_category_id,
        "title": create_request.title,
        "description": create_request.description,
        "recurrence_rule": recurrence,
        "recurrence_display": format_recurrence_display(recurrence),
        "is_active": True
    }

@api_router.get("/permanent-tasks/list")
async def list_permanent_tasks(
    permanent_category_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les tâches permanentes avec leur règle de récurrence"""
    query = {"restaurant_id": current_user["restaurant_id"], "is_active": True}
    
    if permanent_category_id:
        query["permanent_category_id"] = permanent_category_id
    
    tasks = await permanent_tasks_collection.find(query, {"_id": 0}).to_list(500)
    
    # Ajouter l'affichage de la récurrence pour chaque tâche
    for task in tasks:
        task["recurrence_display"] = format_recurrence_display(task.get("recurrence_rule"))
    
    return tasks

@api_router.put("/permanent-tasks/{task_id}")
async def update_permanent_task(
    task_id: str,
    update_request: UpdatePermanentTaskRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {}
    
    if update_request.title is not None:
        update_data["title"] = update_request.title
    if update_request.description is not None:
        update_data["description"] = update_request.description
    if update_request.is_active is not None:
        update_data["is_active"] = update_request.is_active
    if update_request.recurrence_rule is not None:
        update_data["recurrence_rule"] = {
            "type": update_request.recurrence_rule.type,
            "days_of_week": update_request.recurrence_rule.days_of_week,
            "days_of_month": update_request.recurrence_rule.days_of_month
        }
    
    if update_data:
        await permanent_tasks_collection.update_one(
            {"permanent_task_id": task_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    task_doc = await permanent_tasks_collection.find_one(
        {"permanent_task_id": task_id},
        {"_id": 0}
    )
    
    if task_doc:
        task_doc["recurrence_display"] = format_recurrence_display(task_doc.get("recurrence_rule"))
    
    return task_doc

@api_router.delete("/permanent-tasks/{task_id}")
async def delete_permanent_task(
    task_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await permanent_tasks_collection.update_one(
        {"permanent_task_id": task_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Permanent task not found")
    
    return {"message": "Permanent task deleted successfully"}

@api_router.put("/permanent-tasks/{task_id}/complete")
async def complete_permanent_task(
    task_id: str,
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Marquer une tâche permanente comme terminée pour une date spécifique"""
    # Vérifier que la tâche permanente existe
    task_doc = await permanent_tasks_collection.find_one(
        {"permanent_task_id": task_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not task_doc:
        raise HTTPException(status_code=404, detail="Permanent task not found")
    
    # Vérifier si déjà complétée pour cette date
    existing = await permanent_task_completions_collection.find_one({
        "permanent_task_id": task_id,
        "date": date,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if existing:
        return {"message": "Task already completed for this date"}
    
    # Créer l'enregistrement de complétion
    completion = {
        "completion_id": f"comp_{uuid.uuid4().hex[:12]}",
        "permanent_task_id": task_id,
        "restaurant_id": current_user["restaurant_id"],
        "date": date,
        "completed_by": current_user["user_id"],
        "completed_at": datetime.now(timezone.utc)
    }
    await permanent_task_completions_collection.insert_one(completion)
    
    # Historique
    history = {
        "history_id": f"hist_{uuid.uuid4().hex[:12]}",
        "task_id": f"{task_id}_{date}",
        "user_id": current_user["user_id"],
        "user_name": current_user["name"],
        "action": "completed_permanent",
        "timestamp": datetime.now(timezone.utc)
    }
    await history_collection.insert_one(history)
    
    return {
        "message": "Permanent task completed",
        "permanent_task_id": task_id,
        "date": date,
        "status": "completed"
    }

@api_router.put("/permanent-tasks/{task_id}/uncomplete")
async def uncomplete_permanent_task(
    task_id: str,
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Remettre une tâche permanente en attente pour une date spécifique"""
    # Supprimer l'enregistrement de complétion
    result = await permanent_task_completions_collection.delete_one({
        "permanent_task_id": task_id,
        "date": date,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if result.deleted_count == 0:
        return {"message": "Task was not completed for this date"}
    
    return {
        "message": "Permanent task uncompleted",
        "permanent_task_id": task_id,
        "date": date,
        "status": "pending"
    }

# ==================== PERMANENT SUBTASKS ENDPOINTS ====================

@api_router.post("/permanent-subtasks/create")
async def create_permanent_subtask(
    create_request: CreatePermanentSubtaskRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une sous-tâche liée à une tâche permanente"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que la tâche permanente parent existe
    parent_task = await permanent_tasks_collection.find_one({
        "permanent_task_id": create_request.parent_permanent_task_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    })
    
    if not parent_task:
        raise HTTPException(status_code=404, detail="Parent permanent task not found")
    
    subtask_id = f"psub_{uuid.uuid4().hex[:12]}"
    subtask = {
        "subtask_id": subtask_id,
        "parent_permanent_task_id": create_request.parent_permanent_task_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "quantity": create_request.quantity,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await permanent_subtasks_collection.insert_one(subtask)
    
    return {
        "subtask_id": subtask_id,
        "parent_permanent_task_id": create_request.parent_permanent_task_id,
        "name": create_request.name,
        "quantity": create_request.quantity,
        "is_active": True
    }

@api_router.get("/permanent-subtasks/list")
async def list_permanent_subtasks(
    parent_permanent_task_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les sous-tâches permanentes (optionnellement filtrées par parent)"""
    query = {"restaurant_id": current_user["restaurant_id"], "is_active": True}
    
    if parent_permanent_task_id:
        query["parent_permanent_task_id"] = parent_permanent_task_id
    
    subtasks = await permanent_subtasks_collection.find(query, {"_id": 0}).to_list(500)
    return subtasks

@api_router.get("/permanent-subtasks/by-task/{task_id}")
async def get_permanent_subtasks_by_task(
    task_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer toutes les sous-tâches d'une tâche permanente"""
    subtasks = await permanent_subtasks_collection.find({
        "parent_permanent_task_id": task_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    }, {"_id": 0}).to_list(100)
    
    return subtasks

@api_router.put("/permanent-subtasks/{subtask_id}")
async def update_permanent_subtask(
    subtask_id: str,
    update_request: UpdatePermanentSubtaskRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour une sous-tâche permanente"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await permanent_subtasks_collection.update_one(
            {"subtask_id": subtask_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    subtask_doc = await permanent_subtasks_collection.find_one(
        {"subtask_id": subtask_id},
        {"_id": 0}
    )
    
    return subtask_doc

@api_router.delete("/permanent-subtasks/{subtask_id}")
async def delete_permanent_subtask(
    subtask_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une sous-tâche permanente"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await permanent_subtasks_collection.update_one(
        {"subtask_id": subtask_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Permanent subtask not found")
    
    return {"message": "Permanent subtask deleted successfully"}

@api_router.put("/permanent-subtasks/{subtask_id}/update-quantity")
async def update_permanent_subtask_quantity(
    subtask_id: str,
    date: str,
    quantity: int,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour la quantité d'une sous-tâche permanente pour une date donnée"""
    # Vérifier que la sous-tâche existe
    subtask = await permanent_subtasks_collection.find_one({
        "subtask_id": subtask_id,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if not subtask:
        raise HTTPException(status_code=404, detail="Permanent subtask not found")
    
    # Mettre à jour ou créer l'entrée de complétion
    existing = await permanent_subtask_completions_collection.find_one({
        "subtask_id": subtask_id,
        "date": date,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if existing:
        await permanent_subtask_completions_collection.update_one(
            {"subtask_id": subtask_id, "date": date, "restaurant_id": current_user["restaurant_id"]},
            {"$set": {"quantity": quantity, "updated_at": datetime.now(timezone.utc)}}
        )
    else:
        completion = {
            "completion_id": f"pscomp_{uuid.uuid4().hex[:12]}",
            "subtask_id": subtask_id,
            "parent_permanent_task_id": subtask["parent_permanent_task_id"],
            "restaurant_id": current_user["restaurant_id"],
            "date": date,
            "quantity": quantity,
            "completed_by": current_user["user_id"],
            "created_at": datetime.now(timezone.utc)
        }
        await permanent_subtask_completions_collection.insert_one(completion)
    
    return {
        "message": "Quantity updated",
        "subtask_id": subtask_id,
        "date": date,
        "quantity": quantity
    }

@api_router.get("/permanent-subtasks/completions")
async def get_permanent_subtask_completions(
    date: str,
    parent_permanent_task_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer les complétions de sous-tâches permanentes pour une date"""
    query = {
        "restaurant_id": current_user["restaurant_id"],
        "date": date
    }
    
    if parent_permanent_task_id:
        query["parent_permanent_task_id"] = parent_permanent_task_id
    
    completions = await permanent_subtask_completions_collection.find(query, {"_id": 0}).to_list(500)
    return completions

# ==================== TASK TEMPLATE ENDPOINTS ====================

@api_router.post("/task-templates/create")
async def create_task_template(
    create_request: CreateTaskTemplateRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    category_doc = await categories_collection.find_one(
        {"category_id": create_request.category_id, "restaurant_id": current_user["restaurant_id"]}
    )
    if not category_doc:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Préparer la règle de récurrence si fournie
    recurrence = None
    if create_request.recurrence_rule:
        recurrence = {
            "type": create_request.recurrence_rule.type,
            "days_of_week": create_request.recurrence_rule.days_of_week,
            "days_of_month": create_request.recurrence_rule.days_of_month
        }
    
    template_id = f"tmpl_{uuid.uuid4().hex[:12]}"
    template = {
        "template_id": template_id,
        "category_id": create_request.category_id,
        "restaurant_id": current_user["restaurant_id"],
        "title": create_request.title,
        "description": create_request.description,
        "task_type": create_request.task_type,  # "manual" ou "permanent"
        "recurrence_rule": recurrence,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await task_templates_collection.insert_one(template)
    
    return {
        "template_id": template_id,
        "category_id": create_request.category_id,
        "title": create_request.title,
        "description": create_request.description,
        "task_type": create_request.task_type,
        "recurrence_rule": recurrence,
        "recurrence_display": format_recurrence_display(recurrence) if recurrence else None,
        "is_active": True
    }

@api_router.get("/task-templates/list")
async def list_task_templates(
    category_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les tâches modèles - Staff voit seulement ses catégories"""
    query = {"restaurant_id": current_user["restaurant_id"], "is_active": True}
    
    if category_id:
        query["category_id"] = category_id
    
    # Staff voit seulement les templates de ses catégories assignées
    if current_user["role"] == "staff":
        assigned = current_user.get("assigned_categories", [])
        if not assigned:
            return []
        query["category_id"] = {"$in": assigned}
    
    templates = await task_templates_collection.find(query, {"_id": 0}).to_list(500)
    
    # Ajouter recurrence_display pour les tâches permanentes
    for template in templates:
        if template.get("task_type") == "permanent" and template.get("recurrence_rule"):
            template["recurrence_display"] = format_recurrence_display(template.get("recurrence_rule"))
        else:
            template["recurrence_display"] = None
    
    return templates

@api_router.put("/task-templates/{template_id}")
async def update_task_template(
    template_id: str,
    update_request: UpdateTaskTemplateRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await task_templates_collection.update_one(
            {"template_id": template_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    template_doc = await task_templates_collection.find_one(
        {"template_id": template_id},
        {"_id": 0}
    )
    
    return template_doc

@api_router.delete("/task-templates/{template_id}")
async def delete_task_template(
    template_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await task_templates_collection.update_one(
        {"template_id": template_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Task template not found")
    
    return {"message": "Task template deleted successfully"}

# ==================== SUBTASKS ENDPOINTS ====================

@api_router.post("/subtasks/create")
async def create_subtask(
    create_request: CreateSubtaskRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une sous-tâche liée à un template parent"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que le template parent existe
    parent_template = await task_templates_collection.find_one({
        "template_id": create_request.parent_template_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    })
    
    if not parent_template:
        raise HTTPException(status_code=404, detail="Parent template not found")
    
    subtask_id = f"sub_{uuid.uuid4().hex[:12]}"
    subtask = {
        "subtask_id": subtask_id,
        "parent_template_id": create_request.parent_template_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "quantity": create_request.quantity,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await subtasks_collection.insert_one(subtask)
    
    return {
        "subtask_id": subtask_id,
        "parent_template_id": create_request.parent_template_id,
        "name": create_request.name,
        "quantity": create_request.quantity,
        "is_active": True
    }

@api_router.get("/subtasks/list")
async def list_subtasks(
    parent_template_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les sous-tâches (optionnellement filtrées par parent)"""
    query = {"restaurant_id": current_user["restaurant_id"], "is_active": True}
    
    if parent_template_id:
        query["parent_template_id"] = parent_template_id
    
    subtasks = await subtasks_collection.find(query, {"_id": 0}).to_list(500)
    return subtasks

@api_router.get("/subtasks/by-template/{template_id}")
async def get_subtasks_by_template(
    template_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer toutes les sous-tâches d'un template"""
    subtasks = await subtasks_collection.find({
        "parent_template_id": template_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    }, {"_id": 0}).to_list(100)
    
    return subtasks

@api_router.put("/subtasks/{subtask_id}")
async def update_subtask(
    subtask_id: str,
    update_request: UpdateSubtaskRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour une sous-tâche"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await subtasks_collection.update_one(
            {"subtask_id": subtask_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    subtask_doc = await subtasks_collection.find_one(
        {"subtask_id": subtask_id},
        {"_id": 0}
    )
    
    return subtask_doc

@api_router.delete("/subtasks/{subtask_id}")
async def delete_subtask(
    subtask_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une sous-tâche"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await subtasks_collection.update_one(
        {"subtask_id": subtask_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Subtask not found")
    
    return {"message": "Subtask deleted successfully"}

class SubtaskCompleteBody(BaseModel):
    quantity: Optional[int] = None

@api_router.put("/subtasks/{subtask_id}/complete")
async def complete_subtask(
    subtask_id: str,
    date: str,
    body: SubtaskCompleteBody = Body(default=SubtaskCompleteBody()),
    current_user: dict = Depends(get_current_user)
):
    """Marquer une sous-tâche comme terminée pour une date donnée (avec quantité optionnelle)"""
    # Vérifier que la sous-tâche existe
    subtask = await subtasks_collection.find_one({
        "subtask_id": subtask_id,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if not subtask:
        raise HTTPException(status_code=404, detail="Subtask not found")
    
    # Vérifier si déjà complétée
    existing = await subtask_completions_collection.find_one({
        "subtask_id": subtask_id,
        "date": date,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if existing:
        # Si déjà complétée, mettre à jour la quantité si fournie
        if body.quantity is not None:
            await subtask_completions_collection.update_one(
                {"subtask_id": subtask_id, "date": date, "restaurant_id": current_user["restaurant_id"]},
                {"$set": {"quantity": body.quantity, "updated_at": datetime.now(timezone.utc)}}
            )
            return {"message": "Quantity updated", "subtask_id": subtask_id, "date": date, "quantity": body.quantity, "status": "completed"}
        return {"message": "Subtask already completed", "status": "completed", "quantity": existing.get("quantity")}
    
    # Créer l'entrée de complétion
    completion = {
        "completion_id": f"scomp_{uuid.uuid4().hex[:12]}",
        "subtask_id": subtask_id,
        "parent_template_id": subtask["parent_template_id"],
        "restaurant_id": current_user["restaurant_id"],
        "date": date,
        "quantity": body.quantity,
        "completed_by": current_user["user_id"],
        "completed_at": datetime.now(timezone.utc)
    }
    
    await subtask_completions_collection.insert_one(completion)
    
    return {
        "message": "Subtask completed",
        "subtask_id": subtask_id,
        "date": date,
        "quantity": body.quantity,
        "status": "completed"
    }

@api_router.put("/subtasks/{subtask_id}/uncomplete")
async def uncomplete_subtask(
    subtask_id: str,
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Remettre une sous-tâche en attente"""
    result = await subtask_completions_collection.delete_one({
        "subtask_id": subtask_id,
        "date": date,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if result.deleted_count == 0:
        return {"message": "Subtask was not completed", "status": "pending"}
    
    return {
        "message": "Subtask uncompleted",
        "subtask_id": subtask_id,
        "date": date,
        "status": "pending"
    }

@api_router.get("/subtasks/completions")
async def get_subtask_completions(
    date: str,
    parent_template_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer les complétions de sous-tâches pour une date"""
    query = {
        "restaurant_id": current_user["restaurant_id"],
        "date": date
    }
    
    if parent_template_id:
        query["parent_template_id"] = parent_template_id
    
    completions = await subtask_completions_collection.find(query, {"_id": 0}).to_list(500)
    return completions

@api_router.get("/templates/{template_id}/can-complete")
async def check_template_can_complete(
    template_id: str,
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Vérifier si un template peut être marqué comme complet (toutes sous-tâches complètes)"""
    # Récupérer toutes les sous-tâches actives de ce template
    subtasks = await subtasks_collection.find({
        "parent_template_id": template_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    }, {"subtask_id": 1, "_id": 0}).to_list(100)
    
    if not subtasks:
        return {"can_complete": True, "total_subtasks": 0, "completed_subtasks": 0}
    
    subtask_ids = [s["subtask_id"] for s in subtasks]
    
    # Compter combien sont complétées pour cette date
    completed_count = await subtask_completions_collection.count_documents({
        "subtask_id": {"$in": subtask_ids},
        "date": date,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    total = len(subtasks)
    can_complete = completed_count >= total
    
    return {
        "can_complete": can_complete,
        "total_subtasks": total,
        "completed_subtasks": completed_count
    }

# ==================== DAILY TASKS ENDPOINTS ====================

@api_router.post("/daily-tasks/select")
async def select_tasks_for_day(
    select_request: SelectTasksForDayRequest,
    current_user: dict = Depends(get_current_user)
):
    """Sélectionner les tâches modèles pour un jour - Admin ET Staff peuvent le faire"""
    # Staff peut seulement sélectionner des tâches de ses catégories
    assigned_categories = current_user.get("assigned_categories", [])
    
    created_tasks = []
    
    for selection in select_request.selections:
        template = await task_templates_collection.find_one({
            "template_id": selection.template_id,
            "restaurant_id": current_user["restaurant_id"],
            "is_active": True
        })
        
        if not template:
            continue
        
        # Staff ne peut sélectionner que ses catégories
        if current_user["role"] == "staff" and template["category_id"] not in assigned_categories:
            continue
        
        # Vérifier si la tâche existe déjà
        existing = await daily_tasks_collection.find_one({
            "template_id": template["template_id"],
            "date": select_request.date,
            "restaurant_id": current_user["restaurant_id"]
        })
        
        if not existing:
            task_id = f"task_{uuid.uuid4().hex[:12]}"
            daily_task = {
                "task_id": task_id,
                "template_id": template["template_id"],
                "category_id": template["category_id"],
                "restaurant_id": current_user["restaurant_id"],
                "title": template["title"],
                "description": template.get("description"),
                "date": select_request.date,
                "status": "pending",
                "is_recurring": True,
                "is_sent": False,
                "assigned_user_id": selection.assigned_user_id,  # Personne taguée
                "created_by": current_user["user_id"],
                "completed_by": None,
                "completed_at": None,
                "created_at": datetime.now(timezone.utc)
            }
            await daily_tasks_collection.insert_one(daily_task)
            created_tasks.append({
                "task_id": task_id,
                "title": template["title"],
                "category_id": template["category_id"],
                "assigned_user_id": selection.assigned_user_id
            })
    
    return {
        "message": f"{len(created_tasks)} tasks selected for {select_request.date}",
        "tasks_created": created_tasks
    }

@api_router.post("/daily-tasks/create-punctual")
async def create_punctual_task(
    create_request: CreatePunctualTaskRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une tâche ponctuelle - Admin uniquement"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required for punctual tasks")
    
    category_doc = await categories_collection.find_one(
        {"category_id": create_request.category_id, "restaurant_id": current_user["restaurant_id"]}
    )
    if not category_doc:
        raise HTTPException(status_code=404, detail="Category not found")
    
    task_id = f"task_{uuid.uuid4().hex[:12]}"
    daily_task = {
        "task_id": task_id,
        "template_id": None,
        "category_id": create_request.category_id,
        "restaurant_id": current_user["restaurant_id"],
        "title": create_request.title,
        "description": create_request.description,
        "date": create_request.date,
        "status": "pending",
        "is_recurring": False,
        "is_sent": False,
        "assigned_user_id": create_request.assigned_user_id,
        "created_by": current_user["user_id"],
        "completed_by": None,
        "completed_at": None,
        "created_at": datetime.now(timezone.utc)
    }
    await daily_tasks_collection.insert_one(daily_task)
    
    # Envoyer une notification push si la tâche est assignée à quelqu'un
    if create_request.assigned_user_id:
        await send_push_notification_to_user(
            user_id=create_request.assigned_user_id,
            title="Nouvelle tâche assignée",
            body=f"Tâche: {create_request.title}",
            data={"type": "task", "task_id": task_id}
        )
    else:
        # Sinon notifier tout le restaurant
        await send_push_notification_to_restaurant(
            restaurant_id=current_user["restaurant_id"],
            title="Nouvelle tâche",
            body=f"Tâche: {create_request.title}",
            data={"type": "task", "task_id": task_id},
            exclude_user=current_user["user_id"]
        )
    
    return {
        "task_id": task_id,
        "title": create_request.title,
        "category_id": create_request.category_id,
        "date": create_request.date,
        "is_recurring": False,
        "assigned_user_id": create_request.assigned_user_id
    }

@api_router.post("/daily-tasks/send")
async def send_daily_tasks(
    send_request: SendDailyTasksRequest,
    current_user: dict = Depends(get_current_user)
):
    """Envoyer les tâches - Admin ET Staff peuvent envoyer"""
    # Staff peut seulement envoyer les tâches de ses catégories
    assigned_categories = current_user.get("assigned_categories", [])
    
    query = {
        "restaurant_id": current_user["restaurant_id"],
        "date": send_request.date,
        "is_sent": False
    }
    
    # Staff ne peut envoyer que ses catégories
    if current_user["role"] == "staff":
        if not assigned_categories:
            raise HTTPException(status_code=403, detail="No categories assigned")
        query["category_id"] = {"$in": assigned_categories}
    
    tasks = await daily_tasks_collection.find(query, {"_id": 0}).to_list(500)
    
    if not tasks:
        raise HTTPException(status_code=400, detail="No tasks to send for this date")
    
    # Grouper les tâches par catégorie
    tasks_by_category = {}
    for task in tasks:
        cat_id = task["category_id"]
        if cat_id not in tasks_by_category:
            tasks_by_category[cat_id] = []
        tasks_by_category[cat_id].append(task)
    
    # Récupérer tous les utilisateurs staff
    staff_users = await users_collection.find({
        "restaurant_id": current_user["restaurant_id"],
        "role": "staff"
    }, {"_id": 0}).to_list(100)
    
    # Créer les notifications
    notifications_created = []
    users_notified = set()
    
    for task in tasks:
        # Si la tâche est assignée à une personne spécifique
        if task.get("assigned_user_id"):
            target_user = await users_collection.find_one(
                {"user_id": task["assigned_user_id"]},
                {"_id": 0}
            )
            if target_user and target_user["user_id"] not in users_notified:
                users_notified.add(target_user["user_id"])
        else:
            # Sinon, notifier tous les utilisateurs de la catégorie
            for user in staff_users:
                if task["category_id"] in user.get("assigned_categories", []):
                    users_notified.add(user["user_id"])
    
    # Créer une notification globale
    for user_id in users_notified:
        user = await users_collection.find_one({"user_id": user_id}, {"_id": 0})
        if user:
            notification_id = f"notif_{uuid.uuid4().hex[:12]}"
            notification = {
                "notification_id": notification_id,
                "user_id": user_id,
                "user_name": user["name"],
                "restaurant_id": current_user["restaurant_id"],
                "date": send_request.date,
                "message": f"Nouvelles tâches pour le {send_request.date}",
                "status": "sent",
                "sent_at": datetime.now(timezone.utc),
                "sent_by": current_user["user_id"]
            }
            await notifications_collection.insert_one(notification)
            notifications_created.append({
                "user_name": user["name"],
                "user_email": user["email"]
            })
    
    # Marquer les tâches comme envoyées
    task_ids = [t["task_id"] for t in tasks]
    await daily_tasks_collection.update_many(
        {"task_id": {"$in": task_ids}},
        {"$set": {"is_sent": True, "sent_at": datetime.now(timezone.utc)}}
    )
    
    # Historique
    history = {
        "history_id": f"hist_{uuid.uuid4().hex[:12]}",
        "task_id": "batch_send",
        "user_id": current_user["user_id"],
        "user_name": current_user["name"],
        "action": f"sent_tasks_{send_request.date}",
        "details": {
            "date": send_request.date,
            "task_count": len(tasks),
            "users_notified": len(users_notified)
        },
        "timestamp": datetime.now(timezone.utc)
    }
    await history_collection.insert_one(history)
    
    return {
        "message": f"Tasks sent for {send_request.date}",
        "tasks_sent": len(tasks),
        "users_notified": len(users_notified),
        "notifications": notifications_created
    }

@api_router.get("/daily-tasks/list")
async def list_daily_tasks(
    date: str,
    category_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les tâches du jour (inclut les tâches permanentes automatiquement)"""
    # Staff voit seulement les tâches envoyées de ses catégories OU celles qui lui sont assignées
    if current_user["role"] == "staff":
        assigned = current_user.get("assigned_categories", [])
        query = {
            "restaurant_id": current_user["restaurant_id"],
            "date": date,
            "is_sent": True,
            "$or": [
                {"category_id": {"$in": assigned}},
                {"assigned_user_id": current_user["user_id"]}
            ]
        }
        if category_id:
            query["category_id"] = category_id
    else:
        # Admin voit tout
        query = {
            "restaurant_id": current_user["restaurant_id"],
            "date": date
        }
        if category_id:
            query["category_id"] = category_id
    
    tasks = await daily_tasks_collection.find(query, {"_id": 0}).to_list(500)
    
    # Enrichir avec le nom de l'utilisateur assigné si applicable
    for task in tasks:
        if task.get("assigned_user_id"):
            user = await users_collection.find_one(
                {"user_id": task["assigned_user_id"]},
                {"name": 1, "_id": 0}
            )
            task["assigned_user_name"] = user["name"] if user else None
    
    # ========== AJOUTER LES TÂCHES PERMANENTES ==========
    # Récupérer les catégories permanentes
    permanent_categories = await permanent_categories_collection.find(
        {"restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    
    # Récupérer les tâches permanentes actives
    permanent_tasks = await permanent_tasks_collection.find(
        {"restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).to_list(500)
    
    # Récupérer les complétions des tâches permanentes pour cette date
    completions = await permanent_task_completions_collection.find(
        {"restaurant_id": current_user["restaurant_id"], "date": date},
        {"_id": 0}
    ).to_list(500)
    completions_dict = {c["permanent_task_id"]: c for c in completions}
    
    # Créer un dictionnaire des catégories permanentes pour lookup rapide
    pcat_dict = {pc["permanent_category_id"]: pc for pc in permanent_categories}
    
    # Convertir les tâches permanentes au format des daily tasks
    for ptask in permanent_tasks:
        pcat = pcat_dict.get(ptask["permanent_category_id"])
        if pcat:
            # VÉRIFIER SI LA TÂCHE DOIT APPARAÎTRE POUR CETTE DATE (selon recurrence_rule)
            recurrence_rule = ptask.get("recurrence_rule")
            if not task_matches_date(recurrence_rule, date):
                continue  # Cette tâche ne doit pas apparaître pour cette date
            
            # Vérifier si la tâche a été complétée pour ce jour
            completion = completions_dict.get(ptask["permanent_task_id"])
            status = "completed" if completion else "pending"
            completed_by = completion.get("completed_by") if completion else None
            completed_at = completion.get("completed_at") if completion else None
            
            permanent_task_item = {
                "task_id": f"{ptask['permanent_task_id']}_{date}",  # ID unique par jour
                "permanent_task_id": ptask["permanent_task_id"],
                "permanent_category_id": ptask["permanent_category_id"],
                "permanent_category_name": pcat["name"],
                "restaurant_id": ptask["restaurant_id"],
                "title": ptask["title"],
                "description": ptask.get("description"),
                "date": date,
                "status": status,
                "is_permanent": True,  # Flag pour identifier les tâches permanentes
                "is_recurring": True,
                "is_sent": True,
                "completed_by": completed_by,
                "completed_at": completed_at,
                "recurrence_display": format_recurrence_display(recurrence_rule)
            }
            tasks.append(permanent_task_item)
    
    # ========== AJOUTER LES TÂCHES PERMANENTES DES TEMPLATES ==========
    # Récupérer les templates avec task_type="permanent"
    template_query = {
        "restaurant_id": current_user["restaurant_id"],
        "task_type": "permanent",
        "is_active": True
    }
    
    # Staff voit seulement ses catégories
    if current_user["role"] == "staff":
        assigned = current_user.get("assigned_categories", [])
        if assigned:
            template_query["category_id"] = {"$in": assigned}
    
    permanent_templates = await task_templates_collection.find(template_query, {"_id": 0}).to_list(500)
    
    # Récupérer les catégories pour lookup
    categories = await categories_collection.find(
        {"restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    ).to_list(100)
    cat_dict = {c["category_id"]: c for c in categories}
    
    # Récupérer les complétions des templates permanents pour cette date
    # Note: Les complétions sont stockées dans mep_template_completions (par /recurring-tasks endpoints)
    template_completions = await db.mep_template_completions.find(
        {"restaurant_id": current_user["restaurant_id"], "date": date},
        {"_id": 0}
    ).to_list(500)
    template_completions_dict = {c["template_id"]: c for c in template_completions}
    
    for template in permanent_templates:
        cat = cat_dict.get(template["category_id"])
        if cat:
            # VÉRIFIER SI LA TÂCHE DOIT APPARAÎTRE POUR CETTE DATE
            recurrence_rule = template.get("recurrence_rule")
            if not task_matches_date(recurrence_rule, date):
                continue
            
            # Vérifier si complétée
            completion = template_completions_dict.get(template["template_id"])
            status = "completed" if completion else "pending"
            completed_by = completion.get("completed_by") if completion else None
            completed_at = completion.get("completed_at") if completion else None
            
            template_task_item = {
                "task_id": f"{template['template_id']}_{date}",
                "template_id": template["template_id"],
                "category_id": template["category_id"],
                "category_name": cat["name"],
                "restaurant_id": template["restaurant_id"],
                "title": template["title"],
                "description": template.get("description"),
                "date": date,
                "status": status,
                "is_permanent": True,
                "is_recurring": True,
                "is_sent": True,
                "completed_by": completed_by,
                "completed_at": completed_at,
                "recurrence_display": format_recurrence_display(recurrence_rule)
            }
            tasks.append(template_task_item)
    
    return tasks

@api_router.get("/daily-tasks/pending")
async def get_pending_tasks(
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Voir les tâches sélectionnées mais pas encore envoyées"""
    query = {
        "restaurant_id": current_user["restaurant_id"],
        "date": date,
        "is_sent": False
    }
    
    # Staff voit seulement ses catégories
    if current_user["role"] == "staff":
        assigned = current_user.get("assigned_categories", [])
        if not assigned:
            return []
        query["category_id"] = {"$in": assigned}
    
    tasks = await daily_tasks_collection.find(query, {"_id": 0}).to_list(500)
    
    # Enrichir avec le nom de l'utilisateur assigné
    for task in tasks:
        if task.get("assigned_user_id"):
            user = await users_collection.find_one(
                {"user_id": task["assigned_user_id"]},
                {"name": 1, "_id": 0}
            )
            task["assigned_user_name"] = user["name"] if user else None
    
    return tasks

@api_router.put("/daily-tasks/{task_id}/complete")
async def complete_daily_task(
    task_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Marquer une tâche comme terminée"""
    task_doc = await daily_tasks_collection.find_one(
        {"task_id": task_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not task_doc:
        raise HTTPException(status_code=404, detail="Task not found")
    
    # Vérifier l'accès
    if current_user["role"] == "staff":
        assigned = current_user.get("assigned_categories", [])
        is_assigned_to_user = task_doc.get("assigned_user_id") == current_user["user_id"]
        is_in_category = task_doc["category_id"] in assigned
        
        if not (is_assigned_to_user or is_in_category):
            raise HTTPException(status_code=403, detail="Access denied")
    
    await daily_tasks_collection.update_one(
        {"task_id": task_id},
        {"$set": {
            "status": "completed",
            "completed_by": current_user["user_id"],
            "completed_at": datetime.now(timezone.utc)
        }}
    )
    
    # Historique
    history = {
        "history_id": f"hist_{uuid.uuid4().hex[:12]}",
        "task_id": task_id,
        "user_id": current_user["user_id"],
        "user_name": current_user["name"],
        "action": "completed",
        "timestamp": datetime.now(timezone.utc)
    }
    await history_collection.insert_one(history)
    
    updated_task = await daily_tasks_collection.find_one({"task_id": task_id}, {"_id": 0})
    return updated_task

@api_router.put("/daily-tasks/{task_id}/uncomplete")
async def uncomplete_daily_task(
    task_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Remettre une tâche en attente"""
    task_doc = await daily_tasks_collection.find_one(
        {"task_id": task_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not task_doc:
        raise HTTPException(status_code=404, detail="Task not found")
    
    await daily_tasks_collection.update_one(
        {"task_id": task_id},
        {"$set": {
            "status": "pending",
            "completed_by": None,
            "completed_at": None
        }}
    )
    
    updated_task = await daily_tasks_collection.find_one({"task_id": task_id}, {"_id": 0})
    return updated_task

@api_router.delete("/daily-tasks/{task_id}")
async def delete_daily_task(
    task_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une tâche du jour"""
    query = {"task_id": task_id, "restaurant_id": current_user["restaurant_id"]}
    
    # Staff peut supprimer seulement les tâches non envoyées de ses catégories
    if current_user["role"] == "staff":
        assigned = current_user.get("assigned_categories", [])
        query["category_id"] = {"$in": assigned}
        query["is_sent"] = False
    
    result = await daily_tasks_collection.delete_one(query)
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found or cannot be deleted")
    
    return {"message": "Task deleted successfully"}

# ==================== RECURRING TASKS COMPLETION ====================

@api_router.put("/recurring-tasks/{template_id}/complete")
async def complete_recurring_task(
    template_id: str,
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Marquer une tâche récurrente comme terminée pour une date donnée"""
    # Vérifier que le template existe
    template = await task_templates_collection.find_one({
        "template_id": template_id,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Vérifier l'accès pour le staff
    if current_user["role"] == "staff":
        assigned = current_user.get("assigned_categories", [])
        if template["category_id"] not in assigned:
            raise HTTPException(status_code=403, detail="Access denied")
    
    # Vérifier si déjà complété
    existing = await db.mep_template_completions.find_one({
        "template_id": template_id,
        "date": date,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if existing:
        return {"message": "Task already completed for this date", "status": "completed"}
    
    # Créer l'entrée de complétion
    completion = {
        "completion_id": f"comp_{uuid.uuid4().hex[:12]}",
        "template_id": template_id,
        "restaurant_id": current_user["restaurant_id"],
        "date": date,
        "completed_by": current_user["user_id"],
        "completed_at": datetime.now(timezone.utc)
    }
    
    await db.mep_template_completions.insert_one(completion)
    
    # Historique
    history = {
        "history_id": f"hist_{uuid.uuid4().hex[:12]}",
        "template_id": template_id,
        "user_id": current_user["user_id"],
        "user_name": current_user["name"],
        "action": "completed_recurring",
        "date": date,
        "timestamp": datetime.now(timezone.utc)
    }
    await history_collection.insert_one(history)
    
    return {
        "message": "Recurring task completed",
        "template_id": template_id,
        "date": date,
        "status": "completed"
    }

@api_router.put("/recurring-tasks/{template_id}/uncomplete")
async def uncomplete_recurring_task(
    template_id: str,
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Remettre une tâche récurrente en attente pour une date donnée"""
    result = await db.mep_template_completions.delete_one({
        "template_id": template_id,
        "date": date,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if result.deleted_count == 0:
        return {"message": "Task was not completed for this date", "status": "pending"}
    
    return {
        "message": "Recurring task uncompleted",
        "template_id": template_id,
        "date": date,
        "status": "pending"
    }

# ==================== HISTORY ENDPOINT ====================

@api_router.get("/history")
async def get_history(
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Obtenir l'historique des actions"""
    query = {}
    
    if date:
        start_of_day = datetime.fromisoformat(f"{date}T00:00:00+00:00")
        end_of_day = datetime.fromisoformat(f"{date}T23:59:59+00:00")
        query["timestamp"] = {"$gte": start_of_day, "$lte": end_of_day}
    
    tasks = await daily_tasks_collection.find(
        {"restaurant_id": current_user["restaurant_id"]},
        {"task_id": 1, "_id": 0}
    ).to_list(1000)
    task_ids = [t["task_id"] for t in tasks] + ["batch_send"]
    
    query["task_id"] = {"$in": task_ids}
    
    history = await history_collection.find(
        query,
        {"_id": 0}
    ).sort("timestamp", -1).to_list(500)
    
    return history

# ==================== MENU GROUPE MODELS ====================

class CreateMenuSectionRequest(BaseModel):
    name: str
    description: Optional[str] = None
    order: Optional[int] = 0
    price: Optional[float] = None  # Prix de la sous-section (ex: 9.90 pour Entrée 1)
    parent_section_id: Optional[str] = None  # Si défini, c'est une sous-section

class UpdateMenuSectionRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    order: Optional[int] = None
    price: Optional[float] = None  # Prix de la sous-section
    parent_section_id: Optional[str] = None  # Pour déplacer une sous-section

class CreateMenuItemRequest(BaseModel):
    section_id: str
    name: str
    description: Optional[str] = None
    order: Optional[int] = 0
    excel_status: Optional[str] = "added"  # "added", "deleted", "modified", or "normal"
    modified_fields: Optional[List[str]] = None  # ["name", "price", "description"]

class UpdateMenuItemRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    order: Optional[int] = None
    excel_status: Optional[str] = None  # "added", "deleted", "modified", or "normal"
    modified_fields: Optional[List[str]] = None  # ["name", "price", "description"]

class CustomOption(BaseModel):
    name: str
    price: Optional[float] = None  # Prix unitaire TTC
    quantity: Optional[int] = 1  # Quantité (défaut: 1)
    tva_rate: Optional[float] = 20.0  # Taux de TVA: 0, 10 ou 20 (défaut: 20%)

class CreateGroupReservationRequest(BaseModel):
    client_name: str
    client_surname: str
    client_company: Optional[str] = None  # Société (optionnel)
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    # Adresse client (optionnel - pour facturation)
    client_address_street: Optional[str] = None
    client_address_postal_code: Optional[str] = None
    client_address_city: Optional[str] = None
    # Crédit client
    is_credit_client: Optional[bool] = False  # Si true, le client paie plus tard
    num_people: int
    date: str  # Format: YYYY-MM-DD
    time: str  # Format: HH:MM
    selected_sections: Optional[List[str]] = None  # section_ids (optionnel si custom_options uniquement)
    selected_items: Optional[Dict[str, List[str]]] = None  # {section_id: [item_ids]} (optionnel si custom_options uniquement)
    price_per_person: Optional[float] = None
    custom_options: Optional[List[CustomOption]] = None  # Options personnalisées (ex: Privatisation 300€)
    # Statut de la proposition (pour les menus avec proposition)
    # Valeurs: to_send, sent, validated, to_invoice, invoiced, paid
    proposal_status: Optional[str] = "to_send"

class UpdateGroupReservationRequest(BaseModel):
    client_name: Optional[str] = None
    client_surname: Optional[str] = None
    client_company: Optional[str] = None  # Société (optionnel)
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    # Adresse client (optionnel - pour facturation)
    client_address_street: Optional[str] = None
    client_address_postal_code: Optional[str] = None
    client_address_city: Optional[str] = None
    # Crédit client
    is_credit_client: Optional[bool] = None  # Si true, le client paie plus tard
    num_people: Optional[int] = None
    date: Optional[str] = None
    time: Optional[str] = None
    selected_sections: Optional[List[str]] = None
    selected_items: Optional[Dict[str, List[str]]] = None
    price_per_person: Optional[float] = None
    client_selections: Optional[Dict] = None  # Permet de modifier les sélections du client
    status: Optional[str] = None
    custom_options: Optional[List[CustomOption]] = None  # Options personnalisées
    # Statut de la proposition (pour les menus avec proposition)
    # Valeurs: to_send, sent, validated, to_invoice, invoiced, paid
    proposal_status: Optional[str] = None
    # Réponse du client en cas de refus
    client_rejection_reason: Optional[str] = None

# ==================== GROUP OPTIONS MODELS (Options configurables par restaurant) ====================

class CreateGroupOptionRequest(BaseModel):
    """Option configurable pour les réservations groupe (DJ, Gâteau, etc.)"""
    name: str  # Nom de l'option (ex: DJ, Gâteau anniversaire)
    description: Optional[str] = None  # Description optionnelle
    price: Optional[float] = None  # Prix de l'option
    is_free_text: bool = False  # Si true, le client peut entrer du texte libre (ex: "Gâteau personnalisé")
    order: Optional[int] = None

class UpdateGroupOptionRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    price: Optional[float] = None
    is_free_text: Optional[bool] = None
    order: Optional[int] = None
    is_active: Optional[bool] = None

# ==================== PUBLIC GROUP RESERVATION REQUEST (Demande client via lien) ====================

class ClientOptionSelection(BaseModel):
    """Option sélectionnée par le client"""
    option_id: str
    option_name: str
    quantity: int = 1
    free_text: Optional[str] = None  # Texte libre si l'option le permet

class PublicGroupReservationRequest(BaseModel):
    """Demande de réservation faite par un client via le lien public"""
    client_name: str
    client_surname: str
    client_company: Optional[str] = None
    client_email: str  # Email obligatoire pour le suivi
    client_phone: Optional[str] = None
    num_people: int
    preferred_date: Optional[str] = None  # Date souhaitée (format YYYY-MM-DD)
    preferred_time: Optional[str] = None  # Heure souhaitée (format HH:MM)
    selected_sections: List[str] = []  # Formules sélectionnées (Entrée, Plat, Dessert, etc.)
    selected_options: List[ClientOptionSelection] = []  # Options sélectionnées (DJ, Gâteau, etc.)
    message: Optional[str] = None  # Message ou demande particulière

class ClientProposalResponse(BaseModel):
    """Réponse du client à une proposition (accepter ou refuser)"""
    accepted: bool
    rejection_reason: Optional[str] = None  # Raison du refus si accepted=False

# ==================== INVOICE/QUOTE MODELS (Factures et Devis) ====================

class InvoiceItem(BaseModel):
    """Ligne de facture/devis"""
    name: str  # Nom du produit
    description: Optional[str] = None
    quantity: int = 1
    unit_price_ttc: float  # Prix unitaire TTC
    tva_rate: float = 10.0  # Taux TVA (10 ou 20)

class InvoiceClient(BaseModel):
    """Informations client pour la facture"""
    name: str  # Nom ou Prénom Nom
    company: Optional[str] = None  # Nom de société
    address: Optional[str] = None  # Adresse complète
    email: Optional[str] = None
    phone: Optional[str] = None

class CreateInvoiceRequest(BaseModel):
    """Créer un devis ou une facture"""
    type: str  # "quote" (devis) ou "invoice" (facture)
    client: InvoiceClient
    items: List[InvoiceItem]
    notes: Optional[str] = None  # Notes ou conditions
    valid_until: Optional[str] = None  # Date de validité (pour devis)
    due_date: Optional[str] = None  # Date d'échéance (pour facture)

class UpdateInvoiceRequest(BaseModel):
    """Mettre à jour un devis ou une facture"""
    client: Optional[InvoiceClient] = None
    items: Optional[List[InvoiceItem]] = None
    notes: Optional[str] = None
    valid_until: Optional[str] = None
    due_date: Optional[str] = None
    status: Optional[str] = None  # draft, sent, accepted, rejected, paid

# ==================== PRIVATISATION SPACES MODELS ====================

class CreatePrivatisationSpaceRequest(BaseModel):
    """Créer un espace de privatisation (Bibliothèque, Pergola, etc.)"""
    name: str  # Ex: Bibliothèque, Pergola, Terrasse
    description: Optional[str] = None
    capacity_min: Optional[int] = None  # Capacité minimum
    capacity_max: Optional[int] = None  # Capacité maximum
    photos: Optional[List[str]] = []  # Liste de photos en base64
    amenities: Optional[List[str]] = []  # Équipements: écran, sono, etc.
    price_info: Optional[str] = None  # Info sur le prix (ex: "Sur devis", "500€/soirée")
    price_under_minimum: Optional[float] = None  # Prix si nombre personnes < capacity_min
    is_active: bool = True

class UpdatePrivatisationSpaceRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    capacity_min: Optional[int] = None
    capacity_max: Optional[int] = None
    photos: Optional[List[str]] = None
    amenities: Optional[List[str]] = None
    price_info: Optional[str] = None
    price_under_minimum: Optional[float] = None  # Prix si nombre personnes < capacity_min
    is_active: Optional[bool] = None

# ==================== MENU SECTIONS ENDPOINTS ====================

@api_router.post("/menu-sections/create")
async def create_menu_section(
    create_request: CreateMenuSectionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une section de menu (admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Déterminer le parent pour calculer l'ordre
    parent_filter = {"restaurant_id": current_user["restaurant_id"]}
    if create_request.parent_section_id:
        parent_filter["parent_section_id"] = create_request.parent_section_id
    else:
        parent_filter["parent_section_id"] = None
    
    # Get max order for sections at same level
    max_order_doc = await menu_sections_collection.find_one(
        parent_filter,
        sort=[("order", -1)]
    )
    max_order = max_order_doc["order"] + 1 if max_order_doc and max_order_doc.get("order") is not None else 0
    
    section = {
        "section_id": f"sec_{uuid.uuid4().hex[:12]}",
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "description": create_request.description,
        "price": create_request.price,  # Prix de la sous-section uniquement
        "parent_section_id": create_request.parent_section_id,  # None = section principale
        "order": create_request.order if create_request.order else max_order,
        "created_at": datetime.now(timezone.utc)
    }
    
    await menu_sections_collection.insert_one(section)
    del section["_id"]
    
    return section

@api_router.get("/menu-sections/list")
async def list_menu_sections(current_user: dict = Depends(get_current_user)):
    """Lister toutes les sections de menu"""
    sections = await menu_sections_collection.find(
        {"restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    
    return sections

@api_router.put("/menu-sections/{section_id}")
async def update_menu_section(
    section_id: str,
    update_request: UpdateMenuSectionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour une section de menu"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Construire les données de mise à jour
    # Ne pas filtrer description même si elle est None (pour permettre de la vider)
    update_data = {}
    request_dict = update_request.dict()
    
    if request_dict.get("name") is not None:
        update_data["name"] = request_dict["name"]
    
    # Toujours inclure description si elle est présente dans la requête
    # (même si elle est None, ce qui signifie qu'on veut la vider)
    if "description" in request_dict:
        update_data["description"] = request_dict["description"]
    
    if request_dict.get("price") is not None:
        update_data["price"] = request_dict["price"]
    elif "price" in request_dict and request_dict["price"] is None:
        # Permettre de supprimer le prix
        update_data["price"] = None
    
    if update_data:
        await menu_sections_collection.update_one(
            {"section_id": section_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    section_doc = await menu_sections_collection.find_one(
        {"section_id": section_id},
        {"_id": 0}
    )
    
    return section_doc

@api_router.delete("/menu-sections/{section_id}")
async def delete_menu_section(
    section_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une section de menu et ses items"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await menu_sections_collection.delete_one(
        {"section_id": section_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Section not found")
    
    # Supprimer tous les items de cette section
    await menu_items_collection.delete_many({"section_id": section_id})
    
    return {"message": "Section deleted successfully"}

@api_router.put("/menu-sections/{section_id}/reorder")
async def reorder_menu_section(
    section_id: str,
    direction: str,  # "up" ou "down"
    current_user: dict = Depends(get_current_user)
):
    """Réordonner une section de menu (monter ou descendre)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Récupérer la section actuelle
    current_section = await menu_sections_collection.find_one({
        "section_id": section_id,
        "restaurant_id": current_user["restaurant_id"]
    })
    if not current_section:
        raise HTTPException(status_code=404, detail="Section not found")
    
    current_order = current_section.get("order", 0)
    
    # Récupérer toutes les sections triées par ordre
    all_sections = await menu_sections_collection.find(
        {"restaurant_id": current_user["restaurant_id"]}
    ).sort("order", 1).to_list(100)
    
    # Trouver l'index de la section actuelle
    current_index = next((i for i, s in enumerate(all_sections) if s["section_id"] == section_id), -1)
    
    if current_index == -1:
        raise HTTPException(status_code=404, detail="Section not found")
    
    # Déterminer l'index de la section à échanger
    if direction == "up":
        if current_index == 0:
            return {"message": "Section already at top", "sections": []}
        swap_index = current_index - 1
    elif direction == "down":
        if current_index == len(all_sections) - 1:
            return {"message": "Section already at bottom", "sections": []}
        swap_index = current_index + 1
    else:
        raise HTTPException(status_code=400, detail="Direction must be 'up' or 'down'")
    
    swap_section = all_sections[swap_index]
    swap_order = swap_section.get("order", 0)
    
    # Échanger les ordres
    await menu_sections_collection.update_one(
        {"section_id": section_id},
        {"$set": {"order": swap_order}}
    )
    await menu_sections_collection.update_one(
        {"section_id": swap_section["section_id"]},
        {"$set": {"order": current_order}}
    )
    
    # Retourner les sections mises à jour
    updated_sections = await menu_sections_collection.find(
        {"restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    
    return {
        "message": f"Section moved {direction}",
        "sections": updated_sections
    }

# ==================== MENU ITEMS ENDPOINTS ====================

@api_router.post("/menu-items/create")
async def create_menu_item(
    create_request: CreateMenuItemRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer un item de menu dans une section"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que la section existe
    section = await menu_sections_collection.find_one({
        "section_id": create_request.section_id,
        "restaurant_id": current_user["restaurant_id"]
    })
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")
    
    # Get max order for this section
    max_order_doc = await menu_items_collection.find_one(
        {"section_id": create_request.section_id},
        sort=[("order", -1)]
    )
    max_order = max_order_doc["order"] + 1 if max_order_doc else 0
    
    item = {
        "item_id": f"item_{uuid.uuid4().hex[:12]}",
        "section_id": create_request.section_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "description": create_request.description,
        "order": create_request.order if create_request.order else max_order,
        "created_at": datetime.now(timezone.utc)
    }
    
    await menu_items_collection.insert_one(item)
    del item["_id"]
    
    return item

@api_router.get("/menu-items/list")
async def list_menu_items(
    section_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les items de menu (optionnellement filtré par section)"""
    query = {"restaurant_id": current_user["restaurant_id"]}
    if section_id:
        query["section_id"] = section_id
    
    items = await menu_items_collection.find(
        query,
        {"_id": 0}
    ).sort("order", 1).to_list(500)
    
    return items

@api_router.put("/menu-items/{item_id}")
async def update_menu_item(
    item_id: str,
    update_request: UpdateMenuItemRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour un item de menu"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await menu_items_collection.update_one(
            {"item_id": item_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    item_doc = await menu_items_collection.find_one(
        {"item_id": item_id},
        {"_id": 0}
    )
    
    return item_doc

@api_router.delete("/menu-items/{item_id}")
async def delete_menu_item(
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer un item de menu"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await menu_items_collection.delete_one(
        {"item_id": item_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    
    return {"message": "Item deleted successfully"}

# ==================== GROUP RESERVATIONS ENDPOINTS ====================

@api_router.post("/group-reservations/create")
async def create_group_reservation(
    create_request: CreateGroupReservationRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une réservation de groupe"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Déterminer si un menu est sélectionné
    has_menu = bool(create_request.selected_sections and len(create_request.selected_sections) > 0)
    
    # Générer un token uniquement si un menu est sélectionné
    client_token = secrets.token_urlsafe(32) if has_menu else None
    
    # Si pas de menu, la réservation est automatiquement confirmée (facture directe)
    initial_status = "pending" if has_menu else "confirmed"
    
    # Statut de la proposition (pour les menus avec proposition)
    # Valeurs: to_send, sent, validated, to_invoice, invoiced, paid
    initial_proposal_status = create_request.proposal_status if create_request.proposal_status else "to_send"
    
    reservation = {
        "reservation_id": f"grp_{uuid.uuid4().hex[:12]}",
        "restaurant_id": current_user["restaurant_id"],
        "client_name": create_request.client_name,
        "client_surname": create_request.client_surname,
        "client_company": create_request.client_company,  # Société (optionnel)
        "client_email": create_request.client_email,
        "client_phone": create_request.client_phone,
        # Adresse client (optionnel - pour facturation)
        "client_address_street": create_request.client_address_street,
        "client_address_postal_code": create_request.client_address_postal_code,
        "client_address_city": create_request.client_address_city,
        # Crédit client
        "is_credit_client": create_request.is_credit_client or False,
        "num_people": create_request.num_people,
        "date": create_request.date,
        "time": create_request.time,
        "selected_sections": create_request.selected_sections or [],
        "selected_items": create_request.selected_items or {},
        "price_per_person": create_request.price_per_person,
        "custom_options": [opt.model_dump() for opt in create_request.custom_options] if create_request.custom_options else [],
        "client_token": client_token,
        "client_selections": None,  # Sera rempli par le client (si menu)
        "status": initial_status,
        "proposal_status": initial_proposal_status,  # Statut de la proposition
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user["user_id"]
    }
    
    await group_reservations_collection.insert_one(reservation)
    del reservation["_id"]
    
    # Générer le lien pour le client uniquement si un menu est sélectionné
    if has_menu:
        reservation["client_link"] = f"{FRONTEND_URL}?group_token={client_token}"
    else:
        reservation["client_link"] = None  # Pas de lien client, facture directe
    
    # Envoyer une notification push à tout le restaurant
    await send_push_notification_to_restaurant(
        restaurant_id=current_user["restaurant_id"],
        title="Nouvelle réservation groupe",
        body=f"{create_request.client_name} - {create_request.num_people} personnes le {create_request.date}",
        data={"type": "reservation", "reservation_id": reservation["reservation_id"]},
        exclude_user=current_user["user_id"]
    )
    
    return reservation

@api_router.get("/group-reservations/list")
async def list_group_reservations(
    include_archived: bool = False,
    include_deleted: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Lister toutes les réservations de groupe
    Par défaut, exclut les réservations archivées et supprimées
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Build filter
    filter_query = {"restaurant_id": current_user["restaurant_id"]}
    
    if not include_archived and not include_deleted:
        # Default: exclude archived and deleted
        filter_query["$or"] = [
            {"is_archived": {"$ne": True}, "is_deleted": {"$ne": True}},
            {"is_archived": {"$exists": False}, "is_deleted": {"$exists": False}}
        ]
        # Simplified: just exclude both
        filter_query = {
            "restaurant_id": current_user["restaurant_id"],
            "is_archived": {"$ne": True},
            "is_deleted": {"$ne": True}
        }
    elif include_archived and not include_deleted:
        # Only archived
        filter_query["is_archived"] = True
        filter_query["is_deleted"] = {"$ne": True}
    elif include_deleted and not include_archived:
        # Only deleted
        filter_query["is_deleted"] = True
    # else: include all
    
    reservations = await group_reservations_collection.find(
        filter_query,
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return reservations

@api_router.put("/group-reservations/{reservation_id}/archive")
async def toggle_archive_reservation(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Archiver ou désarchiver une réservation"""
    print(f"[ARCHIVE] User: {current_user.get('email')}, role: {current_user.get('role')}, restaurant: {current_user.get('restaurant_id')}")
    
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    print(f"[ARCHIVE] Looking for reservation_id={reservation_id}, restaurant_id={current_user['restaurant_id']}")
    
    # Get the full reservation without projection to check if it exists
    reservation = await group_reservations_collection.find_one(
        {"reservation_id": reservation_id, "restaurant_id": current_user["restaurant_id"]},
        {"_id": 0, "is_archived": 1, "client_name": 1}
    )
    
    print(f"[ARCHIVE] Found: {reservation}")
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    new_archived_status = not reservation.get("is_archived", False)
    
    await group_reservations_collection.update_one(
        {"reservation_id": reservation_id},
        {"$set": {"is_archived": new_archived_status, "updated_at": datetime.now(timezone.utc)}}
    )
    
    action = "archivée" if new_archived_status else "désarchivée"
    return {"message": f"Réservation {action}", "is_archived": new_archived_status}

@api_router.put("/group-reservations/{reservation_id}/restore")
async def restore_deleted_reservation(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Restaurer une réservation supprimée"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await group_reservations_collection.update_one(
        {"reservation_id": reservation_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_deleted": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    return {"message": "Réservation restaurée"}

@api_router.delete("/group-reservations/{reservation_id}/permanent")
async def permanently_delete_reservation(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer définitivement une réservation (depuis la corbeille)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await group_reservations_collection.delete_one(
        {"reservation_id": reservation_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    return {"message": "Réservation supprimée définitivement"}

@api_router.get("/group-reservations/{reservation_id}")
async def get_group_reservation(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtenir les détails d'une réservation"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    reservation = await group_reservations_collection.find_one(
        {"reservation_id": reservation_id, "restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    )
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    return reservation

@api_router.delete("/group-reservations/{reservation_id}")
async def delete_group_reservation(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Marquer une réservation comme supprimée (soft delete)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await group_reservations_collection.update_one(
        {"reservation_id": reservation_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    return {"message": "Réservation déplacée vers la corbeille"}

@api_router.put("/group-reservations/{reservation_id}")
async def update_group_reservation(
    reservation_id: str,
    update_request: UpdateGroupReservationRequest,
    current_user: dict = Depends(get_current_user)
):
    """Modifier une réservation de groupe (même après validation client)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que la réservation existe
    existing = await group_reservations_collection.find_one(
        {"reservation_id": reservation_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not existing:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    # Construire les champs à mettre à jour
    update_fields = {}
    if update_request.client_name is not None:
        update_fields["client_name"] = update_request.client_name
    if update_request.client_surname is not None:
        update_fields["client_surname"] = update_request.client_surname
    if update_request.client_company is not None:
        update_fields["client_company"] = update_request.client_company
    if update_request.client_email is not None:
        update_fields["client_email"] = update_request.client_email
    if update_request.client_phone is not None:
        update_fields["client_phone"] = update_request.client_phone
    # Adresse client
    if update_request.client_address_street is not None:
        update_fields["client_address_street"] = update_request.client_address_street
    if update_request.client_address_postal_code is not None:
        update_fields["client_address_postal_code"] = update_request.client_address_postal_code
    if update_request.client_address_city is not None:
        update_fields["client_address_city"] = update_request.client_address_city
    # Crédit client
    if update_request.is_credit_client is not None:
        update_fields["is_credit_client"] = update_request.is_credit_client
    if update_request.num_people is not None:
        update_fields["num_people"] = update_request.num_people
    if update_request.date is not None:
        update_fields["date"] = update_request.date
    if update_request.time is not None:
        update_fields["time"] = update_request.time
    if update_request.selected_sections is not None:
        update_fields["selected_sections"] = update_request.selected_sections
    if update_request.selected_items is not None:
        update_fields["selected_items"] = update_request.selected_items
    if update_request.price_per_person is not None:
        update_fields["price_per_person"] = update_request.price_per_person
    if update_request.client_selections is not None:
        update_fields["client_selections"] = update_request.client_selections
    if update_request.status is not None:
        update_fields["status"] = update_request.status
    # Statut de la proposition
    if update_request.proposal_status is not None:
        update_fields["proposal_status"] = update_request.proposal_status
    if update_request.custom_options is not None:
        update_fields["custom_options"] = [opt.model_dump() for opt in update_request.custom_options]
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_fields["updated_at"] = datetime.now(timezone.utc)
    update_fields["updated_by"] = current_user["user_id"]
    
    await group_reservations_collection.update_one(
        {"reservation_id": reservation_id},
        {"$set": update_fields}
    )
    
    # Récupérer la réservation mise à jour
    updated = await group_reservations_collection.find_one(
        {"reservation_id": reservation_id},
        {"_id": 0}
    )
    
    return updated

# ==================== CLIENT PUBLIC ENDPOINTS (No Auth) ====================

@api_router.get("/public/group/{client_token}")
async def get_group_for_client(client_token: str):
    """Obtenir les détails d'un groupe pour le client (via lien)"""
    reservation = await group_reservations_collection.find_one(
        {"client_token": client_token},
        {"_id": 0, "client_token": 0}
    )
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Group not found")
    
    # Récupérer les informations du restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": reservation["restaurant_id"]},
        {"_id": 0}
    )
    
    # Récupérer les sections et items sélectionnés
    sections = await menu_sections_collection.find(
        {"section_id": {"$in": reservation["selected_sections"]}},
        {"_id": 0}
    ).to_list(100)
    
    all_item_ids = []
    for section_items in reservation["selected_items"].values():
        all_item_ids.extend(section_items)
    
    items = await menu_items_collection.find(
        {"item_id": {"$in": all_item_ids}},
        {"_id": 0}
    ).to_list(500)
    
    return {
        "reservation": reservation,
        "restaurant": restaurant,
        "sections": sections,
        "items": items
    }

@api_router.post("/public/group/{client_token}/submit")
async def submit_client_selections(
    client_token: str,
    request: Request
):
    """Soumettre les sélections du client"""
    # Récupérer les sélections du body JSON
    selections = await request.json()
    
    reservation = await group_reservations_collection.find_one(
        {"client_token": client_token}
    )
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Group not found")
    
    if reservation["status"] == "client_submitted":
        raise HTTPException(status_code=400, detail="Selections already submitted")
    
    # Mettre à jour avec les sélections du client
    await group_reservations_collection.update_one(
        {"client_token": client_token},
        {
            "$set": {
                "client_selections": selections,
                "status": "client_submitted",
                "submitted_at": datetime.now(timezone.utc)
            }
        }
    )
    
    # TODO: Envoyer une notification au restaurant
    
    return {"message": "Selections submitted successfully"}

# ==================== GROUP OPTIONS ENDPOINTS (Options configurables par restaurant) ====================

@api_router.post("/group-options/create")
async def create_group_option(
    create_request: CreateGroupOptionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une option de groupe configurable (DJ, Gâteau, etc.)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get max order
    max_order_doc = await group_options_collection.find_one(
        {"restaurant_id": current_user["restaurant_id"]},
        sort=[("order", -1)]
    )
    max_order = max_order_doc["order"] + 1 if max_order_doc else 0
    
    option = {
        "option_id": f"gopt_{uuid.uuid4().hex[:12]}",
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "description": create_request.description,
        "price": create_request.price,
        "is_free_text": create_request.is_free_text,
        "order": create_request.order if create_request.order is not None else max_order,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    
    await group_options_collection.insert_one(option)
    del option["_id"]
    
    return option

@api_router.get("/group-options/list")
async def list_group_options(current_user: dict = Depends(get_current_user)):
    """Lister les options de groupe"""
    options = await group_options_collection.find(
        {"restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    
    return options

@api_router.put("/group-options/{option_id}")
async def update_group_option(
    option_id: str,
    update_request: UpdateGroupOptionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour une option de groupe"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await group_options_collection.update_one(
            {"option_id": option_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    option = await group_options_collection.find_one(
        {"option_id": option_id},
        {"_id": 0}
    )
    
    return option

@api_router.delete("/group-options/{option_id}")
async def delete_group_option(
    option_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une option de groupe"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    await group_options_collection.update_one(
        {"option_id": option_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    return {"message": "Option deleted"}

# ==================== PUBLIC GROUP RESERVATION LINK ====================

@api_router.get("/public/restaurant/{restaurant_id}/group-form")
async def get_public_group_form_data(restaurant_id: str):
    """Obtenir les données du formulaire public de réservation groupe"""
    # Récupérer le restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "restaurant_id": 1, "name": 1, "logo_base64": 1, "primary_color": 1, "secondary_color": 1}
    )
    
    if not restaurant:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    # Récupérer TOUTES les sections de menu avec leurs infos complètes (incluant parent_section_id)
    all_sections = await menu_sections_collection.find(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "section_id": 1, "name": 1, "description": 1, "price": 1, "order": 1, "parent_section_id": 1}
    ).sort("order", 1).to_list(100)
    
    # Récupérer les items (plats) de toutes les sections
    all_items = await menu_items_collection.find(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "item_id": 1, "section_id": 1, "name": 1, "description": 1, "order": 1}
    ).sort("order", 1).to_list(500)
    
    # Grouper les items par section
    items_by_section = {}
    for item in all_items:
        section_id = item.get("section_id")
        if section_id not in items_by_section:
            items_by_section[section_id] = []
        items_by_section[section_id].append(item)
    
    # Séparer les sections parents (sans parent_section_id ou avec parent_section_id = None/ROOT) 
    # des sous-sections (avec parent_section_id valide)
    parent_sections = []
    sub_sections_by_parent = {}
    
    for section in all_sections:
        parent_id = section.get("parent_section_id")
        # Si pas de parent OU parent = None OU parent = "ROOT" => c'est une section parent
        if not parent_id or parent_id == "ROOT":
            parent_sections.append(section)
        else:
            # C'est une sous-section
            if parent_id not in sub_sections_by_parent:
                sub_sections_by_parent[parent_id] = []
            sub_sections_by_parent[parent_id].append(section)
    
    # Construire la structure hiérarchique
    hierarchical_sections = []
    for parent in sorted(parent_sections, key=lambda x: x.get("order", 0)):
        parent_id = parent.get("section_id")
        sub_sections = sub_sections_by_parent.get(parent_id, [])
        
        # Trier les sous-sections par order
        sub_sections = sorted(sub_sections, key=lambda x: x.get("order", 0))
        
        # Ajouter les items à chaque sous-section
        for sub in sub_sections:
            sub["items"] = items_by_section.get(sub["section_id"], [])
        
        # Ajouter les items directement dans la section parent (si elle n'a pas de sous-sections)
        parent["items"] = items_by_section.get(parent_id, [])
        parent["sub_sections"] = sub_sections
        
        hierarchical_sections.append(parent)
    
    # Récupérer les options configurables (DJ, Gâteau, etc.) - AVEC PRIX
    options = await group_options_collection.find(
        {"restaurant_id": restaurant_id, "is_active": True},
        {"_id": 0, "option_id": 1, "name": 1, "description": 1, "price": 1, "is_free_text": 1, "order": 1}
    ).sort("order", 1).to_list(100)
    
    return {
        "restaurant": restaurant,
        "sections": hierarchical_sections,
        "options": options
    }

@api_router.post("/public/restaurant/{restaurant_id}/group-request")
async def create_public_group_request(
    restaurant_id: str,
    request: PublicGroupReservationRequest
):
    """Créer une demande de réservation groupe depuis le formulaire public"""
    # Vérifier que le restaurant existe
    restaurant = await restaurants_collection.find_one({"restaurant_id": restaurant_id})
    if not restaurant:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    # Valider les sections sélectionnées
    valid_sections = await menu_sections_collection.find(
        {"restaurant_id": restaurant_id, "section_id": {"$in": request.selected_sections}},
        {"_id": 0}
    ).to_list(100)
    
    # Calculer le prix estimé par personne (basé sur les sections sélectionnées)
    estimated_price = sum(s.get("price", 0) or 0 for s in valid_sections)
    
    # Créer un token unique pour le suivi
    client_token = secrets.token_urlsafe(32)
    
    # Statut initial: demande à traiter
    reservation = {
        "reservation_id": f"grp_{uuid.uuid4().hex[:12]}",
        "restaurant_id": restaurant_id,
        "client_name": request.client_name,
        "client_surname": request.client_surname,
        "client_company": request.client_company,
        "client_email": request.client_email,
        "client_phone": request.client_phone,
        "num_people": request.num_people,
        "date": request.preferred_date,
        "time": request.preferred_time,
        "selected_sections": request.selected_sections,
        "selected_items": {},  # Sera rempli par le restaurant
        "price_per_person": estimated_price if estimated_price > 0 else None,
        # Options sélectionnées par le client (SANS prix, le restaurant ajoutera les prix)
        "client_requested_options": [opt.dict() for opt in request.selected_options],
        "custom_options": [],  # Sera rempli par le restaurant avec les prix
        "client_message": request.message,
        "client_token": client_token,
        "client_selections": None,
        # Statuts
        "status": "request_pending",  # demande à traiter
        "proposal_status": "to_send",  # Proposition à envoyer
        # Historique des échanges
        "history": [{
            "action": "client_request",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "message": "Demande de réservation reçue"
        }],
        "created_at": datetime.now(timezone.utc),
        "created_by": "client_public"  # Créé par le client via le formulaire public
    }
    
    await group_reservations_collection.insert_one(reservation)
    del reservation["_id"]
    
    # Générer le lien de suivi pour le client
    reservation["tracking_link"] = f"{FRONTEND_URL}?track_group={client_token}"
    
    return {
        "message": "Votre demande a été envoyée ! Le restaurant va traiter votre demande et vous recevrez une proposition.",
        "reservation_id": reservation["reservation_id"],
        "tracking_link": reservation["tracking_link"]
    }

@api_router.post("/group-reservations/{reservation_id}/send-proposal")
async def send_proposal_to_client(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Envoyer la proposition au client (après modifications du restaurant)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    reservation = await group_reservations_collection.find_one(
        {"reservation_id": reservation_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    # Mettre à jour le statut
    await group_reservations_collection.update_one(
        {"reservation_id": reservation_id},
        {
            "$set": {
                "status": "proposal_sent",
                "proposal_status": "sent",
                "proposal_sent_at": datetime.now(timezone.utc)
            },
            "$push": {
                "history": {
                    "action": "proposal_sent",
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "message": "Proposition envoyée au client",
                    "by": current_user["name"]
                }
            }
        }
    )
    
    # TODO: Envoyer un email au client avec le lien de validation
    
    return {"message": "Proposition envoyée au client"}

@api_router.get("/public/track/{client_token}")
async def get_reservation_tracking(client_token: str):
    """Obtenir le statut de suivi d'une réservation (pour le client)"""
    reservation = await group_reservations_collection.find_one(
        {"client_token": client_token},
        {"_id": 0}
    )
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    # Récupérer les informations du restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": reservation["restaurant_id"]},
        {"_id": 0, "name": 1, "logo_base64": 1, "primary_color": 1, "secondary_color": 1, "phone": 1, "email": 1}
    )
    
    # Récupérer les sections sélectionnées
    sections = await menu_sections_collection.find(
        {"section_id": {"$in": reservation.get("selected_sections", [])}},
        {"_id": 0}
    ).to_list(100)
    
    # Ne pas exposer le token dans la réponse
    reservation.pop("client_token", None)
    
    return {
        "reservation": reservation,
        "restaurant": restaurant,
        "sections": sections
    }

@api_router.post("/public/track/{client_token}/respond")
async def client_respond_to_proposal(
    client_token: str,
    response: ClientProposalResponse
):
    """Le client accepte ou refuse la proposition"""
    reservation = await group_reservations_collection.find_one(
        {"client_token": client_token}
    )
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    if reservation.get("status") != "proposal_sent":
        raise HTTPException(status_code=400, detail="Aucune proposition en attente de réponse")
    
    if response.accepted:
        # Client accepte
        new_status = "accepted"
        new_proposal_status = "validated"
        history_message = "Client a accepté la proposition"
    else:
        # Client refuse
        new_status = "rejected"
        new_proposal_status = "rejected"
        history_message = f"Client a refusé la proposition. Raison: {response.rejection_reason or 'Non spécifiée'}"
    
    await group_reservations_collection.update_one(
        {"client_token": client_token},
        {
            "$set": {
                "status": new_status,
                "proposal_status": new_proposal_status,
                "client_response_at": datetime.now(timezone.utc),
                "client_rejection_reason": response.rejection_reason if not response.accepted else None
            },
            "$push": {
                "history": {
                    "action": "client_response",
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "message": history_message,
                    "accepted": response.accepted
                }
            }
        }
    )
    
    if response.accepted:
        return {"message": "Merci ! Votre réservation est confirmée. Le restaurant vous contactera pour la suite."}
    else:
        return {"message": "Votre réponse a été envoyée. Le restaurant va revoir la proposition."}

@api_router.post("/group-reservations/{reservation_id}/mark-invoice")
async def mark_reservation_to_invoice(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Marquer une réservation comme à facturer"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    await group_reservations_collection.update_one(
        {"reservation_id": reservation_id, "restaurant_id": current_user["restaurant_id"]},
        {
            "$set": {
                "proposal_status": "to_invoice"
            },
            "$push": {
                "history": {
                    "action": "status_change",
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "message": "Marqué comme à facturer",
                    "by": current_user["name"]
                }
            }
        }
    )
    
    return {"message": "Réservation marquée comme à facturer"}

@api_router.post("/group-reservations/{reservation_id}/mark-invoiced")
async def mark_reservation_invoiced(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Marquer une réservation comme facturée"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    await group_reservations_collection.update_one(
        {"reservation_id": reservation_id, "restaurant_id": current_user["restaurant_id"]},
        {
            "$set": {
                "proposal_status": "invoiced",
                "invoiced_at": datetime.now(timezone.utc)
            },
            "$push": {
                "history": {
                    "action": "invoiced",
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "message": "Facture générée",
                    "by": current_user["name"]
                }
            }
        }
    )
    
    return {"message": "Réservation facturée"}

@api_router.post("/group-reservations/{reservation_id}/mark-paid")
async def mark_reservation_paid(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Marquer une réservation comme payée"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    await group_reservations_collection.update_one(
        {"reservation_id": reservation_id, "restaurant_id": current_user["restaurant_id"]},
        {
            "$set": {
                "proposal_status": "paid",
                "paid_at": datetime.now(timezone.utc)
            },
            "$push": {
                "history": {
                    "action": "paid",
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "message": "Paiement reçu",
                    "by": current_user["name"]
                }
            }
        }
    )
    
    return {"message": "Réservation marquée comme payée"}

# ==================== PDF GENERATION ENDPOINTS ====================

def safe_text(text):
    """Convertir le texte en format compatible latin-1 pour FPDF"""
    if text is None:
        return ""
    # Remplacer les caractères spéciaux non supportés par latin-1
    text = str(text)
    # Ligatures françaises
    text = text.replace('œ', 'oe').replace('Œ', 'OE')
    text = text.replace('æ', 'ae').replace('Æ', 'AE')
    # Autres caractères problématiques
    text = text.replace('…', '...').replace('–', '-').replace('—', '-')
    # Apostrophes et guillemets courbes - utiliser codes Unicode explicites
    text = text.replace('\u2019', "'")  # Right single quotation mark
    text = text.replace('\u2018', "'")  # Left single quotation mark
    text = text.replace('\u201C', '"')  # Left double quotation mark
    text = text.replace('\u201D', '"')  # Right double quotation mark
    text = text.replace('\u0027', "'")  # Apostrophe standard
    text = text.replace('\u2032', "'")  # Prime
    text = text.replace('\u00B4', "'")  # Acute accent
    text = text.replace('\u0060', "'")  # Grave accent
    return text.encode('latin-1', 'replace').decode('latin-1')

def hex_to_rgb(hex_color: str) -> tuple:
    """Convertir une couleur hexadécimale en RGB"""
    hex_color = hex_color.lstrip('#')
    if len(hex_color) == 6:
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    return (38, 37, 45)  # Défaut: #26252D

def apply_blue_tint_to_logo(logo_data: bytes, target_color: tuple = (26, 58, 92)) -> bytes:
    """Appliquer une teinte bleue RAL 5008 au logo
    Transforme toutes les couleurs non-transparentes vers la couleur cible en préservant la luminosité"""
    try:
        img = Image.open(BytesIO(logo_data))
        # Convertir en RGBA pour gérer la transparence
        img = img.convert("RGBA")
        pixels = img.load()
        
        for i in range(img.width):
            for j in range(img.height):
                r, g, b, a = pixels[i, j]
                if a > 0:  # Pixel non transparent
                    # Calculer la luminosité originale
                    luminosity = (r * 0.299 + g * 0.587 + b * 0.114) / 255
                    # Appliquer la couleur cible en modulant par la luminosité
                    new_r = int(target_color[0] * luminosity)
                    new_g = int(target_color[1] * luminosity)
                    new_b = int(target_color[2] * luminosity)
                    pixels[i, j] = (new_r, new_g, new_b, a)
        
        # Sauvegarder en PNG
        output = BytesIO()
        img.save(output, format='PNG')
        output.seek(0)
        return output.getvalue()
    except Exception:
        # En cas d'erreur, retourner l'original
        return logo_data

def create_menu_pdf(restaurant: dict, reservation: dict, sections: list, items: list, include_selections: bool = False, client_selections: dict = None, doc_type: str = "proposition", invoice_number: str = None, invoice_date: str = None):
    """Créer un PDF de menu groupe - format facture avec TVA"""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Récupérer les couleurs du thème du restaurant
    primary_color = restaurant.get("primary_color", "#26252D")
    secondary_color = restaurant.get("secondary_color", "#EAE6CA")
    primary_rgb = hex_to_rgb(primary_color)
    secondary_rgb = hex_to_rgb(secondary_color)
    
    # Utiliser les couleurs du restaurant pour tous les éléments
    # (plus de RAL_5008_BLUE hardcodé)
    accent_text_rgb = primary_rgb  # Texte couleur primaire du restaurant
    accent_bg_rgb = secondary_rgb  # Fond couleur secondaire du restaurant
    
    # Marges
    left_margin = 15
    right_margin = 15
    pdf.set_left_margin(left_margin)
    pdf.set_right_margin(right_margin)
    
    # Largeur totale utilisable
    page_width = 210 - left_margin - right_margin  # 180mm
    
    # ========== EN-TÊTE: Restaurant (gauche) + Client (droite) ==========
    start_y = 10
    pdf.set_y(start_y)
    
    # Côté GAUCHE: Restaurant avec logo
    col_width = 85
    
    # Logo du restaurant (petit, à gauche) - teinte avec couleur primaire du restaurant
    if restaurant.get("logo_base64"):
        try:
            logo_data = base64.b64decode(restaurant["logo_base64"])
            # Appliquer la teinte couleur primaire au logo
            logo_data = apply_blue_tint_to_logo(logo_data, primary_rgb)
            logo_io = BytesIO(logo_data)
            pdf.image(logo_io, x=left_margin, y=start_y, w=25)
        except:
            pass
    
    # Infos restaurant (à côté du logo)
    pdf.set_xy(left_margin + 30, start_y)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(*primary_rgb)
    pdf.cell(col_width - 30, 5, safe_text(restaurant.get("name", "Restaurant")), ln=True)
    
    pdf.set_x(left_margin + 30)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(0, 0, 0)  # Reset to black for addresses
    
    # Adresse restaurant
    address_parts = []
    if restaurant.get("address_street"):
        address_parts.append(restaurant.get("address_street"))
    postal_city = []
    if restaurant.get("address_postal_code"):
        postal_city.append(restaurant.get("address_postal_code"))
    if restaurant.get("address_city"):
        postal_city.append(restaurant.get("address_city"))
    if postal_city:
        address_parts.append(" ".join(postal_city))
    
    for line in address_parts:
        pdf.cell(col_width - 30, 4, safe_text(line), ln=True)
        pdf.set_x(left_margin + 30)
    
    if restaurant.get("email"):
        pdf.cell(col_width - 30, 4, safe_text(restaurant.get("email")), ln=True)
        pdf.set_x(left_margin + 30)
    if restaurant.get("phone"):
        pdf.cell(col_width - 30, 4, safe_text(f"Tel: {restaurant.get('phone')}"), ln=True)
        pdf.set_x(left_margin + 30)
    
    # Informations légales (SIRET et RCS)
    if restaurant.get("siret"):
        pdf.cell(col_width - 30, 4, safe_text(f"SIRET: {restaurant.get('siret')}"), ln=True)
        pdf.set_x(left_margin + 30)
    if restaurant.get("rcs"):
        pdf.cell(col_width - 30, 4, safe_text(f"RCS: {restaurant.get('rcs')}"), ln=True)
    
    # Côté DROIT: Informations client - aligné vraiment à DROITE
    right_col_width = 80
    right_x = 210 - right_margin - right_col_width  # Position pour alignement à droite
    
    pdf.set_xy(right_x, start_y)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(right_col_width, 5, "CLIENT", align="R", ln=True)
    
    pdf.set_x(right_x)
    pdf.set_font("Helvetica", "", 9)
    client_name = f"{reservation.get('client_name', '')} {reservation.get('client_surname', '')}".strip()
    pdf.cell(right_col_width, 4, safe_text(client_name), align="R", ln=True)
    
    # Société (optionnel)
    pdf.set_x(right_x)
    if reservation.get("client_company"):
        pdf.cell(right_col_width, 4, safe_text(reservation.get("client_company")), align="R", ln=True)
        pdf.set_x(right_x)
    
    if reservation.get("client_email"):
        pdf.cell(right_col_width, 4, safe_text(reservation.get("client_email")), align="R", ln=True)
        pdf.set_x(right_x)
    if reservation.get("client_phone"):
        pdf.cell(right_col_width, 4, safe_text(f"Tel: {reservation.get('client_phone')}"), align="R", ln=True)
        pdf.set_x(right_x)
    
    # Adresse client si disponible
    client_address_parts = []
    if reservation.get("client_address_street"):
        client_address_parts.append(reservation.get("client_address_street"))
    client_postal_city = []
    if reservation.get("client_address_postal_code"):
        client_postal_city.append(reservation.get("client_address_postal_code"))
    if reservation.get("client_address_city"):
        client_postal_city.append(reservation.get("client_address_city"))
    if client_postal_city:
        client_address_parts.append(" ".join(client_postal_city))
    
    for line in client_address_parts:
        pdf.cell(right_col_width, 4, safe_text(line), align="R", ln=True)
        pdf.set_x(right_x)
    
    # Ligne de séparation après l'en-tête (espacement réduit)
    pdf.ln(4)
    header_end_y = max(pdf.get_y(), start_y + 35)
    pdf.set_y(header_end_y)
    pdf.set_draw_color(200, 200, 200)
    pdf.line(left_margin, pdf.get_y(), 210 - right_margin, pdf.get_y())
    pdf.ln(3)
    
    # ========== TITRE ET INFOS DE RÉSERVATION ==========
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(*primary_rgb)
    
    # Titre selon le type de document
    if doc_type == "facture" and invoice_number:
        pdf.cell(0, 8, f"FACTURE N. {invoice_number}", align="C", ln=True)
    elif doc_type == "facture":
        pdf.cell(0, 8, "FACTURE", align="C", ln=True)
    else:
        pdf.cell(0, 8, "PROPOSITION MENU GROUPE", align="C", ln=True)
    
    pdf.set_text_color(0, 0, 0)  # Reset to black
    
    pdf.set_font("Helvetica", "", 10)
    
    # Affichage de la date selon le type de document
    if doc_type == "facture":
        # Pour les factures: date de facturation (sans heure)
        if invoice_date:
            pdf.cell(0, 5, safe_text(f"Date de facturation: {invoice_date}"), align="C", ln=True)
        else:
            pdf.cell(0, 5, safe_text(f"Date: {reservation.get('date', '')}"), align="C", ln=True)
    else:
        # Pour les propositions: date et heure de réservation
        pdf.cell(0, 5, safe_text(f"Date: {reservation.get('date', '')} a {reservation.get('time', '')}"), align="C", ln=True)
    
    # Nombre de personnes: seulement pour les propositions (pas les factures)
    if doc_type != "facture":
        pdf.cell(0, 5, safe_text(f"Nombre de personnes: {reservation.get('num_people', '')}"), align="C", ln=True)
    
    pdf.ln(5)
    pdf.line(left_margin, pdf.get_y(), 210 - right_margin, pdf.get_y())
    pdf.ln(5)
    
    # ========== MENU: Sections et Items - Organisation hiérarchique ==========
    items_dict = {item["item_id"]: item for item in items}
    items_by_section = {}
    for item in items:
        sid = item.get("section_id")
        if sid not in items_by_section:
            items_by_section[sid] = []
        items_by_section[sid].append(item)
    
    selected_sections_ids = reservation.get("selected_sections", [])
    selected_items_dict = reservation.get("selected_items", {})
    
    # Séparer sections parents et sous-sections
    parent_sections = []
    sub_sections_by_parent = {}
    
    for section in sections:
        parent_id = section.get("parent_section_id")
        if not parent_id or parent_id == "ROOT":
            parent_sections.append(section)
        else:
            if parent_id not in sub_sections_by_parent:
                sub_sections_by_parent[parent_id] = []
            sub_sections_by_parent[parent_id].append(section)
    
    # Afficher les sections dans l'ordre hiérarchique
    has_displayed_sections = False
    
    for parent in sorted(parent_sections, key=lambda x: x.get("order", 0)):
        parent_id = parent.get("section_id")
        parent_name = parent.get("name", "Section")
        sub_sections = sub_sections_by_parent.get(parent_id, [])
        sub_sections = sorted(sub_sections, key=lambda x: x.get("order", 0))
        
        # Filtrer les sous-sections sélectionnées
        selected_subs = [s for s in sub_sections if s.get("section_id") in selected_sections_ids]
        
        # Ne pas afficher la section parent si aucune sous-section n'est sélectionnée
        if not selected_subs:
            continue
        
        has_displayed_sections = True
        
        # Titre de la section PARENT (Entrées, Plats, Desserts)
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(*accent_bg_rgb)
        pdf.set_text_color(*accent_text_rgb)
        pdf.cell(0, 8, safe_text(parent_name.upper()), ln=True, fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)
        
        # Afficher les sous-sections sélectionnées avec leurs items
        for sub in selected_subs:
            sub_id = sub.get("section_id")
            sub_name = sub.get("name", "Formule")
            sub_price = sub.get("price", 0)
            sub_items = items_by_section.get(sub_id, [])
            
            # Titre de la sous-section avec prix
            pdf.set_font("Helvetica", "B", 10)
            sub_title = f"  {sub_name}"
            if sub_price and sub_price > 0:
                sub_title += f" - {sub_price:.2f} EUR/pers"
            pdf.cell(0, 6, safe_text(sub_title), ln=True)
            
            # Items de la sous-section
            pdf.set_font("Helvetica", "", 9)
            for item in sorted(sub_items, key=lambda x: x.get("order", 0)):
                item_name = item.get('name', 'Item')
                
                # Afficher quantités en rouge si disponibles
                if doc_type == "proposition" and client_selections:
                    item_id = item.get("item_id")
                    qty = client_selections.get(item_id, 0)
                    if qty > 0:
                        item_text = f"      - {item_name} "
                        text_width = pdf.get_string_width(safe_text(item_text))
                        pdf.cell(text_width, 5, safe_text(item_text), ln=False)
                        pdf.set_text_color(255, 0, 0)
                        pdf.cell(0, 5, safe_text(f"x{qty}"), ln=True)
                        pdf.set_text_color(0, 0, 0)
                    else:
                        pdf.cell(0, 5, safe_text(f"      - {item_name}"), ln=True)
                else:
                    pdf.cell(0, 5, safe_text(f"      - {item_name}"), ln=True)
            
            pdf.ln(2)
        
        pdf.ln(3)
    
    # ========== CALCULS TVA ET TOTAUX ==========
    # Récupérer les valeurs nécessaires d'abord
    num_people = reservation.get("num_people", 0) or 0
    price_per_person = reservation.get("price_per_person", 0) or 0
    custom_options = reservation.get("custom_options", [])
    
    # Si aucune section n'a été affichée mais qu'il y a un prix par personne, 
    # afficher un message récapitulatif
    if not has_displayed_sections and price_per_person > 0:
        pdf.set_font("Helvetica", "I", 10)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, safe_text(f"Menu groupe a {price_per_person:.2f} EUR par personne"), ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(3)
    
    pdf.ln(5)
    pdf.line(left_margin, pdf.get_y(), 210 - right_margin, pdf.get_y())
    pdf.ln(5)
    
    # Produits: Prix par personne x nombre de personnes (TVA 10%)
    produits_ttc = price_per_person * num_people
    produits_ht = produits_ttc / 1.10  # TTC -> HT (retirer 10%)
    tva_produits = produits_ttc - produits_ht
    
    # Options supplémentaires avec quantité et TVA variable
    # Calculer les totaux par taux de TVA
    tva_by_rate = {0: 0, 10: 0, 20: 0}  # TVA par taux
    options_total_ht = 0
    options_total_ttc = 0
    
    for opt in custom_options:
        opt_price_unit = float(opt.get("price", 0) or 0)
        opt_quantity = int(opt.get("quantity", 1) or 1)
        # TVA: 0 est une valeur valide, donc on ne fait pas "or 20"
        opt_tva_rate_raw = opt.get("tva_rate")
        opt_tva_rate = float(opt_tva_rate_raw) if opt_tva_rate_raw is not None else 20.0
        
        opt_ttc = opt_price_unit * opt_quantity
        opt_ht = opt_ttc / (1 + opt_tva_rate / 100) if opt_tva_rate > 0 else opt_ttc
        opt_tva = opt_ttc - opt_ht
        
        options_total_ttc += opt_ttc
        options_total_ht += opt_ht
        
        # Accumuler la TVA par taux
        rate_key = int(opt_tva_rate)
        if rate_key in tva_by_rate:
            tva_by_rate[rate_key] += opt_tva
    
    # Totaux
    total_ht = produits_ht + options_total_ht
    total_ttc = produits_ttc + options_total_ttc
    
    # Affichage des totaux - style facture
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "RECAPITULATIF", ln=True)
    pdf.ln(2)
    
    # Tableau récapitulatif - aligné sur la largeur totale (comme les sections)
    # Largeur totale = 180mm (210 - 15 - 15)
    col1 = page_width - 70  # Description (110mm)
    col2 = 35   # HT
    col3 = 35   # TTC
    
    # En-tête tableau avec couleur RAL 5008
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(*accent_bg_rgb)
    pdf.set_text_color(*accent_text_rgb)
    pdf.cell(col1, 6, "Description", border=1, fill=True)
    pdf.cell(col2, 6, "HT (EUR)", border=1, align="C", fill=True)
    pdf.cell(col3, 6, "TTC (EUR)", border=1, align="C", fill=True, ln=True)
    pdf.set_text_color(0, 0, 0)  # Reset to black
    
    pdf.set_font("Helvetica", "", 9)
    
    # Ligne Produits (menu) - Afficher uniquement si un menu est sélectionné (price_per_person > 0)
    has_menu = price_per_person > 0
    if has_menu:
        pdf.cell(col1, 6, safe_text(f"Menu ({num_people} pers. x {price_per_person:.2f} EUR) - TVA 10%"), border=1)
        pdf.cell(col2, 6, f"{produits_ht:.2f}", border=1, align="R")
        pdf.cell(col3, 6, f"{produits_ttc:.2f}", border=1, align="R", ln=True)
    
    # Lignes Options supplémentaires avec quantité et TVA variable
    if custom_options:
        for opt in custom_options:
            opt_name = opt.get("name", "Option")
            opt_price_unit = float(opt.get("price", 0) or 0)
            opt_quantity = int(opt.get("quantity", 1) or 1)
            # TVA: 0 est une valeur valide, donc on ne fait pas "or 20"
            opt_tva_rate_raw = opt.get("tva_rate")
            opt_tva_rate = float(opt_tva_rate_raw) if opt_tva_rate_raw is not None else 20.0
            
            opt_ttc = opt_price_unit * opt_quantity
            opt_ht = opt_ttc / (1 + opt_tva_rate / 100) if opt_tva_rate > 0 else opt_ttc
            
            # Description avec quantité si > 1
            if opt_quantity > 1:
                desc = f"{opt_name} x{opt_quantity} ({opt_price_unit:.2f} EUR) - TVA {int(opt_tva_rate)}%"
            else:
                desc = f"{opt_name} - TVA {int(opt_tva_rate)}%"
            
            pdf.cell(col1, 6, safe_text(desc), border=1)
            pdf.cell(col2, 6, f"{opt_ht:.2f}", border=1, align="R")
            pdf.cell(col3, 6, f"{opt_ttc:.2f}", border=1, align="R", ln=True)
    
    # Ligne TVA détaillée - Afficher uniquement si un menu est sélectionné
    pdf.set_font("Helvetica", "I", 8)
    if has_menu:
        pdf.cell(col1, 5, safe_text(f"   dont TVA 10% (produits): {tva_produits:.2f} EUR"), border=0)
        pdf.cell(col2 + col3, 5, "", border=0, ln=True)
    
    # Afficher les TVA des options par taux (si > 0)
    for rate, tva_amount in tva_by_rate.items():
        if tva_amount > 0:
            pdf.cell(col1, 5, safe_text(f"   dont TVA {rate}% (options): {tva_amount:.2f} EUR"), border=0)
            pdf.cell(col2 + col3, 5, "", border=0, ln=True)
    
    pdf.ln(2)
    
    # TOTAUX en gras avec couleur RAL 5008
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(*accent_bg_rgb)
    pdf.set_text_color(*accent_text_rgb)
    pdf.cell(col1, 7, "TOTAL HT", border=1, fill=True)
    pdf.cell(col2, 7, f"{total_ht:.2f}", border=1, align="R", fill=True)
    pdf.cell(col3, 7, "", border=1, fill=True, ln=True)
    
    pdf.cell(col1, 7, "TOTAL TTC", border=1, fill=True)
    pdf.cell(col2, 7, "", border=1, fill=True)
    pdf.cell(col3, 7, f"{total_ttc:.2f}", border=1, align="R", fill=True, ln=True)
    pdf.set_text_color(0, 0, 0)  # Reset to black
    
    return pdf.output()

@api_router.get("/group-reservations/{reservation_id}/pdf")
async def generate_manager_pdf(
    reservation_id: str,
    pdf_type: str = "auto",  # auto, proposition, facture
    token: Optional[str] = None,  # Token en query param pour ouverture directe
    request: Request = None,
    current_user: dict = Depends(get_current_user_optional)
):
    """Générer un PDF du menu groupe pour le manager
    pdf_type: 'auto' (selon statut), 'proposition' (menu vierge), 'facture' (avec sélections)
    Supporte l'authentification via header Authorization ou query param token
    """
    # Vérifier l'authentification (header ou query param)
    user = current_user
    if not user and token:
        # Essayer avec le token en query param (session token)
        user = await get_user_from_token(token)
    
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    reservation = await group_reservations_collection.find_one(
        {"reservation_id": reservation_id, "restaurant_id": user["restaurant_id"]},
        {"_id": 0}
    )
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": user["restaurant_id"]},
        {"_id": 0}
    )
    
    # Récupérer TOUTES les sections triées par order pour organiser hiérarchiquement
    sections = await menu_sections_collection.find(
        {"restaurant_id": user["restaurant_id"]},
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    
    all_item_ids = []
    for section_items in reservation.get("selected_items", {}).values():
        all_item_ids.extend(section_items)
    
    # Récupérer tous les items des sections sélectionnées 
    selected_section_ids = reservation.get("selected_sections", [])
    all_section_items = await menu_items_collection.find(
        {"section_id": {"$in": selected_section_ids}},
        {"_id": 0}
    ).to_list(500)
    
    items = all_section_items
    
    # Déterminer si on inclut les sélections client
    invoice_number = None
    invoice_date = None
    if pdf_type == "proposition":
        include_selections = False
        # Passer client_selections pour afficher les quantités en rouge sur les propositions
        client_selections = reservation.get("client_selections")
        doc_type = "proposition"
    elif pdf_type == "facture":
        include_selections = True
        client_selections = reservation.get("client_selections")
        doc_type = "facture"
        # Utiliser le numéro de facture et la date existants ou en générer de nouveaux
        if reservation.get("invoice_number"):
            invoice_number = reservation.get("invoice_number")
            invoice_date = reservation.get("invoice_date")
            # Si la date n'existe pas (migration), la créer maintenant
            if not invoice_date:
                invoice_date = datetime.now(timezone.utc).strftime("%d/%m/%Y")
                await group_reservations_collection.update_one(
                    {"reservation_id": reservation_id},
                    {"$set": {"invoice_date": invoice_date}}
                )
        else:
            invoice_number = await get_next_invoice_number(user["restaurant_id"])
            invoice_date = datetime.now(timezone.utc).strftime("%d/%m/%Y")
            # Sauvegarder le numéro de facture et la date dans la réservation
            await group_reservations_collection.update_one(
                {"reservation_id": reservation_id},
                {"$set": {"invoice_number": invoice_number, "invoice_date": invoice_date}}
            )
    else:  # auto
        include_selections = reservation.get("status") == "client_submitted"
        client_selections = reservation.get("client_selections") if include_selections else None
        doc_type = "facture" if include_selections else "proposition"
        if doc_type == "facture":
            # Utiliser le numéro de facture et la date existants ou en générer de nouveaux
            if reservation.get("invoice_number"):
                invoice_number = reservation.get("invoice_number")
                invoice_date = reservation.get("invoice_date")
                # Si la date n'existe pas (migration), la créer maintenant
                if not invoice_date:
                    invoice_date = datetime.now(timezone.utc).strftime("%d/%m/%Y")
                    await group_reservations_collection.update_one(
                        {"reservation_id": reservation_id},
                        {"$set": {"invoice_date": invoice_date}}
                    )
            else:
                invoice_number = await get_next_invoice_number(user["restaurant_id"])
                invoice_date = datetime.now(timezone.utc).strftime("%d/%m/%Y")
                # Sauvegarder le numéro de facture et la date dans la réservation
                await group_reservations_collection.update_one(
                    {"reservation_id": reservation_id},
                    {"$set": {"invoice_number": invoice_number, "invoice_date": invoice_date}}
                )
    
    pdf_content = create_menu_pdf(restaurant, reservation, sections, items, include_selections, client_selections, doc_type, invoice_number, invoice_date)
    
    filename = f"{doc_type}_{reservation.get('client_name', 'client')}_{reservation.get('date', '')}.pdf"
    
    return Response(
        content=bytes(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/public/group/{client_token}/pdf")
async def generate_client_pdf(client_token: str):
    """Générer un PDF du menu groupe pour le client"""
    reservation = await group_reservations_collection.find_one(
        {"client_token": client_token},
        {"_id": 0}
    )
    
    if not reservation:
        raise HTTPException(status_code=404, detail="Group not found")
    
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": reservation["restaurant_id"]},
        {"_id": 0}
    )
    
    sections = await menu_sections_collection.find(
        {"section_id": {"$in": reservation.get("selected_sections", [])}},
        {"_id": 0}
    ).to_list(100)
    
    all_item_ids = []
    for section_items in reservation.get("selected_items", {}).values():
        all_item_ids.extend(section_items)
    
    items = await menu_items_collection.find(
        {"item_id": {"$in": all_item_ids}},
        {"_id": 0}
    ).to_list(500)
    
    include_selections = reservation.get("status") == "client_submitted"
    client_selections = reservation.get("client_selections") if include_selections else None
    doc_type = "proposition"  # Client PDF is always a proposition
    
    pdf_content = create_menu_pdf(restaurant, reservation, sections, items, include_selections, client_selections, doc_type, None)
    
    filename = f"menu_groupe_{reservation.get('client_name', 'client')}_{reservation.get('date', '')}.pdf"
    
    return Response(
        content=bytes(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== INVOICES/QUOTES ENDPOINTS (Factures et Devis) ====================

@api_router.post("/invoices/create")
async def create_invoice(
    create_request: CreateInvoiceRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer un devis ou une facture"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    restaurant_id = current_user["restaurant_id"]
    
    # Générer le numéro de facture/devis
    prefix = "DEV" if create_request.type == "quote" else "FAC"
    year = datetime.now().year
    
    # Incrémenter le compteur
    counter = await invoice_counters_collection.find_one_and_update(
        {"restaurant_id": restaurant_id, "year": year, "type": create_request.type},
        {"$inc": {"count": 1}},
        upsert=True,
        return_document=True
    )
    count = counter.get("count", 1)
    invoice_number = f"{prefix}-{year}-{count:04d}"
    
    # Calculer les totaux
    items_data = []
    total_ht = 0
    tva_10_amount = 0
    tva_20_amount = 0
    
    for item in create_request.items:
        # Calcul du HT à partir du TTC
        unit_price_ht = item.unit_price_ttc / (1 + item.tva_rate / 100)
        line_total_ht = unit_price_ht * item.quantity
        line_tva = (item.unit_price_ttc - unit_price_ht) * item.quantity
        
        total_ht += line_total_ht
        if item.tva_rate == 10:
            tva_10_amount += line_tva
        else:
            tva_20_amount += line_tva
        
        items_data.append({
            "name": item.name,
            "description": item.description,
            "quantity": item.quantity,
            "unit_price_ttc": item.unit_price_ttc,
            "unit_price_ht": round(unit_price_ht, 2),
            "tva_rate": item.tva_rate,
            "line_total_ht": round(line_total_ht, 2),
            "line_total_ttc": round(item.unit_price_ttc * item.quantity, 2)
        })
    
    total_ttc = total_ht + tva_10_amount + tva_20_amount
    
    invoice = {
        "invoice_id": f"inv_{uuid.uuid4().hex[:12]}",
        "restaurant_id": restaurant_id,
        "type": create_request.type,  # "quote" ou "invoice"
        "number": invoice_number,
        "client": create_request.client.dict(),
        "items": items_data,
        "totals": {
            "total_ht": round(total_ht, 2),
            "tva_10_amount": round(tva_10_amount, 2),
            "tva_20_amount": round(tva_20_amount, 2),
            "total_ttc": round(total_ttc, 2)
        },
        "notes": create_request.notes,
        "valid_until": create_request.valid_until,
        "due_date": create_request.due_date,
        "status": "draft",  # draft, sent, accepted, rejected, paid
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user["user_id"]
    }
    
    await invoices_collection.insert_one(invoice)
    del invoice["_id"]
    
    return invoice

@api_router.get("/invoices/list")
async def list_invoices(
    type: Optional[str] = None,  # "quote" ou "invoice"
    current_user: dict = Depends(get_current_user)
):
    """Lister les devis et factures"""
    query = {"restaurant_id": current_user["restaurant_id"]}
    if type:
        query["type"] = type
    
    invoices = await invoices_collection.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    
    return invoices

@api_router.get("/invoices/{invoice_id}")
async def get_invoice(
    invoice_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtenir un devis ou une facture"""
    invoice = await invoices_collection.find_one(
        {"invoice_id": invoice_id, "restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    )
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return invoice

@api_router.put("/invoices/{invoice_id}")
async def update_invoice(
    invoice_id: str,
    update_request: UpdateInvoiceRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour un devis ou une facture"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    invoice = await invoices_collection.find_one(
        {"invoice_id": invoice_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    update_data = {}
    
    if update_request.client:
        update_data["client"] = update_request.client.dict()
    
    if update_request.items:
        # Recalculer les totaux
        items_data = []
        total_ht = 0
        tva_10_amount = 0
        tva_20_amount = 0
        
        for item in update_request.items:
            unit_price_ht = item.unit_price_ttc / (1 + item.tva_rate / 100)
            line_total_ht = unit_price_ht * item.quantity
            line_tva = (item.unit_price_ttc - unit_price_ht) * item.quantity
            
            total_ht += line_total_ht
            if item.tva_rate == 10:
                tva_10_amount += line_tva
            else:
                tva_20_amount += line_tva
            
            items_data.append({
                "name": item.name,
                "description": item.description,
                "quantity": item.quantity,
                "unit_price_ttc": item.unit_price_ttc,
                "unit_price_ht": round(unit_price_ht, 2),
                "tva_rate": item.tva_rate,
                "line_total_ht": round(line_total_ht, 2),
                "line_total_ttc": round(item.unit_price_ttc * item.quantity, 2)
            })
        
        update_data["items"] = items_data
        update_data["totals"] = {
            "total_ht": round(total_ht, 2),
            "tva_10_amount": round(tva_10_amount, 2),
            "tva_20_amount": round(tva_20_amount, 2),
            "total_ttc": round(total_ht + tva_10_amount + tva_20_amount, 2)
        }
    
    if update_request.notes is not None:
        update_data["notes"] = update_request.notes
    if update_request.valid_until is not None:
        update_data["valid_until"] = update_request.valid_until
    if update_request.due_date is not None:
        update_data["due_date"] = update_request.due_date
    if update_request.status is not None:
        update_data["status"] = update_request.status
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        await invoices_collection.update_one(
            {"invoice_id": invoice_id},
            {"$set": update_data}
        )
    
    updated = await invoices_collection.find_one(
        {"invoice_id": invoice_id},
        {"_id": 0}
    )
    
    return updated

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(
    invoice_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer un devis ou une facture"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await invoices_collection.delete_one(
        {"invoice_id": invoice_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return {"message": "Invoice deleted"}

@api_router.post("/invoices/{invoice_id}/convert-to-invoice")
async def convert_quote_to_invoice(
    invoice_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Convertir un devis en facture"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    quote = await invoices_collection.find_one(
        {"invoice_id": invoice_id, "restaurant_id": current_user["restaurant_id"], "type": "quote"}
    )
    
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    # Générer un nouveau numéro de facture
    year = datetime.now().year
    counter = await invoice_counters_collection.find_one_and_update(
        {"restaurant_id": current_user["restaurant_id"], "year": year, "type": "invoice"},
        {"$inc": {"count": 1}},
        upsert=True,
        return_document=True
    )
    count = counter.get("count", 1)
    invoice_number = f"FAC-{year}-{count:04d}"
    
    # Créer la nouvelle facture basée sur le devis
    invoice = {
        "invoice_id": f"inv_{uuid.uuid4().hex[:12]}",
        "restaurant_id": current_user["restaurant_id"],
        "type": "invoice",
        "number": invoice_number,
        "client": quote["client"],
        "items": quote["items"],
        "totals": quote["totals"],
        "notes": quote.get("notes"),
        "due_date": quote.get("due_date"),
        "from_quote_id": quote["invoice_id"],
        "from_quote_number": quote["number"],
        "status": "draft",
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user["user_id"]
    }
    
    await invoices_collection.insert_one(invoice)
    
    # Marquer le devis comme accepté
    await invoices_collection.update_one(
        {"invoice_id": invoice_id},
        {"$set": {"status": "accepted", "converted_to_invoice_id": invoice["invoice_id"]}}
    )
    
    del invoice["_id"]
    return invoice

@api_router.get("/invoices/{invoice_id}/pdf")
async def get_invoice_pdf(
    invoice_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Générer le PDF d'un devis ou d'une facture"""
    invoice = await invoices_collection.find_one(
        {"invoice_id": invoice_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    )
    
    # Générer le PDF
    pdf_buffer = BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4
    
    # En-tête avec logo et infos restaurant
    y_pos = height - 50
    
    # Logo (si disponible)
    if restaurant.get("logo_base64"):
        try:
            logo_data = base64.b64decode(restaurant["logo_base64"])
            logo_image = PILImage.open(io.BytesIO(logo_data))
            logo_temp = io.BytesIO()
            logo_image.save(logo_temp, format='PNG')
            logo_temp.seek(0)
            c.drawImage(ImageReader(logo_temp), 50, y_pos - 50, width=60, height=60, preserveAspectRatio=True, mask='auto')
        except:
            pass
    
    # Infos restaurant
    c.setFont("Helvetica-Bold", 14)
    c.drawString(130, y_pos, restaurant.get("name", "Restaurant"))
    y_pos -= 15
    c.setFont("Helvetica", 9)
    if restaurant.get("address"):
        c.drawString(130, y_pos, restaurant.get("address", ""))
        y_pos -= 12
    if restaurant.get("phone"):
        c.drawString(130, y_pos, f"Tél: {restaurant.get('phone', '')}")
        y_pos -= 12
    if restaurant.get("email"):
        c.drawString(130, y_pos, f"Email: {restaurant.get('email', '')}")
        y_pos -= 12
    if restaurant.get("siret"):
        c.drawString(130, y_pos, f"SIRET: {restaurant.get('siret', '')}")
    
    # Type et numéro du document
    doc_type = "DEVIS" if invoice["type"] == "quote" else "FACTURE"
    c.setFont("Helvetica-Bold", 18)
    c.drawRightString(width - 50, height - 50, f"{doc_type} N° {invoice['number']}")
    c.setFont("Helvetica", 10)
    c.drawRightString(width - 50, height - 70, f"Date: {invoice['created_at'].strftime('%d/%m/%Y') if isinstance(invoice['created_at'], datetime) else str(invoice['created_at'])[:10]}")
    
    if invoice["type"] == "quote" and invoice.get("valid_until"):
        c.drawRightString(width - 50, height - 85, f"Valide jusqu'au: {invoice['valid_until']}")
    elif invoice["type"] == "invoice" and invoice.get("due_date"):
        c.drawRightString(width - 50, height - 85, f"Échéance: {invoice['due_date']}")
    
    # Infos client
    y_pos = height - 150
    c.setFont("Helvetica-Bold", 11)
    c.drawString(350, y_pos, "Client:")
    y_pos -= 15
    c.setFont("Helvetica", 10)
    client = invoice.get("client", {})
    if client.get("company"):
        c.drawString(350, y_pos, client["company"])
        y_pos -= 12
    c.drawString(350, y_pos, client.get("name", ""))
    y_pos -= 12
    if client.get("address"):
        for line in client["address"].split("\n"):
            c.drawString(350, y_pos, line)
            y_pos -= 12
    if client.get("email"):
        c.drawString(350, y_pos, client["email"])
        y_pos -= 12
    if client.get("phone"):
        c.drawString(350, y_pos, f"Tél: {client['phone']}")
    
    # Tableau des produits
    y_pos = height - 280
    c.setFont("Helvetica-Bold", 10)
    c.drawString(50, y_pos, "Désignation")
    c.drawString(280, y_pos, "Qté")
    c.drawString(330, y_pos, "Prix unit. TTC")
    c.drawString(420, y_pos, "TVA")
    c.drawString(470, y_pos, "Total TTC")
    
    y_pos -= 5
    c.line(50, y_pos, width - 50, y_pos)
    y_pos -= 15
    
    c.setFont("Helvetica", 9)
    for item in invoice.get("items", []):
        c.drawString(50, y_pos, item["name"][:35])
        c.drawString(280, y_pos, str(item["quantity"]))
        c.drawString(330, y_pos, f"{item['unit_price_ttc']:.2f}€")
        c.drawString(420, y_pos, f"{int(item['tva_rate'])}%")
        c.drawString(470, y_pos, f"{item['line_total_ttc']:.2f}€")
        y_pos -= 15
        if item.get("description"):
            c.setFont("Helvetica-Oblique", 8)
            c.drawString(60, y_pos, item["description"][:50])
            y_pos -= 12
            c.setFont("Helvetica", 9)
    
    # Totaux
    y_pos -= 20
    c.line(350, y_pos + 15, width - 50, y_pos + 15)
    totals = invoice.get("totals", {})
    
    c.setFont("Helvetica", 10)
    c.drawString(350, y_pos, "Total HT:")
    c.drawRightString(width - 50, y_pos, f"{totals.get('total_ht', 0):.2f}€")
    y_pos -= 15
    
    if totals.get("tva_10_amount", 0) > 0:
        c.drawString(350, y_pos, "TVA 10%:")
        c.drawRightString(width - 50, y_pos, f"{totals.get('tva_10_amount', 0):.2f}€")
        y_pos -= 15
    
    if totals.get("tva_20_amount", 0) > 0:
        c.drawString(350, y_pos, "TVA 20%:")
        c.drawRightString(width - 50, y_pos, f"{totals.get('tva_20_amount', 0):.2f}€")
        y_pos -= 15
    
    c.setFont("Helvetica-Bold", 12)
    c.drawString(350, y_pos, "Total TTC:")
    c.drawRightString(width - 50, y_pos, f"{totals.get('total_ttc', 0):.2f}€")
    
    # Notes
    if invoice.get("notes"):
        y_pos -= 40
        c.setFont("Helvetica-Oblique", 9)
        c.drawString(50, y_pos, "Notes:")
        y_pos -= 12
        for line in invoice["notes"].split("\n")[:3]:
            c.drawString(50, y_pos, line[:80])
            y_pos -= 12
    
    c.save()
    pdf_buffer.seek(0)
    
    filename = f"{doc_type.lower()}_{invoice['number']}.pdf"
    
    return Response(
        content=pdf_buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== PRIVATISATION SPACES ENDPOINTS ====================

@api_router.post("/privatisation-spaces/create")
async def create_privatisation_space(
    create_request: CreatePrivatisationSpaceRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer un espace de privatisation"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    space = {
        "space_id": f"priv_{uuid.uuid4().hex[:12]}",
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "description": create_request.description,
        "capacity_min": create_request.capacity_min,
        "capacity_max": create_request.capacity_max,
        "photos": create_request.photos or [],
        "amenities": create_request.amenities or [],
        "price_info": create_request.price_info,
        "price_under_minimum": create_request.price_under_minimum,
        "is_active": create_request.is_active,
        "created_at": datetime.now(timezone.utc)
    }
    
    await privatisation_spaces_collection.insert_one(space)
    del space["_id"]
    
    return space

@api_router.get("/privatisation-spaces/list")
async def list_privatisation_spaces(
    current_user: dict = Depends(get_current_user)
):
    """Lister les espaces de privatisation"""
    spaces = await privatisation_spaces_collection.find(
        {"restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    return spaces

@api_router.get("/public/restaurant/{restaurant_id}/privatisation-spaces")
async def get_public_privatisation_spaces(restaurant_id: str):
    """Obtenir les espaces de privatisation pour le formulaire public"""
    spaces = await privatisation_spaces_collection.find(
        {"restaurant_id": restaurant_id, "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    return spaces

@api_router.put("/privatisation-spaces/{space_id}")
async def update_privatisation_space(
    space_id: str,
    update_request: UpdatePrivatisationSpaceRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour un espace de privatisation"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await privatisation_spaces_collection.update_one(
            {"space_id": space_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    space = await privatisation_spaces_collection.find_one(
        {"space_id": space_id},
        {"_id": 0}
    )
    
    return space

@api_router.delete("/privatisation-spaces/{space_id}")
async def delete_privatisation_space(
    space_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer un espace de privatisation"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    await privatisation_spaces_collection.update_one(
        {"space_id": space_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    return {"message": "Space deleted"}

# ==================== NOTIFICATIONS ENDPOINT ====================

@api_router.get("/notifications/my")
async def get_my_notifications(
    current_user: dict = Depends(get_current_user)
):
    """Obtenir mes notifications"""
    notifications = await notifications_collection.find(
        {"user_id": current_user["user_id"]},
        {"_id": 0}
    ).sort("sent_at", -1).to_list(50)
    
    return notifications

# ==================== SUPPLIER / ORDER PREPARATION ENDPOINTS ====================

def get_day_name(day_num: int) -> str:
    """Convertir un numéro de jour en nom français"""
    days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    return days[day_num] if 0 <= day_num <= 6 else ""

def get_next_delivery_info(delivery_schedule: dict) -> dict:
    """Calculer les prochaines dates de commande et livraison"""
    if not delivery_schedule:
        return {"next_order_deadline": None, "next_delivery": None}
    
    from datetime import date
    today = datetime.now(timezone.utc)
    today_weekday = today.weekday()  # 0=Lundi, 6=Dimanche
    
    delivery_days = delivery_schedule.get("delivery_days", [])
    order_deadline_days = delivery_schedule.get("order_deadline_days", [])
    order_deadline_time = delivery_schedule.get("order_deadline_time", "19:00")
    
    if not delivery_days:
        return {"next_order_deadline": None, "next_delivery": None}
    
    # Trouver le prochain jour de livraison
    next_delivery_day = None
    days_until_delivery = None
    for i in range(1, 8):  # Chercher dans les 7 prochains jours
        check_day = (today_weekday + i) % 7
        if check_day in delivery_days:
            next_delivery_day = check_day
            days_until_delivery = i
            break
    
    if next_delivery_day is None:
        return {"next_order_deadline": None, "next_delivery": None}
    
    # Calculer la date de livraison
    next_delivery_date = today + timedelta(days=days_until_delivery)
    
    # Calculer la prochaine date limite de commande parmi les jours configurés
    next_order_deadline_day = None
    days_until_deadline = None
    
    if order_deadline_days:
        for i in range(0, 8):  # Chercher dans les 7 prochains jours (en commençant par aujourd'hui)
            check_day = (today_weekday + i) % 7
            if check_day in order_deadline_days:
                if i == 0:
                    # C'est aujourd'hui, vérifier l'heure
                    try:
                        deadline_hour, deadline_minute = map(int, order_deadline_time.split(":"))
                        deadline_time = today.replace(hour=deadline_hour, minute=deadline_minute)
                        if today > deadline_time:
                            continue  # Passer au jour suivant
                    except:
                        pass
                next_order_deadline_day = check_day
                days_until_deadline = i
                break
    
    if next_order_deadline_day is None or days_until_deadline is None:
        return {
            "next_order_deadline": None,
            "next_delivery": f"{get_day_name(next_delivery_day)} {next_delivery_date.strftime('%d/%m')}",
            "next_delivery_date": next_delivery_date.strftime("%Y-%m-%d"),
            "next_order_deadline_date": None
        }
    
    next_order_deadline_date = today + timedelta(days=days_until_deadline)
    
    return {
        "next_order_deadline": f"{get_day_name(next_order_deadline_day)} {next_order_deadline_date.strftime('%d/%m')} avant {order_deadline_time}",
        "next_delivery": f"{get_day_name(next_delivery_day)} {next_delivery_date.strftime('%d/%m')}",
        "next_delivery_date": next_delivery_date.strftime("%Y-%m-%d"),
        "next_order_deadline_date": next_order_deadline_date.strftime("%Y-%m-%d")
    }

@api_router.post("/suppliers/create")
async def create_supplier(
    create_request: CreateSupplierRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer un fournisseur (Metro, Transgourmet, etc.)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    supplier_id = f"sup_{uuid.uuid4().hex[:12]}"
    
    delivery_schedule = None
    if create_request.delivery_schedule:
        delivery_schedule = {
            "delivery_days": create_request.delivery_schedule.delivery_days,
            "order_deadline_days": create_request.delivery_schedule.order_deadline_days,
            "order_deadline_time": create_request.delivery_schedule.order_deadline_time,
            "delivery_time": create_request.delivery_schedule.delivery_time
        }
    
    supplier = {
        "supplier_id": supplier_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "phone": create_request.phone,
        "delivery_schedule": delivery_schedule,
        "notes": create_request.notes,
        "supplier_category": create_request.supplier_category,  # "bar", "cuisine", "both"
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await suppliers_collection.insert_one(supplier)
    
    # Calculer les infos de livraison
    delivery_info = get_next_delivery_info(delivery_schedule)
    
    return {
        "supplier_id": supplier_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "phone": create_request.phone,
        "delivery_schedule": delivery_schedule,
        "notes": create_request.notes,
        "supplier_category": create_request.supplier_category,
        "is_active": True,
        **delivery_info
    }

@api_router.get("/suppliers/list")
async def list_suppliers(current_user: dict = Depends(get_current_user)):
    """Lister tous les fournisseurs actifs"""
    suppliers = await suppliers_collection.find(
        {"restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).sort("name", 1).to_list(100)
    
    # Ajouter les infos de livraison pour chaque fournisseur
    for supplier in suppliers:
        delivery_info = get_next_delivery_info(supplier.get("delivery_schedule"))
        supplier.update(delivery_info)
    
    return suppliers

@api_router.get("/suppliers/{supplier_id}")
async def get_supplier(
    supplier_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtenir les détails d'un fournisseur"""
    supplier = await suppliers_collection.find_one(
        {"supplier_id": supplier_id, "restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    )
    
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Ajouter les infos de livraison
    delivery_info = get_next_delivery_info(supplier.get("delivery_schedule"))
    supplier.update(delivery_info)
    
    return supplier

@api_router.put("/suppliers/{supplier_id}")
async def update_supplier(
    supplier_id: str,
    update_request: UpdateSupplierRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour un fournisseur"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {}
    
    if update_request.name is not None:
        update_data["name"] = update_request.name
    if update_request.phone is not None:
        update_data["phone"] = update_request.phone
    if update_request.notes is not None:
        update_data["notes"] = update_request.notes
    if update_request.is_active is not None:
        update_data["is_active"] = update_request.is_active
    if update_request.supplier_category is not None:
        update_data["supplier_category"] = update_request.supplier_category
    if update_request.delivery_schedule is not None:
        update_data["delivery_schedule"] = {
            "delivery_days": update_request.delivery_schedule.delivery_days,
            "order_deadline_days": update_request.delivery_schedule.order_deadline_days,
            "order_deadline_time": update_request.delivery_schedule.order_deadline_time,
            "delivery_time": update_request.delivery_schedule.delivery_time
        }
    
    if update_data:
        await suppliers_collection.update_one(
            {"supplier_id": supplier_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    supplier = await suppliers_collection.find_one(
        {"supplier_id": supplier_id},
        {"_id": 0}
    )
    
    if supplier:
        delivery_info = get_next_delivery_info(supplier.get("delivery_schedule"))
        supplier.update(delivery_info)
    
    return supplier

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(
    supplier_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer un fournisseur (soft delete)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await suppliers_collection.update_one(
        {"supplier_id": supplier_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Désactiver aussi les produits du fournisseur
    await supplier_products_collection.update_many(
        {"supplier_id": supplier_id},
        {"$set": {"is_active": False}}
    )
    
    return {"message": "Supplier deleted successfully"}

@api_router.get("/suppliers/{supplier_id}/orders")
async def get_supplier_orders(
    supplier_id: str,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer l'historique des commandes pour un fournisseur spécifique"""
    # Vérifier que le fournisseur existe et appartient au restaurant
    supplier = await suppliers_collection.find_one({
        "supplier_id": supplier_id,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Récupérer les commandes de ce fournisseur
    orders = await supplier_orders_collection.find(
        {"supplier_id": supplier_id, "restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(limit)
    
    # Convertir les dates en strings
    for order in orders:
        if "created_at" in order and isinstance(order["created_at"], datetime):
            order["created_at"] = order["created_at"].isoformat()
    
    return {
        "supplier_id": supplier_id,
        "supplier_name": supplier.get("name", ""),
        "orders": orders
    }

# ==================== SUPPLIER PRODUCTS ENDPOINTS ====================

@api_router.post("/supplier-products/create")
async def create_supplier_product(
    create_request: CreateSupplierProductRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer un produit pour un fournisseur"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que le fournisseur existe
    supplier = await suppliers_collection.find_one({
        "supplier_id": create_request.supplier_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    })
    
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Calculer l'ordre
    existing_products = await supplier_products_collection.find(
        {"supplier_id": create_request.supplier_id}
    ).to_list(500)
    max_order = max([p.get("order", 0) for p in existing_products], default=-1)
    
    # Valider le type de produit
    product_type = create_request.product_type
    if product_type not in ["product", "consigne"]:
        product_type = "product"
    
    product_id = f"prod_{uuid.uuid4().hex[:12]}"
    product = {
        "product_id": product_id,
        "supplier_id": create_request.supplier_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "product_type": product_type,
        "order": create_request.order if create_request.order is not None else max_order + 1,
        "price_ht": create_request.price_ht,  # Prix HT pour consignes
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await supplier_products_collection.insert_one(product)
    
    return {
        "product_id": product_id,
        "supplier_id": create_request.supplier_id,
        "name": create_request.name,
        "product_type": product_type,
        "order": product["order"],
        "price_ht": create_request.price_ht,
        "is_active": True
    }

@api_router.get("/supplier-products/list")
async def list_supplier_products(
    supplier_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les produits (optionnellement filtrés par fournisseur)"""
    query = {"restaurant_id": current_user["restaurant_id"], "is_active": True}
    
    if supplier_id:
        query["supplier_id"] = supplier_id
    
    products = await supplier_products_collection.find(query, {"_id": 0}).sort("order", 1).to_list(1000)
    return products

@api_router.get("/supplier-products/by-supplier/{supplier_id}")
async def get_products_by_supplier(
    supplier_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer tous les produits d'un fournisseur"""
    products = await supplier_products_collection.find({
        "supplier_id": supplier_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    }, {"_id": 0}).sort("order", 1).to_list(500)
    
    return products

@api_router.put("/supplier-products/{product_id}")
async def update_supplier_product(
    product_id: str,
    update_request: UpdateSupplierProductRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour un produit"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await supplier_products_collection.update_one(
            {"product_id": product_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    product = await supplier_products_collection.find_one(
        {"product_id": product_id},
        {"_id": 0}
    )
    
    return product

@api_router.delete("/supplier-products/{product_id}")
async def delete_supplier_product(
    product_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer un produit (soft delete)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await supplier_products_collection.update_one(
        {"product_id": product_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Product deleted successfully"}

# ==================== SUPPLIER ORDERS ENDPOINTS ====================

@api_router.post("/supplier-orders/create")
async def create_supplier_order(
    create_request: CreateSupplierOrderRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une commande fournisseur ou une réclamation"""
    # Vérifier que le fournisseur existe
    supplier = await suppliers_collection.find_one({
        "supplier_id": create_request.supplier_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    })
    
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    if not create_request.items or len(create_request.items) == 0:
        raise HTTPException(status_code=400, detail="At least one item is required")
    
    order_id = f"ord_{uuid.uuid4().hex[:12]}"
    
    # Déterminer le type de commande et le statut initial
    order_type = create_request.order_type or "order"
    
    # Déterminer si c'est une commande de consignes uniquement
    has_products = any(item.product_type not in ["consigne", "reclamation"] for item in create_request.items)
    has_consignes = any(item.product_type == "consigne" for item in create_request.items)
    is_only_consignes = has_consignes and not has_products and order_type != "reclamation"
    
    # Mettre à jour order_type si c'est uniquement des consignes
    if is_only_consignes and order_type == "order":
        order_type = "consigne"
    
    # Statut initial selon le type
    if order_type == "reclamation":
        initial_status = "to_request"  # À demander
    elif is_only_consignes:
        initial_status = "sent"  # Envoyé (pour consignes)
    else:
        initial_status = "to_order"  # À commander
    
    # Ajouter is_refunded=False à chaque item
    items_with_refund = []
    for item in create_request.items:
        item_dict = item.dict()
        item_dict["is_refunded"] = False
        items_with_refund.append(item_dict)
    
    order = {
        "order_id": order_id,
        "supplier_id": create_request.supplier_id,
        "supplier_name": supplier["name"],
        "restaurant_id": current_user["restaurant_id"],
        "items": items_with_refund,
        "notes": create_request.notes,
        "order_type": order_type,  # "order", "consigne", "reclamation"
        "is_only_consignes": is_only_consignes,
        "total_ht": create_request.total_ht,
        "status": initial_status,
        "created_by": current_user["user_id"],
        "created_by_name": current_user["name"],
        "created_at": datetime.now(timezone.utc)
    }
    await supplier_orders_collection.insert_one(order)
    
    return {
        "order_id": order_id,
        "supplier_id": create_request.supplier_id,
        "supplier_name": supplier["name"],
        "items": items_with_refund,
        "notes": create_request.notes,
        "order_type": order_type,
        "is_only_consignes": is_only_consignes,
        "total_ht": create_request.total_ht,
        "status": initial_status,
        "created_by_name": current_user["name"],
        "created_at": order["created_at"].isoformat()
    }

@api_router.get("/supplier-orders/list")
async def list_supplier_orders(
    supplier_id: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Lister les commandes (historique)"""
    query = {"restaurant_id": current_user["restaurant_id"]}
    
    if supplier_id:
        query["supplier_id"] = supplier_id
    
    if status:
        query["status"] = status
    
    orders = await supplier_orders_collection.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    
    # Convertir les dates en strings
    for order in orders:
        if "created_at" in order and isinstance(order["created_at"], datetime):
            order["created_at"] = order["created_at"].isoformat()
    
    return orders

@api_router.get("/supplier-orders/{order_id}")
async def get_supplier_order(
    order_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtenir les détails d'une commande"""
    order = await supplier_orders_collection.find_one(
        {"order_id": order_id, "restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    )
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if "created_at" in order and isinstance(order["created_at"], datetime):
        order["created_at"] = order["created_at"].isoformat()
    
    return order

@api_router.put("/supplier-orders/{order_id}/status")
async def update_order_status(
    order_id: str,
    update_request: UpdateOrderStatusRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Mettre à jour le statut d'une commande/consigne/réclamation
    
    Commandes: to_order → ordered → delivered
    Consignes: sent → partial_refund → full_refund
    Réclamations: to_request → requested → partial_refund → full_refund
    """
    # Récupérer la commande existante
    order = await supplier_orders_collection.find_one(
        {"order_id": order_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    order_type = order.get("order_type", "order")
    is_only_consignes = order.get("is_only_consignes", False)
    
    # Valider le statut selon le type
    valid_statuses = []
    if order_type == "reclamation":
        valid_statuses = ["to_request", "requested", "partial_refund", "full_refund"]
    elif is_only_consignes:
        valid_statuses = ["sent", "partial_refund", "full_refund"]
    else:
        valid_statuses = ["to_order", "ordered", "delivered"]
    
    if update_request.status not in valid_statuses:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid status '{update_request.status}' for this order type. Valid: {valid_statuses}"
        )
    
    # Préparer la mise à jour
    update_data = {
        "status": update_request.status,
        "updated_at": datetime.now(timezone.utc)
    }
    
    # Gérer le remboursement manuel (tout remboursé)
    if update_request.manual_full_refund:
        update_data["status"] = "full_refund"
        # Marquer tous les items comme remboursés
        items = order.get("items", [])
        for item in items:
            item["is_refunded"] = True
        update_data["items"] = items
    
    # Gérer le remboursement partiel par sélection d'items
    elif update_request.refunded_items:
        items = order.get("items", [])
        refunded_count = 0
        for item in items:
            if item.get("product_name") in update_request.refunded_items:
                item["is_refunded"] = True
                refunded_count += 1
        
        update_data["items"] = items
        
        # Déterminer le statut automatiquement
        total_items = len(items)
        if refunded_count == 0:
            if order_type == "reclamation":
                update_data["status"] = "to_request"
            elif is_only_consignes:
                update_data["status"] = "sent"
        elif refunded_count < total_items:
            update_data["status"] = "partial_refund"
        else:
            update_data["status"] = "full_refund"
    
    # Effectuer la mise à jour
    result = await supplier_orders_collection.update_one(
        {"order_id": order_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return {
        "message": f"Order status updated to {update_data['status']}", 
        "status": update_data["status"],
        "items": update_data.get("items", order.get("items", []))
    }

@api_router.delete("/supplier-orders/{order_id}")
async def delete_supplier_order(
    order_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une commande"""
    restaurant_id = current_user["restaurant_id"]
    
    # Vérifier que la commande existe
    order = await supplier_orders_collection.find_one(
        {"order_id": order_id, "restaurant_id": restaurant_id},
        {"_id": 0}
    )
    
    if not order:
        raise HTTPException(status_code=404, detail="Commande non trouvée")
    
    # Supprimer la commande
    result = await supplier_orders_collection.delete_one(
        {"order_id": order_id, "restaurant_id": restaurant_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="Erreur lors de la suppression")
    
    return {"message": "Commande supprimée avec succès"}

@api_router.get("/supplier-orders/{order_id}/pdf")
async def generate_order_pdf(
    order_id: str,
    token: Optional[str] = None,
    authorization: Optional[str] = Header(None)
):
    """Générer un PDF pour le bon de commande"""
    # Authentification via header ou query parameter
    current_user = None
    
    # Essayer d'abord le header Authorization
    if authorization and authorization.startswith("Bearer "):
        auth_token = authorization.replace("Bearer ", "")
        session_doc = await sessions_collection.find_one({"session_token": auth_token}, {"_id": 0})
        if session_doc:
            expires_at = session_doc["expires_at"]
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at >= datetime.now(timezone.utc):
                current_user = await users_collection.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    
    # Sinon, essayer le query parameter
    if current_user is None and token:
        session_doc = await sessions_collection.find_one({"session_token": token}, {"_id": 0})
        if session_doc:
            expires_at = session_doc["expires_at"]
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at >= datetime.now(timezone.utc):
                current_user = await users_collection.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    
    if current_user is None:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Récupérer la commande
    order = await supplier_orders_collection.find_one(
        {"order_id": order_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Récupérer le restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    )
    
    # Récupérer le fournisseur
    supplier = await suppliers_collection.find_one(
        {"supplier_id": order["supplier_id"]},
        {"_id": 0}
    )
    
    # Déterminer le type de document (commande, consigne ou réclamation)
    items = order.get("items", [])
    order_type = order.get("order_type", "order")
    is_reclamation = order_type == "reclamation"
    has_products = any(item.get("product_type") not in ["consigne", "reclamation"] for item in items)
    has_consignes = any(item.get("product_type") == "consigne" for item in items)
    is_only_consignes = has_consignes and not has_products and not is_reclamation
    
    # Titres et labels selon le type
    if is_reclamation:
        doc_title = "RÉCLAMATION"
        date_label = "Date de livraison"
    elif is_only_consignes:
        doc_title = "CONSIGNE RENDUE"
        date_label = "Date"
    else:
        doc_title = "BON DE COMMANDE"
        date_label = "Date de commande"
    
    # Créer le PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Couleur RAL 5008 pour les éléments (marron pour consignes, rouge pour réclamations)
    RAL_5008 = (26, 58, 92)
    CONSIGNE_COLOR = (121, 85, 72)  # Marron pour consignes
    RECLAMATION_COLOR = (211, 47, 47)  # Rouge pour réclamations
    
    if is_reclamation:
        doc_color = RECLAMATION_COLOR
    elif is_only_consignes:
        doc_color = CONSIGNE_COLOR
    else:
        doc_color = RAL_5008
    
    # En-tête avec nom du restaurant
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(*doc_color)
    pdf.cell(0, 12, restaurant.get("name", "Restaurant"), ln=True, align="C")
    
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 100, 100)
    
    # Adresse du restaurant si disponible
    address_parts = []
    if restaurant.get("address_street"):
        address_parts.append(restaurant["address_street"])
    if restaurant.get("address_postal_code") or restaurant.get("address_city"):
        address_parts.append(f"{restaurant.get('address_postal_code', '')} {restaurant.get('address_city', '')}".strip())
    if address_parts:
        pdf.cell(0, 5, " - ".join(address_parts), ln=True, align="C")
    
    pdf.ln(8)
    
    # Titre du document (BON DE COMMANDE, CONSIGNE RENDUE ou RÉCLAMATION)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*doc_color)
    pdf.cell(0, 10, doc_title, ln=True, align="C")
    
    pdf.ln(5)
    
    # Informations de la commande
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(0, 0, 0)
    
    # Date
    created_at = order.get("created_at")
    if isinstance(created_at, datetime):
        date_str = created_at.strftime("%d/%m/%Y")
    else:
        date_str = str(created_at)[:10] if created_at else ""
    
    pdf.cell(0, 7, f"{date_label}: {date_str}", ln=True)
    
    # Fournisseur
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, f"Fournisseur: {order.get('supplier_name', '')}", ln=True)
    
    pdf.ln(8)
    
    # Tableau des produits
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_fill_color(*doc_color)
    pdf.set_text_color(255, 255, 255)
    
    # En-tête du tableau - différent pour réclamations (avec prix)
    if is_reclamation:
        col_width_product = 80
        col_width_qty = 30
        col_width_price = 40
        col_width_total = 40
        
        pdf.cell(col_width_product, 8, "Produit", border=1, fill=True)
        pdf.cell(col_width_qty, 8, "Qté", border=1, fill=True, align="C")
        pdf.cell(col_width_price, 8, "Prix HT", border=1, fill=True, align="C")
        pdf.cell(col_width_total, 8, "Total", border=1, fill=True, align="C")
        pdf.ln()
        
        # Lignes du tableau avec prix
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(0, 0, 0)
        
        total_ht = 0
        for item in order.get("items", []):
            quantity = item.get("quantity", 0)
            price_ht = item.get("price_ht", 0) or 0
            line_total = quantity * price_ht
            total_ht += line_total
            
            pdf.cell(col_width_product, 7, item.get("product_name", ""), border=1)
            pdf.cell(col_width_qty, 7, str(quantity), border=1, align="C")
            pdf.cell(col_width_price, 7, f"{price_ht:.2f} EUR", border=1, align="R")
            pdf.cell(col_width_total, 7, f"{line_total:.2f} EUR", border=1, align="R")
            pdf.ln()
        
        # Total général
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(col_width_product + col_width_qty + col_width_price, 8, "TOTAL HT:", border=1, align="R")
        pdf.set_fill_color(255, 235, 238)
        pdf.cell(col_width_total, 8, f"{total_ht:.2f} EUR", border=1, fill=True, align="R")
        pdf.ln()
    else:
        # Vérifier si les consignes ont des prix
        has_prices = any(item.get("price_ht", 0) > 0 for item in order.get("items", []))
        
        if is_only_consignes and has_prices:
            # Consignes avec prix - afficher comme réclamation (avec tableau des prix)
            col_width_product = 80
            col_width_qty = 30
            col_width_price = 40
            col_width_total = 40
            
            pdf.cell(col_width_product, 8, "Consigne", border=1, fill=True)
            pdf.cell(col_width_qty, 8, "Qté", border=1, fill=True, align="C")
            pdf.cell(col_width_price, 8, "Prix HT", border=1, fill=True, align="C")
            pdf.cell(col_width_total, 8, "Total", border=1, fill=True, align="C")
            pdf.ln()
            
            # Lignes du tableau avec prix
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(0, 0, 0)
            
            total_ht = 0
            for item in order.get("items", []):
                quantity = item.get("quantity", 0)
                price_ht = item.get("price_ht", 0) or 0
                line_total = quantity * price_ht
                total_ht += line_total
                
                pdf.cell(col_width_product, 7, item.get("product_name", ""), border=1)
                pdf.cell(col_width_qty, 7, str(quantity), border=1, align="C")
                pdf.cell(col_width_price, 7, f"{price_ht:.2f} EUR", border=1, align="R")
                pdf.cell(col_width_total, 7, f"{line_total:.2f} EUR", border=1, align="R")
                pdf.ln()
            
            # Total général
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(col_width_product + col_width_qty + col_width_price, 8, "TOTAL À REMBOURSER:", border=1, align="R")
            pdf.set_fill_color(232, 245, 233)  # Vert clair pour consignes
            pdf.cell(col_width_total, 8, f"{total_ht:.2f} EUR", border=1, fill=True, align="R")
            pdf.ln()
        else:
            # Consignes sans prix ou commandes normales
            col_width_product = 140
            col_width_qty = 40
            
            # Label selon le type
            item_label = "Consigne" if is_only_consignes else "Produit"
            
            pdf.cell(col_width_product, 8, item_label, border=1, fill=True)
            pdf.cell(col_width_qty, 8, "Quantité", border=1, fill=True, align="C")
            pdf.ln()
            
            # Lignes du tableau
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(0, 0, 0)
            
            for item in order.get("items", []):
                pdf.cell(col_width_product, 7, item.get("product_name", ""), border=1)
                pdf.cell(col_width_qty, 7, str(item.get("quantity", "")), border=1, align="C")
                pdf.ln()
    
    pdf.ln(8)
    
    # Notes
    if order.get("notes"):
        pdf.set_font("Helvetica", "I", 10)
        pdf.set_text_color(80, 80, 80)
        pdf.multi_cell(0, 5, f"Notes: {order['notes']}")
    
    # Pied de page
    pdf.ln(15)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(120, 120, 120)
    if is_reclamation:
        action_label = "Réclamation enregistrée par"
    elif is_only_consignes:
        action_label = "Consigne enregistrée par"
    else:
        action_label = "Commande passée par"
    pdf.cell(0, 5, f"{action_label}: {order.get('created_by_name', '')}", ln=True, align="R")
    
    # Générer le PDF
    pdf_content = pdf.output()
    
    # Créer le nom du fichier
    supplier_name_clean = order.get('supplier_name', 'commande').replace(' ', '_')
    if is_reclamation:
        file_prefix = "reclamation"
    elif is_only_consignes:
        file_prefix = "consigne"
    else:
        file_prefix = "commande"
    filename = f"{file_prefix}_{supplier_name_clean}_{date_str.replace('/', '-')}.pdf"
    
    return Response(
        content=bytes(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ==================== PUSH NOTIFICATIONS ENDPOINTS ====================

class PushTokenRegister(BaseModel):
    token: str
    platform: str = "web"  # web, ios, android

@api_router.post("/push-tokens/register")
async def register_push_token(
    data: PushTokenRegister,
    current_user: dict = Depends(get_current_user)
):
    """Enregistrer un push token pour l'utilisateur actuel"""
    try:
        # Check if token already exists for this user
        existing = await push_tokens_collection.find_one({
            "user_id": current_user["user_id"],
            "token": data.token
        })
        
        if existing:
            # Update token
            await push_tokens_collection.update_one(
                {"_id": existing["_id"]},
                {"$set": {"platform": data.platform, "updated_at": datetime.now(timezone.utc), "is_active": True}}
            )
        else:
            # Create new token
            await push_tokens_collection.insert_one({
                "user_id": current_user["user_id"],
                "restaurant_id": current_user.get("restaurant_id"),
                "token": data.token,
                "platform": data.platform,
                "is_active": True,
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            })
        
        return {"status": "success", "message": "Token enregistré"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/push-tokens/unregister")
async def unregister_push_token(
    token: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer un push token"""
    await push_tokens_collection.delete_one({
        "user_id": current_user["user_id"],
        "token": token
    })
    return {"status": "success"}

async def send_push_notification_to_user(user_id: str, title: str, body: str, data: dict = None):
    """Envoyer une notification push à un utilisateur spécifique"""
    try:
        from exponent_server_sdk import PushClient, PushMessage, PushServerError
        
        # Get user's active push tokens
        tokens = await push_tokens_collection.find({
            "user_id": user_id,
            "is_active": True
        }).to_list(10)
        
        if not tokens:
            return
        
        client = PushClient()
        messages = []
        
        for token_doc in tokens:
            push_token = token_doc.get("token")
            if push_token and push_token.startswith("ExponentPushToken"):
                messages.append(PushMessage(
                    to=push_token,
                    title=title,
                    body=body,
                    data=data or {},
                    sound="default"
                ))
        
        if messages:
            try:
                client.publish_multiple(messages)
            except PushServerError as e:
                logging.error(f"Push notification error: {e}")
    except Exception as e:
        logging.error(f"Error sending push notification: {e}")

async def send_push_notification_to_restaurant(restaurant_id: str, title: str, body: str, data: dict = None, exclude_user: str = None):
    """Envoyer une notification push à tous les utilisateurs d'un restaurant"""
    try:
        from exponent_server_sdk import PushClient, PushMessage, PushServerError
        
        # Get all active push tokens for the restaurant
        query = {"restaurant_id": restaurant_id, "is_active": True}
        if exclude_user:
            query["user_id"] = {"$ne": exclude_user}
        
        tokens = await push_tokens_collection.find(query).to_list(100)
        
        if not tokens:
            return
        
        client = PushClient()
        messages = []
        
        for token_doc in tokens:
            push_token = token_doc.get("token")
            if push_token and push_token.startswith("ExponentPushToken"):
                messages.append(PushMessage(
                    to=push_token,
                    title=title,
                    body=body,
                    data=data or {},
                    sound="default"
                ))
        
        if messages:
            try:
                client.publish_multiple(messages)
            except PushServerError as e:
                logging.error(f"Push notification error: {e}")
    except Exception as e:
        logging.error(f"Error sending push notification: {e}")

# ==================== FICHE TECHNIQUE ENDPOINTS ====================

def calculate_ingredient_cost(ingredient: dict) -> float:
    """Calcule le prix de revient d'un ingrédient"""
    if not ingredient:
        return 0.0
    quantity_used = ingredient.get("quantity_used", 0)
    quantity_purchased = ingredient.get("quantity_purchased", 1)
    purchase_price = ingredient.get("purchase_price", 0)
    
    # Convertir les unités si nécessaire
    unit_used = ingredient.get("unit_used", "g")
    unit_purchased = ingredient.get("unit_purchased", "g")
    
    # Conversion ratios (tout vers grammes ou ml)
    conversions = {
        "g": 1, "kg": 1000,
        "ml": 1, "cl": 10, "l": 1000
    }
    
    # Convertir en unité de base
    used_in_base = quantity_used * conversions.get(unit_used, 1)
    purchased_in_base = quantity_purchased * conversions.get(unit_purchased, 1)
    
    if purchased_in_base == 0:
        return 0.0
    
    return (used_in_base / purchased_in_base) * purchase_price

def calculate_total_cost(ingredients: list) -> float:
    """Calcule le prix de revient total d'un produit"""
    return sum(calculate_ingredient_cost(ing) for ing in (ingredients or []))

# --- Fiche Technique Sections ---

@api_router.post("/fiche-sections/create")
async def create_fiche_section(
    create_request: CreateFicheSectionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une section (ex: Cocktails, Entrée, Préparations) pour Bar ou Cuisine"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if create_request.category not in ["bar", "cuisine"]:
        raise HTTPException(status_code=400, detail="Category must be 'bar' or 'cuisine'")
    
    # Calculer l'ordre
    existing = await fiche_technique_sections_collection.find({
        "restaurant_id": current_user["restaurant_id"],
        "category": create_request.category
    }).to_list(100)
    max_order = max([s.get("order", 0) for s in existing], default=-1)
    
    section_id = f"fsec_{uuid.uuid4().hex[:12]}"
    section = {
        "section_id": section_id,
        "restaurant_id": current_user["restaurant_id"],
        "category": create_request.category,
        "name": create_request.name,
        "order": create_request.order if create_request.order is not None else max_order + 1,
        "is_preparations": create_request.is_preparations,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await fiche_technique_sections_collection.insert_one(section)
    
    return {
        "section_id": section_id,
        "category": create_request.category,
        "name": create_request.name,
        "order": section["order"],
        "is_preparations": create_request.is_preparations,
        "is_active": True
    }

@api_router.get("/fiche-sections/list")
async def list_fiche_sections(
    category: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les sections (optionnel: filtrer par catégorie bar/cuisine)"""
    query = {"restaurant_id": current_user["restaurant_id"], "is_active": True}
    if category:
        query["category"] = category
    
    sections = await fiche_technique_sections_collection.find(
        query, {"_id": 0}
    ).sort("order", 1).to_list(100)
    
    return sections

@api_router.put("/fiche-sections/{section_id}")
async def update_fiche_section(
    section_id: str,
    update_request: UpdateFicheSectionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Modifier une section"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await fiche_technique_sections_collection.update_one(
            {"section_id": section_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    section = await fiche_technique_sections_collection.find_one(
        {"section_id": section_id},
        {"_id": 0}
    )
    return section

@api_router.delete("/fiche-sections/{section_id}")
async def delete_fiche_section(
    section_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une section (soft delete)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await fiche_technique_sections_collection.update_one(
        {"section_id": section_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Section not found")
    
    # Also deactivate products in this section
    await fiche_technique_products_collection.update_many(
        {"section_id": section_id},
        {"$set": {"is_active": False}}
    )
    
    return {"message": "Section deleted successfully"}

@api_router.put("/fiche-sections/{section_id}/margin-thresholds")
async def update_section_margin_thresholds(
    section_id: str,
    thresholds: UpdateMarginThresholdsRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour les seuils de marge d'une section"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Valider les seuils
    if thresholds.low < 0 or thresholds.high < 0:
        raise HTTPException(status_code=400, detail="Les seuils doivent être positifs")
    if thresholds.low >= thresholds.high:
        raise HTTPException(status_code=400, detail="Le seuil 'faible' doit être inférieur au seuil 'bon'")
    
    result = await fiche_technique_sections_collection.update_one(
        {"section_id": section_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"margin_thresholds": {"low": thresholds.low, "high": thresholds.high}}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Section non trouvée")
    
    section = await fiche_technique_sections_collection.find_one(
        {"section_id": section_id},
        {"_id": 0}
    )
    return section

# ==================== FICHE TECHNIQUE REORDERING ENDPOINTS ====================

@api_router.put("/fiche-sections/{section_id}/set-order")
async def set_fiche_section_order(
    section_id: str,
    request: SetOrderRequest,
    current_user: dict = Depends(get_current_user)
):
    """Définir directement l'ordre d'une section Fiche Technique"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await fiche_technique_sections_collection.update_one(
        {"section_id": section_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"order": request.order}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Section non trouvée")
    
    section_doc = await fiche_technique_sections_collection.find_one(
        {"section_id": section_id},
        {"_id": 0}
    )
    return section_doc

@api_router.put("/fiche-products/{product_id}/set-order")
async def set_fiche_product_order(
    product_id: str,
    request: SetOrderRequest,
    current_user: dict = Depends(get_current_user)
):
    """Définir directement l'ordre d'un produit Fiche Technique"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await fiche_technique_products_collection.update_one(
        {"product_id": product_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"order": request.order}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Produit non trouvé")
    
    product_doc = await fiche_technique_products_collection.find_one(
        {"product_id": product_id},
        {"_id": 0}
    )
    return product_doc

@api_router.get("/fiche-products/margin-analysis")
async def get_margin_analysis(
    current_user: dict = Depends(get_current_user)
):
    """Obtenir l'analyse des marges de tous les produits avec catégorisation par couleur"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Récupérer toutes les sections actives avec leurs seuils
    sections = await fiche_technique_sections_collection.find(
        {"restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    # Créer un dictionnaire pour accès rapide aux seuils
    section_thresholds = {}
    section_names = {}
    for section in sections:
        default_thresholds = {"low": 20.0, "high": 50.0}
        section_thresholds[section["section_id"]] = section.get("margin_thresholds", default_thresholds)
        section_names[section["section_id"]] = {
            "name": section["name"],
            "category": section["category"]
        }
    
    # Récupérer tous les produits actifs
    products = await fiche_technique_products_collection.find(
        {"restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).to_list(500)
    
    # Analyser chaque produit
    analysis_results = []
    for product in products:
        section_id = product.get("section_id")
        if section_id not in section_thresholds:
            continue
        
        thresholds = section_thresholds[section_id]
        section_info = section_names.get(section_id, {})
        product_type = product.get("product_type", "standard")
        
        # Calculer la marge selon le type de produit
        if product_type == "boisson_multi":
            # Pour les boissons multi, calculer la marge moyenne des formats
            selling_formats = product.get("selling_formats", [])
            purchase_info = product.get("purchase_info", {})
            purchase_qty = purchase_info.get("quantity", 0)
            purchase_price = purchase_info.get("price", 0)
            
            if selling_formats and purchase_qty > 0:
                margins = []
                for fmt in selling_formats:
                    format_size = fmt.get("size", 0)
                    selling_price = fmt.get("selling_price", 0)
                    if format_size > 0 and selling_price > 0:
                        cost = (format_size / purchase_qty) * purchase_price
                        margin = ((selling_price - cost) / selling_price) * 100
                        margins.append({
                            "format_name": fmt.get("name", ""),
                            "cost": round(cost, 2),
                            "selling_price": selling_price,
                            "margin_percent": round(margin, 1)
                        })
                
                if margins:
                    avg_margin = sum(m["margin_percent"] for m in margins) / len(margins)
                    analysis_results.append({
                        "product_id": product["product_id"],
                        "name": product["name"],
                        "product_type": product_type,
                        "section_id": section_id,
                        "section_name": section_info.get("name", ""),
                        "category": section_info.get("category", ""),
                        "margin_percent": round(avg_margin, 1),
                        "margin_category": _get_margin_category(avg_margin, thresholds),
                        "formats_detail": margins,
                        "total_cost": None,
                        "selling_price": None
                    })
        
        elif product_type == "preparation":
            # Les préparations n'ont pas de marge de vente directe (elles sont utilisées comme ingrédients)
            continue
        
        else:  # standard
            total_cost = product.get("total_cost", 0)
            selling_price = product.get("selling_price", 0) or product.get("selling_price_override", 0)
            
            if total_cost is None:
                total_cost = 0
            
            # Calculer avec le multiplicateur si pas de prix de vente direct
            if not selling_price and product.get("multiplier"):
                selling_price = total_cost * product.get("multiplier", 1)
            
            if selling_price and selling_price > 0:
                margin = ((selling_price - total_cost) / selling_price) * 100
                analysis_results.append({
                    "product_id": product["product_id"],
                    "name": product["name"],
                    "product_type": product_type,
                    "section_id": section_id,
                    "section_name": section_info.get("name", ""),
                    "category": section_info.get("category", ""),
                    "margin_percent": round(margin, 1),
                    "margin_category": _get_margin_category(margin, thresholds),
                    "total_cost": round(total_cost, 2),
                    "selling_price": round(selling_price, 2),
                    "formats_detail": None
                })
            elif total_cost > 0:
                # Produit avec coût mais sans prix de vente défini
                analysis_results.append({
                    "product_id": product["product_id"],
                    "name": product["name"],
                    "product_type": product_type,
                    "section_id": section_id,
                    "section_name": section_info.get("name", ""),
                    "category": section_info.get("category", ""),
                    "margin_percent": None,
                    "margin_category": "undefined",
                    "total_cost": round(total_cost, 2),
                    "selling_price": None,
                    "formats_detail": None
                })
    
    # Grouper par section pour le frontend
    sections_with_products = {}
    for result in analysis_results:
        sid = result["section_id"]
        if sid not in sections_with_products:
            sections_with_products[sid] = {
                "section_id": sid,
                "section_name": result["section_name"],
                "category": result["category"],
                "thresholds": section_thresholds.get(sid, {"low": 20.0, "high": 50.0}),
                "products": []
            }
        sections_with_products[sid]["products"].append(result)
    
    # Statistiques globales
    total_products = len(analysis_results)
    products_with_margin = [r for r in analysis_results if r["margin_percent"] is not None]
    low_margin = len([r for r in products_with_margin if r["margin_category"] == "faible"])
    medium_margin = len([r for r in products_with_margin if r["margin_category"] == "moyen"])
    high_margin = len([r for r in products_with_margin if r["margin_category"] == "bon"])
    undefined_margin = len([r for r in analysis_results if r["margin_category"] == "undefined"])
    
    return {
        "sections": list(sections_with_products.values()),
        "all_products": analysis_results,
        "statistics": {
            "total": total_products,
            "faible": low_margin,
            "moyen": medium_margin,
            "bon": high_margin,
            "undefined": undefined_margin
        }
    }

def _get_margin_category(margin: float, thresholds: dict) -> str:
    """Retourne la catégorie de marge: faible (rouge), moyen (jaune), bon (vert)"""
    low = thresholds.get("low", 20.0)
    high = thresholds.get("high", 50.0)
    
    if margin < low:
        return "faible"
    elif margin >= high:
        return "bon"
    else:
        return "moyen"

# --- Fiche Technique Products ---

async def get_preparation_cost_per_unit(preparation_id: str, restaurant_id: str) -> tuple:
    """Récupère le coût par unité d'une préparation"""
    prep = await fiche_technique_products_collection.find_one({
        "product_id": preparation_id,
        "restaurant_id": restaurant_id,
        "product_type": "preparation",
        "is_active": True
    })
    if not prep:
        return None, None, None
    cost_per_unit = prep.get("cost_per_unit", 0)
    yield_unit = prep.get("yield_unit", "unité")
    name = prep.get("name", "")
    return cost_per_unit, yield_unit, name

def calculate_boisson_format_cost(purchase_info: dict, format_size: float, format_unit: str) -> float:
    """Calcule le coût d'un format de vente pour une boisson"""
    if not purchase_info:
        return 0.0
    
    purchase_qty = purchase_info.get("quantity", 0)
    purchase_unit = purchase_info.get("unit", "cl")
    purchase_price = purchase_info.get("price", 0)
    
    # Conversion en unité de base (ml)
    conversions = {"ml": 1, "cl": 10, "l": 1000}
    
    purchase_in_ml = purchase_qty * conversions.get(purchase_unit, 1)
    format_in_ml = format_size * conversions.get(format_unit, 1)
    
    if purchase_in_ml == 0:
        return 0.0
    
    return (format_in_ml / purchase_in_ml) * purchase_price

@api_router.post("/fiche-products/create")
async def create_fiche_product(
    create_request: CreateFicheProductRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer un produit/recette dans une section - supporte 3 types: standard, preparation, boisson_multi"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que la section existe
    section = await fiche_technique_sections_collection.find_one({
        "section_id": create_request.section_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    })
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")
    
    product_type = create_request.product_type or "standard"
    product_id = f"fprod_{uuid.uuid4().hex[:12]}"
    
    # Base du produit
    product = {
        "product_id": product_id,
        "section_id": create_request.section_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": create_request.name,
        "photo_base64": create_request.photo_base64,
        "notes": create_request.notes,  # Notes de préparation optionnelles
        "product_type": product_type,
        "is_active": True,
        "archived": False,  # Par défaut non archivé
        "created_at": datetime.now(timezone.utc)
    }
    
    # === TYPE: STANDARD ou PREPARATION ===
    if product_type in ["standard", "preparation"]:
        # Préparer les ingrédients avec calcul du prix de revient
        ingredients_with_cost = []
        if create_request.ingredients:
            for ing in create_request.ingredients:
                ing_dict = ing.dict()
                
                # Si c'est une préparation comme ingrédient
                if ing_dict.get("ingredient_type") == "preparation" and ing_dict.get("preparation_id"):
                    cost_per_unit, yield_unit, prep_name = await get_preparation_cost_per_unit(
                        ing_dict["preparation_id"], 
                        current_user["restaurant_id"]
                    )
                    if cost_per_unit is not None:
                        ing_dict["cost"] = cost_per_unit * ing_dict.get("quantity_used", 1)
                        ing_dict["preparation_name"] = prep_name
                        ing_dict["unit_used"] = yield_unit
                else:
                    # Ingrédient standard
                    ing_dict["ingredient_type"] = "standard"
                    ing_dict["cost"] = calculate_ingredient_cost(ing_dict)
                
                ingredients_with_cost.append(ing_dict)
        
        total_cost = sum(ing.get("cost", 0) for ing in ingredients_with_cost)
        
        # Calcul du prix de vente
        multiplier = create_request.multiplier
        selling_price_override = create_request.selling_price_override
        
        if selling_price_override is not None:
            # Prix manuel → recalculer le multiplicateur
            selling_price = selling_price_override
            multiplier = selling_price / total_cost if total_cost > 0 else None
        elif multiplier and total_cost > 0:
            selling_price = total_cost * multiplier
        else:
            selling_price = None
        
        product.update({
            "recipe_unit": create_request.recipe_unit,
            "multiplier": round(multiplier, 2) if multiplier else None,
            "selling_price_override": selling_price_override,
            "ingredients": ingredients_with_cost,
            "total_cost": total_cost
        })
        
        # Champs spécifiques pour PREPARATION
        if product_type == "preparation":
            yield_qty = create_request.yield_quantity or 1
            cost_per_unit = total_cost / yield_qty if yield_qty > 0 else 0
            product.update({
                "yield_quantity": yield_qty,
                "yield_unit": create_request.yield_unit or "portion",
                "cost_per_unit": cost_per_unit,
                "selling_price": None  # Les préparations n'ont pas de prix de vente
            })
        else:
            product["selling_price"] = selling_price
    
    # === TYPE: BOISSON_MULTI ===
    elif product_type == "boisson_multi":
        purchase_info = create_request.purchase_info.dict() if create_request.purchase_info else None
        
        # Calculer le coût pour chaque format de vente
        selling_formats_with_cost = []
        if create_request.selling_formats:
            for sf in create_request.selling_formats:
                sf_dict = sf.dict()
                sf_dict["format_id"] = sf_dict.get("format_id") or f"sf_{uuid.uuid4().hex[:8]}"
                sf_dict["cost"] = calculate_boisson_format_cost(
                    purchase_info, 
                    sf_dict.get("size", 0), 
                    sf_dict.get("unit", "cl")
                )
                selling_formats_with_cost.append(sf_dict)
        
        product.update({
            "purchase_info": purchase_info,
            "selling_formats": selling_formats_with_cost
        })
    
    await fiche_technique_products_collection.insert_one(product)
    
    # Préparer la réponse
    response = {k: v for k, v in product.items() if k != "_id"}
    
    # Arrondir les valeurs
    if response.get("total_cost"):
        response["total_cost"] = round(response["total_cost"], 2)
    if response.get("selling_price"):
        response["selling_price"] = round(response["selling_price"], 2)
    if response.get("cost_per_unit"):
        response["cost_per_unit"] = round(response["cost_per_unit"], 4)
    for ing in response.get("ingredients", []):
        if ing.get("cost"):
            ing["cost"] = round(ing["cost"], 2)
    for sf in response.get("selling_formats", []):
        if sf.get("cost"):
            sf["cost"] = round(sf["cost"], 2)
    
    return response

@api_router.get("/fiche-products/list")
async def list_fiche_products(
    section_id: Optional[str] = None,
    include_archived: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Lister les produits (optionnel: filtrer par section). Par défaut exclut les archivés."""
    query = {"restaurant_id": current_user["restaurant_id"], "is_active": True}
    if not include_archived:
        query["$or"] = [{"archived": False}, {"archived": {"$exists": False}}]
    if section_id:
        query["section_id"] = section_id
    
    products = await fiche_technique_products_collection.find(
        query, {"_id": 0}
    ).to_list(500)
    
    # Arrondir les valeurs pour l'affichage
    for p in products:
        if p.get("total_cost"):
            p["total_cost"] = round(p["total_cost"], 2)
        if p.get("selling_price"):
            p["selling_price"] = round(p["selling_price"], 2)
        for ing in p.get("ingredients", []):
            if ing.get("cost"):
                ing["cost"] = round(ing["cost"], 2)
    
    return products

@api_router.get("/fiche-products/archived")
async def list_archived_products(
    current_user: dict = Depends(get_current_user)
):
    """Lister tous les produits archivés"""
    products = await fiche_technique_products_collection.find({
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True,
        "archived": True
    }, {"_id": 0}).to_list(500)
    
    # Ajouter les infos de section pour chaque produit
    for p in products:
        section = await fiche_technique_sections_collection.find_one(
            {"section_id": p.get("section_id")},
            {"_id": 0, "name": 1, "category": 1}
        )
        if section:
            p["section_name"] = section.get("name", "")
            p["section_category"] = section.get("category", "")
        
        if p.get("total_cost"):
            p["total_cost"] = round(p["total_cost"], 2)
        if p.get("selling_price"):
            p["selling_price"] = round(p["selling_price"], 2)
    
    return products

@api_router.put("/fiche-products/{product_id}/archive")
async def archive_fiche_product(
    product_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Archiver un produit"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await fiche_technique_products_collection.update_one(
        {"product_id": product_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"archived": True, "archived_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Produit archivé avec succès", "product_id": product_id}

@api_router.put("/fiche-products/{product_id}/restore")
async def restore_fiche_product(
    product_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Restaurer un produit archivé"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await fiche_technique_products_collection.update_one(
        {"product_id": product_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"archived": False}, "$unset": {"archived_at": ""}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Produit restauré avec succès", "product_id": product_id}

@api_router.delete("/fiche-products/{product_id}/permanent")
async def permanently_delete_fiche_product(
    product_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer définitivement un produit (uniquement les produits archivés)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que le produit est archivé avant suppression définitive
    product = await fiche_technique_products_collection.find_one({
        "product_id": product_id,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    if not product.get("archived"):
        raise HTTPException(status_code=400, detail="Seuls les produits archivés peuvent être supprimés définitivement. Archivez d'abord le produit.")
    
    result = await fiche_technique_products_collection.delete_one({
        "product_id": product_id,
        "restaurant_id": current_user["restaurant_id"]
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Produit supprimé définitivement", "product_id": product_id}

@api_router.get("/fiche-products/by-section/{section_id}")
async def get_products_by_section(
    section_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer tous les produits d'une section"""
    products = await fiche_technique_products_collection.find({
        "section_id": section_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    }, {"_id": 0}).to_list(500)
    
    for p in products:
        if p.get("total_cost"):
            p["total_cost"] = round(p["total_cost"], 2)
        if p.get("selling_price"):
            p["selling_price"] = round(p["selling_price"], 2)
    
    return products

@api_router.get("/fiche-products/{product_id}")
async def get_fiche_product(
    product_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer les détails d'un produit"""
    product = await fiche_technique_products_collection.find_one({
        "product_id": product_id,
        "restaurant_id": current_user["restaurant_id"]
    }, {"_id": 0})
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return product

@api_router.put("/fiche-products/{product_id}")
async def update_fiche_product(
    product_id: str,
    update_request: UpdateFicheProductRequest,
    current_user: dict = Depends(get_current_user)
):
    """Modifier un produit et recalculer les coûts - supporte tous les types de produits"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Récupérer le produit existant
    existing = await fiche_technique_products_collection.find_one({
        "product_id": product_id,
        "restaurant_id": current_user["restaurant_id"]
    })
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")
    
    product_type = update_request.product_type or existing.get("product_type", "standard")
    update_data = {"product_type": product_type}
    
    if update_request.name is not None:
        update_data["name"] = update_request.name
    if update_request.photo_base64 is not None:
        update_data["photo_base64"] = update_request.photo_base64
    if update_request.notes is not None:
        update_data["notes"] = update_request.notes
    if update_request.is_active is not None:
        update_data["is_active"] = update_request.is_active
    if update_request.archived is not None:
        update_data["archived"] = update_request.archived
    
    # === TYPE: STANDARD ou PREPARATION ===
    if product_type in ["standard", "preparation"]:
        if update_request.recipe_unit is not None:
            update_data["recipe_unit"] = update_request.recipe_unit
        
        # Traiter les ingrédients
        ingredients = update_request.ingredients if update_request.ingredients is not None else existing.get("ingredients", [])
        
        if update_request.ingredients is not None:
            ingredients_with_cost = []
            for ing in update_request.ingredients:
                ing_dict = ing.dict() if hasattr(ing, 'dict') else ing
                
                # Si c'est une préparation comme ingrédient
                if ing_dict.get("ingredient_type") == "preparation" and ing_dict.get("preparation_id"):
                    cost_per_unit, yield_unit, prep_name = await get_preparation_cost_per_unit(
                        ing_dict["preparation_id"], 
                        current_user["restaurant_id"]
                    )
                    if cost_per_unit is not None:
                        ing_dict["cost"] = cost_per_unit * ing_dict.get("quantity_used", 1)
                        ing_dict["preparation_name"] = prep_name
                        ing_dict["unit_used"] = yield_unit
                else:
                    ing_dict["ingredient_type"] = "standard"
                    ing_dict["cost"] = calculate_ingredient_cost(ing_dict)
                
                ingredients_with_cost.append(ing_dict)
            update_data["ingredients"] = ingredients_with_cost
            ingredients = ingredients_with_cost
        
        # Calculer le coût total
        total_cost = sum(ing.get("cost", 0) for ing in ingredients)
        update_data["total_cost"] = total_cost
        
        # Gestion du prix de vente / multiplicateur
        multiplier = update_request.multiplier if update_request.multiplier is not None else existing.get("multiplier")
        selling_price_override = update_request.selling_price_override
        
        if selling_price_override is not None:
            # Prix manuel → recalculer le multiplicateur
            update_data["selling_price"] = selling_price_override
            update_data["selling_price_override"] = selling_price_override
            update_data["multiplier"] = round(selling_price_override / total_cost, 2) if total_cost > 0 else None
        elif update_request.multiplier is not None:
            update_data["multiplier"] = update_request.multiplier
            if total_cost > 0:
                update_data["selling_price"] = total_cost * update_request.multiplier
            update_data["selling_price_override"] = None
        elif multiplier and total_cost > 0:
            update_data["selling_price"] = total_cost * multiplier
        
        # Champs spécifiques pour PREPARATION
        if product_type == "preparation":
            if update_request.yield_quantity is not None:
                update_data["yield_quantity"] = update_request.yield_quantity
            if update_request.yield_unit is not None:
                update_data["yield_unit"] = update_request.yield_unit
            
            yield_qty = update_data.get("yield_quantity", existing.get("yield_quantity", 1))
            update_data["cost_per_unit"] = total_cost / yield_qty if yield_qty > 0 else 0
            update_data["selling_price"] = None  # Préparations n'ont pas de prix de vente
    
    # === TYPE: BOISSON_MULTI ===
    elif product_type == "boisson_multi":
        purchase_info = None
        if update_request.purchase_info is not None:
            purchase_info = update_request.purchase_info.dict() if hasattr(update_request.purchase_info, 'dict') else update_request.purchase_info
            update_data["purchase_info"] = purchase_info
        else:
            purchase_info = existing.get("purchase_info")
        
        if update_request.selling_formats is not None:
            selling_formats_with_cost = []
            for sf in update_request.selling_formats:
                sf_dict = sf.dict() if hasattr(sf, 'dict') else sf
                sf_dict["format_id"] = sf_dict.get("format_id") or f"sf_{uuid.uuid4().hex[:8]}"
                sf_dict["cost"] = calculate_boisson_format_cost(
                    purchase_info, 
                    sf_dict.get("size", 0), 
                    sf_dict.get("unit", "cl")
                )
                selling_formats_with_cost.append(sf_dict)
            update_data["selling_formats"] = selling_formats_with_cost
    
    if update_data:
        await fiche_technique_products_collection.update_one(
            {"product_id": product_id},
            {"$set": update_data}
        )
    
    product = await fiche_technique_products_collection.find_one(
        {"product_id": product_id},
        {"_id": 0}
    )
    
    # Arrondir les valeurs
    if product.get("total_cost"):
        product["total_cost"] = round(product["total_cost"], 2)
    if product.get("selling_price"):
        product["selling_price"] = round(product["selling_price"], 2)
    if product.get("cost_per_unit"):
        product["cost_per_unit"] = round(product["cost_per_unit"], 4)
    for ing in product.get("ingredients", []):
        if ing.get("cost"):
            ing["cost"] = round(ing["cost"], 2)
    for sf in product.get("selling_formats", []):
        if sf.get("cost"):
            sf["cost"] = round(sf["cost"], 2)
    
    return product

@api_router.get("/fiche-products/preparations/list")
async def list_preparations(
    category: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister toutes les préparations disponibles (pour les utiliser comme ingrédients)"""
    query = {
        "restaurant_id": current_user["restaurant_id"],
        "product_type": "preparation",
        "is_active": True
    }
    
    # Si on veut filtrer par catégorie (bar/cuisine)
    if category:
        # Récupérer les sections de cette catégorie
        sections = await fiche_technique_sections_collection.find({
            "restaurant_id": current_user["restaurant_id"],
            "category": category,
            "is_active": True
        }).to_list(100)
        section_ids = [s["section_id"] for s in sections]
        query["section_id"] = {"$in": section_ids}
    
    preparations = await fiche_technique_products_collection.find(
        query, {"_id": 0}
    ).to_list(500)
    
    # Retourner les infos essentielles pour la sélection
    result = []
    for p in preparations:
        result.append({
            "product_id": p["product_id"],
            "name": p["name"],
            "yield_quantity": p.get("yield_quantity", 1),
            "yield_unit": p.get("yield_unit", "portion"),
            "cost_per_unit": round(p.get("cost_per_unit", 0), 4),
            "total_cost": round(p.get("total_cost", 0), 2),
            "section_id": p["section_id"]
        })
    
    return result

@api_router.delete("/fiche-products/{product_id}")
async def delete_fiche_product(
    product_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer un produit (soft delete)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await fiche_technique_products_collection.update_one(
        {"product_id": product_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Product deleted successfully"}

# --- Fiche Technique Export (PDF/Excel) ---

@api_router.post("/fiche-products/export-pdf")
async def export_fiche_pdf(
    product_ids: List[str] = Body(...),
    include_prices: bool = Body(default=True),
    current_user: dict = Depends(get_current_user)
):
    """Exporter les fiches techniques sélectionnées en PDF
    - AVEC prix: Ingrédient, Quantité utilisée, Prix achat, Coût + Récapitulatif
    - SANS prix: Ingrédient, Quantité utilisée uniquement + Notes
    Note: La colonne "quantité achetée" a été retirée comme demandé
    """
    # Récupérer les produits
    products = await fiche_technique_products_collection.find({
        "product_id": {"$in": product_ids},
        "restaurant_id": current_user["restaurant_id"]
    }, {"_id": 0}).to_list(100)
    
    if not products:
        raise HTTPException(status_code=404, detail="No products found")
    
    # Récupérer les infos du restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    )
    
    # Obtenir les couleurs du restaurant
    primary_color = restaurant.get("primary_color", "#26252D") if restaurant else "#26252D"
    secondary_color = restaurant.get("secondary_color", "#EAE6CA") if restaurant else "#EAE6CA"
    primary_rgb = hex_to_rgb(primary_color)
    secondary_rgb = hex_to_rgb(secondary_color)
    
    # Créer le PDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    for product in products:
        pdf.add_page()
        
        # Titre du produit
        pdf.set_font('Helvetica', 'B', 18)
        pdf.cell(0, 10, safe_text(product.get("name", "")), ln=True, align='C')
        pdf.ln(5)
        
        # Section
        section = await fiche_technique_sections_collection.find_one(
            {"section_id": product.get("section_id")},
            {"_id": 0}
        )
        if section:
            pdf.set_font('Helvetica', '', 12)
            category_label = "Bar" if section.get("category") == "bar" else "Cuisine"
            pdf.cell(0, 8, safe_text(f"{category_label} - {section.get('name', '')}"), ln=True, align='C')
        
        pdf.ln(10)
        
        # Tableau des ingrédients
        ingredients = product.get("ingredients", [])
        if ingredients:
            pdf.set_font('Helvetica', 'B', 11)
            pdf.set_fill_color(*primary_rgb)  # Couleur primaire du restaurant
            pdf.set_text_color(*secondary_rgb)  # Couleur secondaire du restaurant
            
            # En-têtes selon include_prices (sans "quantité achetée")
            if include_prices:
                col_widths = [60, 40, 40, 30]
                headers = ["Ingredient", "Qte utilisee", "Prix achat", "Cout"]
                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], 8, header, 1, 0 if i < len(headers)-1 else 1, 'C', True)
            else:
                # Sans prix: seulement Ingrédient et Quantité utilisée
                col_widths = [100, 70]
                headers = ["Ingredient", "Quantite"]
                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], 8, header, 1, 0 if i < len(headers)-1 else 1, 'C', True)
            
            # Lignes
            pdf.set_font('Helvetica', '', 10)
            pdf.set_text_color(0, 0, 0)
            
            for ing in ingredients:
                if include_prices:
                    pdf.cell(col_widths[0], 7, safe_text(ing.get("name", "")[:25]), 1, 0, 'L')
                    pdf.cell(col_widths[1], 7, safe_text(f"{ing.get('quantity_used', '')} {ing.get('unit_used', '')}"), 1, 0, 'C')
                    pdf.cell(col_widths[2], 7, safe_text(f"{ing.get('purchase_price', 0):.2f} EUR"), 1, 0, 'C')
                    pdf.cell(col_widths[3], 7, safe_text(f"{ing.get('cost', 0):.2f} EUR"), 1, 1, 'C')
                else:
                    # Sans prix: seulement nom et quantité utilisée
                    pdf.cell(col_widths[0], 7, safe_text(ing.get("name", "")[:40]), 1, 0, 'L')
                    pdf.cell(col_widths[1], 7, safe_text(f"{ing.get('quantity_used', '')} {ing.get('unit_used', '')}"), 1, 1, 'C')
        
        pdf.ln(5)
        
        # Notes de préparation (toujours affichées si présentes)
        notes = product.get("notes")
        if notes:
            pdf.set_font('Helvetica', 'B', 11)
            pdf.cell(0, 8, "Notes / Instructions:", ln=True)
            pdf.set_font('Helvetica', '', 10)
            # Utiliser multi_cell pour les notes multi-lignes
            pdf.multi_cell(0, 6, safe_text(notes))
            pdf.ln(3)
        
        # Récapitulatif (seulement si include_prices)
        if include_prices:
            pdf.set_font('Helvetica', 'B', 12)
            total_cost = product.get("total_cost", 0)
            multiplier = product.get("multiplier")
            selling_price = product.get("selling_price")
            
            pdf.cell(0, 8, f"Prix de revient: {total_cost:.2f} EUR", ln=True)
            if multiplier:
                pdf.cell(0, 8, f"Multiplicateur: x{multiplier}", ln=True)
            if selling_price:
                pdf.cell(0, 8, f"Prix de vente: {selling_price:.2f} EUR", ln=True)
    
    # Retourner le PDF
    pdf_bytes = pdf.output()
    
    return Response(
        content=bytes(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=fiches_techniques.pdf"}
    )

@api_router.post("/fiche-products/export-excel")
async def export_fiche_excel(
    product_ids: List[str] = Body(...),
    include_prices: bool = Body(default=True),
    current_user: dict = Depends(get_current_user)
):
    """Exporter les fiches techniques sélectionnées en Excel (CSV)
    - AVEC prix: Produit, Section, Catégorie, Ingrédient, Qte utilisée, Prix achat, Coût, Prix revient, etc.
    - SANS prix: Produit, Ingrédient, Quantité uniquement + Notes
    Note: La colonne "quantité achetée" a été retirée comme demandé
    """
    # Récupérer les produits
    products = await fiche_technique_products_collection.find({
        "product_id": {"$in": product_ids},
        "restaurant_id": current_user["restaurant_id"]
    }, {"_id": 0}).to_list(100)
    
    if not products:
        raise HTTPException(status_code=404, detail="No products found")
    
    # Créer le CSV
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # En-têtes selon include_prices (sans "quantité achetée")
    if include_prices:
        writer.writerow([
            "Produit", "Section", "Categorie", 
            "Ingredient", "Qte utilisee", "Unite",
            "Prix achat", "Cout ingredient", 
            "Prix revient total", "Multiplicateur", "Prix vente", "Notes"
        ])
    else:
        # Sans prix: seulement Produit, Ingrédient, Quantité et Notes
        writer.writerow([
            "Produit", "Ingredient", "Quantite", "Notes"
        ])
    
    for product in products:
        section = await fiche_technique_sections_collection.find_one(
            {"section_id": product.get("section_id")},
            {"_id": 0}
        )
        section_name = section.get("name", "") if section else ""
        category = section.get("category", "") if section else ""
        notes = product.get("notes", "") or ""
        
        ingredients = product.get("ingredients", [])
        if ingredients:
            for i, ing in enumerate(ingredients):
                if include_prices:
                    writer.writerow([
                        product.get("name", ""),
                        section_name,
                        "Bar" if category == "bar" else "Cuisine",
                        ing.get("name", ""),
                        ing.get("quantity_used", ""),
                        ing.get("unit_used", ""),
                        ing.get("purchase_price", ""),
                        f"{ing.get('cost', 0):.2f}",
                        f"{product.get('total_cost', 0):.2f}" if i == 0 else "",
                        product.get("multiplier", "") if i == 0 else "",
                        f"{product.get('selling_price', 0):.2f}" if product.get("selling_price") and i == 0 else "",
                        notes if i == 0 else ""  # Notes seulement sur la première ligne
                    ])
                else:
                    # Sans prix: format simplifié
                    writer.writerow([
                        product.get("name", ""),
                        ing.get("name", ""),
                        f"{ing.get('quantity_used', '')} {ing.get('unit_used', '')}",
                        notes if i == 0 else ""  # Notes seulement sur la première ligne
                    ])
        else:
            # Produit sans ingrédients
            if include_prices:
                writer.writerow([
                    product.get("name", ""),
                    section_name,
                    "Bar" if category == "bar" else "Cuisine",
                    "", "", "", "", "",
                    f"{product.get('total_cost', 0):.2f}",
                    product.get("multiplier", ""),
                    f"{product.get('selling_price', 0):.2f}" if product.get("selling_price") else "",
                    notes
                ])
            else:
                writer.writerow([
                    product.get("name", ""),
                    "", "", notes
                ])
    
    csv_content = output.getvalue()
    
    return Response(
        content=csv_content.encode('utf-8-sig'),  # UTF-8 with BOM for Excel
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=fiches_techniques.csv"}
    )
    
    csv_content = output.getvalue()
    
    return Response(
        content=csv_content.encode('utf-8-sig'),  # UTF-8 with BOM for Excel
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=fiches_techniques.csv"}
    )

@api_router.post("/fiche-products/import-from-pdf")
async def import_fiche_products_from_pdf(
    import_request: ImportFicheProductsRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Importer des produits en masse depuis des données extraites d'un PDF.
    Crée automatiquement les sections manquantes si create_sections=True.
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    restaurant_id = current_user["restaurant_id"]
    imported_count = 0
    skipped_count = 0
    created_sections = []
    errors = []
    
    # Regrouper les produits par section et catégorie
    for product_data in import_request.products:
        try:
            category = product_data.category.lower()
            if category not in ["bar", "cuisine"]:
                category = "cuisine"  # Par défaut cuisine si non spécifié
            
            section_name = product_data.section_name.strip()
            
            # Chercher ou créer la section
            section = await fiche_technique_sections_collection.find_one({
                "restaurant_id": restaurant_id,
                "category": category,
                "name": {"$regex": f"^{section_name}$", "$options": "i"},
                "is_active": True
            })
            
            if not section:
                if import_request.create_sections:
                    # Créer la nouvelle section
                    existing_sections = await fiche_technique_sections_collection.find({
                        "restaurant_id": restaurant_id,
                        "category": category
                    }).to_list(100)
                    max_order = max([s.get("order", 0) for s in existing_sections], default=-1)
                    
                    section_id = f"fsec_{uuid.uuid4().hex[:12]}"
                    new_section = {
                        "section_id": section_id,
                        "restaurant_id": restaurant_id,
                        "category": category,
                        "name": section_name,
                        "order": max_order + 1,
                        "is_preparations": False,
                        "is_active": True,
                        "created_at": datetime.now(timezone.utc)
                    }
                    await fiche_technique_sections_collection.insert_one(new_section)
                    section = new_section
                    created_sections.append({"name": section_name, "category": category})
                else:
                    errors.append(f"Section '{section_name}' non trouvée pour '{product_data.product_name}'")
                    skipped_count += 1
                    continue
            
            # Vérifier si le produit existe déjà
            existing_product = await fiche_technique_products_collection.find_one({
                "restaurant_id": restaurant_id,
                "section_id": section["section_id"],
                "name": {"$regex": f"^{product_data.product_name}$", "$options": "i"},
                "is_active": True
            })
            
            if existing_product:
                skipped_count += 1
                continue  # Skip existing products
            
            # Préparer les ingrédients
            ingredients_with_cost = []
            if product_data.ingredients:
                for ing in product_data.ingredients:
                    ing_name = ing.get("name", "")
                    if not ing_name:
                        continue
                    
                    # Créer un ingrédient avec des valeurs par défaut
                    ingredient = {
                        "ingredient_type": "standard",
                        "name": ing_name,
                        "quantity_used": ing.get("quantity", 1) if ing.get("quantity") else 1,
                        "unit_used": ing.get("unit", "pièce") if ing.get("unit") else "pièce",
                        "quantity_purchased": 1,
                        "unit_purchased": "pièce",
                        "purchase_price": 0,
                        "cost": 0
                    }
                    ingredients_with_cost.append(ingredient)
            
            # Créer le produit
            product_id = f"fprod_{uuid.uuid4().hex[:12]}"
            product = {
                "product_id": product_id,
                "section_id": section["section_id"],
                "restaurant_id": restaurant_id,
                "name": product_data.product_name,
                "photo_base64": None,
                "notes": product_data.instructions_or_notes,
                "product_type": "standard",
                "is_active": True,
                "archived": False,
                "created_at": datetime.now(timezone.utc),
                "recipe_unit": None,
                "multiplier": None,
                "selling_price_override": product_data.selling_price,
                "ingredients": ingredients_with_cost,
                "total_cost": 0,
                "selling_price": product_data.selling_price
            }
            
            await fiche_technique_products_collection.insert_one(product)
            imported_count += 1
            
        except Exception as e:
            errors.append(f"Erreur pour '{product_data.product_name}': {str(e)}")
            skipped_count += 1
    
    return {
        "success": True,
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "created_sections": created_sections,
        "errors": errors[:10]  # Limiter les erreurs retournées
    }

# ==================== MENU RESTAURANT ENDPOINTS ====================

@api_router.post("/menu-restaurant/sections/create")
async def create_menu_restaurant_section(
    create_request: CreateMenuRestaurantSectionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une section dans le Menu Restaurant (Food ou Boisson)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Valider le type de menu
    if create_request.menu_type not in ["food", "boisson"]:
        raise HTTPException(status_code=400, detail="menu_type doit être 'food' ou 'boisson'")
    
    # Si c'est une sous-section, vérifier que le parent existe
    if create_request.parent_section_id:
        parent = await menu_restaurant_sections_collection.find_one({
            "section_id": create_request.parent_section_id,
            "restaurant_id": current_user["restaurant_id"]
        })
        if not parent:
            raise HTTPException(status_code=404, detail="Section parente non trouvée")
    
    # Calculer l'ordre
    existing_sections = await menu_restaurant_sections_collection.find(
        {
            "restaurant_id": current_user["restaurant_id"],
            "menu_type": create_request.menu_type,
            "parent_section_id": create_request.parent_section_id
        }
    ).to_list(100)
    max_order = max([s.get("order", 0) for s in existing_sections], default=-1)
    
    section_id = f"mrs_{uuid.uuid4().hex[:12]}"
    section = {
        "section_id": section_id,
        "restaurant_id": current_user["restaurant_id"],
        "menu_type": create_request.menu_type,
        "name": create_request.name,
        "parent_section_id": create_request.parent_section_id,
        "order": create_request.order if create_request.order is not None else max_order + 1,
        "color": create_request.color,
        "has_happy_hour": create_request.has_happy_hour,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await menu_restaurant_sections_collection.insert_one(section)
    
    return {
        "section_id": section_id,
        "restaurant_id": current_user["restaurant_id"],
        "menu_type": create_request.menu_type,
        "name": create_request.name,
        "parent_section_id": create_request.parent_section_id,
        "order": section["order"],
        "color": create_request.color,
        "has_happy_hour": create_request.has_happy_hour,
        "is_active": True
    }
    
    # Trigger automatic translation regeneration
    trigger_translation_regeneration(current_user["restaurant_id"])
    
    return {
        "section_id": section_id,
        "menu_type": create_request.menu_type,
        "name": create_request.name,
        "parent_section_id": create_request.parent_section_id,
        "order": section["order"],
        "color": create_request.color,
        "has_happy_hour": create_request.has_happy_hour,
        "is_active": True
    }

@api_router.get("/menu-restaurant/sections/list")
async def list_menu_restaurant_sections(
    menu_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les sections du Menu Restaurant"""
    query = {"restaurant_id": current_user["restaurant_id"], "is_active": True}
    if menu_type:
        query["menu_type"] = menu_type
    
    sections = await menu_restaurant_sections_collection.find(query, {"_id": 0}).sort("order", 1).to_list(200)
    return sections

@api_router.put("/menu-restaurant/sections/{section_id}")
async def update_menu_restaurant_section(
    section_id: str,
    update_request: UpdateMenuRestaurantSectionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour une section du Menu Restaurant"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await menu_restaurant_sections_collection.update_one(
            {"section_id": section_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    # Trigger automatic translation regeneration
    trigger_translation_regeneration(current_user["restaurant_id"])
    
    section_doc = await menu_restaurant_sections_collection.find_one(
        {"section_id": section_id},
        {"_id": 0}
    )
    return section_doc

@api_router.delete("/menu-restaurant/sections/{section_id}")
async def delete_menu_restaurant_section(
    section_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une section du Menu Restaurant"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Soft delete
    result = await menu_restaurant_sections_collection.update_one(
        {"section_id": section_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Section non trouvée")
    
    # Also soft delete all items in this section
    await menu_restaurant_items_collection.update_many(
        {"section_id": section_id},
        {"$set": {"is_active": False}}
    )
    
    # Trigger automatic translation regeneration
    trigger_translation_regeneration(current_user["restaurant_id"])
    
    return {"message": "Section supprimée avec succès"}

@api_router.post("/menu-restaurant/items/create")
async def create_menu_restaurant_item(
    create_request: CreateMenuRestaurantItemRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer un item dans une section du Menu Restaurant"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Vérifier que la section existe
    section = await menu_restaurant_sections_collection.find_one({
        "section_id": create_request.section_id,
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    })
    if not section:
        raise HTTPException(status_code=404, detail="Section non trouvée")
    
    # Calculer l'ordre
    existing_items = await menu_restaurant_items_collection.find(
        {"section_id": create_request.section_id, "is_active": True}
    ).to_list(500)
    max_order = max([i.get("order", 0) for i in existing_items], default=-1)
    
    # Préparer les formats avec IDs
    formats = []
    if create_request.formats:
        for fmt in create_request.formats:
            formats.append({
                "format_id": f"fmt_{uuid.uuid4().hex[:8]}",
                "name": fmt.name,
                "price": fmt.price,
                "happy_hour_price": fmt.happy_hour_price
            })
    
    # Préparer les suggestions avec IDs
    suggestions = []
    if create_request.suggestions:
        for sug in create_request.suggestions:
            suggestions.append({
                "suggestion_id": f"sug_{uuid.uuid4().hex[:8]}",
                "name": sug.name,
                "price": sug.price
            })
    
    # Préparer les suppléments avec IDs
    supplements = []
    if create_request.supplements:
        for sup in create_request.supplements:
            supplements.append({
                "supplement_id": f"sup_{uuid.uuid4().hex[:8]}",
                "name": sup.name,
                "price": sup.price
            })
    
    # Préparer les options payantes avec IDs
    options = []
    if create_request.options:
        for opt in create_request.options:
            options.append({
                "option_id": f"opt_{uuid.uuid4().hex[:8]}",
                "name": opt.name,
                "price": opt.price
            })
    
    item_id = f"mri_{uuid.uuid4().hex[:12]}"
    item = {
        "item_id": item_id,
        "section_id": create_request.section_id,
        "restaurant_id": current_user["restaurant_id"],
        "fiche_technique_product_id": create_request.fiche_technique_product_id,
        "name": create_request.name,
        "descriptions": create_request.descriptions,
        "price": create_request.price,
        "formats": formats,
        "suggestions": suggestions,
        "supplements": supplements,
        "options": options,
        "order": create_request.order if create_request.order is not None else max_order + 1,
        "is_active": True,
        "created_at": datetime.now(timezone.utc),
        # Prix et TVA
        "happy_hour_price": create_request.happy_hour_price,
        "tva_rate": create_request.tva_rate or 10.0,
        # Nouveaux champs
        "allergens": create_request.allergens or [],
        "tags": create_request.tags or [],
        "excel_status": create_request.excel_status or "added",  # Nouveaux items = ajoutés (vert)
        "modified_fields": create_request.modified_fields or ["name", "price", "description"]  # Tous en vert
    }
    await menu_restaurant_items_collection.insert_one(item)
    
    # Trigger automatic translation regeneration
    trigger_translation_regeneration(current_user["restaurant_id"])
    
    return {
        "item_id": item_id,
        "section_id": create_request.section_id,
        "name": create_request.name,
        "descriptions": create_request.descriptions,
        "price": create_request.price,
        "happy_hour_price": create_request.happy_hour_price,
        "tva_rate": create_request.tva_rate or 10.0,
        "formats": formats,
        "suggestions": suggestions,
        "supplements": supplements,
        "options": options,
        "order": item["order"],
        "is_active": True,
        "allergens": create_request.allergens or [],
        "tags": create_request.tags or []
    }

@api_router.get("/menu-restaurant/items/list")
async def list_menu_restaurant_items(
    section_id: Optional[str] = None,
    menu_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les items du Menu Restaurant"""
    query = {"restaurant_id": current_user["restaurant_id"], "is_active": True}
    
    if section_id:
        query["section_id"] = section_id
    elif menu_type:
        # Get all section IDs for this menu type
        sections = await menu_restaurant_sections_collection.find(
            {"restaurant_id": current_user["restaurant_id"], "menu_type": menu_type, "is_active": True},
            {"section_id": 1}
        ).to_list(200)
        section_ids = [s["section_id"] for s in sections]
        query["section_id"] = {"$in": section_ids}
    
    items = await menu_restaurant_items_collection.find(query, {"_id": 0}).sort("order", 1).to_list(1000)
    return items

@api_router.put("/menu-restaurant/items/{item_id}")
async def update_menu_restaurant_item(
    item_id: str,
    update_request: UpdateMenuRestaurantItemRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour un item du Menu Restaurant"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {}
    
    if update_request.name is not None:
        update_data["name"] = update_request.name
    if update_request.descriptions is not None:
        update_data["descriptions"] = update_request.descriptions
    if update_request.price is not None:
        update_data["price"] = update_request.price
    if update_request.order is not None:
        update_data["order"] = update_request.order
    if update_request.is_active is not None:
        update_data["is_active"] = update_request.is_active
    
    # Handle formats with IDs
    if update_request.formats is not None:
        formats = []
        for fmt in update_request.formats:
            formats.append({
                "format_id": fmt.format_id or f"fmt_{uuid.uuid4().hex[:8]}",
                "name": fmt.name,
                "price": fmt.price,
                "happy_hour_price": fmt.happy_hour_price
            })
        update_data["formats"] = formats
    
    # Handle suggestions with IDs
    if update_request.suggestions is not None:
        suggestions = []
        for sug in update_request.suggestions:
            suggestions.append({
                "suggestion_id": sug.suggestion_id or f"sug_{uuid.uuid4().hex[:8]}",
                "name": sug.name,
                "price": sug.price
            })
        update_data["suggestions"] = suggestions
    
    # Handle supplements with IDs
    if update_request.supplements is not None:
        supplements = []
        for sup in update_request.supplements:
            supplements.append({
                "supplement_id": sup.supplement_id or f"sup_{uuid.uuid4().hex[:8]}",
                "name": sup.name,
                "price": sup.price
            })
        update_data["supplements"] = supplements
    
    # Handle options with IDs
    if update_request.options is not None:
        options = []
        for opt in update_request.options:
            options.append({
                "option_id": opt.option_id or f"opt_{uuid.uuid4().hex[:8]}",
                "name": opt.name,
                "price": opt.price
            })
        update_data["options"] = options
    
    # Handle happy_hour_price and tva_rate
    if update_request.happy_hour_price is not None:
        update_data["happy_hour_price"] = update_request.happy_hour_price
    if update_request.tva_rate is not None:
        update_data["tva_rate"] = update_request.tva_rate
    
    # Handle allergens, tags and excel status
    if update_request.allergens is not None:
        update_data["allergens"] = update_request.allergens
    if update_request.tags is not None:
        update_data["tags"] = update_request.tags
    if update_request.excel_status is not None:
        update_data["excel_status"] = update_request.excel_status
    if update_request.modified_fields is not None:
        update_data["modified_fields"] = update_request.modified_fields
    
    if update_data:
        await menu_restaurant_items_collection.update_one(
            {"item_id": item_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    # Trigger automatic translation regeneration
    trigger_translation_regeneration(current_user["restaurant_id"])
    
    item_doc = await menu_restaurant_items_collection.find_one(
        {"item_id": item_id},
        {"_id": 0}
    )
    return item_doc

@api_router.delete("/menu-restaurant/items/{item_id}")
async def delete_menu_restaurant_item(
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer un item du Menu Restaurant"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await menu_restaurant_items_collection.update_one(
        {"item_id": item_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item non trouvé")
    
    # Trigger automatic translation regeneration
    trigger_translation_regeneration(current_user["restaurant_id"])
    
    return {"message": "Item supprimé avec succès"}

@api_router.post("/menu-restaurant/notes/create")
async def create_menu_restaurant_note(
    create_request: CreateMenuRestaurantNoteRequest,
    current_user: dict = Depends(get_current_user)
):
    """Créer une note de menu (ex: Happy Hour -20%)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Valider le type de menu
    if create_request.menu_type not in ["food", "boisson"]:
        raise HTTPException(status_code=400, detail="menu_type doit être 'food' ou 'boisson'")
    
    # Calculer l'ordre
    existing_notes = await menu_restaurant_notes_collection.find(
        {"restaurant_id": current_user["restaurant_id"], "menu_type": create_request.menu_type, "is_active": True}
    ).to_list(100)
    max_order = max([n.get("order", 0) for n in existing_notes], default=-1)
    
    note_id = f"mrn_{uuid.uuid4().hex[:12]}"
    note = {
        "note_id": note_id,
        "restaurant_id": current_user["restaurant_id"],
        "menu_type": create_request.menu_type,
        "content": create_request.content,
        "order": create_request.order if create_request.order is not None else max_order + 1,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    await menu_restaurant_notes_collection.insert_one(note)
    
    return {
        "note_id": note_id,
        "menu_type": create_request.menu_type,
        "content": create_request.content,
        "order": note["order"],
        "is_active": True
    }

@api_router.get("/menu-restaurant/notes/list")
async def list_menu_restaurant_notes(
    menu_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lister les notes du Menu Restaurant"""
    query = {"restaurant_id": current_user["restaurant_id"], "is_active": True}
    if menu_type:
        query["menu_type"] = menu_type
    
    notes = await menu_restaurant_notes_collection.find(query, {"_id": 0}).sort("order", 1).to_list(100)
    return notes

@api_router.put("/menu-restaurant/notes/{note_id}")
async def update_menu_restaurant_note(
    note_id: str,
    update_request: UpdateMenuRestaurantNoteRequest,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour une note du Menu Restaurant"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in update_request.dict().items() if v is not None}
    
    if update_data:
        await menu_restaurant_notes_collection.update_one(
            {"note_id": note_id, "restaurant_id": current_user["restaurant_id"]},
            {"$set": update_data}
        )
    
    note_doc = await menu_restaurant_notes_collection.find_one(
        {"note_id": note_id},
        {"_id": 0}
    )
    return note_doc

@api_router.delete("/menu-restaurant/notes/{note_id}")
async def delete_menu_restaurant_note(
    note_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une note du Menu Restaurant"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await menu_restaurant_notes_collection.update_one(
        {"note_id": note_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Note non trouvée")
    
    return {"message": "Note supprimée avec succès"}

# ==================== MENU RESTAURANT REORDERING ENDPOINTS ====================

@api_router.put("/menu-restaurant/sections/{section_id}/set-order")
async def set_section_order(
    section_id: str,
    request: SetOrderRequest,
    current_user: dict = Depends(get_current_user)
):
    """Définir directement l'ordre d'une section"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await menu_restaurant_sections_collection.update_one(
        {"section_id": section_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"order": request.order}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Section non trouvée")
    
    section_doc = await menu_restaurant_sections_collection.find_one(
        {"section_id": section_id},
        {"_id": 0}
    )
    return section_doc

@api_router.put("/menu-restaurant/items/{item_id}/set-order")
async def set_item_order(
    item_id: str,
    request: SetOrderRequest,
    current_user: dict = Depends(get_current_user)
):
    """Définir directement l'ordre d'un item"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await menu_restaurant_items_collection.update_one(
        {"item_id": item_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"order": request.order}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item non trouvé")
    
    item_doc = await menu_restaurant_items_collection.find_one(
        {"item_id": item_id},
        {"_id": 0}
    )
    return item_doc

@api_router.put("/menu-restaurant/bulk-reorder")
async def bulk_reorder_menu(
    reorder_data: Dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """Réordonner en masse les sections et/ou items du menu
    
    Body format:
    {
        "sections": [{"section_id": "...", "order": 1}, ...],
        "items": [{"item_id": "...", "order": 1}, ...]
    }
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    restaurant_id = current_user["restaurant_id"]
    updated_sections = 0
    updated_items = 0
    
    # Update sections
    if "sections" in reorder_data:
        for section_data in reorder_data["sections"]:
            result = await menu_restaurant_sections_collection.update_one(
                {"section_id": section_data["section_id"], "restaurant_id": restaurant_id},
                {"$set": {"order": section_data["order"]}}
            )
            if result.modified_count > 0:
                updated_sections += 1
    
    # Update items
    if "items" in reorder_data:
        for item_data in reorder_data["items"]:
            result = await menu_restaurant_items_collection.update_one(
                {"item_id": item_data["item_id"], "restaurant_id": restaurant_id},
                {"$set": {"order": item_data["order"]}}
            )
            if result.modified_count > 0:
                updated_items += 1
    
    return {
        "message": "Réordonnancement terminé",
        "updated_sections": updated_sections,
        "updated_items": updated_items
    }

@api_router.get("/menu-restaurant/fiche-products/search")
async def search_fiche_products_for_menu(
    query: Optional[str] = None,
    category: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Rechercher des produits Fiche Technique pour les ajouter au menu"""
    search_query = {
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True,
        "archived": {"$ne": True}
    }
    
    if query:
        search_query["name"] = {"$regex": query, "$options": "i"}
    
    if category:
        # Get sections for this category
        sections = await fiche_technique_sections_collection.find(
            {"restaurant_id": current_user["restaurant_id"], "category": category, "is_active": True},
            {"section_id": 1}
        ).to_list(100)
        section_ids = [s["section_id"] for s in sections]
        search_query["section_id"] = {"$in": section_ids}
    
    products = await fiche_technique_products_collection.find(
        search_query,
        {"_id": 0, "product_id": 1, "name": 1, "section_id": 1, "product_type": 1}
    ).to_list(100)
    
    # Enrich with section info
    for product in products:
        section = await fiche_technique_sections_collection.find_one(
            {"section_id": product.get("section_id")},
            {"_id": 0, "name": 1, "category": 1}
        )
        if section:
            product["section_name"] = section.get("name")
            product["category"] = section.get("category")
    
    return products

# ==================== MENU RESTAURANT EXPORT ENDPOINTS ====================

@api_router.get("/menu-restaurant/export-pdf/{menu_type}")
async def export_menu_restaurant_pdf(
    menu_type: str,
    current_user: dict = Depends(get_current_user)
):
    """Exporter la carte (food ou boisson) en PDF avec gestion des descriptions longues et sections non coupées"""
    if menu_type not in ["food", "boisson"]:
        raise HTTPException(status_code=400, detail="menu_type doit être 'food' ou 'boisson'")
    
    restaurant_id = current_user["restaurant_id"]
    
    # Get restaurant info
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    
    # Get all sections for this menu type
    sections = await menu_restaurant_sections_collection.find({
        "restaurant_id": restaurant_id,
        "menu_type": menu_type,
        "is_active": True
    }, {"_id": 0}).sort("order", 1).to_list(100)
    
    # Get all items
    all_items = await menu_restaurant_items_collection.find({
        "restaurant_id": restaurant_id,
        "is_active": True
    }, {"_id": 0}).sort("order", 1).to_list(500)
    
    # Get notes
    notes = await menu_restaurant_notes_collection.find({
        "restaurant_id": restaurant_id,
        "menu_type": menu_type,
        "is_active": True
    }, {"_id": 0}).sort("order", 1).to_list(50)
    
    # Create PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=False)  # Disable auto page break to control manually
    
    page_width = 210
    page_height = 297
    margin = 15
    content_width = page_width - 2 * margin
    price_width = 45  # Largeur réservée pour le prix
    name_width = content_width - price_width  # Largeur pour le nom
    desc_width = content_width - price_width - 5  # Descriptions s'arrêtent AVANT les prix
    
    def check_page_break(required_height):
        """Vérifier si on doit passer à une nouvelle page"""
        if pdf.get_y() + required_height > page_height - margin:
            pdf.add_page()
            return True
        return False
    
    def estimate_section_height(section_items, sub_sections_list, all_items_list):
        """Estimer la hauteur d'une section complète"""
        height = 12  # Titre section
        for item in section_items:
            height += 8  # Nom + prix
            for desc in item.get("descriptions", []):
                if desc:
                    # Estimer le nombre de lignes pour la description
                    desc_chars = len(desc)
                    lines = max(1, desc_chars // 60 + 1)  # ~60 chars par ligne
                    height += lines * 5
            if item.get("formats"):
                height += len(item.get("formats", [])) * 5
            if item.get("suggestions"):
                height += len(item.get("suggestions", [])) * 5
        for sub in sub_sections_list:
            height += 10  # Titre sous-section
            sub_items = [i for i in all_items_list if i.get("section_id") == sub.get("section_id")]
            for item in sub_items:
                height += 7
                for desc in item.get("descriptions", []):
                    if desc:
                        desc_chars = len(desc)
                        lines = max(1, desc_chars // 55 + 1)
                        height += lines * 5
        return height + 10  # Marge
    
    # Labels pour allergènes et tags (version PDF sans emojis pour compatibilité latin-1)
    ALLERGENS_PDF = {
        'gluten': 'Gluten',
        'crustaces': 'Crustaces',
        'oeufs': 'Oeufs',
        'poissons': 'Poissons',
        'arachides': 'Arachides',
        'soja': 'Soja',
        'lactose': 'Lactose',
        'fruits_a_coque': 'Fruits a coque',
        'fruits_coque': 'Fruits a coque',
        'celeri': 'Celeri',
        'moutarde': 'Moutarde',
        'sesame': 'Sesame',
        'sulfites': 'Sulfites',
        'lupin': 'Lupin',
        'mollusques': 'Mollusques'
    }
    
    TAGS_PDF = {
        'vegetarien': 'Vegetarien',
        'vegan': 'Vegan',
        'epice': 'Epice'
    }
    
    def render_item(item, indent=0):
        """Render un item avec gestion des descriptions longues"""
        indent_str = "  " * indent
        
        # Vérifier le status de l'item pour la couleur
        item_status = item.get("status")
        status_colors = {
            "added": (46, 125, 50),      # Vert
            "a_ajouter": (46, 125, 50),  # Vert
            "deleted": (200, 75, 49),     # Rouge
            "a_supprimer": (200, 75, 49), # Rouge
            "modified": (128, 0, 128),    # Violet
            "a_modifier": (128, 0, 128)   # Violet
        }
        
        # Couleur du nom selon status (noir par défaut)
        if item_status and item_status in status_colors:
            r, g, b = status_colors[item_status]
            pdf.set_text_color(r, g, b)
        else:
            pdf.set_text_color(0, 0, 0)
        
        pdf.set_font("Helvetica", "B", 11)
        item_name = item.get("name", "")
        x_start = pdf.get_x() + indent * 4
        
        if item.get("price"):
            price_str = f"{float(item['price']):.2f}e"
            # Nom à gauche (coloré selon status)
            pdf.set_xy(margin + indent * 4, pdf.get_y())
            pdf.cell(name_width - indent * 4, 6, safe_text(item_name), ln=False)
            # Prix à droite (toujours noir)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(price_width, 6, safe_text(price_str), ln=True, align="R")
        elif item.get("formats"):
            pdf.set_xy(margin + indent * 4, pdf.get_y())
            pdf.cell(name_width, 6, safe_text(item_name), ln=True)
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(0, 0, 0)  # Formats en noir
            for fmt in item.get("formats", []):
                fmt_name = fmt.get("name", "")
                fmt_price = f"{float(fmt.get('price', 0)):.2f}e"
                pdf.set_xy(margin + indent * 4 + 4, pdf.get_y())
                pdf.cell(name_width - indent * 4 - 4, 5, safe_text(fmt_name), ln=False)
                pdf.cell(price_width, 5, safe_text(fmt_price), ln=True, align="R")
        else:
            pdf.set_xy(margin + indent * 4, pdf.get_y())
            pdf.cell(name_width, 6, safe_text(item_name), ln=True)
        
        # Descriptions - colorées selon status ou marron par défaut
        pdf.set_font("Helvetica", "I", 9)
        if item_status and item_status in status_colors:
            r, g, b = status_colors[item_status]
            pdf.set_text_color(r, g, b)
        else:
            pdf.set_text_color(139, 90, 43)  # Marron par défaut
        
        for desc in item.get("descriptions", []):
            if desc:
                pdf.set_xy(margin + indent * 4 + 4, pdf.get_y())
                # multi_cell pour wrap automatique - limiter à desc_width
                pdf.multi_cell(desc_width - indent * 4 - 20, 4, safe_text(desc), ln=True)
        
        # Allergènes - texte bleu, limité en largeur pour ne pas dépasser le prix
        item_allergens = item.get("allergens", [])
        if item_allergens:
            allergens_text = "Allergenes: " + ", ".join([ALLERGENS_PDF.get(a, a) for a in item_allergens])
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(41, 98, 180)  # Texte bleu
            pdf.set_xy(margin + indent * 4 + 4, pdf.get_y())
            # Utiliser multi_cell avec largeur limitée pour éviter de dépasser le prix
            pdf.multi_cell(desc_width - indent * 4 - 20, 3.5, safe_text(allergens_text), ln=True)
        
        # Tags - texte orange
        item_tags = item.get("tags", [])
        if item_tags:
            tags_text = ", ".join([TAGS_PDF.get(t, t) for t in item_tags])
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(255, 140, 0)  # Orange
            pdf.set_xy(margin + indent * 4 + 4, pdf.get_y())
            pdf.cell(desc_width - indent * 4 - 20, 3.5, safe_text(tags_text), ln=True)
        
        # Pas de badge "[À ajouter]" - la couleur du nom/desc indique déjà le status
        
        # Suggestions
        if item.get("suggestions"):
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(200, 75, 49)  # Rouge
            for sug in item.get("suggestions", []):
                sug_str = f"Suggestion: {sug.get('name', '')} +{float(sug.get('price', 0)):.2f}e"
                pdf.set_xy(margin + indent * 4 + 4, pdf.get_y())
                pdf.cell(desc_width, 4, safe_text(sug_str), ln=True)
        
        pdf.ln(1)
    
    # Title
    pdf.set_font("Helvetica", "B", 20)
    title = "Carte Food" if menu_type == "food" else "Carte Boisson"
    pdf.cell(0, 15, safe_text(title), ln=True, align="C")
    
    if restaurant:
        pdf.set_font("Helvetica", "", 12)
        pdf.cell(0, 8, safe_text(restaurant.get("name", "")), ln=True, align="C")
    
    pdf.ln(8)
    
    # Organize sections hierarchically
    parent_sections = [s for s in sections if not s.get("parent_section_id")]
    
    for parent in parent_sections:
        # Get items and sub-sections for this section
        section_items = [i for i in all_items if i.get("section_id") == parent.get("section_id")]
        sub_sections_list = [s for s in sections if s.get("parent_section_id") == parent.get("section_id")]
        
        # Estimer la hauteur de la section complète
        estimated_height = estimate_section_height(section_items, sub_sections_list, all_items)
        
        # Espace restant sur la page actuelle (page A4 = 297mm, marge bas = 15mm)
        current_y = pdf.get_y()
        remaining_space = 297 - 15 - current_y
        
        # Logique de saut de page pour garder les sections entières:
        # Si la section entière ne tient pas dans l'espace restant ET
        # qu'on n'est pas au tout début de la page (y > 60) -> nouvelle page
        # Cela évite de couper une section entre deux pages
        if estimated_height > remaining_space and current_y > 60:
            pdf.add_page()
        
        # Section header
        pdf.set_font("Helvetica", "B", 14)
        section_color = (44, 95, 45) if menu_type == "food" else (27, 73, 101)
        pdf.set_text_color(*section_color)
        pdf.cell(0, 10, safe_text(parent.get("name", "").upper()), ln=True)
        
        # Items de la section
        for item in section_items:
            check_page_break(25)  # Hauteur min pour un item
            render_item(item, indent=0)
        
        # Sub-sections
        for sub in sub_sections_list:
            sub_items = [i for i in all_items if i.get("section_id") == sub.get("section_id")]
            
            # Vérifier si on peut afficher la sous-section en entier
            sub_height = 10 + len(sub_items) * 15
            check_page_break(min(sub_height, 40))
            
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(91, 94, 166)  # Purple
            pdf.cell(0, 8, safe_text(f"  {sub.get('name', '')}"), ln=True)
            
            for item in sub_items:
                check_page_break(20)
                render_item(item, indent=1)
        
        # Happy Hour table if applicable
        if parent.get("has_happy_hour"):
            hh_items = []
            for item in section_items:
                for fmt in item.get("formats", []):
                    if fmt.get("happy_hour_price"):
                        hh_items.append({
                            "name": f"{item.get('name', '')} ({fmt.get('name', '')})",
                            "price": fmt.get("price", 0),
                            "hh_price": fmt.get("happy_hour_price", 0)
                        })
            
            if hh_items:
                check_page_break(len(hh_items) * 5 + 15)
                pdf.ln(3)
                pdf.set_font("Helvetica", "B", 11)
                pdf.set_text_color(255, 107, 53)  # Orange
                pdf.cell(0, 6, safe_text("Happy Hour"), ln=True)
                pdf.set_font("Helvetica", "", 10)
                for hh in hh_items:
                    pdf.cell(0, 5, safe_text(f"  {hh['name']}: {float(hh['price']):.2f}e -> {float(hh['hh_price']):.2f}e"), ln=True)
        
        pdf.ln(5)
    
    # Notes at the end
    if notes:
        check_page_break(len(notes) * 6 + 15)
        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(139, 0, 0)  # Dark red
        pdf.cell(0, 8, "Notes", ln=True)
        pdf.set_font("Helvetica", "I", 10)
        for note in notes:
            pdf.multi_cell(content_width, 5, safe_text(f"  {note.get('content', '')}"), ln=True)
    
    # Generate PDF bytes
    pdf_bytes = pdf.output()
    filename = f"carte_{menu_type}.pdf"
    
    return Response(
        content=bytes(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/menu-restaurant/export-excel/{menu_type}")
async def export_menu_restaurant_excel(
    menu_type: str,
    current_user: dict = Depends(get_current_user)
):
    """Exporter la carte (food ou boisson) en Excel (.xlsx) avec couleurs et allergènes/tags"""
    if menu_type not in ["food", "boisson"]:
        raise HTTPException(status_code=400, detail="menu_type doit être 'food' ou 'boisson'")
    
    restaurant_id = current_user["restaurant_id"]
    
    # Récupérer les couleurs du restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "primary_color": 1, "secondary_color": 1}
    )
    primary_color = restaurant.get("primary_color", "#26252D") if restaurant else "#26252D"
    secondary_color = restaurant.get("secondary_color", "#EAE6CA") if restaurant else "#EAE6CA"
    
    # Get all sections for this menu type
    sections = await menu_restaurant_sections_collection.find({
        "restaurant_id": restaurant_id,
        "menu_type": menu_type,
        "is_active": True
    }, {"_id": 0}).sort("order", 1).to_list(100)
    
    # Get all items
    all_items = await menu_restaurant_items_collection.find({
        "restaurant_id": restaurant_id,
        "is_active": True
    }, {"_id": 0}).sort("order", 1).to_list(500)
    
    # Create section lookup
    section_lookup = {s["section_id"]: s for s in sections}
    
    # Import openpyxl for real Excel with colors
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    # Allergens mapping for display
    ALLERGENS_LABELS = {
        'gluten': '🌾 Gluten',
        'crustaces': '🦐 Crustacés',
        'oeufs': '🥚 Œufs',
        'poissons': '🐟 Poissons',
        'arachides': '🥜 Arachides',
        'soja': '🌱 Soja',
        'lactose': '🥛 Lactose',
        'fruits_coque': '🌰 Fruits à coque',
        'celeri': '🌿 Céleri',
        'moutarde': '🌭 Moutarde',
        'sesame': '🥖 Sésame',
        'sulfites': '🍷 Sulfites',
        'lupin': '🌼 Lupin',
        'mollusques': '🐚 Mollusques'
    }
    
    TAGS_LABELS = {
        'vegetarien': '🥬 Végétarien',
        'vegan': '🥦 Végan',
        'epice': '🌶️ Épicé'
    }
    
    # Status colors for Excel (ARGB format: alpha + RGB)
    EXCEL_STATUS_COLORS = {
        'added': 'FF28A745',      # Vert - pour les nouveaux items
        'deleted': 'FFDC3545',    # Rouge - pour les items à supprimer
        'modified': 'FF9B59B6',   # Violet - pour les items modifiés
        'normal': None            # Pas de couleur
    }
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Carte {menu_type.capitalize()}"
    
    # Style for headers - utiliser les couleurs du restaurant
    primary_hex = primary_color.lstrip('#')
    secondary_hex = secondary_color.lstrip('#')
    header_fill = PatternFill(start_color=f"FF{primary_hex}", end_color=f"FF{primary_hex}", fill_type="solid")
    header_font = Font(bold=True, color=f"FF{secondary_hex}", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Section", "Sous-section", "Produit", "Description", 
        "Format", "Prix", "Prix HH",
        "Tags", "Allergènes"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column widths
    ws.column_dimensions['A'].width = 20  # Section
    ws.column_dimensions['B'].width = 18  # Sous-section
    ws.column_dimensions['C'].width = 30  # Produit
    ws.column_dimensions['D'].width = 45  # Description
    ws.column_dimensions['E'].width = 12  # Format
    ws.column_dimensions['F'].width = 10  # Prix
    ws.column_dimensions['G'].width = 10  # Prix HH
    ws.column_dimensions['H'].width = 25  # Tags
    ws.column_dimensions['I'].width = 50  # Allergènes
    
    # Filter items by menu type (through section)
    section_ids = [s["section_id"] for s in sections]
    menu_items = [i for i in all_items if i.get("section_id") in section_ids]
    
    row_num = 2
    for item in menu_items:
        section = section_lookup.get(item.get("section_id"), {})
        section_name = section.get("name", "")
        
        # Check if it's a sub-section
        parent_section_id = section.get("parent_section_id")
        if parent_section_id:
            parent_section = section_lookup.get(parent_section_id, {})
            parent_name = parent_section.get("name", "")
            sub_section_name = section_name
        else:
            parent_name = section_name
            sub_section_name = ""
        
        descriptions = " | ".join(item.get("descriptions", []))
        
        # Get tags and allergens labels
        tags_list = item.get("tags", [])
        tags_display = ", ".join([TAGS_LABELS.get(t, t) for t in tags_list])
        
        allergens_list = item.get("allergens", [])
        allergens_display = ", ".join([ALLERGENS_LABELS.get(a, a) for a in allergens_list])
        
        # Get excel_status and modified_fields for coloring
        excel_status = item.get("excel_status", "normal")
        modified_fields = item.get("modified_fields", []) or []
        
        # Couleur de base
        brown_color = "FF8B5A2B"  # Marron pour description normale
        
        # Déterminer les couleurs pour chaque champ (Produit=name, Prix=price, Description=description)
        def get_field_font(field_name):
            """Retourne la police colorée selon le statut et les champs modifiés"""
            if excel_status == 'added':
                # Vert pour TOUS les champs si ajouté
                return Font(color=EXCEL_STATUS_COLORS['added'])
            elif excel_status == 'deleted':
                # Rouge pour TOUS les champs si supprimé
                return Font(color=EXCEL_STATUS_COLORS['deleted'])
            elif excel_status == 'modified':
                # Violet UNIQUEMENT pour les champs sélectionnés
                if field_name in modified_fields:
                    return Font(color=EXCEL_STATUS_COLORS['modified'])
                elif field_name == 'description':
                    return Font(color=brown_color)  # Marron par défaut
                return None
            else:
                # Normal - description en marron
                if field_name == 'description':
                    return Font(color=brown_color)
                return None
        
        def write_row_with_color(section_val, sub_val, name_val, desc_val, format_val, price_val, hh_val):
            nonlocal row_num
            row_data = [
                section_val, sub_val, name_val, desc_val, format_val, price_val, hh_val,
                tags_display, allergens_display
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True if col in [4, 8, 9] else False)
                
                # Appliquer les couleurs selon le champ
                if col == 3:  # Produit (name)
                    font = get_field_font('name')
                    if font:
                        cell.font = font
                elif col == 4:  # Description
                    font = get_field_font('description')
                    if font:
                        cell.font = font
                elif col in [5, 6, 7]:  # Format, Prix, Prix HH
                    font = get_field_font('price')
                    if font:
                        cell.font = font
            row_num += 1
        
        if item.get("price"):
            # Simple price
            write_row_with_color(
                parent_name, sub_section_name, item.get("name", ""),
                descriptions, "", f"{item['price']:.2f}€", ""
            )
        elif item.get("formats"):
            # Multiple formats
            for idx, fmt in enumerate(item.get("formats", [])):
                hh_price = f"{fmt.get('happy_hour_price', 0):.2f}€" if fmt.get("happy_hour_price") else ""
                write_row_with_color(
                    parent_name if idx == 0 else "",
                    sub_section_name if idx == 0 else "",
                    item.get("name", "") if idx == 0 else "",
                    descriptions if idx == 0 else "",
                    fmt.get("name", ""),
                    f"{fmt.get('price', 0):.2f}€",
                    hh_price
                )
        else:
            # No price
            write_row_with_color(
                parent_name, sub_section_name, item.get("name", ""),
                descriptions, "", "", ""
            )
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"carte_{menu_type}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/menu-restaurant/export-csv/{menu_type}")
async def export_menu_restaurant_csv(
    menu_type: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Exporter la carte en format Excel/CSV avec colonnes:
    Food: A=Nom, B=Description, C=Prix unique, D=Petite, E=Grande + Allergènes + Tags
    Boisson: A=Nom, B=Description, C=Prix unique, D=Taille1, E=Taille2, F=Taille3 + Prix HH + Allergènes + Tags
    """
    if menu_type not in ["food", "boisson"]:
        raise HTTPException(status_code=400, detail="menu_type doit être 'food' ou 'boisson'")
    
    restaurant_id = current_user.get("restaurant_id")
    if not restaurant_id:
        raise HTTPException(status_code=400, detail="Restaurant non trouvé")
    
    # Get restaurant name
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "name": 1}
    )
    restaurant_name = restaurant.get("name", "Restaurant") if restaurant else "Restaurant"
    
    # Get all sections for this menu type
    sections = await menu_restaurant_sections_collection.find({
        "restaurant_id": restaurant_id,
        "menu_type": menu_type,
        "is_active": True
    }, {"_id": 0}).sort("order", 1).to_list(100)
    
    # Get all items
    all_items = await menu_restaurant_items_collection.find({
        "restaurant_id": restaurant_id,
        "is_active": True
    }, {"_id": 0}).sort("order", 1).to_list(500)
    
    # Allergens list (in order)
    ALLERGENS = ['gluten', 'crustaces', 'oeufs', 'poissons', 'arachides', 'soja', 
                 'lactose', 'fruits_coque', 'celeri', 'moutarde', 'sesame', 
                 'sulfites', 'lupin', 'mollusques']
    
    ALLERGENS_LABELS = {
        'gluten': 'Gluten', 'crustaces': 'Crustacés', 'oeufs': 'Œufs',
        'poissons': 'Poissons', 'arachides': 'Arachides', 'soja': 'Soja',
        'lactose': 'Lactose', 'fruits_coque': 'Fruits à coque', 'celeri': 'Céleri',
        'moutarde': 'Moutarde', 'sesame': 'Sésame', 'sulfites': 'Sulfites',
        'lupin': 'Lupin', 'mollusques': 'Mollusques'
    }
    
    TAGS = ['vegetarien', 'vegan', 'epice']
    TAGS_LABELS = {'vegetarien': 'Végétarien', 'vegan': 'Végan', 'epice': 'Épicé'}
    
    # Create Excel workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Carte {menu_type.capitalize()}"
    
    # Define colors
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    price_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    allergen_header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    tag_header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    check_font = Font(color="00B050", bold=True)
    brown_font = Font(color="8B5A2B")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Row 1: Restaurant name and date
    ws.merge_cells('A1:E1')
    ws['A1'] = f"{restaurant_name} - {menu_type.upper()} MENU"
    ws['A1'].font = white_font
    ws['A1'].fill = header_fill
    ws['F1'] = datetime.now().strftime("%d/%m/%Y")
    
    row = 3
    
    # Define columns based on menu type
    if menu_type == "food":
        # Food: A=Nom, B=Description, C=Prix, D=Prix HH, E=Petite, F=HH Petite, G=Grande, H=HH Grande + Allergènes + Tags
        price_cols = 6  # C, D, E, F, G, H
        allergen_start_col = 9  # I
        tag_start_col = allergen_start_col + len(ALLERGENS)  # Après allergènes
        include_allergens_tags = True
    else:
        # Boisson: A=Nom, B=Description, C=Prix unique, D=T1, E=T2, F=T3, G=HH1, H=HH2, I=HH3 (PAS d'allergènes/tags)
        price_cols = 7  # C, D, E, F, G, H, I pour prix + HH
        include_allergens_tags = False
    
    # Build hierarchical structure
    parent_sections = [s for s in sections if not s.get("parent_section_id")]
    
    for parent in parent_sections:
        # Section header row
        ws[f'A{row}'] = parent.get('name', '').upper()
        ws[f'A{row}'].font = white_font
        ws[f'A{row}'].fill = section_fill
        
        ws[f'B{row}'] = "Description"
        ws[f'B{row}'].font = white_font
        ws[f'B{row}'].fill = section_fill
        
        if menu_type == "food":
            ws[f'C{row}'] = "Prix"
            ws[f'C{row}'].font = white_font
            ws[f'C{row}'].fill = section_fill
            
            ws[f'D{row}'] = "Prix HH"
            ws[f'D{row}'].font = white_font
            ws[f'D{row}'].fill = price_fill
            
            ws[f'E{row}'] = "Petite"
            ws[f'E{row}'].font = white_font
            ws[f'E{row}'].fill = section_fill
            
            ws[f'F{row}'] = "HH Petite"
            ws[f'F{row}'].font = white_font
            ws[f'F{row}'].fill = price_fill
            
            ws[f'G{row}'] = "Grande"
            ws[f'G{row}'].font = white_font
            ws[f'G{row}'].fill = section_fill
            
            ws[f'H{row}'] = "HH Grande"
            ws[f'H{row}'].font = white_font
            ws[f'H{row}'].fill = price_fill
        else:
            ws[f'C{row}'] = "Prix"
            ws[f'C{row}'].font = white_font
            ws[f'C{row}'].fill = section_fill
            
            ws[f'D{row}'] = "Prix HH"
            ws[f'D{row}'].font = white_font
            ws[f'D{row}'].fill = price_fill
            
            ws[f'E{row}'] = "Taille 1"
            ws[f'E{row}'].font = white_font
            ws[f'E{row}'].fill = section_fill
            
            ws[f'F{row}'] = "Taille 2"
            ws[f'F{row}'].font = white_font
            ws[f'F{row}'].fill = section_fill
            
            ws[f'G{row}'] = "Taille 3"
            ws[f'G{row}'].font = white_font
            ws[f'G{row}'].fill = section_fill
            
            ws[f'H{row}'] = "HH T1"
            ws[f'H{row}'].font = white_font
            ws[f'H{row}'].fill = price_fill
            
            ws[f'I{row}'] = "HH T2"
            ws[f'I{row}'].font = white_font
            ws[f'I{row}'].fill = price_fill
            
            ws[f'J{row}'] = "HH T3"
            ws[f'J{row}'].font = white_font
            ws[f'J{row}'].fill = price_fill
        
        # Allergens header - SEULEMENT pour Food
        if include_allergens_tags:
            ws.merge_cells(f'{get_column_letter(allergen_start_col)}{row}:{get_column_letter(allergen_start_col + len(ALLERGENS) - 1)}{row}')
            ws[f'{get_column_letter(allergen_start_col)}{row}'] = "ALLERGÈNES"
            ws[f'{get_column_letter(allergen_start_col)}{row}'].font = white_font
            ws[f'{get_column_letter(allergen_start_col)}{row}'].fill = allergen_header_fill
            ws[f'{get_column_letter(allergen_start_col)}{row}'].alignment = Alignment(horizontal='center')
            
            # Tags header
            ws.merge_cells(f'{get_column_letter(tag_start_col)}{row}:{get_column_letter(tag_start_col + len(TAGS) - 1)}{row}')
            ws[f'{get_column_letter(tag_start_col)}{row}'] = "TAGS"
            ws[f'{get_column_letter(tag_start_col)}{row}'].font = white_font
            ws[f'{get_column_letter(tag_start_col)}{row}'].fill = tag_header_fill
            ws[f'{get_column_letter(tag_start_col)}{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Sub-header with allergen names - SEULEMENT pour Food
        if include_allergens_tags:
            for i, allergen in enumerate(ALLERGENS):
                col = allergen_start_col + i
                ws[f'{get_column_letter(col)}{row}'] = ALLERGENS_LABELS.get(allergen, allergen)
                ws[f'{get_column_letter(col)}{row}'].font = Font(size=8, bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center', text_rotation=90)
            
            for i, tag in enumerate(TAGS):
                col = tag_start_col + i
                ws[f'{get_column_letter(col)}{row}'] = TAGS_LABELS.get(tag, tag)
                ws[f'{get_column_letter(col)}{row}'].font = Font(size=8, bold=True)
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center', text_rotation=90)
        
        row += 1
        
        # Get items for this section
        section_items = [i for i in all_items if i.get("section_id") == parent.get("section_id")]
        
        for item in section_items:
            # Item name (A)
            ws[f'A{row}'] = item.get("name", "")
            
            # Description (B)
            descriptions = " | ".join(item.get("descriptions", []))
            ws[f'B{row}'] = descriptions
            ws[f'B{row}'].font = brown_font
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            
            # Prices
            price = item.get("price")
            happy_hour_price = item.get("happy_hour_price")
            formats = item.get("formats", [])
            
            if menu_type == "food":
                if price:
                    ws[f'C{row}'] = f"{float(price):.2f}€"
                if happy_hour_price:
                    ws[f'D{row}'] = f"{float(happy_hour_price):.2f}€"
                if formats:
                    for fmt in formats:
                        fmt_name = fmt.get("name", "").lower()
                        fmt_price = fmt.get("price", 0)
                        fmt_hh = fmt.get("happy_hour_price", 0)
                        if fmt_price:
                            if "petite" in fmt_name:
                                ws[f'E{row}'] = f"{float(fmt_price):.2f}€"
                                if fmt_hh:
                                    ws[f'F{row}'] = f"{float(fmt_hh):.2f}€"
                            elif "grande" in fmt_name:
                                ws[f'G{row}'] = f"{float(fmt_price):.2f}€"
                                if fmt_hh:
                                    ws[f'H{row}'] = f"{float(fmt_hh):.2f}€"
            else:
                # Boisson
                if price:
                    ws[f'C{row}'] = f"{float(price):.2f}€"
                # Prix HH pour boisson simple (sans formats)
                if happy_hour_price and not formats:
                    ws[f'D{row}'] = f"{float(happy_hour_price):.2f}€"
                if formats:
                    for i, fmt in enumerate(formats[:3]):
                        col = get_column_letter(5 + i)  # E, F, G pour tailles
                        fmt_name = fmt.get("name", "")
                        fmt_price = fmt.get("price", 0)
                        hh_price = fmt.get("happy_hour_price", 0)
                        if fmt_price:
                            ws[f'{col}{row}'] = f"{fmt_name}: {float(fmt_price):.2f}€"
                        # Prix HH dans H, I, J
                        if hh_price:
                            hh_col = get_column_letter(8 + i)  # H, I, J
                            ws[f'{hh_col}{row}'] = f"{float(hh_price):.2f}€"
            
            # Options (ajouter après description si présentes)
            options = item.get("options", [])
            if options:
                options_str = " | ".join([f"+{o.get('price', 0):.2f}€ {o.get('name', '')}" for o in options])
                current_desc = ws[f'B{row}'].value or ""
                ws[f'B{row}'] = f"{current_desc}\n[Options: {options_str}]" if current_desc else f"[Options: {options_str}]"
            
            # Allergens - SEULEMENT pour Food
            if include_allergens_tags:
                item_allergens = item.get("allergens", [])
                for i, allergen in enumerate(ALLERGENS):
                    col = allergen_start_col + i
                    if allergen in item_allergens:
                        ws[f'{get_column_letter(col)}{row}'] = "✓"
                        ws[f'{get_column_letter(col)}{row}'].font = check_font
                        ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                
                # Tags
                item_tags = item.get("tags", [])
                for i, tag in enumerate(TAGS):
                    col = tag_start_col + i
                    if tag in item_tags:
                        ws[f'{get_column_letter(col)}{row}'] = "✓"
                        ws[f'{get_column_letter(col)}{row}'].font = check_font
                        ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Sub-sections
        sub_sections = [s for s in sections if s.get("parent_section_id") == parent.get("section_id")]
        for sub in sub_sections:
            ws[f'A{row}'] = f"  {sub.get('name', '')}"
            ws[f'A{row}'].font = Font(italic=True, bold=True)
            row += 1
            
            sub_items = [i for i in all_items if i.get("section_id") == sub.get("section_id")]
            for item in sub_items:
                ws[f'A{row}'] = f"    {item.get('name', '')}"
                
                descriptions = " | ".join(item.get("descriptions", []))
                ws[f'B{row}'] = descriptions
                ws[f'B{row}'].font = brown_font
                
                price = item.get("price")
                formats = item.get("formats", [])
                
                if menu_type == "food":
                    if price:
                        ws[f'C{row}'] = f"{float(price):.2f}€"
                    if formats:
                        for fmt in formats:
                            fmt_name = fmt.get("name", "").lower()
                            fmt_price = fmt.get("price", 0)
                            if fmt_price:
                                if "petite" in fmt_name:
                                    ws[f'D{row}'] = f"{float(fmt_price):.2f}€"
                                elif "grande" in fmt_name:
                                    ws[f'E{row}'] = f"{float(fmt_price):.2f}€"
                else:
                    if price:
                        ws[f'C{row}'] = f"{float(price):.2f}€"
                    if formats:
                        for i, fmt in enumerate(formats[:3]):
                            col = get_column_letter(4 + i)
                            fmt_name = fmt.get("name", "")
                            fmt_price = fmt.get("price", 0)
                            hh_price = fmt.get("happy_hour_price", 0)
                            if fmt_price:
                                ws[f'{col}{row}'] = f"{fmt_name}: {float(fmt_price):.2f}€"
                            if hh_price:
                                hh_col = get_column_letter(7 + i)
                                ws[f'{hh_col}{row}'] = f"{float(hh_price):.2f}€"
                
                # Options
                options = item.get("options", [])
                if options:
                    options_str = " | ".join([f"+{o.get('price', 0):.2f}€ {o.get('name', '')}" for o in options])
                    current_desc = ws[f'B{row}'].value or ""
                    ws[f'B{row}'] = f"{current_desc}\n[Options: {options_str}]" if current_desc else f"[Options: {options_str}]"
                
                # Allergens
                item_allergens = item.get("allergens", [])
                for i, allergen in enumerate(ALLERGENS):
                    col = allergen_start_col + i
                    if allergen in item_allergens:
                        ws[f'{get_column_letter(col)}{row}'] = "✓"
                        ws[f'{get_column_letter(col)}{row}'].font = check_font
                        ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                
                item_tags = item.get("tags", [])
                for i, tag in enumerate(TAGS):
                    col = tag_start_col + i
                    if tag in item_tags:
                        ws[f'{get_column_letter(col)}{row}'] = "✓"
                        ws[f'{get_column_letter(col)}{row}'].font = check_font
                        ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center')
                
                row += 1
        
        row += 1  # Empty row between sections
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    if menu_type == "boisson":
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"carte_{menu_type}_{sanitize_filename(restaurant_name)}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Import CSV endpoint for Menu Restaurant
class MenuRestaurantCSVImportRequest(BaseModel):
    menu_type: str  # 'food' or 'boisson'
    csv_content: str  # CSV content as string (base64 or raw)
    clear_existing: Optional[bool] = False
    update_existing: Optional[bool] = True  # Update existing items by name

# Reverse mapping for allergens (from display label to ID)
ALLERGENS_REVERSE = {
    '🌾 Gluten': 'gluten', 'Gluten': 'gluten', 'gluten': 'gluten',
    '🦐 Crustacés': 'crustaces', 'Crustacés': 'crustaces', 'crustaces': 'crustaces',
    '🥚 Œufs': 'oeufs', 'Œufs': 'oeufs', 'Oeufs': 'oeufs', 'oeufs': 'oeufs',
    '🐟 Poissons': 'poissons', 'Poissons': 'poissons', 'poissons': 'poissons',
    '🥜 Arachides': 'arachides', 'Arachides': 'arachides', 'arachides': 'arachides',
    '🌱 Soja': 'soja', 'Soja': 'soja', 'soja': 'soja',
    '🥛 Lactose': 'lactose', 'Lactose': 'lactose', 'lactose': 'lactose',
    '🌰 Fruits à coque': 'fruits_coque', 'Fruits à coque': 'fruits_coque', 'fruits_coque': 'fruits_coque',
    '🌿 Céleri': 'celeri', 'Céleri': 'celeri', 'celeri': 'celeri',
    '🌭 Moutarde': 'moutarde', 'Moutarde': 'moutarde', 'moutarde': 'moutarde',
    '🥖 Sésame': 'sesame', 'Sésame': 'sesame', 'sesame': 'sesame',
    '🍷 Sulfites': 'sulfites', 'Sulfites': 'sulfites', 'sulfites': 'sulfites',
    '🌼 Lupin': 'lupin', 'Lupin': 'lupin', 'lupin': 'lupin',
    '🐚 Mollusques': 'mollusques', 'Mollusques': 'mollusques', 'mollusques': 'mollusques'
}

# Reverse mapping for tags
TAGS_REVERSE = {
    '🥬 Végétarien': 'vegetarien', 'Végétarien': 'vegetarien', 'vegetarien': 'vegetarien',
    '🥦 Végan': 'vegan', 'Végan': 'vegan', 'vegan': 'vegan',
    '🌶️ Épicé': 'epice', 'Épicé': 'epice', 'epice': 'epice'
}

# Reverse mapping for status
STATUS_REVERSE = {
    '✓ À ajouter': 'a_ajouter', 'À ajouter': 'a_ajouter', 'a_ajouter': 'a_ajouter',
    '✏️ En modification': 'a_modifier', 'En modification': 'a_modifier', 'a_modifier': 'a_modifier',
    '✗ À supprimer': 'a_supprimer', 'À supprimer': 'a_supprimer', 'a_supprimer': 'a_supprimer',
    '': 'normal', 'Normal': 'normal', 'normal': 'normal'
}

def parse_allergens(allergens_str: str) -> list:
    """Parse allergens string from CSV to list of allergen IDs"""
    if not allergens_str:
        return []
    allergens = []
    for a in allergens_str.split(','):
        a = a.strip()
        # Try exact match first
        if a in ALLERGENS_REVERSE:
            allergens.append(ALLERGENS_REVERSE[a])
        else:
            # Try partial match (check if key contains the allergen name)
            for key, value in ALLERGENS_REVERSE.items():
                if key.lower() in a.lower() or a.lower() in key.lower():
                    allergens.append(value)
                    break
    return list(set(allergens))  # Remove duplicates

def parse_tags(tags_str: str) -> list:
    """Parse tags string from CSV to list of tag IDs"""
    if not tags_str:
        return []
    tags = []
    for t in tags_str.split(','):
        t = t.strip()
        # Try exact match first
        if t in TAGS_REVERSE:
            tags.append(TAGS_REVERSE[t])
        else:
            # Try partial match
            for key, value in TAGS_REVERSE.items():
                if key.lower() in t.lower() or t.lower() in key.lower():
                    tags.append(value)
                    break
    return list(set(tags))  # Remove duplicates

def parse_status(status_str: str) -> str:
    """Parse status string from CSV to status ID"""
    if not status_str:
        return 'normal'
    status_str = status_str.strip()
    # Try exact match first
    if status_str in STATUS_REVERSE:
        return STATUS_REVERSE[status_str]
    # Try partial match
    status_lower = status_str.lower()
    if 'ajouter' in status_lower:
        return 'a_ajouter'
    elif 'modif' in status_lower:
        return 'a_modifier'
    elif 'supprimer' in status_lower:
        return 'a_supprimer'
    return 'normal'

@api_router.post("/menu-restaurant/import-csv")
async def import_menu_restaurant_csv(
    import_request: MenuRestaurantCSVImportRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Importer une carte depuis un fichier CSV/Excel.
    Format attendu (même format que l'export):
    Section;Sous-section;Produit;Description;Format;Prix;Prix HH;Tags;Allergènes;Statut
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    menu_type = import_request.menu_type
    if menu_type not in ["food", "boisson"]:
        raise HTTPException(status_code=400, detail="menu_type doit être 'food' ou 'boisson'")
    
    restaurant_id = current_user["restaurant_id"]
    
    import csv
    from io import StringIO
    import base64
    
    # Decode CSV content if base64
    csv_content = import_request.csv_content
    try:
        # Try to decode as base64 first
        decoded = base64.b64decode(csv_content).decode('utf-8-sig')
        csv_content = decoded
    except:
        # Already plain text
        pass
    
    # Parse CSV - try semicolon first, then comma
    csv_content_io = StringIO(csv_content)
    first_line = csv_content_io.readline()
    csv_content_io.seek(0)
    
    delimiter = ';' if ';' in first_line else ','
    reader = csv.DictReader(csv_content_io, delimiter=delimiter)
    
    stats = {
        "sections_created": 0,
        "sections_updated": 0,
        "items_created": 0,
        "items_updated": 0,
        "errors": []
    }
    
    # Optionally clear existing data
    if import_request.clear_existing:
        existing_sections = await menu_restaurant_sections_collection.find(
            {"restaurant_id": restaurant_id, "menu_type": menu_type},
            {"section_id": 1}
        ).to_list(500)
        section_ids = [s["section_id"] for s in existing_sections]
        
        await menu_restaurant_items_collection.delete_many({
            "restaurant_id": restaurant_id,
            "section_id": {"$in": section_ids}
        })
        await menu_restaurant_sections_collection.delete_many({
            "restaurant_id": restaurant_id,
            "menu_type": menu_type
        })
    
    # Cache for sections
    section_cache = {}
    
    async def get_or_create_section(section_name: str, parent_id: Optional[str] = None) -> str:
        cache_key = f"{section_name}|{parent_id or ''}"
        if cache_key in section_cache:
            return section_cache[cache_key]
        
        # Search existing section
        query = {
            "restaurant_id": restaurant_id,
            "menu_type": menu_type,
            "name": section_name,
            "is_active": True
        }
        if parent_id:
            query["parent_section_id"] = parent_id
        else:
            query["parent_section_id"] = {"$exists": False}
        
        existing = await menu_restaurant_sections_collection.find_one(query)
        
        if existing:
            section_cache[cache_key] = existing["section_id"]
            return existing["section_id"]
        
        # Create new section
        max_order_doc = await menu_restaurant_sections_collection.find_one(
            {"restaurant_id": restaurant_id, "menu_type": menu_type},
            sort=[("order", -1)]
        )
        new_order = (max_order_doc.get("order", 0) if max_order_doc else 0) + 1
        
        section_id = str(uuid.uuid4())
        new_section = {
            "section_id": section_id,
            "restaurant_id": restaurant_id,
            "menu_type": menu_type,
            "name": section_name,
            "order": new_order,
            "has_happy_hour": False,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        if parent_id:
            new_section["parent_section_id"] = parent_id
        
        await menu_restaurant_sections_collection.insert_one(new_section)
        stats["sections_created"] += 1
        section_cache[cache_key] = section_id
        return section_id
    
    # Process each row
    current_item = None
    current_item_formats = []
    
    for row in reader:
        try:
            section_name = row.get("Section", "").strip()
            sub_section_name = row.get("Sous-section", "").strip()
            product_name = row.get("Produit", "").strip()
            description = row.get("Description", "").strip()
            format_name = row.get("Format", "").strip()
            prix = row.get("Prix", "").strip().replace("€", "")
            prix_hh = row.get("Prix Happy Hour", row.get("Prix HH", "")).strip().replace("€", "")
            
            # New columns for allergens, tags, status
            tags_str = row.get("Tags", "").strip()
            allergens_str = row.get("Allergènes", row.get("Allergens", "")).strip()
            status_str = row.get("Statut", row.get("Status", "")).strip()
            
            # Parse new fields
            tags = parse_tags(tags_str)
            allergens = parse_allergens(allergens_str)
            status = parse_status(status_str)
            
            # Skip empty rows
            if not section_name and not product_name and not format_name:
                continue
            
            # If this is a format line (no product name, but has format)
            if not product_name and format_name and current_item:
                # Add format to current item
                try:
                    price_val = float(prix.replace(",", ".")) if prix else 0
                    hh_val = float(prix_hh.replace(",", ".")) if prix_hh else None
                    current_item_formats.append({
                        "name": format_name,
                        "price": price_val,
                        "happy_hour_price": hh_val
                    })
                except ValueError:
                    pass
                continue
            
            # Save previous item if exists
            if current_item and current_item.get("name"):
                if current_item_formats:
                    current_item["formats"] = current_item_formats
                
                # Check if item exists
                existing_item = await menu_restaurant_items_collection.find_one({
                    "restaurant_id": restaurant_id,
                    "section_id": current_item["section_id"],
                    "name": current_item["name"],
                    "is_active": True
                })
                
                if existing_item and import_request.update_existing:
                    # Update existing item with all fields including allergens, tags, status
                    update_data = {
                        "descriptions": current_item.get("descriptions", []),
                        "formats": current_item.get("formats", []),
                        "allergens": current_item.get("allergens", []),
                        "tags": current_item.get("tags", []),
                        "status": current_item.get("status", "normal")
                    }
                    if current_item.get("price"):
                        update_data["price"] = current_item["price"]
                    
                    await menu_restaurant_items_collection.update_one(
                        {"item_id": existing_item["item_id"]},
                        {"$set": update_data}
                    )
                    stats["items_updated"] += 1
                elif not existing_item:
                    # Create new item
                    current_item["item_id"] = str(uuid.uuid4())
                    current_item["restaurant_id"] = restaurant_id
                    current_item["is_active"] = True
                    current_item["created_at"] = datetime.now(timezone.utc).isoformat()
                    
                    max_order = await menu_restaurant_items_collection.find_one(
                        {"restaurant_id": restaurant_id, "section_id": current_item["section_id"]},
                        sort=[("order", -1)]
                    )
                    current_item["order"] = (max_order.get("order", 0) if max_order else 0) + 1
                    
                    await menu_restaurant_items_collection.insert_one(current_item)
                    stats["items_created"] += 1
            
            # Start new item
            if product_name:
                # Get or create section
                parent_section_id = None
                if section_name:
                    parent_section_id = await get_or_create_section(section_name, None)
                
                # Get or create sub-section if exists
                if sub_section_name and parent_section_id:
                    section_id = await get_or_create_section(sub_section_name, parent_section_id)
                elif parent_section_id:
                    section_id = parent_section_id
                else:
                    stats["errors"].append(f"Pas de section pour le produit: {product_name}")
                    current_item = None
                    current_item_formats = []
                    continue
                
                # Parse price
                price_val = None
                if prix and not format_name:
                    try:
                        price_val = float(prix.replace(",", "."))
                    except ValueError:
                        pass
                
                # Parse descriptions (split by |)
                descriptions = [d.strip() for d in description.split("|") if d.strip()] if description else []
                
                current_item = {
                    "section_id": section_id,
                    "name": product_name,
                    "descriptions": descriptions,
                    "price": price_val,
                    "formats": [],
                    "suggestions": [],
                    "supplements": [],
                    "allergens": allergens,
                    "tags": tags,
                    "status": status
                }
                current_item_formats = []
                
                # If has format on same line
                if format_name:
                    try:
                        price_val = float(prix.replace(",", ".")) if prix else 0
                        hh_val = float(prix_hh.replace(",", ".")) if prix_hh else None
                        current_item_formats.append({
                            "name": format_name,
                            "price": price_val,
                            "happy_hour_price": hh_val
                        })
                    except ValueError:
                        pass
                    current_item["price"] = None  # Has formats, no single price
                    
        except Exception as e:
            stats["errors"].append(f"Erreur ligne: {str(e)}")
    
    # Save last item
    if current_item and current_item.get("name"):
        if current_item_formats:
            current_item["formats"] = current_item_formats
        
        existing_item = await menu_restaurant_items_collection.find_one({
            "restaurant_id": restaurant_id,
            "section_id": current_item["section_id"],
            "name": current_item["name"],
            "is_active": True
        })
        
        if existing_item and import_request.update_existing:
            update_data = {
                "descriptions": current_item.get("descriptions", []),
                "formats": current_item.get("formats", []),
                "allergens": current_item.get("allergens", []),
                "tags": current_item.get("tags", []),
                "status": current_item.get("status", "normal")
            }
            if current_item.get("price"):
                update_data["price"] = current_item["price"]
            
            await menu_restaurant_items_collection.update_one(
                {"item_id": existing_item["item_id"]},
                {"$set": update_data}
            )
            stats["items_updated"] += 1
        elif not existing_item:
            current_item["item_id"] = str(uuid.uuid4())
            current_item["restaurant_id"] = restaurant_id
            current_item["is_active"] = True
            current_item["created_at"] = datetime.now(timezone.utc).isoformat()
            current_item["order"] = 1
            await menu_restaurant_items_collection.insert_one(current_item)
            stats["items_created"] += 1
    
    return {
        "success": True,
        "stats": stats,
        "message": f"Import terminé: {stats['sections_created']} sections créées, {stats['items_created']} items créés, {stats['items_updated']} items mis à jour"
    }

# ==================== MENU RESTAURANT IMPORT PDF ====================

class MenuRestaurantPDFImportRequest(BaseModel):
    menu_type: str  # 'food' or 'boisson'
    pdf_base64: str  # PDF encoded in base64
    clear_existing: bool = False
    update_existing: bool = True

@api_router.post("/menu-restaurant/import-pdf")
async def import_menu_restaurant_pdf(
    import_request: MenuRestaurantPDFImportRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Importer une carte depuis un fichier PDF.
    Utilise pdfplumber pour extraire le texte structuré.
    """
    import pdfplumber
    import re
    
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    menu_type = import_request.menu_type
    if menu_type not in ["food", "boisson"]:
        raise HTTPException(status_code=400, detail="menu_type doit être 'food' ou 'boisson'")
    
    restaurant_id = current_user["restaurant_id"]
    
    # Decode PDF from base64
    try:
        pdf_bytes = base64.b64decode(import_request.pdf_base64)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur décodage PDF: {str(e)}")
    
    stats = {
        "sections_created": 0,
        "items_created": 0,
        "items_updated": 0,
        "errors": [],
        "extracted_data": []
    }
    
    # Extract text from PDF
    extracted_items = []
    current_section = None
    
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                if not text:
                    continue
                
                lines = text.split('\n')
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Detect section headers (usually uppercase or with specific patterns)
                    # Common patterns: "NOS ENTRÉES", "VIANDES", "DESSERTS", etc.
                    if (line.isupper() and len(line) > 2 and len(line) < 50) or \
                       (line.startswith("NOS ") or line.startswith("LES ")):
                        current_section = line.title()
                        continue
                    
                    # Try to extract item with price
                    # Pattern: "Item name ... 12,90€" or "Item name 12.90"
                    price_pattern = r'^(.+?)\s+(\d+[,\.]\d{2})\s*€?$'
                    match = re.match(price_pattern, line)
                    
                    if match and current_section:
                        item_name = match.group(1).strip()
                        price_str = match.group(2).replace(',', '.')
                        try:
                            price = float(price_str)
                            extracted_items.append({
                                "section": current_section,
                                "name": item_name,
                                "price": price,
                                "description": ""
                            })
                        except ValueError:
                            pass
                    elif current_section and len(line) > 3:
                        # Could be a description for the previous item
                        if extracted_items and not any(c.isdigit() for c in line[-5:]):
                            if not extracted_items[-1].get("description"):
                                extracted_items[-1]["description"] = line
                            else:
                                extracted_items[-1]["description"] += " " + line
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lecture PDF: {str(e)}")
    
    if not extracted_items:
        return {
            "success": False,
            "stats": stats,
            "message": "Aucun item extrait du PDF. Le format n'est peut-être pas compatible.",
            "preview": []
        }
    
    # Return preview without importing (user will confirm)
    stats["extracted_data"] = extracted_items[:100]  # Limit to 100 items for preview
    
    return {
        "success": True,
        "stats": stats,
        "message": f"{len(extracted_items)} item(s) extrait(s) du PDF",
        "preview": extracted_items[:100]
    }

@api_router.post("/menu-restaurant/import-pdf/confirm")
async def confirm_import_menu_restaurant_pdf(
    items: List[dict] = Body(...),
    menu_type: str = Body(...),
    clear_existing: bool = Body(False),
    current_user: dict = Depends(get_current_user)
):
    """Confirmer l'import des items extraits du PDF"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    restaurant_id = current_user["restaurant_id"]
    
    stats = {
        "sections_created": 0,
        "items_created": 0,
        "items_updated": 0,
        "errors": []
    }
    
    # Clear existing if requested
    if clear_existing:
        existing_sections = await menu_restaurant_sections_collection.find(
            {"restaurant_id": restaurant_id, "menu_type": menu_type},
            {"section_id": 1}
        ).to_list(500)
        section_ids = [s["section_id"] for s in existing_sections]
        
        await menu_restaurant_items_collection.delete_many({
            "restaurant_id": restaurant_id,
            "section_id": {"$in": section_ids}
        })
        await menu_restaurant_sections_collection.delete_many({
            "restaurant_id": restaurant_id,
            "menu_type": menu_type
        })
    
    # Cache for sections
    section_cache = {}
    
    async def get_or_create_section(section_name: str) -> str:
        if section_name in section_cache:
            return section_cache[section_name]
        
        existing = await menu_restaurant_sections_collection.find_one({
            "restaurant_id": restaurant_id,
            "menu_type": menu_type,
            "name": section_name,
            "is_active": True
        })
        
        if existing:
            section_cache[section_name] = existing["section_id"]
            return existing["section_id"]
        
        # Create new section
        max_order_doc = await menu_restaurant_sections_collection.find_one(
            {"restaurant_id": restaurant_id, "menu_type": menu_type},
            sort=[("order", -1)]
        )
        new_order = (max_order_doc.get("order", 0) if max_order_doc else 0) + 1
        
        section_id = str(uuid.uuid4())
        new_section = {
            "section_id": section_id,
            "restaurant_id": restaurant_id,
            "menu_type": menu_type,
            "name": section_name,
            "order": new_order,
            "has_happy_hour": False,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await menu_restaurant_sections_collection.insert_one(new_section)
        stats["sections_created"] += 1
        section_cache[section_name] = section_id
        return section_id
    
    # Import items
    for item_data in items:
        try:
            section_name = item_data.get("section", "Sans Section")
            section_id = await get_or_create_section(section_name)
            
            item_name = item_data.get("name", "").strip()
            if not item_name:
                continue
            
            # Check existing
            existing_item = await menu_restaurant_items_collection.find_one({
                "restaurant_id": restaurant_id,
                "section_id": section_id,
                "name": item_name,
                "is_active": True
            })
            
            if existing_item:
                # Update existing
                await menu_restaurant_items_collection.update_one(
                    {"item_id": existing_item["item_id"]},
                    {"$set": {
                        "price": item_data.get("price"),
                        "descriptions": [item_data.get("description", "")] if item_data.get("description") else [],
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                stats["items_updated"] += 1
            else:
                # Create new
                max_order_doc = await menu_restaurant_items_collection.find_one(
                    {"section_id": section_id},
                    sort=[("order", -1)]
                )
                new_order = (max_order_doc.get("order", 0) if max_order_doc else 0) + 1
                
                new_item = {
                    "item_id": str(uuid.uuid4()),
                    "restaurant_id": restaurant_id,
                    "section_id": section_id,
                    "name": item_name,
                    "descriptions": [item_data.get("description", "")] if item_data.get("description") else [],
                    "price": item_data.get("price"),
                    "formats": [],
                    "suggestions": [],
                    "supplements": [],
                    "allergens": [],
                    "tags": [],
                    "status": "normal",
                    "order": new_order,
                    "is_active": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await menu_restaurant_items_collection.insert_one(new_item)
                stats["items_created"] += 1
        
        except Exception as e:
            stats["errors"].append(f"Erreur item {item_data.get('name', 'unknown')}: {str(e)}")
    
    return {
        "success": True,
        "stats": stats,
        "message": f"Import terminé: {stats['sections_created']} sections créées, {stats['items_created']} items créés, {stats['items_updated']} items mis à jour"
    }

# ==================== MENU RESTAURANT BULK IMPORT ====================

class BulkImportFormat(BaseModel):
    name: str
    price: float
    happy_hour_price: Optional[float] = None

class BulkImportItem(BaseModel):
    name: str
    descriptions: Optional[List[str]] = []
    price: Optional[float] = None
    formats: Optional[List[BulkImportFormat]] = []

class BulkImportSection(BaseModel):
    name: str
    has_happy_hour: Optional[bool] = False
    items: Optional[List[BulkImportItem]] = []
    sub_sections: Optional[List['BulkImportSection']] = []

# Need to update forward reference
BulkImportSection.model_rebuild()

class BulkImportRequest(BaseModel):
    menu_type: str  # 'food' or 'boisson'
    sections: List[BulkImportSection]
    clear_existing: bool = True  # Par défaut, supprime les données existantes avant import

@api_router.post("/menu-restaurant/bulk-import")
async def bulk_import_menu_restaurant(
    import_request: BulkImportRequest,
    current_user: dict = Depends(get_current_user)
):
    """Import en masse des sections et items du Menu Restaurant"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    menu_type = import_request.menu_type
    if menu_type not in ["food", "boisson"]:
        raise HTTPException(status_code=400, detail="menu_type doit être 'food' ou 'boisson'")
    
    restaurant_id = current_user["restaurant_id"]
    stats = {"sections_created": 0, "items_created": 0, "sub_sections_created": 0}
    
    # Optionally clear existing data
    if import_request.clear_existing:
        # Get all section IDs for this menu type
        existing_sections = await menu_restaurant_sections_collection.find(
            {"restaurant_id": restaurant_id, "menu_type": menu_type},
            {"section_id": 1}
        ).to_list(500)
        section_ids = [s["section_id"] for s in existing_sections]
        
        # Delete items in these sections
        await menu_restaurant_items_collection.delete_many({
            "restaurant_id": restaurant_id,
            "section_id": {"$in": section_ids}
        })
        
        # Delete sections
        await menu_restaurant_sections_collection.delete_many({
            "restaurant_id": restaurant_id,
            "menu_type": menu_type
        })
    
    # Get current max order
    max_order_doc = await menu_restaurant_sections_collection.find_one(
        {"restaurant_id": restaurant_id, "menu_type": menu_type},
        sort=[("order", -1)]
    )
    current_order = (max_order_doc.get("order", 0) if max_order_doc else 0) + 1
    
    async def create_section_with_items(section_data: BulkImportSection, parent_id: Optional[str], order: int) -> int:
        nonlocal stats
        section_id = str(uuid.uuid4())
        
        section = {
            "section_id": section_id,
            "restaurant_id": restaurant_id,
            "menu_type": menu_type,
            "name": section_data.name,
            "parent_section_id": parent_id,
            "has_happy_hour": section_data.has_happy_hour,
            "order": order,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await menu_restaurant_sections_collection.insert_one(section)
        
        if parent_id:
            stats["sub_sections_created"] += 1
        else:
            stats["sections_created"] += 1
        
        # Create items for this section
        item_order = 1
        for item_data in section_data.items or []:
            item_id = str(uuid.uuid4())
            item = {
                "item_id": item_id,
                "restaurant_id": restaurant_id,
                "section_id": section_id,
                "name": item_data.name,
                "descriptions": item_data.descriptions or [],
                "price": item_data.price,
                "formats": [{"name": f.name, "price": f.price, "happy_hour_price": f.happy_hour_price} for f in (item_data.formats or [])],
                "suggestions": [],
                "supplements": [],
                "order": item_order,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await menu_restaurant_items_collection.insert_one(item)
            stats["items_created"] += 1
            item_order += 1
        
        next_order = order + 1
        
        # Recursively create sub-sections
        for sub_section_data in section_data.sub_sections or []:
            next_order = await create_section_with_items(sub_section_data, section_id, next_order)
        
        return next_order
    
    # Process all top-level sections
    for section_data in import_request.sections:
        current_order = await create_section_with_items(section_data, None, current_order)
    
    return {
        "message": f"Import réussi pour Carte {menu_type.capitalize()}",
        "stats": stats
    }

# ==================== A L'ARDOISE ENDPOINTS ====================
# Menu "A l'Ardoise" avec structure fixe et lien partageable permanent

class ArdoiseItem(BaseModel):
    name: str = ""
    description: str = ""
    price: Optional[float] = None
    quantity_sold: Optional[int] = None  # Quantité vendue à la fin du service

class ArdoiseSection(BaseModel):
    title: str
    items: List[ArdoiseItem]

class FormulePrices(BaseModel):
    plat_du_jour: Optional[float] = 15.90
    entree_plat: Optional[float] = 18.90
    plat_dessert: Optional[float] = 18.90
    entree_plat_dessert: Optional[float] = 23.90

class UpdateArdoiseRequest(BaseModel):
    entree: List[ArdoiseItem]
    plat: List[ArdoiseItem]
    dessert: List[ArdoiseItem]
    formule_prices: Optional[FormulePrices] = None

@api_router.get("/ardoise")
async def get_ardoise(current_user: dict = Depends(get_current_user)):
    """Récupérer l'ardoise du restaurant"""
    restaurant_id = current_user["restaurant_id"]
    
    # Chercher l'ardoise existante
    ardoise = await ardoise_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    
    # Si pas d'ardoise, créer une structure par défaut
    if not ardoise:
        ardoise = {
            "restaurant_id": restaurant_id,
            "share_token": secrets.token_urlsafe(32),
            "entree": [
                {"name": "", "description": "", "price": None, "quantity_sold": None},
                {"name": "", "description": "", "price": None, "quantity_sold": None}
            ],
            "plat": [
                {"name": "", "description": "", "price": None, "quantity_sold": None},
                {"name": "", "description": "", "price": None, "quantity_sold": None}
            ],
            "dessert": [
                {"name": "", "description": "", "price": None, "quantity_sold": None},
                {"name": "", "description": "", "price": None, "quantity_sold": None}
            ],
            "formule_prices": {
                "plat_du_jour": 15.90,
                "entree_plat": 18.90,
                "plat_dessert": 18.90,
                "entree_plat_dessert": 23.90
            },
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await ardoise_collection.insert_one(ardoise)
        ardoise.pop("_id", None)
    else:
        # Ajouter les prix de formules par défaut si absents
        if "formule_prices" not in ardoise:
            ardoise["formule_prices"] = {
                "plat_du_jour": 15.90,
                "entree_plat": 18.90,
                "plat_dessert": 18.90,
                "entree_plat_dessert": 23.90
            }
    
    return ardoise

@api_router.put("/ardoise")
async def update_ardoise(request: UpdateArdoiseRequest, current_user: dict = Depends(get_current_user)):
    """Mettre à jour l'ardoise du restaurant"""
    restaurant_id = current_user["restaurant_id"]
    
    # Vérifier que chaque section a exactement 2 items
    if len(request.entree) != 2 or len(request.plat) != 2 or len(request.dessert) != 2:
        raise HTTPException(status_code=400, detail="Chaque section doit contenir exactement 2 items")
    
    update_data = {
        "entree": [item.dict() for item in request.entree],
        "plat": [item.dict() for item in request.plat],
        "dessert": [item.dict() for item in request.dessert],
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Ajouter les prix de formules si fournis
    if request.formule_prices:
        update_data["formule_prices"] = request.formule_prices.dict()
    
    result = await ardoise_collection.update_one(
        {"restaurant_id": restaurant_id},
        {"$set": update_data},
        upsert=True
    )
    
    # Regenerate translations for Ardoise items
    try:
        asyncio.create_task(_generate_restaurant_translations(restaurant_id))
    except Exception as e:
        print(f"Error regenerating translations after Ardoise update: {e}")
    
    return {"message": "Ardoise mise à jour avec succès"}

@api_router.get("/ardoise/share-link")
async def get_ardoise_share_link(current_user: dict = Depends(get_current_user)):
    """Récupérer le lien de partage permanent de l'ardoise"""
    restaurant_id = current_user["restaurant_id"]
    
    ardoise = await ardoise_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "share_token": 1}
    )
    
    if not ardoise:
        # Créer l'ardoise si elle n'existe pas
        share_token = secrets.token_urlsafe(32)
        await ardoise_collection.insert_one({
            "restaurant_id": restaurant_id,
            "share_token": share_token,
            "entree": [{"name": "", "description": "", "price": None}, {"name": "", "description": "", "price": None}],
            "plat": [{"name": "", "description": "", "price": None}, {"name": "", "description": "", "price": None}],
            "dessert": [{"name": "", "description": "", "price": None}, {"name": "", "description": "", "price": None}],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        })
    else:
        share_token = ardoise.get("share_token")
        if not share_token:
            share_token = secrets.token_urlsafe(32)
            await ardoise_collection.update_one(
                {"restaurant_id": restaurant_id},
                {"$set": {"share_token": share_token}}
            )
    
    # Construire l'URL de partage
    frontend_url = os.environ.get('FRONTEND_URL', '')
    share_url = f"{frontend_url}/ardoise/{share_token}"
    
    return {
        "share_token": share_token,
        "share_url": share_url,
        "message": "Ce lien est permanent et ne change jamais. Partagez-le avec votre équipe."
    }

@api_router.get("/app-page")
async def serve_app_page_via_api():
    """Serve the main application page via API route"""
    from fastapi.responses import HTMLResponse
    index_file = DIST_DIR / "index.html"
    if index_file.exists():
        content = index_file.read_text()
        return HTMLResponse(
            content=content, 
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0"
            }
        )
    raise HTTPException(status_code=404, detail="App not found")

@api_router.get("/ardoise/page/{share_token}")
async def get_ardoise_page(share_token: str):
    """Sert la page HTML de l'ardoise directement (bypass cache CDN)"""
    from fastapi.responses import HTMLResponse
    
    # Verify token exists
    ardoise = await ardoise_collection.find_one({"share_token": share_token})
    if not ardoise:
        raise HTTPException(status_code=404, detail="Ardoise non trouvée")
    
    # Read the HTML file
    html_path = DIST_DIR / "gestion.html"
    if not html_path.exists():
        raise HTTPException(status_code=500, detail="Page non disponible")
    
    html_content = html_path.read_text()
    
    # Return with strong no-cache headers
    return HTMLResponse(
        content=html_content,
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
            "Surrogate-Control": "no-store",
            "X-Accel-Expires": "0"
        }
    )

@api_router.get("/gestion-ardoise")
async def get_gestion_page():
    """Page de gestion ardoise indépendante"""
    from fastapi.responses import HTMLResponse
    
    html_path = DIST_DIR / "gestion.html"
    if not html_path.exists():
        raise HTTPException(status_code=500, detail="Page non disponible")
    
    html_content = html_path.read_text()
    
    return HTMLResponse(
        content=html_content,
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0"
        }
    )

@api_router.get("/ardoise/public/{share_token}")
async def get_public_ardoise(share_token: str):
    """Récupérer l'ardoise via le lien public (sans authentification)"""
    ardoise = await ardoise_collection.find_one(
        {"share_token": share_token},
        {"_id": 0}
    )
    
    if not ardoise:
        raise HTTPException(status_code=404, detail="Ardoise non trouvée")
    
    # Récupérer le nom du restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": ardoise["restaurant_id"]},
        {"_id": 0, "name": 1}
    )
    
    # Prix des formules par défaut si absents
    formule_prices = ardoise.get("formule_prices", {
        "plat_du_jour": 15.90,
        "entree_plat": 18.90,
        "plat_dessert": 18.90,
        "entree_plat_dessert": 23.90
    })
    
    return {
        "restaurant_name": restaurant.get("name", "Restaurant") if restaurant else "Restaurant",
        "entree": ardoise.get("entree", []),
        "plat": ardoise.get("plat", []),
        "dessert": ardoise.get("dessert", []),
        "formule_prices": formule_prices,
        "updated_at": ardoise.get("updated_at")
    }

@api_router.get("/ardoise/by-restaurant/{restaurant_id}")
async def get_ardoise_by_restaurant(restaurant_id: str):
    """Récupérer l'ardoise d'un restaurant (endpoint public pour la vue client)
    Génère automatiquement un share_token si l'ardoise existe mais n'en a pas.
    """
    ardoise = await ardoise_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    
    if not ardoise:
        # Retourner une ardoise vide si elle n'existe pas
        return {
            "entree": [],
            "plat": [],
            "dessert": [],
            "formule_prices": {
                "plat_du_jour": 15.90,
                "entree_plat": 18.90,
                "plat_dessert": 18.90,
                "entree_plat_dessert": 23.90
            },
            "share_token": None
        }
    
    # AUTO-GÉNÉRATION DU SHARE_TOKEN si l'ardoise existe mais n'a pas de token
    share_token = ardoise.get("share_token")
    if not share_token:
        share_token = secrets.token_urlsafe(32)
        await ardoise_collection.update_one(
            {"restaurant_id": restaurant_id},
            {"$set": {"share_token": share_token}}
        )
        logger.info(f"[ARDOISE] Auto-generated share_token for restaurant {restaurant_id}")
    
    return {
        "entree": ardoise.get("entree", []),
        "plat": ardoise.get("plat", []),
        "dessert": ardoise.get("dessert", []),
        "formule_prices": ardoise.get("formule_prices", {
            "plat_du_jour": 15.90,
            "entree_plat": 18.90,
            "plat_dessert": 18.90,
            "entree_plat_dessert": 23.90
        }),
        "share_token": share_token
    }

@api_router.put("/ardoise/public/{share_token}")
async def update_public_ardoise(share_token: str, request: UpdateArdoiseRequest):
    """Mettre à jour l'ardoise via le lien public (sans authentification)"""
    ardoise = await ardoise_collection.find_one({"share_token": share_token})
    
    if not ardoise:
        raise HTTPException(status_code=404, detail="Ardoise non trouvée")
    
    # Vérifier que chaque section a exactement 2 items
    if len(request.entree) != 2 or len(request.plat) != 2 or len(request.dessert) != 2:
        raise HTTPException(status_code=400, detail="Chaque section doit contenir exactement 2 items")
    
    update_data = {
        "entree": [item.dict() for item in request.entree],
        "plat": [item.dict() for item in request.plat],
        "dessert": [item.dict() for item in request.dessert],
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Ajouter les prix de formules si fournis
    if request.formule_prices:
        update_data["formule_prices"] = request.formule_prices.dict()
    
    await ardoise_collection.update_one(
        {"share_token": share_token},
        {"$set": update_data}
    )
    
    return {"message": "Ardoise mise à jour avec succès"}

# ==================== ARDOISE SALES TRACKING ====================

class ArdoiseSalesRecord(BaseModel):
    """Enregistrement des ventes de l'ardoise pour un service"""
    date: str  # Format YYYY-MM-DD
    service: str = "midi"  # "midi" ou "soir"
    entree: List[dict] = []  # [{name, quantity_sold}]
    plat: List[dict] = []
    dessert: List[dict] = []
    formule_prices: Optional[dict] = None
    notes: Optional[str] = None

@api_router.post("/ardoise/sales")
async def save_ardoise_sales(sales: ArdoiseSalesRecord, current_user: dict = Depends(get_current_user)):
    """Enregistrer les ventes de l'ardoise à la fin d'un service"""
    restaurant_id = current_user["restaurant_id"]
    
    sales_record = {
        "restaurant_id": restaurant_id,
        "date": sales.date,
        "service": sales.service,
        "entree": sales.entree,
        "plat": sales.plat,
        "dessert": sales.dessert,
        "formule_prices": sales.formule_prices,
        "notes": sales.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Vérifier si un enregistrement existe déjà pour cette date et ce service
    existing = await ardoise_sales_collection.find_one({
        "restaurant_id": restaurant_id,
        "date": sales.date,
        "service": sales.service
    })
    
    if existing:
        # Mettre à jour l'existant
        await ardoise_sales_collection.update_one(
            {"restaurant_id": restaurant_id, "date": sales.date, "service": sales.service},
            {"$set": sales_record}
        )
        return {"message": "Ventes mises à jour avec succès"}
    else:
        # Créer un nouvel enregistrement
        await ardoise_sales_collection.insert_one(sales_record)
        return {"message": "Ventes enregistrées avec succès"}

@api_router.post("/ardoise/public/{share_token}/sales")
async def save_ardoise_sales_public(share_token: str, sales: ArdoiseSalesRecord):
    """Enregistrer les ventes de l'ardoise via le lien public"""
    ardoise = await ardoise_collection.find_one({"share_token": share_token})
    
    if not ardoise:
        raise HTTPException(status_code=404, detail="Ardoise non trouvée")
    
    restaurant_id = ardoise["restaurant_id"]
    
    sales_record = {
        "restaurant_id": restaurant_id,
        "date": sales.date,
        "service": sales.service,
        "entree": sales.entree,
        "plat": sales.plat,
        "dessert": sales.dessert,
        "formule_prices": sales.formule_prices,
        "notes": sales.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Vérifier si un enregistrement existe déjà
    existing = await ardoise_sales_collection.find_one({
        "restaurant_id": restaurant_id,
        "date": sales.date,
        "service": sales.service
    })
    
    if existing:
        await ardoise_sales_collection.update_one(
            {"restaurant_id": restaurant_id, "date": sales.date, "service": sales.service},
            {"$set": sales_record}
        )
        return {"message": "Ventes mises à jour avec succès"}
    else:
        await ardoise_sales_collection.insert_one(sales_record)
        return {"message": "Ventes enregistrées avec succès"}

@api_router.post("/ardoise/sales/{restaurant_id}")
async def save_ardoise_sales_by_restaurant(restaurant_id: str, sales: ArdoiseSalesRecord):
    """Enregistrer les ventes de l'ardoise par restaurant_id (endpoint pour l'app principale)"""
    # Vérifier que le restaurant existe
    restaurant = await restaurants_collection.find_one({"restaurant_id": restaurant_id})
    if not restaurant:
        raise HTTPException(status_code=404, detail="Restaurant non trouvé")
    
    sales_record = {
        "restaurant_id": restaurant_id,
        "date": sales.date,
        "service": sales.service,
        "entree": sales.entree,
        "plat": sales.plat,
        "dessert": sales.dessert,
        "formule_prices": sales.formule_prices,
        "notes": sales.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Vérifier si un enregistrement existe déjà pour cette date et ce service
    existing = await ardoise_sales_collection.find_one({
        "restaurant_id": restaurant_id,
        "date": sales.date,
        "service": sales.service
    })
    
    if existing:
        await ardoise_sales_collection.update_one(
            {"restaurant_id": restaurant_id, "date": sales.date, "service": sales.service},
            {"$set": sales_record}
        )
        return {"message": "Ventes mises à jour avec succès"}
    else:
        await ardoise_sales_collection.insert_one(sales_record)
        return {"message": "Ventes enregistrées avec succès"}

@api_router.get("/ardoise/suggestions/{restaurant_id}")
async def get_ardoise_suggestions(
    restaurant_id: str,
    query: str,
    category: str  # "entree", "plat", "dessert"
):
    """
    Auto-complétion intelligente basée sur l'historique des ventes.
    Retourne les plats similaires avec leurs dates et quantités vendues.
    """
    if len(query) < 2:
        return {"suggestions": []}
    
    # Rechercher dans l'historique des ventes
    sales = await ardoise_sales_collection.find(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "date": 1, "service": 1, category: 1}
    ).sort("date", -1).to_list(200)
    
    # Agréger les plats par nom
    items_history = {}
    query_lower = query.lower()
    
    for sale in sales:
        date = sale.get("date", "")
        service = sale.get("service", "")
        items = sale.get(category, [])
        
        for item in items:
            name = item.get("name", "")
            qty = item.get("quantity_sold", 0) or 0
            
            if not name:
                continue
            
            # Vérifier si le nom contient la requête (recherche flexible)
            if query_lower in name.lower():
                if name not in items_history:
                    items_history[name] = {
                        "name": name,
                        "total_sold": 0,
                        "dates": [],
                        "last_date": None
                    }
                
                items_history[name]["total_sold"] += qty
                if date and date not in [d["date"] for d in items_history[name]["dates"]]:
                    items_history[name]["dates"].append({
                        "date": date,
                        "service": service,
                        "quantity": qty
                    })
                
                if not items_history[name]["last_date"] or date > items_history[name]["last_date"]:
                    items_history[name]["last_date"] = date
    
    # Trier par total vendu (décroissant) et limiter à 10 suggestions
    suggestions = sorted(
        items_history.values(),
        key=lambda x: (x["total_sold"], x["last_date"] or ""),
        reverse=True
    )[:10]
    
    # Formater les dates pour l'affichage
    for suggestion in suggestions:
        suggestion["dates"] = sorted(suggestion["dates"], key=lambda x: x["date"], reverse=True)[:5]
        # Formater la dernière date
        if suggestion["last_date"]:
            try:
                from datetime import datetime as dt
                d = dt.strptime(suggestion["last_date"], "%Y-%m-%d")
                suggestion["last_date_formatted"] = d.strftime("%d/%m/%Y")
            except:
                suggestion["last_date_formatted"] = suggestion["last_date"]
    
    return {"suggestions": suggestions}

# ==================== ARDOISES PLANIFIÉES ====================

class PlannedArdoiseRequest(BaseModel):
    """Requête pour sauvegarder une ardoise planifiée"""
    date: str  # Format YYYY-MM-DD
    entree: List[dict] = []
    plat: List[dict] = []
    dessert: List[dict] = []
    formule_prices: Optional[dict] = None

@api_router.get("/ardoise/planned/{restaurant_id}")
async def get_planned_ardoise(restaurant_id: str, date: str):
    """Récupérer l'ardoise planifiée pour une date donnée"""
    today = datetime.now(timezone.utc).date().isoformat()
    
    # Toujours récupérer les prix de l'ardoise principale
    current = await ardoise_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    default_prices = current.get("formule_prices", {
        "plat_du_jour": 15.90,
        "entree_plat": 18.90,
        "plat_dessert": 18.90,
        "entree_plat_dessert": 23.90
    }) if current else {
        "plat_du_jour": 15.90,
        "entree_plat": 18.90,
        "plat_dessert": 18.90,
        "entree_plat_dessert": 23.90
    }
    
    planned = await ardoise_planned_collection.find_one(
        {"restaurant_id": restaurant_id, "date": date},
        {"_id": 0}
    )
    
    if not planned:
        # Si c'est aujourd'hui et pas de planification, retourner l'ardoise actuelle
        if date == today:
            if current:
                return {
                    "found": False,
                    "is_current": True,
                    "entree": current.get("entree", []),
                    "plat": current.get("plat", []),
                    "dessert": current.get("dessert", []),
                    "formule_prices": default_prices
                }
        
        # Pour les jours futurs sans planification, retourner VIDE mais AVEC les prix
        return {
            "found": False,
            "is_current": False,
            "entree": [],
            "plat": [],
            "dessert": [],
            "formule_prices": default_prices  # Toujours inclure les prix
        }
    
    return {
        "found": True,
        "is_current": False,
        "date": planned.get("date"),
        "entree": planned.get("entree", []),
        "plat": planned.get("plat", []),
        "dessert": planned.get("dessert", []),
        "formule_prices": planned.get("formule_prices") or default_prices  # Utiliser les prix par défaut si vide
    }

@api_router.post("/ardoise/planned/{restaurant_id}")
async def save_planned_ardoise(restaurant_id: str, request: PlannedArdoiseRequest):
    """Sauvegarder une ardoise planifiée pour une date future"""
    today = datetime.now(timezone.utc).date().isoformat()
    
    planned_data = {
        "restaurant_id": restaurant_id,
        "date": request.date,
        "entree": request.entree,
        "plat": request.plat,
        "dessert": request.dessert,
        "formule_prices": request.formule_prices or {},
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Upsert - mise à jour ou création
    result = await ardoise_planned_collection.update_one(
        {"restaurant_id": restaurant_id, "date": request.date},
        {"$set": planned_data},
        upsert=True
    )
    
    # Si c'est pour aujourd'hui, mettre aussi à jour l'ardoise principale
    if request.date == today:
        await ardoise_collection.update_one(
            {"restaurant_id": restaurant_id},
            {"$set": {
                "entree": request.entree,
                "plat": request.plat,
                "dessert": request.dessert,
                "formule_prices": request.formule_prices or {},
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    return {"message": "Ardoise planifiée sauvegardée", "date": request.date}

@api_router.get("/ardoise/planned/list/{restaurant_id}")
async def list_planned_ardoises(restaurant_id: str):
    """Lister toutes les ardoises planifiées pour les 14 prochains jours"""
    today = datetime.now(timezone.utc).date()
    dates_with_planning = []
    
    # Récupérer les planifications existantes
    planned = await ardoise_planned_collection.find(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "date": 1}
    ).to_list(30)
    
    planned_dates = [p["date"] for p in planned]
    
    # Créer la liste des 14 prochains jours avec indicateur de planification
    for i in range(14):
        d = (today + timedelta(days=i)).isoformat()
        dates_with_planning.append({
            "date": d,
            "has_planning": d in planned_dates,
            "is_today": i == 0
        })
    
    return {"dates": dates_with_planning}

@api_router.get("/ardoise/planned/export-pdf/{restaurant_id}")
async def export_planned_ardoise_pdf(restaurant_id: str, days: int = 7):
    """Exporter les ardoises planifiées pour les X prochains jours en PDF"""
    
    # Fonction pour nettoyer les caractères
    def clean_text(text):
        if not text:
            return ""
        replacements = {"'": "'", "'": "'", """: '"', """: '"', "–": "-", "—": "-", "…": "..."}
        for old, new in replacements.items():
            text = text.replace(old, new)
        return ''.join(c if ord(c) < 128 else '' for c in text)
    
    # Récupérer le nom du restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "name": 1}
    )
    if not restaurant:
        raise HTTPException(status_code=404, detail="Restaurant non trouvé")
    restaurant_name = clean_text(restaurant.get("name", "Restaurant"))
    
    today = datetime.now(timezone.utc).date()
    end_date = (today + timedelta(days=days)).isoformat()
    
    # Récupérer les ardoises planifiées
    planned = await ardoise_planned_collection.find(
        {
            "restaurant_id": restaurant_id,
            "date": {"$gte": today.isoformat(), "$lte": end_date}
        },
        {"_id": 0}
    ).sort("date", 1).to_list(days)
    
    # Si pas de planification, récupérer l'ardoise actuelle
    if not planned:
        current = await ardoise_collection.find_one({"restaurant_id": restaurant_id}, {"_id": 0})
        if current:
            planned = [{"date": today.isoformat(), "entree": current.get("entree", []), "plat": current.get("plat", []), "dessert": current.get("dessert", [])}]
    
    # Générer le PDF en paysage
    class PlanningPDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 14)
            self.cell(0, 8, f"{restaurant_name} - Planning Ardoise", align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_font("Helvetica", "", 9)
            self.cell(0, 5, f"Du {today.strftime('%d/%m/%Y')} au {(today + timedelta(days=days-1)).strftime('%d/%m/%Y')}", align="C", new_x="LMARGIN", new_y="NEXT")
            self.ln(3)
        
        def footer(self):
            self.set_y(-10)
            self.set_font("Helvetica", "I", 7)
            self.cell(0, 10, f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')} - NeoChef", align="C")
    
    pdf = PlanningPDF(orientation='L')
    pdf.add_page()
    
    # En-têtes
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(255, 209, 102)
    pdf.cell(25, 7, "Date", border=1, fill=True)
    pdf.cell(75, 7, "Entrees", border=1, fill=True)
    pdf.cell(75, 7, "Plats", border=1, fill=True)
    pdf.cell(75, 7, "Desserts", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
    
    pdf.set_font("Helvetica", "", 7)
    
    # Remplir avec les jours (même sans planification)
    for i in range(days):
        d = (today + timedelta(days=i)).isoformat()
        day_data = next((p for p in planned if p.get("date") == d), None)
        
        # Date formatée
        date_obj = today + timedelta(days=i)
        date_str = date_obj.strftime("%a %d/%m")
        
        entrees = day_data.get("entree", []) if day_data else []
        plats = day_data.get("plat", []) if day_data else []
        desserts = day_data.get("dessert", []) if day_data else []
        
        # Filtrer les vides
        entrees = [e for e in entrees if e.get("name")]
        plats = [p for p in plats if p.get("name")]
        desserts = [d for d in desserts if d.get("name")]
        
        max_rows = max(len(entrees), len(plats), len(desserts), 1)
        
        for row_idx in range(max_rows):
            if row_idx == 0:
                pdf.cell(25, 5, date_str, border=1)
            else:
                pdf.cell(25, 5, "", border=1)
            
            # Entrées
            e_text = clean_text(entrees[row_idx].get("name", ""))[:40] if row_idx < len(entrees) else ""
            pdf.cell(75, 5, e_text, border=1)
            
            # Plats
            p_text = clean_text(plats[row_idx].get("name", ""))[:40] if row_idx < len(plats) else ""
            pdf.cell(75, 5, p_text, border=1)
            
            # Desserts
            d_text = clean_text(desserts[row_idx].get("name", ""))[:40] if row_idx < len(desserts) else ""
            pdf.cell(75, 5, d_text, border=1, new_x="LMARGIN", new_y="NEXT")
    
    buffer = BytesIO(pdf.output())
    filename = f"planning_ardoise_{today.isoformat()}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/ardoise/planned/export-excel/{restaurant_id}")
async def export_planned_ardoise_excel(restaurant_id: str, days: int = 7):
    """Exporter les ardoises planifiées pour les X prochains jours en Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "name": 1}
    )
    if not restaurant:
        raise HTTPException(status_code=404, detail="Restaurant non trouvé")
    restaurant_name = restaurant.get("name", "Restaurant")
    
    today = datetime.now(timezone.utc).date()
    end_date = (today + timedelta(days=days)).isoformat()
    
    planned = await ardoise_planned_collection.find(
        {
            "restaurant_id": restaurant_id,
            "date": {"$gte": today.isoformat(), "$lte": end_date}
        },
        {"_id": 0}
    ).sort("date", 1).to_list(days)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Planning Ardoise"
    
    header_fill = PatternFill(start_color="FFD166", end_color="FFD166", fill_type="solid")
    header_font = Font(bold=True)
    
    ws["A1"] = f"{restaurant_name} - Planning Ardoise"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Du {today.strftime('%d/%m/%Y')} au {(today + timedelta(days=days-1)).strftime('%d/%m/%Y')}"
    
    headers = ["Date", "Entrées", "Plats", "Desserts"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    current_row = 5
    for i in range(days):
        d = (today + timedelta(days=i)).isoformat()
        day_data = next((p for p in planned if p.get("date") == d), None)
        
        date_obj = today + timedelta(days=i)
        date_str = date_obj.strftime("%a %d/%m/%Y")
        
        entrees = [e.get("name", "") for e in (day_data.get("entree", []) if day_data else []) if e.get("name")]
        plats = [p.get("name", "") for p in (day_data.get("plat", []) if day_data else []) if p.get("name")]
        desserts = [d.get("name", "") for d in (day_data.get("dessert", []) if day_data else []) if d.get("name")]
        
        max_rows = max(len(entrees), len(plats), len(desserts), 1)
        
        for row_idx in range(max_rows):
            if row_idx == 0:
                ws.cell(row=current_row, column=1, value=date_str)
            ws.cell(row=current_row, column=2, value=entrees[row_idx] if row_idx < len(entrees) else "")
            ws.cell(row=current_row, column=3, value=plats[row_idx] if row_idx < len(plats) else "")
            ws.cell(row=current_row, column=4, value=desserts[row_idx] if row_idx < len(desserts) else "")
            current_row += 1
    
    column_widths = [15, 40, 40, 40]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"planning_ardoise_{today.isoformat()}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/ardoise/sales/by-date/{restaurant_id}")
async def get_ardoise_sales_by_date(
    restaurant_id: str,
    date: str,
    service: Optional[str] = None
):
    """Récupérer les ventes d'une date spécifique (endpoint public pour l'app principale)"""
    query = {"restaurant_id": restaurant_id, "date": date}
    if service:
        query["service"] = service
    
    sales = await ardoise_sales_collection.find(
        query,
        {"_id": 0}
    ).to_list(10)
    
    if not sales:
        return {"found": False, "sales": []}
    
    return {"found": True, "sales": sales}

@api_router.get("/ardoise/sales")
async def get_ardoise_sales(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer l'historique des ventes de l'ardoise"""
    restaurant_id = current_user["restaurant_id"]
    
    query = {"restaurant_id": restaurant_id}
    
    if start_date or end_date:
        query["date"] = {}
        if start_date:
            query["date"]["$gte"] = start_date
        if end_date:
            query["date"]["$lte"] = end_date
    
    sales = await ardoise_sales_collection.find(
        query,
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    return {"sales": sales}

@api_router.get("/ardoise/sales/report")
async def get_ardoise_sales_report(
    period: str = "week",  # "day", "week", "month"
    current_user: dict = Depends(get_current_user)
):
    """Récupérer un rapport agrégé des ventes de l'ardoise"""
    restaurant_id = current_user["restaurant_id"]
    
    # Calculer les dates de début selon la période
    today = datetime.now(timezone.utc).date()
    if period == "day":
        start_date = today.isoformat()
    elif period == "week":
        start_date = (today - timedelta(days=7)).isoformat()
    else:  # month
        start_date = (today - timedelta(days=30)).isoformat()
    
    # Récupérer toutes les ventes de la période
    sales = await ardoise_sales_collection.find(
        {
            "restaurant_id": restaurant_id,
            "date": {"$gte": start_date}
        },
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Agréger les données
    total_entrees = 0
    total_plats = 0
    total_desserts = 0
    items_stats = {}  # {item_name: {total_qty, category}}
    
    for sale in sales:
        for item in sale.get("entree", []):
            qty = item.get("quantity_sold", 0) or 0
            total_entrees += qty
            name = item.get("name", "")
            if name:
                if name not in items_stats:
                    items_stats[name] = {"total_qty": 0, "category": "entree"}
                items_stats[name]["total_qty"] += qty
        
        for item in sale.get("plat", []):
            qty = item.get("quantity_sold", 0) or 0
            total_plats += qty
            name = item.get("name", "")
            if name:
                if name not in items_stats:
                    items_stats[name] = {"total_qty": 0, "category": "plat"}
                items_stats[name]["total_qty"] += qty
        
        for item in sale.get("dessert", []):
            qty = item.get("quantity_sold", 0) or 0
            total_desserts += qty
            name = item.get("name", "")
            if name:
                if name not in items_stats:
                    items_stats[name] = {"total_qty": 0, "category": "dessert"}
                items_stats[name]["total_qty"] += qty
    
    # Trier les items par quantité vendue
    top_items = sorted(
        [{"name": k, **v} for k, v in items_stats.items()],
        key=lambda x: x["total_qty"],
        reverse=True
    )[:10]
    
    return {
        "period": period,
        "start_date": start_date,
        "end_date": today.isoformat(),
        "total_services": len(sales),
        "totals": {
            "entrees": total_entrees,
            "plats": total_plats,
            "desserts": total_desserts,
            "total": total_entrees + total_plats + total_desserts
        },
        "top_items": top_items,
        "daily_breakdown": sales
    }

@api_router.get("/ardoise/sales/report/public/{share_token}")
async def get_ardoise_sales_report_public(
    share_token: str,
    period: str = "week"
):
    """Récupérer un rapport des ventes via le lien public"""
    ardoise = await ardoise_collection.find_one({"share_token": share_token})
    
    if not ardoise:
        raise HTTPException(status_code=404, detail="Ardoise non trouvée")
    
    restaurant_id = ardoise["restaurant_id"]
    
    today = datetime.now(timezone.utc).date()
    if period == "day":
        start_date = today.isoformat()
    elif period == "week":
        start_date = (today - timedelta(days=7)).isoformat()
    else:
        start_date = (today - timedelta(days=30)).isoformat()
    
    sales = await ardoise_sales_collection.find(
        {
            "restaurant_id": restaurant_id,
            "date": {"$gte": start_date}
        },
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Agréger les données
    total_entrees = 0
    total_plats = 0
    total_desserts = 0
    items_stats = {}
    
    for sale in sales:
        for item in sale.get("entree", []):
            qty = item.get("quantity_sold", 0) or 0
            total_entrees += qty
            name = item.get("name", "")
            if name:
                if name not in items_stats:
                    items_stats[name] = {"total_qty": 0, "category": "entree"}
                items_stats[name]["total_qty"] += qty
        
        for item in sale.get("plat", []):
            qty = item.get("quantity_sold", 0) or 0
            total_plats += qty
            name = item.get("name", "")
            if name:
                if name not in items_stats:
                    items_stats[name] = {"total_qty": 0, "category": "plat"}
                items_stats[name]["total_qty"] += qty
        
        for item in sale.get("dessert", []):
            qty = item.get("quantity_sold", 0) or 0
            total_desserts += qty
            name = item.get("name", "")
            if name:
                if name not in items_stats:
                    items_stats[name] = {"total_qty": 0, "category": "dessert"}
                items_stats[name]["total_qty"] += qty
    
    top_items = sorted(
        [{"name": k, **v} for k, v in items_stats.items()],
        key=lambda x: x["total_qty"],
        reverse=True
    )[:10]
    
    return {
        "period": period,
        "start_date": start_date,
        "end_date": today.isoformat(),
        "total_services": len(sales),
        "total_items": total_entrees + total_plats + total_desserts,
        "total_by_category": {
            "entree": total_entrees,
            "plat": total_plats,
            "dessert": total_desserts
        },
        "totals": {
            "entrees": total_entrees,
            "plats": total_plats,
            "desserts": total_desserts,
            "total": total_entrees + total_plats + total_desserts
        },
        "top_items": [{"name": item["name"], "total_sold": item["total_qty"], "category": item.get("category", "")} for item in top_items],
        "daily_details": sales,
        "daily_breakdown": sales
    }

@api_router.get("/ardoise/sales/report/{restaurant_id}")
async def get_ardoise_sales_report_by_restaurant(
    restaurant_id: str,
    period: str = "week",
    current_user: dict = Depends(get_current_user)
):
    """Récupérer un rapport agrégé des ventes de l'ardoise par restaurant_id"""
    # Verify user has access to this restaurant
    if current_user["restaurant_id"] != restaurant_id and current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Accès non autorisé à ce restaurant")
    
    today = datetime.now(timezone.utc).date()
    if period == "day":
        start_date = today.isoformat()
    elif period == "week":
        start_date = (today - timedelta(days=7)).isoformat()
    elif period == "year":
        start_date = (today - timedelta(days=365)).isoformat()
    else:  # month
        start_date = (today - timedelta(days=30)).isoformat()
    
    sales = await ardoise_sales_collection.find(
        {"restaurant_id": restaurant_id, "date": {"$gte": start_date}},
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    total_entrees = 0
    total_plats = 0
    total_desserts = 0
    items_stats = {}
    
    for sale in sales:
        for item in sale.get("entree", []):
            qty = item.get("quantity_sold", 0) or 0
            total_entrees += qty
            name = item.get("name", "")
            if name and qty > 0:
                if name not in items_stats:
                    items_stats[name] = {"total_qty": 0, "category": "entree"}
                items_stats[name]["total_qty"] += qty
        
        for item in sale.get("plat", []):
            qty = item.get("quantity_sold", 0) or 0
            total_plats += qty
            name = item.get("name", "")
            if name and qty > 0:
                if name not in items_stats:
                    items_stats[name] = {"total_qty": 0, "category": "plat"}
                items_stats[name]["total_qty"] += qty
        
        for item in sale.get("dessert", []):
            qty = item.get("quantity_sold", 0) or 0
            total_desserts += qty
            name = item.get("name", "")
            if name and qty > 0:
                if name not in items_stats:
                    items_stats[name] = {"total_qty": 0, "category": "dessert"}
                items_stats[name]["total_qty"] += qty
    
    top_items = sorted(
        [{"name": k, **v} for k, v in items_stats.items()],
        key=lambda x: x["total_qty"],
        reverse=True
    )[:10]
    
    return {
        "period": period,
        "start_date": start_date,
        "end_date": today.isoformat(),
        "total_services": len(sales),
        "total_items": total_entrees + total_plats + total_desserts,
        "total_by_category": {
            "entree": total_entrees,
            "plat": total_plats,
            "dessert": total_desserts
        },
        "totals": {
            "entrees": total_entrees,
            "plats": total_plats,
            "desserts": total_desserts,
            "total": total_entrees + total_plats + total_desserts
        },
        "top_items": [{"name": item["name"], "total_sold": item["total_qty"], "category": item.get("category", "")} for item in top_items],
        "daily_details": sales,
        "daily_breakdown": sales
    }

@api_router.get("/ardoise/sales/export-pdf/{share_token}")
async def export_ardoise_sales_pdf(
    share_token: str,
    period: str = "week"
):
    """Exporter le rapport des ventes en PDF"""
    ardoise = await ardoise_collection.find_one({"share_token": share_token})
    
    if not ardoise:
        raise HTTPException(status_code=404, detail="Ardoise non trouvée")
    
    restaurant_id = ardoise["restaurant_id"]
    
    # Récupérer le nom du restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "name": 1}
    )
    restaurant_name = restaurant.get("name", "Restaurant") if restaurant else "Restaurant"
    
    today = datetime.now(timezone.utc).date()
    if period == "day":
        start_date = today.isoformat()
        period_label = "Aujourd'hui"
    elif period == "week":
        start_date = (today - timedelta(days=7)).isoformat()
        period_label = "7 derniers jours"
    else:
        start_date = (today - timedelta(days=30)).isoformat()
        period_label = "30 derniers jours"
    
    sales = await ardoise_sales_collection.find(
        {
            "restaurant_id": restaurant_id,
            "date": {"$gte": start_date}
        },
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Agréger les données
    items_stats = {}
    for sale in sales:
        for section in ["entree", "plat", "dessert"]:
            for item in sale.get(section, []):
                qty = item.get("quantity_sold", 0) or 0
                name = item.get("name", "")
                if name:
                    if name not in items_stats:
                        items_stats[name] = {"total_qty": 0, "category": section}
                    items_stats[name]["total_qty"] += qty
    
    # Générer le PDF avec FPDF
    class ArdoisePDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 18)
            self.cell(0, 10, restaurant_name, align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_font("Helvetica", "B", 14)
            self.cell(0, 8, "Rapport des Ventes - Ardoise", align="C", new_x="LMARGIN", new_y="NEXT")
            self.ln(5)
        
        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.cell(0, 10, f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')} - NeoChef", align="C")
    
    pdf = ArdoisePDF()
    pdf.add_page()
    
    # Période
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Periode: {period_label}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Du {start_date} au {today.isoformat()}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    
    # Statistiques globales
    total_qty = sum(item["total_qty"] for item in items_stats.values())
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Resume", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Nombre de services: {len(sales)}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Total plats vendus: {total_qty}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    
    # Détail par plat
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Detail par plat", new_x="LMARGIN", new_y="NEXT")
    
    # En-têtes du tableau
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(255, 209, 102)
    pdf.cell(100, 8, "Plat", border=1, fill=True)
    pdf.cell(40, 8, "Categorie", border=1, fill=True)
    pdf.cell(40, 8, "Quantite", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
    
    pdf.set_font("Helvetica", "", 9)
    sorted_items = sorted(items_stats.items(), key=lambda x: x[1]["total_qty"], reverse=True)
    
    for name, data in sorted_items:
        display_name = name[:35] + "..." if len(name) > 35 else name
        cat_label = {"entree": "Entree", "plat": "Plat", "dessert": "Dessert"}.get(data["category"], data["category"])
        pdf.cell(100, 7, display_name, border=1)
        pdf.cell(40, 7, cat_label, border=1)
        pdf.cell(40, 7, str(data["total_qty"]), border=1, new_x="LMARGIN", new_y="NEXT")
    
    pdf.ln(10)
    
    # Détail par jour
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Detail par jour", new_x="LMARGIN", new_y="NEXT")
    
    # En-têtes
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(255, 209, 102)
    pdf.cell(35, 8, "Date", border=1, fill=True)
    pdf.cell(30, 8, "Service", border=1, fill=True)
    pdf.cell(30, 8, "Entrees", border=1, fill=True)
    pdf.cell(30, 8, "Plats", border=1, fill=True)
    pdf.cell(30, 8, "Desserts", border=1, fill=True)
    pdf.cell(30, 8, "Total", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
    
    pdf.set_font("Helvetica", "", 9)
    for sale in sales:
        entree_qty = sum(item.get("quantity_sold", 0) or 0 for item in sale.get("entree", []))
        plat_qty = sum(item.get("quantity_sold", 0) or 0 for item in sale.get("plat", []))
        dessert_qty = sum(item.get("quantity_sold", 0) or 0 for item in sale.get("dessert", []))
        
        pdf.cell(35, 7, sale.get("date", ""), border=1)
        pdf.cell(30, 7, sale.get("service", "").capitalize(), border=1)
        pdf.cell(30, 7, str(entree_qty), border=1, align="C")
        pdf.cell(30, 7, str(plat_qty), border=1, align="C")
        pdf.cell(30, 7, str(dessert_qty), border=1, align="C")
        pdf.cell(30, 7, str(entree_qty + plat_qty + dessert_qty), border=1, align="C", new_x="LMARGIN", new_y="NEXT")
    
    # Générer le buffer
    buffer = BytesIO(pdf.output())
    
    filename = f"rapport_ventes_{start_date}_{today.isoformat()}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/ardoise/sales/export-excel/{share_token}")
async def export_ardoise_sales_excel(
    share_token: str,
    period: str = "week"
):
    """Exporter le rapport des ventes en Excel"""
    ardoise = await ardoise_collection.find_one({"share_token": share_token})
    
    if not ardoise:
        raise HTTPException(status_code=404, detail="Ardoise non trouvée")
    
    restaurant_id = ardoise["restaurant_id"]
    
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "name": 1}
    )
    restaurant_name = restaurant.get("name", "Restaurant") if restaurant else "Restaurant"
    
    today = datetime.now(timezone.utc).date()
    if period == "day":
        start_date = today.isoformat()
    elif period == "week":
        start_date = (today - timedelta(days=7)).isoformat()
    else:
        start_date = (today - timedelta(days=30)).isoformat()
    
    sales = await ardoise_sales_collection.find(
        {
            "restaurant_id": restaurant_id,
            "date": {"$gte": start_date}
        },
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Créer le fichier Excel avec openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    
    # Feuille 1: Résumé
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    
    # Styles
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=18)
    header_fill = PatternFill(start_color="FFD166", end_color="FFD166", fill_type="solid")
    
    ws_summary["A1"] = restaurant_name
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = f"Rapport des ventes - Ardoise"
    ws_summary["A3"] = f"Période: {start_date} au {today.isoformat()}"
    
    ws_summary["A5"] = "Statistiques globales"
    ws_summary["A5"].font = header_font
    ws_summary["A6"] = "Nombre de services:"
    ws_summary["B6"] = len(sales)
    
    # Agréger par plat
    items_stats = {}
    for sale in sales:
        for section in ["entree", "plat", "dessert"]:
            for item in sale.get(section, []):
                qty = item.get("quantity_sold", 0) or 0
                name = item.get("name", "")
                if name:
                    if name not in items_stats:
                        items_stats[name] = {"total_qty": 0, "category": section}
                    items_stats[name]["total_qty"] += qty
    
    total_qty = sum(item["total_qty"] for item in items_stats.values())
    ws_summary["A7"] = "Total plats vendus:"
    ws_summary["B7"] = total_qty
    
    # Feuille 2: Détail par plat
    ws_items = wb.create_sheet("Par Plat")
    ws_items["A1"] = "Plat"
    ws_items["B1"] = "Catégorie"
    ws_items["C1"] = "Quantité vendue"
    
    for col in ["A1", "B1", "C1"]:
        ws_items[col].font = Font(bold=True)
        ws_items[col].fill = header_fill
    
    row = 2
    sorted_items = sorted(items_stats.items(), key=lambda x: x[1]["total_qty"], reverse=True)
    for name, data in sorted_items:
        ws_items[f"A{row}"] = name
        cat_label = {"entree": "Entrée", "plat": "Plat", "dessert": "Dessert"}.get(data["category"], data["category"])
        ws_items[f"B{row}"] = cat_label
        ws_items[f"C{row}"] = data["total_qty"]
        row += 1
    
    # Feuille 3: Détail par jour
    ws_daily = wb.create_sheet("Par Jour")
    ws_daily["A1"] = "Date"
    ws_daily["B1"] = "Service"
    ws_daily["C1"] = "Entrées"
    ws_daily["D1"] = "Plats"
    ws_daily["E1"] = "Desserts"
    ws_daily["F1"] = "Total"
    
    for col in ["A1", "B1", "C1", "D1", "E1", "F1"]:
        ws_daily[col].font = Font(bold=True)
        ws_daily[col].fill = header_fill
    
    row = 2
    for sale in sales:
        entree_qty = sum(item.get("quantity_sold", 0) or 0 for item in sale.get("entree", []))
        plat_qty = sum(item.get("quantity_sold", 0) or 0 for item in sale.get("plat", []))
        dessert_qty = sum(item.get("quantity_sold", 0) or 0 for item in sale.get("dessert", []))
        
        ws_daily[f"A{row}"] = sale.get("date", "")
        ws_daily[f"B{row}"] = sale.get("service", "").capitalize()
        ws_daily[f"C{row}"] = entree_qty
        ws_daily[f"D{row}"] = plat_qty
        ws_daily[f"E{row}"] = dessert_qty
        ws_daily[f"F{row}"] = entree_qty + plat_qty + dessert_qty
        row += 1
    
    # Ajuster la largeur des colonnes
    for ws in [ws_summary, ws_items, ws_daily]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"rapport_ventes_{start_date}_{today.isoformat()}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== EXPORT PAR RESTAURANT_ID ====================

@api_router.get("/ardoise/sales/export-pdf/by-restaurant/{restaurant_id}")
async def export_ardoise_sales_pdf_by_restaurant(
    restaurant_id: str,
    period: str = "week"
):
    """Exporter le rapport des ventes en PDF par restaurant_id"""
    
    # Fonction pour nettoyer les caractères Unicode non supportés
    def clean_text(text):
        if not text:
            return ""
        # Remplacer les apostrophes typographiques et autres caractères spéciaux
        replacements = {
            "'": "'",  # Apostrophe typographique -> apostrophe simple
            "'": "'",  # Autre apostrophe typographique
            """: '"',  # Guillemet typographique
            """: '"',  # Guillemet typographique
            "–": "-",  # Tiret long
            "—": "-",  # Tiret très long
            "…": "...",  # Points de suspension
            "œ": "oe",  # Ligature
            "Œ": "OE",  # Ligature majuscule
            "é": "e",  # Accent (si problème)
            "è": "e",
            "ê": "e",
            "ë": "e",
            "à": "a",
            "â": "a",
            "ù": "u",
            "û": "u",
            "ô": "o",
            "î": "i",
            "ï": "i",
            "ç": "c",
        }
        for old, new in replacements.items():
            text = text.replace(old, new)
        # Supprimer tout caractère non-ASCII restant
        return ''.join(c if ord(c) < 128 else '' for c in text)
    
    # Récupérer le nom du restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "name": 1}
    )
    if not restaurant:
        raise HTTPException(status_code=404, detail="Restaurant non trouvé")
    restaurant_name = clean_text(restaurant.get("name", "Restaurant"))
    
    today = datetime.now(timezone.utc).date()
    if period == "day":
        start_date = today.isoformat()
        period_label = "Aujourd'hui"
    elif period == "week":
        start_date = (today - timedelta(days=7)).isoformat()
        period_label = "7 derniers jours"
    else:
        start_date = (today - timedelta(days=30)).isoformat()
        period_label = "30 derniers jours"
    
    sales = await ardoise_sales_collection.find(
        {
            "restaurant_id": restaurant_id,
            "date": {"$gte": start_date}
        },
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Agréger les données
    items_stats = {}
    for sale in sales:
        for section in ["entree", "plat", "dessert"]:
            for item in sale.get(section, []):
                qty = item.get("quantity_sold", 0) or 0
                name = item.get("name", "")
                if name:
                    if name not in items_stats:
                        items_stats[name] = {"total_qty": 0, "category": section}
                    items_stats[name]["total_qty"] += qty
    
    # Générer le PDF avec FPDF - FORMAT PAYSAGE
    class ArdoisePDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 16)
            self.cell(0, 8, restaurant_name, align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_font("Helvetica", "", 10)
            self.cell(0, 6, f"Du {start_date} au {today.isoformat()}", align="C", new_x="LMARGIN", new_y="NEXT")
            self.ln(5)
        
        def footer(self):
            self.set_y(-10)
            self.set_font("Helvetica", "I", 7)
            self.cell(0, 10, f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')} - NeoChef", align="C")
    
    pdf = ArdoisePDF(orientation='L')  # L = Landscape (Paysage)
    pdf.add_page()
    
    # En-têtes du tableau - colonnes ajustées pour A4 paysage (largeur 277mm avec marges 10mm)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(255, 209, 102)
    pdf.cell(22, 7, "Date", border=1, fill=True)
    pdf.cell(62, 7, "Entree", border=1, fill=True)
    pdf.cell(12, 7, "Qte", border=1, fill=True)
    pdf.cell(62, 7, "Plat", border=1, fill=True)
    pdf.cell(12, 7, "Qte", border=1, fill=True)
    pdf.cell(62, 7, "Dessert", border=1, fill=True)
    pdf.cell(12, 7, "Qte", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
    
    pdf.set_font("Helvetica", "", 7)
    
    for sale in sales:
        date_str = sale.get("date", "")
        
        # Filtrer les items avec quantité > 0
        entrees = [e for e in sale.get("entree", []) if (e.get("quantity_sold") or 0) > 0]
        plats = [p for p in sale.get("plat", []) if (p.get("quantity_sold") or 0) > 0]
        desserts = [d for d in sale.get("dessert", []) if (d.get("quantity_sold") or 0) > 0]
        
        # Nombre max de lignes pour ce jour
        max_rows = max(len(entrees), len(plats), len(desserts), 1)
        
        for row_idx in range(max_rows):
            # Date seulement sur la première ligne
            if row_idx == 0:
                pdf.cell(22, 5, f"{date_str[:10]}", border=1)
            else:
                pdf.cell(22, 5, "", border=1)
            
            # Entrée - nom complet (jusqu'à 35 caractères)
            if row_idx < len(entrees):
                e_name = clean_text(entrees[row_idx].get("name", ""))[:35]
                e_qty = entrees[row_idx].get("quantity_sold", 0)
                pdf.cell(62, 5, e_name, border=1)
                pdf.cell(12, 5, str(e_qty), border=1, align="C")
            else:
                pdf.cell(62, 5, "", border=1)
                pdf.cell(12, 5, "", border=1)
            
            # Plat - nom complet (jusqu'à 35 caractères)
            if row_idx < len(plats):
                p_name = clean_text(plats[row_idx].get("name", ""))[:35]
                p_qty = plats[row_idx].get("quantity_sold", 0)
                pdf.cell(62, 5, p_name, border=1)
                pdf.cell(12, 5, str(p_qty), border=1, align="C")
            else:
                pdf.cell(62, 5, "", border=1)
                pdf.cell(12, 5, "", border=1)
            
            # Dessert - nom complet (jusqu'à 35 caractères)
            if row_idx < len(desserts):
                d_name = clean_text(desserts[row_idx].get("name", ""))[:35]
                d_qty = desserts[row_idx].get("quantity_sold", 0)
                pdf.cell(62, 5, d_name, border=1)
                pdf.cell(12, 5, str(d_qty), border=1, align="C", new_x="LMARGIN", new_y="NEXT")
            else:
                pdf.cell(62, 5, "", border=1)
                pdf.cell(12, 5, "", border=1, new_x="LMARGIN", new_y="NEXT")

    # Générer le buffer
    buffer = BytesIO(pdf.output())
    
    filename = f"rapport_ventes_{start_date}_{today.isoformat()}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/ardoise/sales/export-excel/by-restaurant/{restaurant_id}")
async def export_ardoise_sales_excel_by_restaurant(
    restaurant_id: str,
    period: str = "week"
):
    """Exporter le rapport des ventes en Excel par restaurant_id - Format détaillé jour par jour"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "name": 1}
    )
    if not restaurant:
        raise HTTPException(status_code=404, detail="Restaurant non trouvé")
    restaurant_name = restaurant.get("name", "Restaurant")
    
    today = datetime.now(timezone.utc).date()
    if period == "day":
        start_date = today.isoformat()
    elif period == "week":
        start_date = (today - timedelta(days=7)).isoformat()
    else:
        start_date = (today - timedelta(days=30)).isoformat()
    
    sales = await ardoise_sales_collection.find(
        {
            "restaurant_id": restaurant_id,
            "date": {"$gte": start_date}
        },
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Créer le workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail par jour"
    
    # Style pour les en-têtes
    header_fill = PatternFill(start_color="FFD166", end_color="FFD166", fill_type="solid")
    header_font = Font(bold=True)
    
    # Titre simple
    ws["A1"] = f"{restaurant_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Du {start_date} au {today.isoformat()}"
    ws["A2"].font = Font(italic=True)
    
    # En-têtes du tableau - Ligne 4
    headers = ["Date", "Entree", "Qte", "Plat", "Qte", "Dessert", "Qte"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Données - commencer à la ligne 5
    current_row = 5
    
    for sale in sales:
        date_str = sale.get("date", "")
        
        # Filtrer les items avec quantité > 0
        entrees = [e for e in sale.get("entree", []) if (e.get("quantity_sold") or 0) > 0]
        plats = [p for p in sale.get("plat", []) if (p.get("quantity_sold") or 0) > 0]
        desserts = [d for d in sale.get("dessert", []) if (d.get("quantity_sold") or 0) > 0]
        
        # Nombre max de lignes pour ce jour
        max_rows = max(len(entrees), len(plats), len(desserts), 1)
        
        for row_idx in range(max_rows):
            # Date seulement sur la première ligne
            if row_idx == 0:
                ws.cell(row=current_row, column=1, value=date_str)
            
            # Entrée
            if row_idx < len(entrees):
                ws.cell(row=current_row, column=2, value=entrees[row_idx].get("name", ""))
                ws.cell(row=current_row, column=3, value=entrees[row_idx].get("quantity_sold", 0))
            
            # Plat
            if row_idx < len(plats):
                ws.cell(row=current_row, column=4, value=plats[row_idx].get("name", ""))
                ws.cell(row=current_row, column=5, value=plats[row_idx].get("quantity_sold", 0))
            
            # Dessert
            if row_idx < len(desserts):
                ws.cell(row=current_row, column=6, value=desserts[row_idx].get("name", ""))
                ws.cell(row=current_row, column=7, value=desserts[row_idx].get("quantity_sold", 0))
            
            current_row += 1
    
    # Ajuster les largeurs de colonnes
    column_widths = [12, 35, 6, 35, 6, 35, 6]  # Date, Entree, Qte, Plat, Qte, Dessert, Qte
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"rapport_ventes_{start_date}_{today.isoformat()}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/ardoise/export-pdf")
async def export_ardoise_pdf(current_user: dict = Depends(get_current_user)):
    """Exporter l'ardoise en PDF - une seule page, sans prix par plat, avec formules"""
    restaurant_id = current_user["restaurant_id"]
    
    # Récupérer l'ardoise
    ardoise = await ardoise_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    
    if not ardoise:
        raise HTTPException(status_code=404, detail="Ardoise non trouvée")
    
    # Récupérer le nom du restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "name": 1}
    )
    restaurant_name = restaurant.get("name", "Restaurant") if restaurant else "Restaurant"
    
    # Récupérer les packages (items de la section A L'ARDOISE du menu)
    ardoise_section = await menu_restaurant_sections_collection.find_one(
        {"restaurant_id": restaurant_id, "name": {"$regex": "ardoise", "$options": "i"}}
    )
    
    packages = []
    plat_du_jour_price = None
    if ardoise_section:
        section_id = ardoise_section.get("section_id")
        items_cursor = menu_restaurant_items_collection.find(
            {"section_id": section_id},
            {"_id": 0, "name": 1, "price": 1}
        )
        async for item in items_cursor:
            name = item.get("name", "")
            price = item.get("price")
            # Capturer le prix du plat du jour
            if "plat du jour" in name.lower():
                plat_du_jour_price = price
            else:
                packages.append({"name": name, "price": price})
    
    # Créer le PDF - style ardoise élégant
    pdf = FPDF()
    pdf.set_auto_page_break(auto=False)  # Désactiver la pagination automatique
    pdf.add_page()
    
    # Fond bleu RAL 5008 (Bleu gris)
    pdf.set_fill_color(38, 55, 74)  # #26374A - RAL 5008
    pdf.rect(0, 0, 210, 297, 'F')
    
    # Cadre décoratif principal - beige
    frame_margin = 12
    pdf.set_draw_color(234, 222, 189)  # Beige #EADEBD
    pdf.set_line_width(2)
    pdf.rect(frame_margin, frame_margin, 210 - 2*frame_margin, 297 - 2*frame_margin)
    # Double cadre intérieur
    pdf.set_line_width(0.8)
    pdf.rect(frame_margin + 5, frame_margin + 5, 210 - 2*frame_margin - 10, 297 - 2*frame_margin - 10)
    
    # ============ COINS DÉCORATIFS ART DÉCO ============
    # Positionnés entre les deux cadres (au milieu de l'espace de 5mm)
    # frame_margin = 12, double cadre à frame_margin + 5 = 17
    # Centre de l'espace entre les cadres: 12 + 2.5 = 14.5
    corner_base = frame_margin + 2.5  # Centre entre les deux cadres
    corner_outer = 18
    corner_inner = 12
    
    pdf.set_line_width(1.5)
    pdf.set_draw_color(234, 222, 189)  # Beige
    pdf.set_fill_color(234, 222, 189)  # Beige pour le losange
    
    # Coin haut gauche - Losange exactement centré au point de jointure
    x1, y1 = corner_base + 12, corner_base + 12
    pdf.line(x1, y1, x1 + corner_outer, y1)
    pdf.line(x1, y1, x1, y1 + corner_outer)
    pdf.set_line_width(0.8)
    pdf.line(x1 + 3, y1 + 3, x1 + corner_inner, y1 + 3)
    pdf.line(x1 + 3, y1 + 3, x1 + 3, y1 + corner_inner)
    # Losange centré EXACTEMENT au point x1, y1 (rotation 45°)
    pdf.polygon([(x1, y1 - 2.5), (x1 + 2.5, y1), (x1, y1 + 2.5), (x1 - 2.5, y1)], 'F')
    
    # Coin haut droit - Losange centré exactement
    pdf.set_line_width(1.5)
    x2, y2 = 210 - corner_base - 12, corner_base + 12
    pdf.line(x2, y2, x2 - corner_outer, y2)
    pdf.line(x2, y2, x2, y2 + corner_outer)
    pdf.set_line_width(0.8)
    pdf.line(x2 - 3, y2 + 3, x2 - corner_inner, y2 + 3)
    pdf.line(x2 - 3, y2 + 3, x2 - 3, y2 + corner_inner)
    pdf.polygon([(x2, y2 - 2.5), (x2 + 2.5, y2), (x2, y2 + 2.5), (x2 - 2.5, y2)], 'F')
    
    # Coin bas gauche - Losange centré exactement
    pdf.set_line_width(1.5)
    x3, y3 = corner_base + 12, 297 - corner_base - 12
    pdf.line(x3, y3, x3 + corner_outer, y3)
    pdf.line(x3, y3, x3, y3 - corner_outer)
    pdf.set_line_width(0.8)
    pdf.line(x3 + 3, y3 - 3, x3 + corner_inner, y3 - 3)
    pdf.line(x3 + 3, y3 - 3, x3 + 3, y3 - corner_inner)
    pdf.polygon([(x3, y3 - 2.5), (x3 + 2.5, y3), (x3, y3 + 2.5), (x3 - 2.5, y3)], 'F')
    
    # Coin bas droit - Losange centré exactement
    pdf.set_line_width(1.5)
    x4, y4 = 210 - corner_base - 12, 297 - corner_base - 12
    pdf.line(x4, y4, x4 - corner_outer, y4)
    pdf.line(x4, y4, x4, y4 - corner_outer)
    pdf.set_line_width(0.8)
    pdf.line(x4 - 3, y4 - 3, x4 - corner_inner, y4 - 3)
    pdf.line(x4 - 3, y4 - 3, x4 - 3, y4 - corner_inner)
    pdf.polygon([(x4, y4 - 2.5), (x4 + 2.5, y4), (x4, y4 + 2.5), (x4 - 2.5, y4)], 'F')
    
    center_x = 105  # Centre de la page A4
    
    # ============ LOGO EN HAUT (plus d'espace au-dessus) ============
    logo_y = 35  # Augmenté de 28 à 35 pour descendre le logo
    logo_size = 35  # Un peu plus grand
    logo_x = center_x - logo_size / 2
    logo_displayed = False
    
    # Récupérer le restaurant complet (logo, adresse, email, téléphone)
    full_restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "logo_base64": 1, "address": 1, "email": 1, "phone": 1, 
         "address_street": 1, "address_postal_code": 1, "address_city": 1}
    )
    
    if full_restaurant and full_restaurant.get("logo_base64"):
        try:
            logo_data = base64.b64decode(full_restaurant["logo_base64"])
            logo_io = BytesIO(logo_data)
            
            # Ouvrir l'image pour obtenir ses dimensions et calculer l'aspect ratio
            try:
                from PIL import Image
                logo_img = Image.open(BytesIO(logo_data))
                img_width, img_height = logo_img.size
                aspect_ratio = img_width / img_height if img_height > 0 else 1
                
                # Si le logo est plus haut que large (aspect_ratio < 1), utiliser h=logo_size
                # pour éviter qu'il soit trop grand en hauteur
                if aspect_ratio < 0.8:  # Logo vertical
                    actual_logo_h = logo_size
                    actual_logo_w = logo_size * aspect_ratio
                    actual_logo_x = center_x - actual_logo_w / 2
                    logo_io.seek(0)  # Reset stream
                    pdf.image(logo_io, x=actual_logo_x, y=logo_y, h=actual_logo_h)
                else:  # Logo horizontal ou carré
                    pdf.image(logo_io, x=logo_x, y=logo_y, w=logo_size)
            except Exception:
                # Fallback si PIL échoue
                pdf.image(logo_io, x=logo_x, y=logo_y, w=logo_size)
            
            logo_displayed = True
        except Exception:
            pass
    
    # Si pas de logo, cercle avec initiales
    if not logo_displayed:
        pdf.set_fill_color(200, 170, 100)  # Or/doré
        pdf.ellipse(logo_x, logo_y, logo_size, logo_size, 'F')
        initials = "".join([word[0].upper() for word in restaurant_name.split()[:2]]) if restaurant_name else "AR"
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(45, 52, 54)
        pdf.set_xy(logo_x, logo_y + logo_size/2 - 5)
        pdf.cell(logo_size, 10, initials, align="C")
    
    # Titre principal - plus d'espace après le logo
    pdf.set_y(logo_y + logo_size + 15)  # Augmenté de 8 à 15
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 28)
    pdf.cell(0, 12, "A L'ARDOISE", 0, 1, 'C')
    
    pdf.ln(5)  # Réduit de 12 à 5
    
    # Sections avec plats - plus d'espace pour remplir la page
    sections = [
        ("ENTREE", ardoise.get("entree", [])),
        ("PLAT", ardoise.get("plat", [])),
        ("DESSERT", ardoise.get("dessert", []))
    ]
    
    for section_title, items in sections:
        # Titre de section - centré
        pdf.set_text_color(255, 209, 102)  # Jaune/or
        pdf.set_font('Helvetica', 'B', 16)  # Légèrement réduit
        pdf.cell(0, 10, section_title, 0, 1, 'C')  # Espacement réduit de 14 à 10
        
        # Items - NOM et description côte à côte, CENTRÉ
        for item in items:
            if item.get("name"):
                name = item.get("name", "")
                desc = item.get("description", "")
                
                # Calculer le texte complet pour centrer
                if desc:
                    full_text = f"{name}  -  {desc}"
                else:
                    full_text = name
                
                # Remplacer les caractères spéciaux avant encodage latin-1
                def safe_latin1(text):
                    if not text:
                        return ""
                    # Remplacements manuels pour caractères non-latin1
                    # Apostrophes typographiques et guillemets - codes Unicode explicites
                    text = text.replace('\u2019', "'")  # Right single quotation mark
                    text = text.replace('\u2018', "'")  # Left single quotation mark
                    text = text.replace('\u0060', "'")  # Grave accent
                    text = text.replace('\u00B4', "'")  # Acute accent
                    text = text.replace('\u2032', "'")  # Prime
                    text = text.replace('\u201C', '"')  # Left double quotation mark
                    text = text.replace('\u201D', '"')  # Right double quotation mark
                    text = text.replace('\u00AB', '"')  # Left guillemet
                    text = text.replace('\u00BB', '"')  # Right guillemet
                    # Autres caractères spéciaux
                    text = text.replace('\u0152', 'OE').replace('\u0153', 'oe')  # Œ œ
                    text = text.replace('\u20AC', 'EUR')  # €
                    text = text.replace('\u2026', '...')  # …
                    text = text.replace('\u2013', '-').replace('\u2014', '-')  # – —
                    return text.encode('latin-1', 'replace').decode('latin-1')
                
                name_safe = safe_latin1(name)
                desc_safe = safe_latin1(desc)
                
                # Calculer la largeur totale et position de départ pour centrer
                pdf.set_font('Helvetica', 'B', 12)  # Réduit de 13 à 12
                name_width = pdf.get_string_width(name_safe)
                
                if desc_safe:
                    pdf.set_font('Helvetica', 'I', 10)  # Réduit de 11 à 10
                    desc_width = pdf.get_string_width(f"  -  {desc_safe}")
                    total_width = name_width + desc_width
                else:
                    total_width = name_width
                
                start_x = center_x - total_width / 2
                
                # Nom en blanc gras
                pdf.set_font('Helvetica', 'B', 12)
                pdf.set_text_color(255, 255, 255)
                pdf.set_x(start_x)
                pdf.cell(name_width, 7, name_safe, 0, 0, 'L')  # Réduit de 9 à 7
                
                # Description en beige/marron à côté
                if desc_safe:
                    pdf.set_font('Helvetica', 'I', 10)
                    pdf.set_text_color(210, 180, 140)  # Beige/tan
                    pdf.cell(0, 7, f"  -  {desc_safe}", 0, 1, 'L')  # Réduit de 9 à 7
                else:
                    pdf.ln(7)  # Réduit de 9 à 7
        
        pdf.ln(6)  # Réduit de 15 à 6 - Moins d'espace entre sections
    
    # Ligne décorative dorée avant FORMULES (plus d'espace avant)
    pdf.ln(10)  # Augmenté de 2 à 10 - Plus d'espace entre DESSERT et FORMULES
    pdf.set_draw_color(255, 209, 102)
    pdf.set_line_width(1)
    pdf.line(40, pdf.get_y(), 170, pdf.get_y())
    pdf.ln(8)  # Augmenté de 5 à 8
    
    # Section FORMULES en bas - CENTRÉE - Taille augmentée pour différencier
    pdf.set_text_color(255, 209, 102)
    pdf.set_font('Helvetica', 'B', 18)  # Augmenté de 14 à 18 pour mieux différencier
    pdf.cell(0, 10, "FORMULES", 0, 1, 'C')  # Augmenté de 8 à 10
    
    pdf.set_font('Helvetica', '', 12)  # Augmenté de 11 à 12
    pdf.set_text_color(255, 255, 255)
    
    # Plat du jour d'abord - CENTRÉ avec espacement réduit
    if plat_du_jour_price:
        formule_text = "Plat du jour"
        # Format prix: 15,90€ (pas EUR)
        price_val = float(plat_du_jour_price)
        price_text = f"{price_val:.2f}".replace('.', ',') + chr(128)  # € symbol
        spacing = 10  # Espacement réduit
        total_width = pdf.get_string_width(formule_text) + spacing + pdf.get_string_width(price_text)
        start_x = center_x - total_width / 2
        pdf.set_x(start_x)
        pdf.cell(pdf.get_string_width(formule_text) + spacing, 6, formule_text, 0, 0, 'L')
        pdf.set_text_color(255, 209, 102)  # Prix en doré
        pdf.cell(0, 6, price_text, 0, 1, 'L')
        pdf.set_text_color(255, 255, 255)
    
    # Autres formules - CENTRÉES avec espacement réduit
    for pkg in packages:
        name = pkg.get("name", "")
        price = pkg.get("price")
        if price:
            price_val = float(price)
            price_text = f"{price_val:.2f}".replace('.', ',') + chr(128)  # € symbol
            spacing = 10
            total_width = pdf.get_string_width(name) + spacing + pdf.get_string_width(price_text)
            start_x = center_x - total_width / 2
            pdf.set_x(start_x)
            pdf.cell(pdf.get_string_width(name) + spacing, 6, name.encode('latin-1', 'replace').decode('latin-1'), 0, 0, 'L')
            pdf.set_text_color(255, 209, 102)  # Prix en doré
            pdf.cell(0, 6, price_text, 0, 1, 'L')
            pdf.set_text_color(255, 255, 255)
    
    # PAS de deuxième page - on ne met pas la date
    
    # ============ COORDONNÉES DU RESTAURANT EN BAS ============
    footer_y = 268  # Remonté de 275 à 268 pour ne pas être coincé dans les barres
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(234, 222, 189)  # Beige
    
    # Construire l'adresse complète
    address_parts = []
    if full_restaurant:
        # Préférer les nouveaux champs structurés
        if full_restaurant.get("address_street"):
            address_parts.append(full_restaurant.get("address_street"))
        if full_restaurant.get("address_postal_code") or full_restaurant.get("address_city"):
            city_part = ""
            if full_restaurant.get("address_postal_code"):
                city_part += full_restaurant.get("address_postal_code")
            if full_restaurant.get("address_city"):
                if city_part:
                    city_part += " "
                city_part += full_restaurant.get("address_city")
            if city_part:
                address_parts.append(city_part)
        # Fallback à l'ancien champ address si pas de nouveaux champs
        if not address_parts and full_restaurant.get("address"):
            address_parts.append(full_restaurant.get("address"))
    
    address = " - ".join(address_parts) if address_parts else ""
    if address:
        pdf.set_xy(frame_margin + 10, footer_y)
        pdf.cell(210 - 2*frame_margin - 20, 4, safe_text(address).encode('latin-1', 'replace').decode('latin-1'), 0, 1, 'C')
    
    # Email et téléphone sur une ligne
    email = full_restaurant.get("email", "") if full_restaurant else ""
    phone = full_restaurant.get("phone", "") if full_restaurant else ""
    contact_line = ""
    if email:
        contact_line += email
    if phone:
        if contact_line:
            contact_line += "  |  "
        contact_line += phone
    
    if contact_line:
        pdf.set_xy(frame_margin + 10, footer_y + 4)
        pdf.cell(210 - 2*frame_margin - 20, 4, safe_text(contact_line), 0, 1, 'C')
    
    # Générer le PDF
    pdf_output = pdf.output(dest='S')
    if isinstance(pdf_output, str):
        pdf_output = pdf_output.encode('latin-1')
    
    output = BytesIO(pdf_output)
    output.seek(0)
    
    filename = f"ardoise_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== SOCIAL MEDIA MENU IMAGE ENDPOINT ====================
@api_router.get("/ardoise/export-social/{token}")
async def export_ardoise_social(
    token: str,
    format: str = "instagram_story"  # instagram_story, instagram_post, facebook
):
    """Exporter le menu Ardoise comme image pour réseaux sociaux"""
    from PIL import Image, ImageDraw, ImageFont
    
    # Dimensions selon le format
    dimensions = {
        "instagram_story": (1080, 1920),
        "instagram_post": (1080, 1080),
        "facebook": (1200, 630)
    }
    
    width, height = dimensions.get(format, (1080, 1920))
    
    # Récupérer l'ardoise
    ardoise = await ardoise_collection.find_one(
        {"share_token": token},
        {"_id": 0}
    )
    
    if not ardoise:
        raise HTTPException(status_code=404, detail="Ardoise non trouvée")
    
    restaurant_id = ardoise.get("restaurant_id")
    
    # Récupérer le restaurant
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0, "name": 1, "address": 1, "phone": 1, "email": 1}
    )
    restaurant_name = restaurant.get("name", "Restaurant") if restaurant else "Restaurant"
    
    # Créer l'image avec fond bleu RAL 5008
    img = Image.new('RGB', (width, height), color=(38, 55, 74))
    draw = ImageDraw.Draw(img)
    
    # Définir les couleurs
    beige = (234, 222, 189)
    white = (255, 255, 255)
    gold = (255, 209, 102)
    
    # Bordure beige
    border = 40
    draw.rectangle(
        [(border, border), (width - border, height - border)],
        outline=beige,
        width=3
    )
    draw.rectangle(
        [(border + 10, border + 10), (width - border - 10, height - border - 10)],
        outline=beige,
        width=1
    )
    
    # Utiliser des polices de base
    try:
        title_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 72)
        subtitle_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 48)
        section_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 42)
        item_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 36)
        desc_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf", 28)
        price_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 38)
    except:
        title_font = ImageFont.load_default()
        subtitle_font = title_font
        section_font = title_font
        item_font = title_font
        desc_font = title_font
        price_font = title_font
    
    # Position de départ
    y_pos = border + 80
    center_x = width // 2
    
    # Titre "A L'ARDOISE"
    title = "A L'ARDOISE"
    bbox = draw.textbbox((0, 0), title, font=title_font)
    title_width = bbox[2] - bbox[0]
    draw.text((center_x - title_width // 2, y_pos), title, fill=gold, font=title_font)
    y_pos += 90
    
    # Nom du restaurant
    bbox = draw.textbbox((0, 0), restaurant_name, font=subtitle_font)
    name_width = bbox[2] - bbox[0]
    draw.text((center_x - name_width // 2, y_pos), restaurant_name, fill=white, font=subtitle_font)
    y_pos += 100
    
    # Ligne décorative
    draw.line([(center_x - 100, y_pos), (center_x + 100, y_pos)], fill=gold, width=2)
    y_pos += 50
    
    # Afficher les sections
    sections = ["entree", "plat", "dessert"]
    section_labels = {"entree": "ENTRÉE", "plat": "PLAT", "dessert": "DESSERT"}
    
    for section_key in sections:
        section_items = ardoise.get(section_key, [])
        if section_items and len(section_items) > 0:
            # Titre de section
            section_label = section_labels.get(section_key, section_key.upper())
            bbox = draw.textbbox((0, 0), section_label, font=section_font)
            label_width = bbox[2] - bbox[0]
            draw.text((center_x - label_width // 2, y_pos), section_label, fill=gold, font=section_font)
            y_pos += 60
            
            # Items
            for item in section_items[:2]:  # Max 2 items par section
                item_name = item.get("name", "")
                item_desc = item.get("description", "")
                
                if item_name:
                    bbox = draw.textbbox((0, 0), item_name, font=item_font)
                    item_width = bbox[2] - bbox[0]
                    draw.text((center_x - item_width // 2, y_pos), item_name, fill=white, font=item_font)
                    y_pos += 45
                
                if item_desc:
                    # Tronquer si trop long
                    if len(item_desc) > 50:
                        item_desc = item_desc[:47] + "..."
                    bbox = draw.textbbox((0, 0), item_desc, font=desc_font)
                    desc_width = bbox[2] - bbox[0]
                    draw.text((center_x - desc_width // 2, y_pos), item_desc, fill=beige, font=desc_font)
                    y_pos += 40
            
            y_pos += 30
    
    # Ligne séparatrice
    draw.line([(center_x - 150, y_pos), (center_x + 150, y_pos)], fill=gold, width=2)
    y_pos += 40
    
    # Formules (prix fixes)
    formules_title = "FORMULES"
    bbox = draw.textbbox((0, 0), formules_title, font=section_font)
    label_width = bbox[2] - bbox[0]
    draw.text((center_x - label_width // 2, y_pos), formules_title, fill=gold, font=section_font)
    y_pos += 60
    
    formules = [
        ("Plat du jour", "15,90€"),
        ("Entrée + Plat", "18,90€"),
        ("Plat + Dessert", "18,90€"),
        ("Entrée + Plat + Dessert", "22,90€")
    ]
    
    for name, price in formules:
        line = f"{name}  ·  {price}"
        bbox = draw.textbbox((0, 0), line, font=item_font)
        line_width = bbox[2] - bbox[0]
        draw.text((center_x - line_width // 2, y_pos), line, fill=white, font=item_font)
        y_pos += 50
    
    # Convertir en bytes
    output = BytesIO()
    img.save(output, format='PNG', quality=95)
    output.seek(0)
    
    filename = f"menu_social_{format}_{datetime.now().strftime('%Y%m%d')}.png"
    
    return StreamingResponse(
        output,
        media_type="image/png",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== EVENTS ENDPOINTS ====================

def has_event_access(user: dict, permission_level: str = "admin") -> bool:
    """Vérifier si l'utilisateur a accès au module événements"""
    if user.get("role") in ["admin", "manager"]:
        return True
    event_permission = user.get("permissions", {}).get("events", "none")
    if permission_level == "admin":
        return event_permission == "admin"
    elif permission_level == "read_only":
        return event_permission in ["admin", "read_only", "task_status_only"]
    elif permission_level == "task_status_only":
        return event_permission in ["admin", "task_status_only"]
    return False

@api_router.post("/events")
async def create_event(request: CreateEventRequest, current_user: dict = Depends(get_current_user)):
    """Créer un nouvel événement"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    event_id = f"event_{uuid.uuid4().hex[:12]}"
    event = {
        "event_id": event_id,
        "restaurant_id": current_user["restaurant_id"],
        "title": request.title,
        "date": request.date,
        "description": request.description,
        "notes": request.notes,
        "assigned_team": request.assigned_team,  # Équipe assignée
        "is_active": True,
        "archived": False,  # Champ pour l'archivage
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await events_collection.insert_one(event)
    event.pop("_id", None)
    return event

@api_router.get("/events")
async def list_events(include_archived: bool = False, current_user: dict = Depends(get_current_user)):
    """Lister tous les événements du restaurant, triés par date
    
    - Par défaut, seuls les événements non-archivés sont retournés
    - Les événements dont la date est passée de plus d'un jour sont auto-archivés
    """
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    # Date limite pour l'archivage automatique (hier)
    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).strftime("%Y-%m-%d")
    
    # Archiver automatiquement les événements passés (date < hier)
    await events_collection.update_many(
        {
            "restaurant_id": current_user["restaurant_id"],
            "is_active": True,
            "archived": {"$ne": True},
            "date": {"$lt": yesterday}
        },
        {"$set": {"archived": True, "auto_archived": True, "archived_at": datetime.now(timezone.utc)}}
    )
    
    # Construire le filtre
    query_filter = {
        "restaurant_id": current_user["restaurant_id"],
        "is_active": True
    }
    
    if not include_archived:
        query_filter["$or"] = [
            {"archived": {"$ne": True}},
            {"archived": False}
        ]
    
    events = await events_collection.find(
        query_filter,
        {"_id": 0}
    ).sort("date", 1).to_list(length=1000)
    
    return events

@api_router.get("/events/archived")
async def list_archived_events(current_user: dict = Depends(get_current_user)):
    """Lister les événements archivés du restaurant"""
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    events = await events_collection.find(
        {
            "restaurant_id": current_user["restaurant_id"],
            "is_active": True,
            "archived": True
        },
        {"_id": 0}
    ).sort("date", -1).to_list(length=1000)  # Tri par date décroissante pour les archives
    
    return events

@api_router.post("/events/{event_id}/restore")
async def restore_event(event_id: str, current_user: dict = Depends(get_current_user)):
    """Restaurer un événement archivé"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    event = await events_collection.find_one(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not event:
        raise HTTPException(status_code=404, detail="Événement non trouvé")
    
    if not event.get("archived"):
        raise HTTPException(status_code=400, detail="Cet événement n'est pas archivé")
    
    await events_collection.update_one(
        {"event_id": event_id},
        {"$set": {"archived": False, "auto_archived": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    updated_event = await events_collection.find_one({"event_id": event_id}, {"_id": 0})
    return updated_event

@api_router.post("/events/{event_id}/archive")
async def archive_event(event_id: str, current_user: dict = Depends(get_current_user)):
    """Archiver manuellement un événement"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    event = await events_collection.find_one(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not event:
        raise HTTPException(status_code=404, detail="Événement non trouvé")
    
    await events_collection.update_one(
        {"event_id": event_id},
        {"$set": {"archived": True, "auto_archived": False, "archived_at": datetime.now(timezone.utc), "updated_at": datetime.now(timezone.utc)}}
    )
    
    updated_event = await events_collection.find_one({"event_id": event_id}, {"_id": 0})
    return updated_event

@api_router.get("/events/{event_id}")
async def get_event(event_id: str, current_user: dict = Depends(get_current_user)):
    """Obtenir les détails d'un événement"""
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    event = await events_collection.find_one(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    )
    
    if not event:
        raise HTTPException(status_code=404, detail="Événement non trouvé")
    
    return event

@api_router.put("/events/{event_id}")
async def update_event(event_id: str, request: UpdateEventRequest, current_user: dict = Depends(get_current_user)):
    """Modifier un événement"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    update_data = {k: v for k, v in request.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await events_collection.update_one(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Événement non trouvé")
    
    return await get_event(event_id, current_user)

@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer un événement (soft delete)"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    result = await events_collection.update_one(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Événement non trouvé")
    
    return {"message": "Événement supprimé"}

@api_router.post("/events/{event_id}/duplicate")
async def duplicate_event(event_id: str, request: DuplicateEventRequest, current_user: dict = Depends(get_current_user)):
    """Dupliquer un événement avec tous ses prestataires, tâches et menu"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    original_event = await events_collection.find_one(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"_id": 0}
    )
    
    if not original_event:
        raise HTTPException(status_code=404, detail="Événement non trouvé")
    
    new_event_id = f"event_{uuid.uuid4().hex[:12]}"
    new_event = {
        **original_event,
        "event_id": new_event_id,
        "title": request.new_title or f"{original_event['title']} (copie)",
        "date": request.new_date or original_event["date"],
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
        "created_by": current_user["user_id"]
    }
    
    await events_collection.insert_one(new_event)
    
    # Dupliquer les prestataires
    providers = await event_providers_collection.find({"event_id": event_id}, {"_id": 0}).to_list(length=1000)
    provider_count = 0
    for provider in providers:
        new_provider = {
            **provider,
            "provider_id": f"provider_{uuid.uuid4().hex[:12]}",
            "event_id": new_event_id,
            "quote_path": None,
            "quote_status": "pending",
            "invoice_path": None,
            "invoice_status": "pending",
            "payment_proof_path": None,
            "payment_method": None,
            "created_at": datetime.now(timezone.utc)
        }
        await event_providers_collection.insert_one(new_provider)
        provider_count += 1
    
    # Dupliquer les tâches
    tasks = await event_tasks_collection.find({"event_id": event_id}, {"_id": 0}).to_list(length=1000)
    task_count = 0
    for task in tasks:
        new_task = {
            **task,
            "task_id": f"task_{uuid.uuid4().hex[:12]}",
            "event_id": new_event_id,
            "status": "todo",
            "created_at": datetime.now(timezone.utc)
        }
        await event_tasks_collection.insert_one(new_task)
        task_count += 1
    
    # Dupliquer les sections de menu
    sections = await event_menu_sections_collection.find({"event_id": event_id}, {"_id": 0}).to_list(length=1000)
    section_id_map = {}
    section_count = 0
    for section in sections:
        old_id = section["section_id"]
        new_id = f"section_{uuid.uuid4().hex[:12]}"
        section_id_map[old_id] = new_id
        new_section = {
            **section,
            "section_id": new_id,
            "event_id": new_event_id,
            "created_at": datetime.now(timezone.utc)
        }
        await event_menu_sections_collection.insert_one(new_section)
        section_count += 1
    
    # Dupliquer les items de menu
    items = await event_menu_items_collection.find({"event_id": event_id}, {"_id": 0}).to_list(length=1000)
    item_count = 0
    for item in items:
        new_item = {
            **item,
            "item_id": f"item_{uuid.uuid4().hex[:12]}",
            "event_id": new_event_id,
            "section_id": section_id_map.get(item["section_id"], item["section_id"]),
            "created_at": datetime.now(timezone.utc)
        }
        await event_menu_items_collection.insert_one(new_item)
        item_count += 1
    
    # Dupliquer les packages de prix
    packages = await event_price_packages_collection.find({"event_id": event_id}, {"_id": 0}).to_list(length=1000)
    package_count = 0
    for package in packages:
        new_section_ids = [section_id_map.get(sid, sid) for sid in package.get("section_ids", [])]
        new_package = {
            **package,
            "package_id": f"package_{uuid.uuid4().hex[:12]}",
            "event_id": new_event_id,
            "section_ids": new_section_ids,
            "created_at": datetime.now(timezone.utc)
        }
        await event_price_packages_collection.insert_one(new_package)
        package_count += 1
    
    # Dupliquer les options de boissons
    drinks = await event_drink_options_collection.find({"event_id": event_id}, {"_id": 0}).to_list(length=1000)
    drink_count = 0
    for drink in drinks:
        new_drink = {
            **drink,
            "drink_id": f"drink_{uuid.uuid4().hex[:12]}",
            "event_id": new_event_id,
            "is_selected": False,
            "created_at": datetime.now(timezone.utc)
        }
        await event_drink_options_collection.insert_one(new_drink)
        drink_count += 1
    
    new_event.pop("_id", None)
    return {
        "message": "Événement dupliqué avec succès",
        "event": new_event,
        "duplicated": {
            "providers": provider_count,
            "tasks": task_count,
            "menu_sections": section_count,
            "menu_items": item_count,
            "price_packages": package_count,
            "drink_options": drink_count
        }
    }

# ==================== PROVIDERS (PRESTATAIRES) ENDPOINTS ====================

@api_router.post("/events/{event_id}/providers")
async def create_provider(event_id: str, request: CreateProviderRequest, current_user: dict = Depends(get_current_user)):
    """Ajouter un prestataire à un événement"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    event = await events_collection.find_one({"event_id": event_id, "restaurant_id": current_user["restaurant_id"]})
    if not event:
        raise HTTPException(status_code=404, detail="Événement non trouvé")
    
    provider_id = f"provider_{uuid.uuid4().hex[:12]}"
    provider = {
        "provider_id": provider_id,
        "event_id": event_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": request.name,
        "contact_name": request.contact_name,
        "phone": request.phone,
        "email": request.email,
        "notes": request.notes,
        "quote_path": None,
        "quote_status": "pending",
        "invoice_path": None,
        "invoice_status": "pending",
        "payment_method": None,
        "payment_proof_path": None,
        "is_active": True,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await event_providers_collection.insert_one(provider)
    provider.pop("_id", None)
    return provider

@api_router.get("/events/{event_id}/providers")
async def list_providers(event_id: str, current_user: dict = Depends(get_current_user)):
    """Lister les prestataires d'un événement"""
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    providers = await event_providers_collection.find(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).to_list(length=1000)
    
    return providers

@api_router.put("/events/{event_id}/providers/{provider_id}")
async def update_provider(event_id: str, provider_id: str, request: UpdateProviderRequest, current_user: dict = Depends(get_current_user)):
    """Modifier un prestataire"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    update_data = {k: v for k, v in request.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await event_providers_collection.update_one(
        {"provider_id": provider_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Prestataire non trouvé")
    
    provider = await event_providers_collection.find_one({"provider_id": provider_id}, {"_id": 0})
    return provider

@api_router.delete("/events/{event_id}/providers/{provider_id}")
async def delete_provider(event_id: str, provider_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer un prestataire"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    result = await event_providers_collection.update_one(
        {"provider_id": provider_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Prestataire non trouvé")
    
    return {"message": "Prestataire supprimé"}

@api_router.post("/events/{event_id}/providers/{provider_id}/quote")
async def upload_provider_quote(event_id: str, provider_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Uploader un devis PDF pour un prestataire"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Seuls les fichiers PDF sont acceptés")
    
    event_dir = os.path.join(EVENTS_UPLOADS_DIR, event_id, "providers", provider_id)
    os.makedirs(event_dir, exist_ok=True)
    
    file_path = os.path.join(event_dir, f"quote_{uuid.uuid4().hex[:8]}.pdf")
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    await event_providers_collection.update_one(
        {"provider_id": provider_id, "event_id": event_id},
        {"$set": {"quote_path": file_path, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Devis uploadé", "path": file_path}

@api_router.post("/events/{event_id}/providers/{provider_id}/validate-quote")
async def validate_provider_quote(event_id: str, provider_id: str, request: ValidateQuoteRequest, current_user: dict = Depends(get_current_user)):
    """Valider un devis"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    provider = await event_providers_collection.find_one({"provider_id": provider_id, "event_id": event_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestataire non trouvé")
    
    if not provider.get("quote_path"):
        raise HTTPException(status_code=400, detail="Aucun devis uploadé")
    
    status = "validated" if request.validated else "pending"
    await event_providers_collection.update_one(
        {"provider_id": provider_id},
        {"$set": {"quote_status": status, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": f"Devis {'validé' if request.validated else 'remis en attente'}"}

@api_router.post("/events/{event_id}/providers/{provider_id}/invoice")
async def upload_provider_invoice(event_id: str, provider_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Uploader une facture PDF pour un prestataire"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Seuls les fichiers PDF sont acceptés")
    
    event_dir = os.path.join(EVENTS_UPLOADS_DIR, event_id, "providers", provider_id)
    os.makedirs(event_dir, exist_ok=True)
    
    file_path = os.path.join(event_dir, f"invoice_{uuid.uuid4().hex[:8]}.pdf")
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    await event_providers_collection.update_one(
        {"provider_id": provider_id, "event_id": event_id},
        {"$set": {"invoice_path": file_path, "invoice_status": "awaiting_payment", "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Facture uploadée", "path": file_path}

@api_router.post("/events/{event_id}/providers/{provider_id}/invoice-status")
async def update_invoice_status(event_id: str, provider_id: str, request: UpdateInvoiceStatusRequest, current_user: dict = Depends(get_current_user)):
    """Mettre à jour le statut de la facture"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    if request.status not in ["pending", "awaiting_payment", "paid"]:
        raise HTTPException(status_code=400, detail="Statut invalide")
    
    if request.status == "paid" and not request.payment_method:
        raise HTTPException(status_code=400, detail="Méthode de paiement requise pour statut 'payé'")
    
    update_data = {"invoice_status": request.status, "updated_at": datetime.now(timezone.utc)}
    if request.payment_method:
        update_data["payment_method"] = request.payment_method
    
    await event_providers_collection.update_one(
        {"provider_id": provider_id, "event_id": event_id},
        {"$set": update_data}
    )
    
    return {"message": "Statut de facture mis à jour"}

@api_router.post("/events/{event_id}/providers/{provider_id}/payment-proof")
async def upload_payment_proof(event_id: str, provider_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Uploader une preuve de paiement"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    allowed_extensions = ['.pdf', '.png', '.jpg', '.jpeg']
    if not any(file.filename.lower().endswith(ext) for ext in allowed_extensions):
        raise HTTPException(status_code=400, detail="Format de fichier non accepté (PDF, PNG, JPG)")
    
    event_dir = os.path.join(EVENTS_UPLOADS_DIR, event_id, "providers", provider_id)
    os.makedirs(event_dir, exist_ok=True)
    
    ext = os.path.splitext(file.filename)[1].lower()
    file_path = os.path.join(event_dir, f"payment_proof_{uuid.uuid4().hex[:8]}{ext}")
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    await event_providers_collection.update_one(
        {"provider_id": provider_id, "event_id": event_id},
        {"$set": {"payment_proof_path": file_path, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Preuve de paiement uploadée", "path": file_path}

@api_router.get("/events/{event_id}/providers/{provider_id}/download/{file_type}")
async def download_provider_file(event_id: str, provider_id: str, file_type: str, current_user: dict = Depends(get_current_user)):
    """Télécharger un fichier d'un prestataire (quote, invoice, payment_proof)"""
    from fastapi.responses import FileResponse
    
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    if file_type not in ["quote", "invoice", "payment_proof"]:
        raise HTTPException(status_code=400, detail="Type de fichier invalide")
    
    provider = await event_providers_collection.find_one(
        {"provider_id": provider_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]}
    )
    
    if not provider:
        raise HTTPException(status_code=404, detail="Prestataire non trouvé")
    
    file_path = provider.get(f"{file_type}_path")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Fichier non trouvé")
    
    return FileResponse(file_path, filename=os.path.basename(file_path))

# ==================== EVENT TASKS ENDPOINTS ====================

@api_router.post("/events/{event_id}/tasks")
async def create_event_task(event_id: str, request: CreateEventTaskRequest, current_user: dict = Depends(get_current_user)):
    """Créer une tâche pour un événement"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    event = await events_collection.find_one({"event_id": event_id, "restaurant_id": current_user["restaurant_id"]})
    if not event:
        raise HTTPException(status_code=404, detail="Événement non trouvé")
    
    task_id = f"task_{uuid.uuid4().hex[:12]}"
    
    assigned_user_name = None
    if request.assigned_user_id:
        assigned_user = await users_collection.find_one({"user_id": request.assigned_user_id}, {"name": 1})
        if assigned_user:
            assigned_user_name = assigned_user.get("name")
    
    task = {
        "task_id": task_id,
        "event_id": event_id,
        "restaurant_id": current_user["restaurant_id"],
        "title": request.title,
        "description": request.description,
        "due_date": request.due_date,
        "assigned_user_id": request.assigned_user_id,
        "assigned_user_name": assigned_user_name,
        "status": "todo",
        "is_active": True,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await event_tasks_collection.insert_one(task)
    task.pop("_id", None)
    return task

@api_router.get("/events/{event_id}/tasks")
async def list_event_tasks(event_id: str, current_user: dict = Depends(get_current_user)):
    """Lister les tâches d'un événement"""
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    tasks = await event_tasks_collection.find(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).sort("due_date", 1).to_list(length=1000)
    
    return tasks

@api_router.put("/events/{event_id}/tasks/{task_id}")
async def update_event_task(event_id: str, task_id: str, request: UpdateEventTaskRequest, current_user: dict = Depends(get_current_user)):
    """Modifier une tâche d'événement"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    update_data = {k: v for k, v in request.model_dump().items() if v is not None}
    
    if "assigned_user_id" in update_data:
        if update_data["assigned_user_id"]:
            assigned_user = await users_collection.find_one({"user_id": update_data["assigned_user_id"]}, {"name": 1})
            update_data["assigned_user_name"] = assigned_user.get("name") if assigned_user else None
        else:
            update_data["assigned_user_name"] = None
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await event_tasks_collection.update_one(
        {"task_id": task_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tâche non trouvée")
    
    task = await event_tasks_collection.find_one({"task_id": task_id}, {"_id": 0})
    return task

@api_router.put("/events/{event_id}/tasks/{task_id}/status")
async def update_event_task_status(event_id: str, task_id: str, request: UpdateTaskStatusRequest, current_user: dict = Depends(get_current_user)):
    """Mettre à jour uniquement le statut d'une tâche"""
    if not has_event_access(current_user, "task_status_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    if request.status not in ["todo", "in_progress", "completed"]:
        raise HTTPException(status_code=400, detail="Statut invalide")
    
    if not has_event_access(current_user, "admin"):
        task = await event_tasks_collection.find_one({"task_id": task_id, "event_id": event_id})
        if task and task.get("assigned_user_id") != current_user["user_id"]:
            raise HTTPException(status_code=403, detail="Vous pouvez uniquement modifier vos tâches assignées")
    
    result = await event_tasks_collection.update_one(
        {"task_id": task_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"status": request.status, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tâche non trouvée")
    
    return {"message": "Statut mis à jour"}

@api_router.delete("/events/{event_id}/tasks/{task_id}")
async def delete_event_task(event_id: str, task_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer une tâche"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    result = await event_tasks_collection.update_one(
        {"task_id": task_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tâche non trouvée")
    
    return {"message": "Tâche supprimée"}

# ==================== EVENT MENU ENDPOINTS ====================

@api_router.post("/events/{event_id}/menu/sections")
async def create_event_menu_section(event_id: str, request: CreateEventMenuSectionRequest, current_user: dict = Depends(get_current_user)):
    """Créer une section de menu pour l'événement"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    event = await events_collection.find_one({"event_id": event_id, "restaurant_id": current_user["restaurant_id"]})
    if not event:
        raise HTTPException(status_code=404, detail="Événement non trouvé")
    
    if request.order is None:
        last_section = await event_menu_sections_collection.find_one({"event_id": event_id}, sort=[("order", -1)])
        order = (last_section["order"] + 1) if last_section else 0
    else:
        order = request.order
    
    section_id = f"section_{uuid.uuid4().hex[:12]}"
    section = {
        "section_id": section_id,
        "event_id": event_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": request.name,
        "order": order,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    
    await event_menu_sections_collection.insert_one(section)
    section.pop("_id", None)
    return section

@api_router.get("/events/{event_id}/menu/sections")
async def list_event_menu_sections(event_id: str, current_user: dict = Depends(get_current_user)):
    """Lister les sections de menu d'un événement"""
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    sections = await event_menu_sections_collection.find(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).sort("order", 1).to_list(length=1000)
    
    return sections

@api_router.put("/events/{event_id}/menu/sections/{section_id}")
async def update_event_menu_section(event_id: str, section_id: str, request: UpdateEventMenuSectionRequest, current_user: dict = Depends(get_current_user)):
    """Modifier une section de menu"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    update_data = {k: v for k, v in request.model_dump().items() if v is not None}
    
    result = await event_menu_sections_collection.update_one(
        {"section_id": section_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Section non trouvée")
    
    section = await event_menu_sections_collection.find_one({"section_id": section_id}, {"_id": 0})
    return section

@api_router.delete("/events/{event_id}/menu/sections/{section_id}")
async def delete_event_menu_section(event_id: str, section_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer une section de menu"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    result = await event_menu_sections_collection.update_one(
        {"section_id": section_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Section non trouvée")
    
    return {"message": "Section supprimée"}

@api_router.post("/events/{event_id}/menu/items")
async def create_event_menu_item(event_id: str, request: CreateEventMenuItemRequest, current_user: dict = Depends(get_current_user)):
    """Créer un item de menu"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    if request.order is None:
        last_item = await event_menu_items_collection.find_one({"section_id": request.section_id}, sort=[("order", -1)])
        order = (last_item["order"] + 1) if last_item else 0
    else:
        order = request.order
    
    item_id = f"item_{uuid.uuid4().hex[:12]}"
    item = {
        "item_id": item_id,
        "event_id": event_id,
        "section_id": request.section_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": request.name,
        "description": request.description,
        "order": order,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    
    await event_menu_items_collection.insert_one(item)
    item.pop("_id", None)
    return item

@api_router.get("/events/{event_id}/menu/items")
async def list_event_menu_items(event_id: str, section_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Lister les items de menu d'un événement"""
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    query = {"event_id": event_id, "restaurant_id": current_user["restaurant_id"], "is_active": True}
    if section_id:
        query["section_id"] = section_id
    
    items = await event_menu_items_collection.find(query, {"_id": 0}).sort("order", 1).to_list(length=1000)
    return items

@api_router.put("/events/{event_id}/menu/items/{item_id}")
async def update_event_menu_item(event_id: str, item_id: str, request: UpdateEventMenuItemRequest, current_user: dict = Depends(get_current_user)):
    """Modifier un item de menu"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    update_data = {k: v for k, v in request.model_dump().items() if v is not None}
    
    result = await event_menu_items_collection.update_one(
        {"item_id": item_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item non trouvé")
    
    item = await event_menu_items_collection.find_one({"item_id": item_id}, {"_id": 0})
    return item

@api_router.delete("/events/{event_id}/menu/items/{item_id}")
async def delete_event_menu_item(event_id: str, item_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer un item de menu"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    result = await event_menu_items_collection.update_one(
        {"item_id": item_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item non trouvé")
    
    return {"message": "Item supprimé"}

# ==================== PRICE PACKAGES ENDPOINTS ====================

@api_router.post("/events/{event_id}/menu/packages")
async def create_event_price_package(event_id: str, request: CreatePricePackageRequest, current_user: dict = Depends(get_current_user)):
    """Créer un package de prix (ex: Entrée + Plat = 25€)"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    package_id = f"package_{uuid.uuid4().hex[:12]}"
    package = {
        "package_id": package_id,
        "event_id": event_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": request.name,
        "section_ids": request.section_ids,
        "price": request.price,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    
    await event_price_packages_collection.insert_one(package)
    package.pop("_id", None)
    return package

@api_router.get("/events/{event_id}/menu/packages")
async def list_event_price_packages(event_id: str, current_user: dict = Depends(get_current_user)):
    """Lister les packages de prix d'un événement"""
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    packages = await event_price_packages_collection.find(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).to_list(length=1000)
    
    return packages

@api_router.put("/events/{event_id}/menu/packages/{package_id}")
async def update_event_price_package(event_id: str, package_id: str, request: UpdatePricePackageRequest, current_user: dict = Depends(get_current_user)):
    """Modifier un package de prix"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    update_data = {k: v for k, v in request.model_dump().items() if v is not None}
    
    result = await event_price_packages_collection.update_one(
        {"package_id": package_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Package non trouvé")
    
    package = await event_price_packages_collection.find_one({"package_id": package_id}, {"_id": 0})
    return package

@api_router.delete("/events/{event_id}/menu/packages/{package_id}")
async def delete_event_price_package(event_id: str, package_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer un package de prix"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    result = await event_price_packages_collection.update_one(
        {"package_id": package_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Package non trouvé")
    
    return {"message": "Package supprimé"}

# ==================== DRINK OPTIONS ENDPOINTS ====================

@api_router.post("/events/{event_id}/menu/drinks")
async def create_event_drink_option(event_id: str, request: CreateDrinkOptionRequest, current_user: dict = Depends(get_current_user)):
    """Créer une option de boisson"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    drink_id = f"drink_{uuid.uuid4().hex[:12]}"
    drink = {
        "drink_id": drink_id,
        "event_id": event_id,
        "restaurant_id": current_user["restaurant_id"],
        "name": request.name,
        "price": request.price,
        "is_selected": False,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    
    await event_drink_options_collection.insert_one(drink)
    drink.pop("_id", None)
    return drink

@api_router.get("/events/{event_id}/menu/drinks")
async def list_event_drink_options(event_id: str, current_user: dict = Depends(get_current_user)):
    """Lister les options de boissons d'un événement"""
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    drinks = await event_drink_options_collection.find(
        {"event_id": event_id, "restaurant_id": current_user["restaurant_id"], "is_active": True},
        {"_id": 0}
    ).to_list(length=1000)
    
    return drinks

@api_router.put("/events/{event_id}/menu/drinks/{drink_id}")
async def update_event_drink_option(event_id: str, drink_id: str, request: UpdateDrinkOptionRequest, current_user: dict = Depends(get_current_user)):
    """Modifier une option de boisson"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    update_data = {k: v for k, v in request.model_dump().items() if v is not None}
    
    result = await event_drink_options_collection.update_one(
        {"drink_id": drink_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Boisson non trouvée")
    
    drink = await event_drink_options_collection.find_one({"drink_id": drink_id}, {"_id": 0})
    return drink

@api_router.delete("/events/{event_id}/menu/drinks/{drink_id}")
async def delete_event_drink_option(event_id: str, drink_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer une option de boisson"""
    if not has_event_access(current_user, "admin"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    result = await event_drink_options_collection.update_one(
        {"drink_id": drink_id, "event_id": event_id, "restaurant_id": current_user["restaurant_id"]},
        {"$set": {"is_active": False}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Boisson non trouvée")
    
    return {"message": "Boisson supprimée"}

# ==================== EVENT MENU PDF EXPORT ====================

def get_default_section_color(name: str) -> str:
    """Obtenir une couleur par défaut basée sur le nom de la section"""
    lower_name = name.lower()
    if 'entrée' in lower_name or 'entree' in lower_name:
        return '#27ae60'  # Vert
    if 'plat' in lower_name or 'principal' in lower_name:
        return '#8e44ad'  # Violet
    if 'dessert' in lower_name:
        return '#e67e22'  # Orange
    if 'boisson' in lower_name or 'drink' in lower_name:
        return '#3498db'  # Bleu
    if 'fromage' in lower_name:
        return '#f1c40f'  # Jaune
    if 'accompagnement' in lower_name:
        return '#1abc9c'  # Turquoise
    return '#3498db'  # Bleu par défaut

@api_router.get("/events/{event_id}/menu/export-pdf")
async def export_event_menu_pdf(event_id: str, current_user: dict = Depends(get_current_user)):
    """Exporter le menu d'un événement en PDF avec design graphique élégant et centré"""
    if not has_event_access(current_user, "read_only"):
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    restaurant_id = current_user["restaurant_id"]
    
    # Get event details
    event = await events_collection.find_one(
        {"event_id": event_id, "restaurant_id": restaurant_id},
        {"_id": 0}
    )
    if not event:
        raise HTTPException(status_code=404, detail="Événement non trouvé")
    
    # Get restaurant info
    restaurant = await restaurants_collection.find_one(
        {"restaurant_id": restaurant_id},
        {"_id": 0}
    )
    
    # Get menu sections with items
    sections = await event_menu_sections_collection.find(
        {"event_id": event_id, "restaurant_id": restaurant_id, "is_active": True},
        {"_id": 0}
    ).sort("order", 1).to_list(100)
    
    # Get all menu items for this event
    all_menu_items = await event_menu_items_collection.find(
        {"event_id": event_id, "restaurant_id": restaurant_id, "is_active": True},
        {"_id": 0}
    ).sort("order", 1).to_list(500)
    
    # Get packages
    packages = await event_price_packages_collection.find(
        {"event_id": event_id, "restaurant_id": restaurant_id, "is_active": True},
        {"_id": 0}
    ).sort("order", 1).to_list(50)
    
    # Create PDF - design élégant, CENTRÉ et AGRANDI
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=False)
    
    page_width = 210  # A4 width
    page_height = 297  # A4 height
    margin = 15
    content_width = page_width - 2*margin
    center_x = page_width / 2
    
    # ============ FOND ÉLÉGANT AVEC MOTIFS ============
    # Fond RAL 1013 (Blanc perle)
    pdf.set_fill_color(234, 222, 189)  # #EADEBD - RAL 1013
    pdf.rect(0, 0, page_width, page_height, 'F')
    
    # ============ COINS DÉCORATIFS ART DÉCO (LOSANGE CENTRÉ EXACTEMENT) ============
    corner_offset = 15  # Distance du bord
    corner_outer = 20   # Longueur des lignes extérieures
    corner_inner = 14   # Longueur des lignes intérieures
    
    pdf.set_draw_color(38, 55, 74)  # RAL 5008 Bleu
    pdf.set_fill_color(38, 55, 74)  # RAL 5008 Bleu pour le losange
    
    # Coin haut gauche - Losange exactement centré au point de jointure
    x1, y1 = margin + corner_offset, margin + corner_offset
    pdf.set_line_width(1.5)
    pdf.line(x1, y1, x1 + corner_outer, y1)  # Horizontal
    pdf.line(x1, y1, x1, y1 + corner_outer)  # Vertical
    pdf.set_line_width(0.8)
    pdf.line(x1 + 4, y1 + 4, x1 + corner_inner, y1 + 4)  # Horizontal intérieur
    pdf.line(x1 + 4, y1 + 4, x1 + 4, y1 + corner_inner)  # Vertical intérieur
    # Losange centré EXACTEMENT au point x1, y1 (rotation 45°)
    pdf.polygon([(x1, y1 - 3), (x1 + 3, y1), (x1, y1 + 3), (x1 - 3, y1)], 'F')
    
    # Coin haut droit - Losange centré au point de jointure
    x2, y2 = page_width - margin - corner_offset, margin + corner_offset
    pdf.set_line_width(1.5)
    pdf.line(x2, y2, x2 - corner_outer, y2)
    pdf.line(x2, y2, x2, y2 + corner_outer)
    pdf.set_line_width(0.8)
    pdf.line(x2 - 4, y2 + 4, x2 - corner_inner, y2 + 4)
    pdf.line(x2 - 4, y2 + 4, x2 - 4, y2 + corner_inner)
    pdf.polygon([(x2, y2 - 3), (x2 + 3, y2), (x2, y2 + 3), (x2 - 3, y2)], 'F')
    
    # Coin bas gauche - Losange centré au point de jointure
    x3, y3 = margin + corner_offset, page_height - margin - corner_offset
    pdf.set_line_width(1.5)
    pdf.line(x3, y3, x3 + corner_outer, y3)
    pdf.line(x3, y3, x3, y3 - corner_outer)
    pdf.set_line_width(0.8)
    pdf.line(x3 + 4, y3 - 4, x3 + corner_inner, y3 - 4)
    pdf.line(x3 + 4, y3 - 4, x3 + 4, y3 - corner_inner)
    pdf.polygon([(x3, y3 - 3), (x3 + 3, y3), (x3, y3 + 3), (x3 - 3, y3)], 'F')
    
    # Coin bas droit - Losange centré au point de jointure
    x4, y4 = page_width - margin - corner_offset, page_height - margin - corner_offset
    pdf.set_line_width(1.5)
    pdf.line(x4, y4, x4 - corner_outer, y4)
    pdf.line(x4, y4, x4, y4 - corner_outer)
    pdf.set_line_width(0.8)
    pdf.line(x4 - 4, y4 - 4, x4 - corner_inner, y4 - 4)
    pdf.line(x4 - 4, y4 - 4, x4 - 4, y4 - corner_inner)
    pdf.polygon([(x4, y4 - 3), (x4 + 3, y4), (x4, y4 + 3), (x4 - 3, y4)], 'F')
    
    # Bordure double élégante - bleu RAL 5008
    pdf.set_draw_color(38, 55, 74)  # #26374A - RAL 5008
    pdf.set_line_width(0.8)
    pdf.rect(margin, margin, content_width, page_height - 2*margin)
    pdf.set_line_width(0.3)
    pdf.rect(margin + 3, margin + 3, content_width - 6, page_height - 2*margin - 6)
    
    # ============ EN-TÊTE AVEC LOGO - CENTRÉ (logo plus haut) ============
    header_y = margin + 8  # Réduit de 15 à 8 pour remonter le logo
    logo_size = 30  # Légèrement réduit
    logo_x = center_x - logo_size / 2
    
    # Utiliser TOUJOURS le logo bleu RAL 5008 (Le Cercle) pour le PDF Événement
    logo_displayed = False
    try:
        default_logo_path = ROOT_DIR / "logo_bleu_ral5008.png"
        if default_logo_path.exists():
            pdf.image(str(default_logo_path), x=logo_x, y=header_y, w=logo_size)
            logo_displayed = True
    except Exception as e:
        print(f"Logo error: {e}")
    
    # Si le logo n'existe pas, utiliser le logo du restaurant en fallback
    if not logo_displayed and restaurant and restaurant.get("logo_base64"):
        try:
            logo_data = base64.b64decode(restaurant["logo_base64"])
            logo_io = BytesIO(logo_data)
            pdf.image(logo_io, x=logo_x, y=header_y, w=logo_size)
            logo_displayed = True
        except Exception:
            pass
    
    # Si toujours pas de logo, cercle avec initiales
    if not logo_displayed and restaurant:
        logo_center_y = header_y + logo_size / 2
        pdf.set_fill_color(38, 55, 74)  # RAL 5008 Bleu
        pdf.ellipse(logo_x, header_y, logo_size, logo_size, 'F')
        initials = "".join([word[0].upper() for word in restaurant.get("name", "R").split()[:2]])
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_text_color(255, 255, 255)
        pdf.set_xy(logo_x, logo_center_y - 5)
        pdf.cell(logo_size, 10, safe_text(initials), align="C")
    
    # Ligne décorative avec losange central - après le logo (moins d'espace)
    deco_y = header_y + logo_size + 12  # Augmenté de 5 à 12 pour plus d'espace entre logo et ligne
    pdf.set_draw_color(38, 55, 74)  # RAL 5008 Bleu
    pdf.set_line_width(0.4)
    line_width = 70
    pdf.line(center_x - line_width/2, deco_y, center_x + line_width/2, deco_y)
    pdf.set_fill_color(38, 55, 74)  # RAL 5008 Bleu
    diamond_size = 3
    pdf.polygon([(center_x, deco_y - diamond_size), (center_x + diamond_size, deco_y), 
                 (center_x, deco_y + diamond_size), (center_x - diamond_size, deco_y)], 'F')
    
    # ============ TITRE + DATE SUR LA MÊME LIGNE (moins d'espace) ============
    title_y = deco_y + 4  # Réduit de 10 à 4
    event_title = event.get("title", "Menu")
    event_date = event.get("date", "")
    
    # Formater la date
    formatted_date = ""
    if event_date:
        try:
            from datetime import datetime as dt
            date_obj = dt.strptime(event_date, "%Y-%m-%d")
            months_fr = ["janvier", "fevrier", "mars", "avril", "mai", "juin", 
                        "juillet", "aout", "septembre", "octobre", "novembre", "decembre"]
            formatted_date = f"{date_obj.day} {months_fr[date_obj.month - 1]} {date_obj.year}"
        except:
            formatted_date = event_date
    
    # Titre et date sur la même ligne
    pdf.set_font("Helvetica", "B", 24)
    pdf.set_text_color(44, 44, 44)
    title_width = pdf.get_string_width(event_title)
    
    if formatted_date:
        pdf.set_font("Helvetica", "I", 12)
        date_width = pdf.get_string_width(f"  -  {formatted_date}")
        total_width = title_width + date_width
        start_x = center_x - total_width / 2
        
        pdf.set_font("Helvetica", "B", 24)
        pdf.set_xy(start_x, title_y)
        pdf.cell(title_width, 12, safe_text(event_title), ln=False)
        
        pdf.set_font("Helvetica", "I", 12)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(date_width, 12, safe_text(f"  -  {formatted_date}"), ln=True)
    else:
        pdf.set_xy(margin, title_y)
        pdf.cell(content_width, 12, safe_text(event_title), align="C")
    
    # Seconde ligne décorative (espace réduit)
    deco2_y = title_y + 14  # Réduit de 16 à 14
    pdf.set_draw_color(180, 160, 120)
    pdf.set_line_width(0.3)
    pdf.line(margin + 30, deco2_y, page_width - margin - 30, deco2_y)
    
    # ============ SECTIONS DU MENU - CENTRÉES ============
    content_start_y = deco2_y + 6  # Réduit de 10 à 6
    pdf.set_y(content_start_y)
    
    # Espacements ajustés - réduits pour moins d'espace entre sections
    section_spacing = 5  # Réduit de 12 à 5 pour moins d'espace entre sections
    item_height = 8  # Réduit de 9 à 8
    item_spacing = 2  # Réduit de 5 à 2
    
    for section in sections:
        section_name = safe_text(section.get("name", "Section"))
        section_color = section.get("color") or get_default_section_color(section_name)
        r, g, b = hex_to_rgb(section_color)
        current_y = pdf.get_y()
        
        # Titre de section avec couleur - CENTRÉ (sans carré coloré)
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(r, g, b)
        
        # Titre centré sans carré
        title_text = section_name.upper()
        title_width = pdf.get_string_width(title_text)
        
        pdf.set_xy(center_x - title_width/2, current_y)
        pdf.cell(title_width, item_height, title_text, ln=True, align="C")
        
        # Ligne sous le titre - CENTRÉE
        pdf.set_draw_color(r, g, b)
        pdf.set_line_width(0.3)
        underline_y = pdf.get_y() - 1
        pdf.line(center_x - title_width/2 - 5, underline_y, center_x + title_width/2 + 5, underline_y)
        
        pdf.ln(item_spacing)
        
        # Items de la section - NOM + DESCRIPTION SUR LA MÊME LIGNE
        section_id = section.get("section_id")
        items = [item for item in all_menu_items if item.get("section_id") == section_id]
        
        for item in items:
            item_name = safe_text(item.get("name", ""))
            item_desc = safe_text(item.get("description", ""))
            
            # Calculer les dimensions
            pdf.set_font("Helvetica", "B", 12)
            name_width = pdf.get_string_width(f"- {item_name}")
            
            if item_desc:
                pdf.set_font("Helvetica", "I", 10)
                desc_width = pdf.get_string_width(f"  {item_desc}")
                total_width = name_width + desc_width
                start_x = center_x - total_width / 2
                
                # Nom en gras
                pdf.set_font("Helvetica", "B", 12)
                pdf.set_text_color(0, 0, 0)
                pdf.set_xy(start_x, pdf.get_y())
                pdf.cell(name_width, item_height - 1, f"- {item_name}", ln=False)
                
                # Description en italique à côté
                pdf.set_font("Helvetica", "I", 10)
                pdf.set_text_color(139, 90, 43)
                pdf.cell(desc_width, item_height - 1, f"  {item_desc}", ln=True)
            else:
                pdf.set_font("Helvetica", "B", 12)
                pdf.set_text_color(0, 0, 0)
                pdf.set_xy(margin, pdf.get_y())
                pdf.cell(content_width, item_height - 1, f"- {item_name}", align="C")
                pdf.ln()
        
        pdf.ln(section_spacing)
    
    # ============ SECTION FORMULES - CENTRÉE avec prix rapprochés ============
    if packages:
        # Ajouter plus d'espace entre la dernière section et FORMULES
        pdf.ln(8)  # Espace supplémentaire avant FORMULES
        
        current_y = pdf.get_y()
        
        # Titre FORMULES - CENTRÉ (sans carré coloré)
        title_text = "FORMULES"
        title_width = pdf.get_string_width(title_text)
        
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(180, 150, 80)
        pdf.set_xy(center_x - title_width/2, current_y)
        pdf.cell(title_width, item_height, safe_text(title_text), ln=True, align="C")
        
        pdf.set_draw_color(200, 170, 100)
        pdf.set_line_width(0.3)
        underline_y = pdf.get_y() - 1
        pdf.line(center_x - title_width/2 - 5, underline_y, center_x + title_width/2 + 5, underline_y)
        
        pdf.ln(item_spacing + 2)
        
        for pkg in packages:
            pkg_price = pkg.get("price")
            
            # Construire le nom d'affichage à partir des sections (ex: "Entrée + Plat")
            # au lieu d'utiliser le nom "Menu" du package
            components = pkg.get("section_ids", [])
            display_name = pkg.get("name", "")  # Fallback au nom du package
            if components:
                section_names = []
                for sid in components:
                    for s in sections:
                        if s.get("section_id") == sid:
                            section_names.append(s.get("name", "?"))
                            break
                if section_names:
                    display_name = " + ".join(section_names)
            
            # Nom du forfait et prix sur la même ligne - CENTRÉ et RAPPROCHÉ
            pdf.set_font("Helvetica", "B", 14)  # Augmenté de 12 à 14
            pdf.set_text_color(44, 44, 44)
            
            if pkg_price is not None:
                price_str = f"{float(pkg_price):.2f} EUR"
                full_text = safe_text(f"- {display_name}")  # Use - instead of bullet
                text_width = pdf.get_string_width(full_text)
                price_width = pdf.get_string_width(price_str)
                
                # Calculer la position pour centrer nom + prix ensemble
                total_width = text_width + 15 + price_width
                start_x = center_x - total_width/2
                
                pdf.set_xy(start_x, pdf.get_y())
                pdf.cell(text_width + 15, item_height, full_text, ln=False)
                
                pdf.set_font("Helvetica", "B", 13)
                pdf.set_text_color(39, 150, 80)
                pdf.cell(price_width, item_height, safe_text(price_str), ln=True)
            else:
                pdf.set_xy(margin, pdf.get_y())
                pdf.cell(content_width, item_height, safe_text(f"- {display_name}"), align="C")
                pdf.ln()
            
            pdf.ln(1)  # Réduit de 2 à 1 pour moins d'espace entre formules
    
    # ============ PIED DE PAGE ============
    footer_y = page_height - margin - 22  # Remonté pour faire de la place
    
    # Ligne décorative - bleu RAL 5008
    pdf.set_draw_color(38, 55, 74)
    pdf.set_line_width(0.3)
    pdf.line(margin + 30, footer_y, page_width - margin - 30, footer_y)
    
    # Petit losange central - bleu RAL 5008
    pdf.set_fill_color(38, 55, 74)
    pdf.polygon([(center_x, footer_y - 2), (center_x + 2, footer_y), 
                 (center_x, footer_y + 2), (center_x - 2, footer_y)], 'F')
    
    # Informations de contact
    if restaurant:
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(38, 55, 74)  # Bleu RAL 5008
        
        # Adresse
        address = restaurant.get("address", "")
        if not address:
            street = restaurant.get("address_street", "")
            city = restaurant.get("address_city", "")
            if street or city:
                address = f"{street}, {city}".strip(", ")
        
        if address:
            pdf.set_xy(margin, footer_y + 4)
            pdf.cell(content_width, 4, safe_text(address), align="C")
        
        # Email et téléphone
        email = restaurant.get("email", "")
        phone = restaurant.get("phone", "")
        contact_line = ""
        if email:
            contact_line += email
        if phone:
            if contact_line:
                contact_line += "  |  "
            contact_line += phone
        
        if contact_line:
            pdf.set_xy(margin, footer_y + 9)
            pdf.cell(content_width, 4, safe_text(contact_line), align="C")
    
    # Generate PDF bytes
    pdf_content = pdf.output(dest='S')
    if isinstance(pdf_content, str):
        pdf_content = pdf_content.encode('latin-1')
    elif isinstance(pdf_content, bytearray):
        pdf_content = bytes(pdf_content)
    
    # Create filename
    event_title = event.get("title", "menu").replace(" ", "_")
    filename = f"menu_{event_title}_{event_date}.pdf"
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ==================== ORDER TICKET ENDPOINT ====================

class OrderTicketItem(BaseModel):
    name: str
    quantity: int
    price: float
    format_name: Optional[str] = None
    composition: Optional[List[dict]] = None

class OrderTicketRequest(BaseModel):
    restaurant_id: str
    items: List[OrderTicketItem]
    total: float

@api_router.post("/public/order-ticket/pdf")
async def generate_order_ticket_pdf(request: OrderTicketRequest):
    """Generate a PDF ticket for an order - public endpoint for QR code scanning"""
    try:
        # Get restaurant info
        restaurant = await restaurants_collection.find_one(
            {"restaurant_id": request.restaurant_id},
            {"_id": 0}
        )
        
        # Create PDF - small receipt format (80mm width typical for thermal printers)
        pdf = FPDF(orientation='P', unit='mm', format=(80, 200))
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=5)
        
        # Set font
        pdf.set_font("Helvetica", "", 10)
        
        # Header
        pdf.set_font("Helvetica", "B", 14)
        restaurant_name = restaurant.get("name", "Restaurant") if restaurant else "Restaurant"
        pdf.cell(70, 8, safe_text(restaurant_name), align="C", ln=True)
        
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(70, 5, "TICKET DE COMMANDE", align="C", ln=True)
        pdf.cell(70, 5, datetime.now().strftime("%d/%m/%Y %H:%M"), align="C", ln=True)
        
        # Separator line
        pdf.ln(2)
        pdf.set_draw_color(0, 0, 0)
        pdf.line(5, pdf.get_y(), 75, pdf.get_y())
        pdf.ln(3)
        
        # Items
        pdf.set_font("Helvetica", "", 9)
        for item in request.items:
            # Item name and quantity
            item_text = f"{item.quantity}x {item.name}"
            if item.format_name:
                item_text += f" ({item.format_name})"
            
            # Split long names
            pdf.set_font("Helvetica", "B", 9)
            pdf.multi_cell(55, 4, safe_text(item_text), align="L")
            
            # Price on the right
            price_text = f"{(item.price * item.quantity):.2f}E"
            pdf.set_xy(55, pdf.get_y() - 4)
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(20, 4, price_text, align="R", ln=True)
            
            # Composition details if any
            if item.composition:
                pdf.set_font("Helvetica", "", 7)
                for comp in item.composition:
                    cat = comp.get("category", "")
                    name = comp.get("item_name", "")
                    pdf.cell(70, 3, safe_text(f"  - {cat}: {name}"), ln=True)
            
            pdf.ln(1)
        
        # Separator line
        pdf.ln(2)
        pdf.line(5, pdf.get_y(), 75, pdf.get_y())
        pdf.ln(3)
        
        # Total
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(50, 6, "TOTAL", align="L")
        pdf.cell(20, 6, f"{request.total:.2f}E", align="R", ln=True)
        
        # Footer
        pdf.ln(5)
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(70, 4, "Merci de votre visite!", align="C", ln=True)
        
        # Generate PDF
        pdf_content = pdf.output()
        
        return Response(
            content=bytes(pdf_content),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'inline; filename="ticket_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"',
                "Cache-Control": "no-cache"
            }
        )
        
    except Exception as e:
        logging.error(f"Error generating order ticket PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/public/order-ticket/view")
async def view_order_ticket(data: str):
    """View order ticket from encoded data - for QR code scanning"""
    try:
        import json
        import urllib.parse
        
        # Decode the data
        decoded = urllib.parse.unquote(data)
        order_data = json.loads(decoded)
        
        # Generate HTML receipt
        restaurant_name = order_data.get("restaurant", "Restaurant")
        items = order_data.get("items", [])
        total = order_data.get("total", 0)
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Ticket - {restaurant_name}</title>
            <style>
                * {{ box-sizing: border-box; margin: 0; padding: 0; }}
                body {{
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    background: linear-gradient(135deg, #f5f0e8 0%, #e8e0d5 100%);
                    min-height: 100vh;
                    padding: 20px;
                }}
                .ticket {{
                    max-width: 380px;
                    margin: 0 auto;
                    background: #fff;
                    border-radius: 16px;
                    box-shadow: 0 10px 40px rgba(0,0,0,0.15);
                    overflow: hidden;
                }}
                .header {{
                    background: linear-gradient(135deg, #2a3f54 0%, #1a2a3a 100%);
                    color: #fff;
                    padding: 24px 20px;
                    text-align: center;
                }}
                .header h1 {{
                    font-size: 24px;
                    font-weight: 700;
                    margin-bottom: 4px;
                }}
                .header .subtitle {{
                    font-size: 12px;
                    opacity: 0.8;
                    text-transform: uppercase;
                    letter-spacing: 2px;
                }}
                .content {{
                    padding: 20px;
                }}
                .item {{
                    padding: 14px 0;
                    border-bottom: 1px solid #eee;
                }}
                .item:last-child {{
                    border-bottom: none;
                }}
                .item-header {{
                    display: flex;
                    justify-content: space-between;
                    align-items: flex-start;
                    margin-bottom: 6px;
                }}
                .item-name {{
                    font-weight: 600;
                    font-size: 15px;
                    color: #333;
                    flex: 1;
                    padding-right: 10px;
                }}
                .item-price {{
                    font-weight: 700;
                    font-size: 15px;
                    color: #2a3f54;
                    white-space: nowrap;
                }}
                .item-format {{
                    font-size: 12px;
                    color: #888;
                    margin-bottom: 4px;
                }}
                .composition {{
                    font-size: 12px;
                    color: #666;
                    padding-left: 12px;
                    border-left: 2px solid #c9a961;
                    margin-top: 6px;
                }}
                .composition-item {{
                    margin: 3px 0;
                }}
                .total-section {{
                    background: #f8f8f8;
                    padding: 20px;
                    margin-top: 10px;
                }}
                .total {{
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                }}
                .total-label {{
                    font-size: 18px;
                    font-weight: 700;
                    color: #333;
                }}
                .total-amount {{
                    font-size: 24px;
                    font-weight: 700;
                    color: #c9a961;
                }}
                .footer {{
                    text-align: center;
                    padding: 16px 20px;
                    background: #2a3f54;
                    color: #fff;
                    font-size: 12px;
                }}
                .datetime {{
                    text-align: center;
                    font-size: 11px;
                    color: #999;
                    padding: 10px;
                }}
            </style>
        </head>
        <body>
            <div class="ticket">
                <div class="header">
                    <h1>{restaurant_name}</h1>
                    <div class="subtitle">Ticket de Commande</div>
                </div>
                <div class="content">
        """
        
        for item in items:
            name = item.get("name", "")
            qty = item.get("qty", 1)
            price = item.get("price", 0)
            fmt = item.get("format", "")
            composition = item.get("comp", [])
            
            html += f'''
                    <div class="item">
                        <div class="item-header">
                            <span class="item-name">{qty}x {name}</span>
                            <span class="item-price">{price:.2f}€</span>
                        </div>
            '''
            
            if fmt:
                html += f'<div class="item-format">({fmt})</div>'
            
            if composition and len(composition) > 0:
                html += '<div class="composition">'
                for comp in composition:
                    html += f'<div class="composition-item">• {comp}</div>'
                html += '</div>'
            
            html += '</div>'
        
        from datetime import datetime
        now = datetime.now().strftime("%d/%m/%Y à %H:%M")
        
        html += f"""
                </div>
                <div class="total-section">
                    <div class="total">
                        <span class="total-label">TOTAL</span>
                        <span class="total-amount">{total:.2f}€</span>
                    </div>
                </div>
                <div class="datetime">{now}</div>
                <div class="footer">
                    Merci de votre visite !
                </div>
            </div>
        </body>
        </html>
        """
        
        return Response(content=html, media_type="text/html")
        
    except Exception as e:
        logging.error(f"Error viewing order ticket: {e}")
        return Response(content=f"<html><body><h1>Erreur</h1><p>{str(e)}</p></body></html>", media_type="text/html")

# ==================== SIMPLE ORDER STORAGE FOR QR ====================

# Temporary storage for orders (in memory - resets on server restart)
order_storage = {}

@api_router.post("/public/order/create")
async def create_order_for_qr(request: dict):
    """Store order temporarily and return short ID for QR code"""
    import uuid
    import time
    
    # Generate short ID
    order_id = str(uuid.uuid4())[:8]
    
    # Store with timestamp
    order_storage[order_id] = {
        "data": request,
        "created": time.time()
    }
    
    # Clean old orders (older than 24 hours)
    current_time = time.time()
    to_delete = [k for k, v in order_storage.items() if current_time - v["created"] > 86400]
    for k in to_delete:
        del order_storage[k]
    
    return {"order_id": order_id}

@api_router.get("/public/order/{order_id}")
async def get_order_ticket(order_id: str):
    """Display order ticket from stored ID"""
    if order_id not in order_storage:
        return Response(
            content="<html><body style='font-family:sans-serif;text-align:center;padding:50px'><h1>Commande expirée</h1><p>Ce ticket n'est plus disponible.</p></body></html>",
            media_type="text/html"
        )
    
    order_data = order_storage[order_id]["data"]
    restaurant_name = order_data.get("restaurant", "Restaurant")
    items = order_data.get("items", [])
    total = order_data.get("total", 0)
    
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Ticket - {restaurant_name}</title>
    <style>
        *{{box-sizing:border-box;margin:0;padding:0}}
        body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f5f0e8;min-height:100vh;padding:15px}}
        .ticket{{max-width:380px;margin:0 auto;background:#fff;border-radius:16px;box-shadow:0 10px 40px rgba(0,0,0,0.15);overflow:hidden}}
        .header{{background:linear-gradient(135deg,#2a3f54,#1a2a3a);color:#fff;padding:20px;text-align:center}}
        .header h1{{font-size:22px;font-weight:700}}
        .header .sub{{font-size:11px;opacity:0.8;text-transform:uppercase;letter-spacing:2px;margin-top:5px}}
        .content{{padding:15px}}
        .item{{padding:12px 0;border-bottom:1px solid #eee}}
        .item:last-child{{border-bottom:none}}
        .item-row{{display:flex;justify-content:space-between;align-items:flex-start}}
        .item-name{{font-weight:600;font-size:14px;color:#333;flex:1;padding-right:10px}}
        .item-price{{font-weight:700;font-size:14px;color:#2a3f54}}
        .item-fmt{{font-size:11px;color:#888}}
        .comp{{font-size:11px;color:#666;padding-left:10px;border-left:2px solid #c9a961;margin-top:5px}}
        .total-box{{background:#f8f8f8;padding:15px;margin-top:10px}}
        .total{{display:flex;justify-content:space-between;align-items:center}}
        .total-lbl{{font-size:16px;font-weight:700;color:#333}}
        .total-amt{{font-size:22px;font-weight:700;color:#c9a961}}
        .footer{{text-align:center;padding:12px;background:#2a3f54;color:#fff;font-size:11px}}
    </style>
</head>
<body>
    <div class="ticket">
        <div class="header">
            <h1>{restaurant_name}</h1>
            <div class="sub">Ticket de Commande</div>
        </div>
        <div class="content">"""
    
    for item in items:
        name = item.get("name", "")
        qty = item.get("qty", 1)
        price = item.get("price", 0)
        fmt = item.get("format", "")
        comp = item.get("comp", [])
        
        html += f'<div class="item"><div class="item-row"><span class="item-name">{qty}x {name}</span><span class="item-price">{price:.2f}€</span></div>'
        if fmt:
            html += f'<div class="item-fmt">({fmt})</div>'
        if comp:
            html += '<div class="comp">'
            for c in comp:
                html += f'• {c}<br>'
            html += '</div>'
        html += '</div>'
    
    html += f"""</div>
        <div class="total-box">
            <div class="total">
                <span class="total-lbl">TOTAL</span>
                <span class="total-amt">{total:.2f}€</span>
            </div>
        </div>
        <div class="footer">Merci de votre visite !</div>
    </div>
</body>
</html>"""
    
    return Response(content=html, media_type="text/html")

# ==================== ROOT ENDPOINT ====================

from fastapi.responses import RedirectResponse

# Serve PWA at root "/" 
@app.get("/")
async def serve_root():
    """Serve the PWA index.html"""
    # Chercher index.html local
    possible_paths = [
        Path(__file__).parent / "dist" / "index.html",
        Path("/app/backend/dist/index.html"),
    ]
    
    for index_path in possible_paths:
        if index_path.exists():
            return FileResponse(str(index_path), media_type='text/html')
    
    # Fallback si pas de fichiers - page d'erreur simple
    return HTMLResponse(content='''<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="utf-8"/>
    <meta name="viewport" content="width=device-width, initial-scale=1"/>
    <title>NeoChef</title>
    <style>
        body { 
            font-family: -apple-system, sans-serif;
            background: #1a1a2e; color: white;
            display: flex; align-items: center; justify-content: center;
            min-height: 100vh; margin: 0;
        }
        .container { text-align: center; }
        h1 { color: #d4af37; }
    </style>
</head>
<body>
    <div class="container">
        <h1>NeoChef</h1>
        <p>Application en cours de chargement...</p>
        <p>Si ce message persiste, veuillez rafraîchir la page.</p>
    </div>
</body>
</html>''', status_code=200)

# Health check at "/health" for Kubernetes liveness/readiness probes
@app.get("/health")
async def kubernetes_health_check():
    """Health check endpoint for Kubernetes probes"""
    return {"status": "ok", "app": "RestoPilot", "version": "2.1"}

@app.get("/api/")
async def api_health_check():
    """API health check with database connectivity test"""
    try:
        # Test MongoDB connection
        await client.admin.command('ping')
        db_status = "connected"
    except Exception as e:
        db_status = f"error: {str(e)}"
        logging.error(f"Health check - MongoDB error: {e}")
    
    return {
        "status": "ok" if db_status == "connected" else "degraded",
        "app": "RestoPilot",
        "version": "2.1",
        "database": db_status
    }

@app.get("/api/app")
async def serve_app():
    """Serve the main PWA application index.html with /api/ paths"""
    # Utiliser la version avec les chemins /api/
    api_index_path = Path(__file__).parent / "dist" / "index_api.html"
    if api_index_path.exists():
        return FileResponse(str(api_index_path), media_type='text/html')
    
    # Fallback sur l'index normal
    possible_paths = [
        Path(__file__).parent / "dist" / "index.html",
        Path("/app/backend/dist/index.html"),
    ]
    
    for index_path in possible_paths:
        if index_path.exists():
            return FileResponse(str(index_path), media_type='text/html')
    
    return HTMLResponse(content='<h1>NeoChef - Application non trouvée</h1>', status_code=404)

# Routes pour servir les fichiers PWA via /api/
@app.get("/api/manifest.json")
async def get_api_manifest():
    """Serve PWA manifest via /api/"""
    file_path = DIST_DIR / "manifest.json"
    if file_path and file_path.exists():
        return FileResponse(str(file_path), media_type="application/manifest+json")
    raise HTTPException(status_code=404, detail="Manifest not found")

@app.get("/api/sw.js")
async def get_api_sw():
    """Serve Service Worker via /api/"""
    file_path = DIST_DIR / "sw.js"
    if file_path and file_path.exists():
        return FileResponse(str(file_path), media_type="application/javascript")
    raise HTTPException(status_code=404, detail="Service Worker not found")

@app.get("/api/apple-touch-icon.png")
async def get_api_apple_icon():
    """Serve Apple Touch Icon via /api/"""
    file_path = DIST_DIR / "apple-touch-icon.png"
    if file_path and file_path.exists():
        return FileResponse(str(file_path), media_type="image/png")
    raise HTTPException(status_code=404, detail="Icon not found")

@app.get("/api/apple-touch-icon-180x180.png")
async def get_api_apple_icon_180():
    """Serve Apple Touch Icon 180x180 via /api/"""
    file_path = DIST_DIR / "apple-touch-icon-180x180.png"
    if file_path and file_path.exists():
        return FileResponse(str(file_path), media_type="image/png")
    raise HTTPException(status_code=404, detail="Icon not found")

@app.get("/api/favicon.png")
async def get_api_favicon():
    """Serve Favicon via /api/"""
    file_path = DIST_DIR / "favicon.png"
    if file_path and file_path.exists():
        return FileResponse(str(file_path), media_type="image/png")
    raise HTTPException(status_code=404, detail="Favicon not found")

@app.get("/api/health")
async def detailed_health_check():
    """Detailed health check for monitoring"""
    health = {
        "status": "ok",
        "app": "RestoPilot",
        "version": "2.1",
        "checks": {}
    }
    
    # Check MongoDB
    try:
        await client.admin.command('ping')
        health["checks"]["mongodb"] = "ok"
    except Exception as e:
        health["checks"]["mongodb"] = f"error: {str(e)}"
        health["status"] = "degraded"
    
    return health

# Special route to serve homepage (bypass CDN cache)
@app.get("/app")
async def serve_app_homepage():
    """Serve the main application homepage"""
    from fastapi.responses import FileResponse
    index_file = DIST_DIR / "index.html"
    if index_file.exists():
        return FileResponse(
            str(index_file), 
            media_type='text/html',
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0"
            }
        )
    raise HTTPException(status_code=404, detail="App not found")

# ==================== APP SETUP ====================

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== PWA STATIC FILES ====================
PWA_DIR = Path("/app/frontend/public")

# Try multiple locations for dist folder (local dev vs deployment)
DIST_DIR = None
dist_paths = [
    Path(__file__).parent / "dist",  # /app/backend/dist (deployment)
    Path("/app/frontend/dist"),       # /app/frontend/dist (local dev)
    Path("/app/backend/dist"),        # explicit backend path
]
for dist_path in dist_paths:
    if dist_path.exists():
        DIST_DIR = dist_path
        break

if DIST_DIR is None:
    DIST_DIR = Path("/app/backend/dist")  # fallback

# Mount static files from Expo web build
if DIST_DIR.exists():
    # Mount _fresh static assets (new cache-busting path)
    fresh_static = DIST_DIR / "_fresh"
    if fresh_static.exists():
        app.mount("/_fresh", StaticFiles(directory=str(fresh_static)), name="fresh_static")
    
    # Mount _expo static assets first (JS, CSS, etc.)
    expo_static = DIST_DIR / "_expo"
    if expo_static.exists():
        app.mount("/_expo", StaticFiles(directory=str(expo_static)), name="expo_static")
        # Also mount via /api/ path for production deployment
        app.mount("/api/_expo", StaticFiles(directory=str(expo_static)), name="api_expo_static")
    
    # Mount assets folder
    assets_dir = DIST_DIR / "assets"
    if assets_dir.exists():
        app.mount("/assets", StaticFiles(directory=str(assets_dir)), name="assets")
        # Also mount via /api/ path for production deployment
        app.mount("/api/assets", StaticFiles(directory=str(assets_dir)), name="api_assets")
    
    # Mount ardoise folder
    ardoise_dir = DIST_DIR / "ardoise"
    if ardoise_dir.exists():
        app.mount("/api/ardoise", StaticFiles(directory=str(ardoise_dir)), name="api_ardoise")
    
    # Mount client folder
    client_dir = DIST_DIR / "client"
    if client_dir.exists():
        app.mount("/api/client", StaticFiles(directory=str(client_dir)), name="api_client")

@app.get("/manifest.json")
async def get_manifest():
    """Serve PWA manifest"""
    file_path = PWA_DIR / "manifest.json"
    if file_path.exists():
        return FileResponse(file_path, media_type="application/manifest+json")
    raise HTTPException(status_code=404, detail="Manifest not found")

@app.get("/apple-touch-icon.png")
async def get_apple_touch_icon():
    """Serve Apple touch icon"""
    file_path = PWA_DIR / "apple-touch-icon.png"
    if file_path.exists():
        return FileResponse(file_path, media_type="image/png")
    raise HTTPException(status_code=404, detail="Icon not found")

@app.get("/apple-touch-icon-180x180.png")
async def get_apple_touch_icon_180():
    """Serve Apple touch icon 180x180"""
    file_path = PWA_DIR / "apple-touch-icon-180x180.png"
    if file_path.exists():
        return FileResponse(file_path, media_type="image/png")
    raise HTTPException(status_code=404, detail="Icon not found")

@app.get("/favicon.png")
async def get_favicon():
    """Serve favicon"""
    file_path = PWA_DIR / "favicon.png"
    if file_path.exists():
        return FileResponse(file_path, media_type="image/png")
    raise HTTPException(status_code=404, detail="Favicon not found")

@app.get("/icon-192.png")
async def get_icon_192():
    """Serve 192x192 icon"""
    file_path = PWA_DIR / "icon-192.png"
    if file_path.exists():
        return FileResponse(file_path, media_type="image/png")
    raise HTTPException(status_code=404, detail="Icon not found")

@app.get("/icon-512.png")
async def get_icon_512():
    """Serve 512x512 icon"""
    file_path = PWA_DIR / "icon-512.png"
    if file_path.exists():
        return FileResponse(file_path, media_type="image/png")
    raise HTTPException(status_code=404, detail="Icon not found")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# ==================== KEEP PREVIEW ALIVE TASK ====================
# Cette tâche ping le preview URL toutes les 5 minutes pour éviter le sleep
import httpx

_keep_alive_task = None

async def keep_preview_alive():
    """Background task to keep the app awake (self-ping)"""
    # Use environment variable for the URL
    app_url = os.environ.get('APP_URL', os.environ.get('FRONTEND_URL', 'http://localhost:8001'))
    health_url = f"{app_url}/api/health"
    while True:
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.get(health_url)
                logging.info(f"[KEEP-ALIVE] Self ping: {response.status_code}")
        except Exception as e:
            logging.warning(f"[KEEP-ALIVE] Self ping failed: {e}")
        
        # Wait 5 minutes before next ping
        await asyncio.sleep(5 * 60)

@app.on_event("startup")
async def start_keep_alive_task():
    """Start the keep-alive background task"""
    global _keep_alive_task
    _keep_alive_task = asyncio.create_task(keep_preview_alive())
    logging.info("[KEEP-ALIVE] Background task started - app will stay awake")

@app.on_event("shutdown")
async def stop_keep_alive_task():
    """Stop the keep-alive task on shutdown"""
    global _keep_alive_task
    if _keep_alive_task:
        _keep_alive_task.cancel()
        try:
            await _keep_alive_task
        except asyncio.CancelledError:
            pass
        logging.info("[KEEP-ALIVE] Background task stopped")

# ==================== FIX FOR DOUBLE API PREFIX BUG ====================
# Handle /api/api/... by internally rewriting to /api/... (bug from old frontend code)
@app.middleware("http")
async def fix_double_api_prefix(request: Request, call_next):
    """Fix double /api/api/ prefix from cached frontend"""
    path = request.url.path
    if path.startswith("/api/api/"):
        # Rewrite the path to remove the extra /api
        new_path = path.replace("/api/api/", "/api/", 1)
        # Create a new scope with the corrected path
        scope = request.scope.copy()
        scope["path"] = new_path
        # Create a new request with the corrected scope
        from starlette.requests import Request as StarletteRequest
        request = StarletteRequest(scope, request.receive)
    return await call_next(request)

# ==================== CATCH-ALL ROUTE FOR SPA ====================
# This must be at the very end, after all other routes
@app.get("/{full_path:path}")
async def serve_spa(full_path: str):
    """Serve the Expo web build for all non-API routes (SPA fallback)"""
    # Don't serve SPA for API routes
    if full_path.startswith("api/"):
        raise HTTPException(status_code=404, detail="API route not found")
    
    # Check if requesting a static file from dist
    dist_file = DIST_DIR / full_path
    if dist_file.exists() and dist_file.is_file():
        # Determine content type
        suffix = dist_file.suffix.lower()
        content_types = {
            '.html': 'text/html',
            '.js': 'application/javascript',
            '.css': 'text/css',
            '.json': 'application/json',
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.svg': 'image/svg+xml',
            '.ico': 'image/x-icon',
            '.woff': 'font/woff',
            '.woff2': 'font/woff2',
            '.ttf': 'font/ttf',
        }
        media_type = content_types.get(suffix, 'application/octet-stream')
        
        # Add cache headers based on file type - FORCE NO CACHE for all
        headers = {
            "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
            "Surrogate-Control": "no-store"
        }
        
        return FileResponse(str(dist_file), media_type=media_type, headers=headers)
    
    # For gestion-ardoise pages (new path to bypass cache)
    if full_path.startswith("gestion-ardoise/"):
        ardoise_html = DIST_DIR / "gestion-ardoise" / "[token].html"
        if ardoise_html.exists():
            headers = {"Cache-Control": "no-cache, no-store, must-revalidate, max-age=0", "Pragma": "no-cache", "Expires": "0", "Surrogate-Control": "no-store"}
            return FileResponse(str(ardoise_html), media_type='text/html', headers=headers)
    
    # For ardoise pages, serve the gestion.html (nouvelle page de gestion complète)
    if full_path.startswith("ardoise/"):
        gestion_html = DIST_DIR / "gestion.html"
        if gestion_html.exists():
            headers = {"Cache-Control": "no-cache, no-store, must-revalidate, max-age=0", "Pragma": "no-cache", "Expires": "0", "Surrogate-Control": "no-store"}
            return FileResponse(str(gestion_html), media_type='text/html', headers=headers)
    
    # For client pages, serve the [restaurant_id].html template
    if full_path.startswith("client/"):
        client_html = DIST_DIR / "client" / "[restaurant_id].html"
        if client_html.exists():
            headers = {"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"}
            return FileResponse(str(client_html), media_type='text/html', headers=headers)
    
    # For all other routes, serve index.html (SPA routing)
    index_file = DIST_DIR / "index.html"
    if index_file.exists():
        headers = {"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"}
        return FileResponse(str(index_file), media_type='text/html', headers=headers)
    
    raise HTTPException(status_code=404, detail="Not found")

